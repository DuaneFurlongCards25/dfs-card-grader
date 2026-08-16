import streamlit as st
import pandas as pd
import json
import urllib.request
import urllib.parse
import ssl
import random
import string
from datetime import date, datetime
from pathlib import Path
import io
import base64
import csv
import re
import collections
from concurrent.futures import ThreadPoolExecutor, as_completed

# ─── Constants ────────────────────────────────────────────────────────────────
APP_VERSION = "1.5.98"

# Product branding — change APP_NAME on this one line to rebrand the whole app.
APP_NAME = "The CardPulse™"
APP_TAGLINE = "Real-time market intelligence for card sellers."

# Daily cap on live CardHedger look-ups per member (protects the API budget).
DAILY_PRICING_CAP = 50

# eBay File Exchange condition descriptor option IDs for category 261328 (Sports Trading Cards).
# Values MUST be in the format "Display Name - (ID: XXXXXX)" — eBay rejects bare strings.
# Source: Card Dealer Pro export (verified working Aug 2026). Add graders/grades as needed.
EBAY_GRADER_VALUES = {
    "PSA":  "Professional Sports Authenticator (PSA) - (ID: 275010)",
    "BGS":  "Beckett Grading Services (BGS) - (ID: 275011)",
    "SGC":  "SGC - (ID: 275012)",
    "CGC":  "CGC Cards - (ID: 275013)",
    "CSG":  "CSG - (ID: 275014)",
    "HGA":  "Hybrid Grading Approach (HGA) - (ID: 275015)",
}
# Condition descriptor 40001 (ungraded card condition) option IDs for category 261328.
# If upload errors with 21920352 (invalid) or 21920355 (required), call eBay
# GetItemConditionDescriptors API for the real IDs and update these values.
EBAY_CONDITION_DEFAULT = "400011"   # Excellent — confirmed valid from eBay template Aug 2026
EBAY_CONDITION_VALUES = {
    "near mint":   "400010",  # Near mint or better
    "excellent":   "400011",  # Excellent
    "very good":   "400012",  # Very good
    "poor":        "400013",  # Poor
}

# Grade option IDs — 10 confirmed from Card Dealer Pro; others follow sequential pattern.
# If a grade upload errors with 21920352, find its ID in eBay's GetItemConditionDescriptors.
EBAY_GRADE_VALUES = {
    "10":   "10 - (ID: 275020)",
    "9.5":  "9.5 - (ID: 275021)",
    "9":    "9 - (ID: 275022)",
    "8.5":  "8.5 - (ID: 275023)",
    "8":    "8 - (ID: 275024)",
    "7.5":  "7.5 - (ID: 275025)",
    "7":    "7 - (ID: 275026)",
    "6":    "6 - (ID: 275027)",
    "5":    "5 - (ID: 275028)",
    "4":    "4 - (ID: 275029)",
    "3":    "3 - (ID: 275030)",
    "2":    "2 - (ID: 275031)",
    "1.5":  "1.5 - (ID: 275032)",
    "1":    "1 - (ID: 275033)",
    "A":    "Authentic - (ID: 275034)",
}

RELEASE_NOTES = {
    "1.5.30": {
        "emoji": "🤖",
        "title": "AI Batch Scanner — scan multiple cards at once",
        "items": [
            ("🤖", "New 🤖 AI Batch tab: upload 1–100+ card photos, Claude Vision reads every card in sequence with a progress bar."),
            ("📊", "eBay comps fetched automatically after each identification — shows avg sold price per card."),
            ("📥", "Export all results as CSV (player, year, set, card #, parallel, sport, eBay avg, search query)."),
            ("💰", "Cost: ~1¢ per card (Claude Sonnet 5 vision). Zero tokens wasted — processes only the cards you upload."),
        ],
    },
    "1.5.28": {
        "emoji": "📝",
        "title": "Manual sale logger on outstanding cards",
        "items": [
            ("📝", "Each lot with outstanding cards now has a '📝 Log a Sale' expander — pick the card, enter channel/date/net, save."),
            ("✅", "Card flips from 🟡 Outstanding to ✅ Sold instantly after logging."),
            ("🃏", "Works for CollX, eBay, DC Sports, Whatnot, Facebook, or any other channel."),
        ],
    },
    "1.5.27": {
        "emoji": "🃏",
        "title": "CollX import: only completed orders",
        "items": [
            ("🃏", "CollX import now skips non-completed orders — only 'completed' status rows are saved."),
        ],
    },
    "1.5.26": {
        "emoji": "📦",
        "title": "Persistent lot inventory — cross-check sold/outstanding per card",
        "items": [
            ("📦", "Import Cards tab now saves cards to Supabase — inventory persists across sessions."),
            ("✅", "Each lot expander shows a full cross-check table: every card with Sold / Outstanding status, channel, sale date, and net."),
            ("🔄", "Re-uploading a CSV updates existing cards and adds new ones (safe re-import, no duplicates)."),
            ("🟡", "Outstanding cards show which channel they may still be listed on."),
        ],
    },
    "1.5.25": {
        "emoji": "🃏",
        "title": "Individual Card Purchases — Whatnot, FB, small buys",
        "items": [
            ("🃏", "New 🃏 Individual Cards tab under Purchases — track small buys without creating a lot."),
            ("🔗", "SKU auto-links to sales_records when card sells — see cost vs net vs P&L per card."),
            ("📊", "Summary: Total Invested, Net Recovered, Overall P&L, Cards Sold, Win Rate."),
            ("🔍", "Filter by source (Whatnot, Facebook, etc.), status (Sold/Unsold), and result (Profit/Loss)."),
            ("⏱️", "Days Held and Sold Via channel tracked per card."),
        ],
    },
    "1.5.24": {
        "emoji": "📈",
        "title": "Lots: full ROI, turn rate, avg days to sell",
        "items": [
            ("📈", "Each lot now shows ROI %, Projected ROI, $/Card Cost, $/Card Net — see if the purchase was worth it."),
            ("⏱️", "Turn Rate (cards/week) and Avg Days to Sell — how fast the lot is moving."),
            ("📅", "Projected days to clear remaining cards at current pace."),
        ],
    },
    "1.5.23": {
        "emoji": "✅",
        "title": "Lots: sold cards detail table inside each lot",
        "items": [
            ("✅", "Each lot expander now shows a Sold Cards table — SKU, title, date, gross, net for every card sold from that lot."),
            ("📋", "Active listings section labeled separately so sold vs still-listed is easy to compare."),
        ],
    },
    "1.5.22": {
        "emoji": "🔑",
        "title": "Login: block browser autofill, keep access code",
        "items": [
            ("🔑", "JS injection blocks browser email autofill on the login screen — access code stays put."),
        ],
    },
    "1.5.21": {
        "emoji": "📊",
        "title": "Lots: live sold/left/revenue tally on every lot",
        "items": [
            ("📊", "Each lot now shows sold count, cards left, net revenue, and P&L right in the header — no need to go to P&L sub-tab."),
            ("🟢", "Metrics row inside each lot: Cost · Cards · Sold · Left · Net Rev with P&L delta."),
        ],
    },
    "1.5.20": {
        "emoji": "🔑",
        "title": "Login remembers your access code",
        "items": [
            ("🔑", "Access code is remembered in your browser — no more retyping it every session."),
            ("🏷️", "eBay Orders Report SKU/lot tagging fix from v1.5.19 also included."),
        ],
    },
    "1.5.19": {
        "emoji": "🏷️",
        "title": "eBay Orders Report: fix SKU/lot tagging",
        "items": [
            ("🏷️", "Orders Report now correctly reads Custom Label (SKU) → sales link to lots in P&L."),
            ("🔍", "Fixed detection: skips blank preamble row before headers so Orders format is auto-identified correctly."),
            ("🔁", "Dedup key updated to ebay_orders|order|item — safe to re-import without duplicates."),
        ],
    },
    "1.5.18": {
        "emoji": "📦",
        "title": "eBay Orders Report import (Paid & Shipped)",
        "items": [
            ("📦", "Sales & P&L now accepts eBay Orders Report CSV — Seller Hub → Orders → Sold → Download. Real fees, no estimates."),
            ("💰", "Actual Final Value Fee (fixed + variable) pulled directly from the report — more accurate than the 12.35% estimate."),
            ("🔁", "Auto-detected alongside existing Transactions and Financial Ledger formats — just drop it in."),
        ],
    },
    "1.5.17": {
        "emoji": "💎",
        "title": "Gem rate slider feeds GO/NO-GO + Grade vs Flip",
        "items": [
            ("💎", "Card Research now shows a gem rate slider right on the card stats row — set your estimate, it feeds the GO/NO-GO verdict immediately."),
            ("✅", "GO/NO-GO now checks gem rate against your min threshold (sidebar) — not just price target alone."),
            ("🔗", "PSA Pop Report link next to the slider so you can look up the real gem rate in one click."),
            ("⚖️", "Grade vs Flip uses the same gem rate estimate — no more duplicate slider."),
            ("⚡", "Search now goes to CardHedger first (fast) — no more waiting for GemRate timeout on every search."),
        ],
    },
    "1.5.16": {
        "emoji": "📊",
        "title": "Lot P&L: inventory turn, gross revenue, per-card stats, projected P&L",
        "items": [
            ("📊", "P&L by Lot now shows: Gross Revenue, Net Revenue, P&L, Inventory Turn %, $/Card Cost, $/Card Net, Projected Net Revenue, Projected P&L."),
            ("🔄", "Inventory Turn % = cards sold ÷ total cards in lot — see which lots are moving fast vs. sitting."),
            ("📈", "Projected P&L extrapolates from your current avg net/card × remaining cards — rough estimate of where the lot ends up."),
            ("📋", "Header row now shows 5 totals: Total Invested · Gross Revenue · Net Revenue · Overall P&L · Overall Turn %."),
        ],
    },
    "1.5.15": {
        "emoji": "🔍",
        "title": "Scan: Title Search — type a card name, get FMV + trend instantly",
        "items": [
            ("🔍", "New 🔍 Title Search tab in Scan — type any card name (year, set, parallel) and get live FMV, price range, 7/30-day sales volume, and trend signal. No image needed."),
            ("📈", "Trend displayed prominently: 🔥 UP / 🛑 DOWN / ➡️ FLAT with % change over 90 days."),
            ("📋", "Recent sales table expandable below the metrics."),
        ],
    },
    "1.5.14": {
        "emoji": "🌾",
        "title": "Sunday Reprice: Haystack structured queries + parallel matching",
        "items": [
            ("🌾", "Optional Haystack CSV upload in Sunday Reprice — matches on SKU, builds query from Player/Athlete + Season + Manufacturer + Card Number + Parallel/Variety for a much cleaner CardHedger lookup than the truncated eBay title."),
            ("🔴", "Parallel variants (Red, Blue Wave, Orange, etc.) now included in the query — so CAM SCHLITTLER Red pulls Red comps, not base."),
            ("📊", "Table shows 🌾 vs eBay source per row so you can see which listings got Haystack data vs fell back to the eBay title."),
        ],
    },
    "1.5.13": {
        "emoji": "🎯",
        "title": "Sunday Reprice: cleaner title matching + bad-match guard",
        "items": [
            ("🎯", "Title cleaner strips eBay noise (HTA CHOICE, trailing 'R' condition, 2025-26 season spans) before sending to CardHedger — improves match rate on ungraded cards."),
            ("🛡️", "Bad-match guard: if suggested price is <20% or >4× the current price it's flagged ⚠️ and excluded from the eBay upload — no more -94% crash suggestions."),
            ("📊", "Cached summary now shows: X ready to upload · X no match · X suspect (excluded)."),
        ],
    },
    "1.5.12": {
        "emoji": "📅",
        "title": "Sunday Reprice tab — bulk CardHedger FMV for 7–30 day listings",
        "items": [
            ("📅", "New 📅 Sunday Reprice tab in Operations — upload eBay active listings CSV, filter to any age range (default 7–30 days), run CardHedger FMV on every card with a progress bar."),
            ("💾", "Results cached in session — re-running the page after pricing burns zero extra API calls."),
            ("📤", "Downloads eBay Seller Hub bulk-edit CSV (Item number + Start price) and a full reference sheet with comp, trend, and ∆ vs current price."),
            ("⚾", "Cards auto-prioritized: Baseball rookies → Baseball → Soccer → Other, same order as the main Reprice Queue."),
        ],
    },
    "1.5.3": {
        "emoji": "🔍",
        "title": "Show real error message when imports fail",
        "items": [
            ("🔍", "Import failures now show the actual Supabase error (HTTP status + message) so the root cause is visible instead of just a count of failed rows."),
        ],
    },
    "1.5.2": {
        "emoji": "🐛",
        "title": "Fix eBay import crash on blank Quantity column",
        "items": [
            ("🐛", "Fixed: eBay CSV import crashed with 'cannot convert float NaN to integer' when Quantity column was blank. Rows with no quantity now default to 1."),
        ],
    },
    "1.5.1": {
        "emoji": "📊",
        "title": "Full multi-channel import: Whatnot, DC Sports & manual sales",
        "items": [
            ("🎥", "Whatnot importer — upload WhatNot_Seller_Earnings XLSX. Net earnings are already after their fee, imported directly."),
            ("🏷️", "DC Sports importer — upload their CSV export directly to Sales & P&L. Only Paid rows imported; fees and net broken out per card."),
            ("✏️", "Manual sale entry — log social media, show sales, or any off-platform sale with a quick form. Platform, date, gross, fee, net."),
            ("🔒", "All importers deduplicate on re-import — safe to upload the same file multiple times."),
        ],
    },
    "1.5.0": {
        "emoji": "💰",
        "title": "Sales & P&L + DC Sports Consignment tracking",
        "items": [
            ("💰", "New 💰 Sales & P&L tab — import eBay and CollX sales exports. P&L by month, channel mix (eBay vs CollX vs DC Sports), and full sales log."),
            ("🏷️", "New 🏷️ Consignments tab — track DC Sports auction consignments. Import send history and settled results. Lot P&L lives in the Purchases tab."),
            ("📦", "7 DC Sports shipments ready to import — 151 cards, $3,479 net across batches 221184–245869."),
            ("📊", "1,536 eBay sales (Jul 2025–Jul 2026) ready to import into Sales & P&L — $24k gross revenue."),
        ],
    },
    "1.4.0": {
        "emoji": "🚀",
        "title": "CardHedger, maximized — Hot Movers, Scan, Player Demand",
        "items": [
            ("🔥", "New 🔥 Hot Movers tab — the week's biggest price gainers by sport. A buy-radar for what's heating up."),
            ("📷", "New 📷 Scan tab — snap a raw card to ID + price it (AI image match), or look up a graded slab by its cert number."),
            ("📊", "Player Demand on Card Research — weekly sold-volume trend across ALL of a player's cards (rising = growing demand)."),
            ("⚡", "Behind the scenes: Reprice/Operations now use one all-grades price call instead of many — fewer API calls, faster."),
        ],
    },
    "1.3.5": {
        "emoji": "📰",
        "title": "Player Watch — live news & injury alerts (free)",
        "items": [
            ("🔴", "Injury status on the card's player — pulls the live ESPN injury report (MLB/NBA/NFL/NHL) with the actual injury + comment."),
            ("📰", "Recent ESPN headlines for that player, so hot news (or bad news) shows up right next to the price — news moves cards before comps do."),
            ("💸", "Free — uses ESPN's public feeds, no extra subscription on top of CardHedger."),
            ("🌱", "Note: deep prospects may show 'no ESPN match' — coverage is best for established pros."),
        ],
    },
    "1.3.4": {
        "emoji": "📈",
        "title": "Sold Price Trend chart + sales volume",
        "items": [
            ("📈", "New Sold Price Trend chart on Card Research — pick a grade and view the sold-price line over 7 / 30 / 60 / 90 days."),
            ("🔀", "Δ 7d / 30d / 60d / 90d change tiles side by side — see momentum (short term) vs the longer-term trend at a glance."),
            ("🔁", "Sales-volume tiles (7-day & 30-day counts) show how liquid a card is — how easy it'll be to sell."),
            ("🛒", "Note: 'active for sale' shows n/a — CardHedger tracks sold sales only, not live listings."),
        ],
    },
    "1.3.3": {
        "emoji": "⚖️",
        "title": "FMV + Grade vs Flip now work while GemRate is offline",
        "items": [
            ("⚖️", "Grade vs Flip + PASS verdict now appear on the CardHedger screen too (the one you see when GemRate is down) — no longer hidden."),
            ("🎯", "FMV (cleaned value) + confidence grade now show on that screen and pre-fill your cost/sell prices."),
            ("📊", "Set your own gem-rate estimate there to weight the PSA 10 vs PSA 9 outcome while PSA pop data is unavailable."),
        ],
    },
    "1.3.2": {
        "emoji": "❌",
        "title": "Grade vs Flip now warns when to just PASS",
        "items": [
            ("❌", "New PASS verdict — when flipping raw AND grading both lose money at your buy price, the tool tells you to skip the card (or pay less) instead of flipping at a loss."),
            ("🎯", "Grade vs Flip runs on cleaned FMV numbers (not raw averages) for a truer call."),
        ],
    },
    "1.3.1": {
        "emoji": "🎯",
        "title": "Fair Market Value — true numbers, not just averages",
        "items": [
            ("🎯", "New FMV shown beside every comp average (Raw + PSA 10) — a statistically cleaned price that throws out fluke sales the plain average gets fooled by."),
            ("🔤", "Confidence grade (A/B/C) + a low–high price band on each FMV, so you know how much to trust it at a card show."),
            ("✅", "Your cost, sell price, ROI, and Grade-vs-Flip now run on FMV — falls back to the comp average automatically when FMV data is thin."),
        ],
    },
    "1.3.0": {
        "emoji": "🧰",
        "title": "Operations & Inventory — live pricing on your whole list",
        "items": [
            ("🧰", "New Operations tab (full membership) — maintain your inventory, edit costs inline, and pull live market prices across the list."),
            ("✏️", "Editable cost basis + listed price — every margin, reprice, and profit number recalculates the instant you change a cost."),
            ("⚡", "Live pricing is metered — a daily look-up budget keeps pricing fast and protects against runaway API costs. Cached cards re-price free."),
            ("🏷️", "Cleaner brand — the app is now white-labelable; set the name in one place."),
        ],
    },
    "1.2.7": {
        "emoji": "⚖️",
        "title": "Grade vs Flip — the holding-cost decision",
        "items": [
            ("⚖️", "New Grade vs Flip panel in Card Research — see what flipping raw nets you NOW vs grading and waiting ~5 months, side by side."),
            ("🎯", "Probability-weighted: blends the PSA 10 and PSA 9 outcomes by the card's gem rate into one expected-profit number, with the downside spelled out."),
            ("💸", "Holding cost made the headline — exactly how much of your cash gets locked up, for how long, and what that time costs you."),
            ("📐", "Submission Planner in Inventory Check — model a whole batch (e.g. 20 cards): total capital locked, total holding cost, grade-vs-flip totals."),
        ],
    },
    "1.2.6": {
        "emoji": "💰",
        "title": "Reprice Assistant for your inventory",
        "items": [
            ("💰", "New Reprice Assistant in Inventory Check — pulls live sold comps + 90-day trend for every card at once."),
            ("📈", "Trend-following suggested prices: leans the suggestion up when a card is rising, down when it's falling. Strategy and sensitivity are adjustable."),
            ("🚩", "Over/underpriced flags — instantly see which listings are above or below market so you reprice the ones that matter."),
            ("📥", "Download a reprice CSV to update eBay in bulk instead of editing listings one at a time."),
        ],
    },
    "1.2.5": {
        "emoji": "📦",
        "title": "Shipping costs + true total ROI",
        "items": [
            ("📦", "Shipping costs added — enter your true per-card cost (including insurance) in Settings. Baked into every ROI and net profit calculation."),
            ("⏳", "Time cost of capital — see the hidden fee of slow grading tiers. A $500 card at Value tier for 154 days costs you ~$27 you can't deploy elsewhere."),
            ("🔄", "Auto-fallback to CardHedger — when GemRate is offline, live PSA 10 / Raw prices load automatically from CardHedger."),
            ("💎", "Card images now show in search results."),
            ("📋", "PSA fees & turnaround times updated to May 2025 pricing (Super Express $349, all times extended)."),
        ],
    },
}

# PSA current pricing & turnaround (updated May 18, 2026)
# days = midpoint of business-day range; calendar days ≈ days × 1.4
PSA_FEES_ALL = {
    "Value (100-120 days)":      {"fee": 32.99,  "days": 110, "max_insured":  500},
    "Value Plus (60-80 days)":   {"fee": 49.99,  "days":  70, "max_insured":  500},
    "Value Max (40-50 days)":    {"fee": 64.99,  "days":  45, "max_insured": 1000},
    "Regular (30-40 days)":      {"fee": 79.99,  "days":  35, "max_insured": 1500},
    "Express (20-30 days)":      {"fee": 149.00, "days":  25, "max_insured": 2500},
    "Super Express (7-10 days)": {"fee": 349.00, "days":   9, "max_insured": 5000},
    "Walk-Through (5-7 days)":   {"fee": 599.00, "days":   6, "max_insured": 10000},
}
PSA_FEES = {k: v["fee"]  for k, v in PSA_FEES_ALL.items()}
PSA_DAYS = {k: v["days"] for k, v in PSA_FEES_ALL.items()}
EBAY_FEE = 0.1325

# ─── Secrets ──────────────────────────────────────────────────────────────────
def get_secret(section, key, default=""):
    try:
        return st.secrets[section][key]
    except Exception:
        return default

SUPABASE_URL = get_secret("supabase", "url")
SUPABASE_KEY = get_secret("supabase", "key")
DEFAULT_EBAY_KEY = get_secret("ebay", "app_id")
CARDHEDGER_KEY = get_secret("cardhedger", "api_key")
CARDHEDGER_BASE = "https://api.cardhedger.com"
ANTHROPIC_KEY = get_secret("anthropic", "api_key")
IMGBB_KEY = get_secret("imgbb", "api_key") or "e3909e93e7d7962973b65bdf4bf60f52"
WP_PROXY_URL   = "https://duanefurlongstudios.com/wp-admin/admin-ajax.php?action=dfs_gemrate"

# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title=APP_NAME,
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded",
    menu_items={"About": f"{APP_NAME} — research & decision-support tool"},
)

# ─── Mobile CSS ───────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Hide Streamlit toolbar (GitHub, edit, share icons) ── */
[data-testid="stToolbar"] { display: none !important; }
header[data-testid="stHeader"] { background: transparent !important; }

/* ── Force sidebar always visible — overrides browser localStorage collapsed state ── */
section[data-testid="stSidebar"] {
    transform: translateX(0) !important;
    display: block !important;
    visibility: visible !important;
    min-width: 244px !important;
    width: 244px !important;
    left: 0 !important;
    margin-left: 0 !important;
    position: relative !important;
}

/* ── Page padding ── */
.block-container { padding-top: 2.5rem !important; }


@media (max-width: 768px) {
    /* Tighten page padding */
    .block-container {
        padding: 0.75rem 0.75rem 2rem !important;
    }

    /* Stack every column row vertically */
    [data-testid="stHorizontalBlock"] {
        flex-wrap: wrap !important;
        gap: 0.25rem !important;
    }
    [data-testid="column"] {
        width: 100% !important;
        flex: 1 1 100% !important;
        min-width: 100% !important;
    }

    /* Bigger tap targets for buttons */
    .stButton > button {
        min-height: 48px !important;
        font-size: 1rem !important;
        width: 100% !important;
    }

    /* Full-width inputs */
    .stTextInput > div, .stNumberInput > div, .stSelectbox > div {
        width: 100% !important;
    }
    .stTextInput input, .stNumberInput input {
        font-size: 1rem !important;
        min-height: 44px !important;
    }

    /* Shrink tab labels so all 3 fit on one row */
    .stTabs [data-baseweb="tab"] {
        font-size: 0.75rem !important;
        padding: 6px 8px !important;
    }

    /* Metrics: show 2-up on mobile instead of cramped 4-up */
    [data-testid="metric-container"] {
        min-width: 45% !important;
    }

    /* Download button full width */
    .stDownloadButton > button {
        width: 100% !important;
        min-height: 48px !important;
    }

    /* Disclaimer card mobile padding */
    div[data-testid="stMarkdownContainer"] > div[style*="max-width:680px"] {
        margin: 20px auto !important;
        padding: 24px 16px !important;
    }

    /* Dataframes: allow horizontal scroll instead of squishing */
    [data-testid="stDataFrame"] {
        overflow-x: auto !important;
    }

    /* Hide sidebar toggle label on very small screens */
    .st-emotion-cache-eczf16 { font-size: 0 !important; }
}
</style>
""", unsafe_allow_html=True)

# ─── Admin panel ──────────────────────────────────────────────────────────────
ADMIN_PASSWORD = get_secret("admin", "password")

def gen_code():
    chars = string.ascii_uppercase + string.digits
    suffix = "".join(random.choices(chars, k=4)) + "-" + "".join(random.choices(chars, k=4))
    return f"DFS-{suffix}"

def admin_get_codes():
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/access_codes?select=id,code,name,active,usage_count,last_used,created_at,daily_limit&order=id.asc",
        headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"},
    )
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=10) as r:
            return json.loads(r.read().decode())
    except Exception:
        return []

def admin_insert_code(code, name, trial_days=None):
    import datetime as _dt
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    payload = {"code": code, "name": name}
    if trial_days and trial_days > 0:
        exp = _dt.datetime.utcnow() + _dt.timedelta(days=trial_days)
        payload["expires_at"] = exp.strftime("%Y-%m-%dT%H:%M:%SZ")
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/access_codes",
        data=data,
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=10):
            return True
    except Exception:
        return False

def admin_set_expiry(code_id, days):
    import datetime as _dt_exp
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    if days and days > 0:
        exp = _dt_exp.datetime.utcnow() + _dt_exp.timedelta(days=days)
        payload = {"expires_at": exp.strftime("%Y-%m-%dT%H:%M:%SZ")}
    else:
        payload = {"expires_at": None}
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/access_codes?id=eq.{code_id}",
        data=data,
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="PATCH",
    )
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=10):
            return True
    except Exception:
        return False

def admin_set_daily_limit(code_id, limit):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    payload = {"daily_limit": limit if limit and limit > 0 else None}
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/access_codes?id=eq.{code_id}",
        data=data,
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="PATCH",
    )
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=10):
            return True
    except Exception:
        return False

def admin_toggle_code(code_id, active):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    data = json.dumps({"active": active}).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/access_codes?id=eq.{code_id}",
        data=data,
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="PATCH",
    )
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=10):
            return True
    except Exception:
        return False

if st.query_params.get("admin") == "true":
    if not st.session_state.get("admin_authed"):
        st.markdown(
            """
            <div style="max-width:380px; margin:80px auto; padding:36px 28px;
                        background:#1e2130; border-radius:12px;
                        border:1px solid #2e3250; text-align:center;">
                <div style="font-size:2rem; margin-bottom:8px;">🔐</div>
                <h2 style="margin-bottom:16px;">Admin Access</h2>
            </div>
            """,
            unsafe_allow_html=True,
        )
        a1, a2, a3 = st.columns([1, 2, 1])
        with a2:
            pw = st.text_input("Password", type="password", label_visibility="collapsed", placeholder="Admin password")
            if st.button("Unlock", use_container_width=True, type="primary"):
                if pw == ADMIN_PASSWORD:
                    st.session_state.admin_authed = True
                    st.rerun()
                else:
                    st.error("Incorrect password.")
        st.stop()

    # ── Admin dashboard ──
    st.markdown(f"## 🔐 {APP_NAME} — Admin")
    st.caption(f"v{APP_VERSION} · Access code management")
    st.markdown("---")

    codes = admin_get_codes()

    st.markdown(f"### 👥 Access Codes ({len(codes)} total)")

    import datetime as _dt
    now_utc = _dt.datetime.utcnow().replace(tzinfo=_dt.timezone.utc)

    for row in codes:
        c1, c2, c3, c4, c5, c6 = st.columns([2, 2, 1, 1, 1, 2])
        c1.markdown(f"**{row['name']}**")
        c2.code(row["code"])
        c3.markdown("🟢 Active" if row["active"] else "🔴 Inactive")
        c4.markdown(f"Uses: **{row['usage_count']}**")
        last = (row.get("last_used") or "Never")[:10]
        c5.markdown(f"Last: {last}")
        # Expiry display
        exp_raw = row.get("expires_at")
        if exp_raw:
            try:
                exp_dt = _dt.datetime.fromisoformat(exp_raw.replace("Z", "+00:00"))
                days_left = (exp_dt - now_utc).days
                if days_left < 0:
                    c6.markdown("🔴 **Expired**")
                elif days_left == 0:
                    c6.markdown("🟡 **Expires today**")
                else:
                    c6.markdown(f"⏳ {days_left}d left ({exp_dt.strftime('%b %d')})")
            except Exception:
                c6.markdown(f"Expires: {exp_raw[:10]}")
        else:
            c6.markdown("♾️ No expiry")

        if row["active"]:
            if st.button(f"Revoke", key=f"rev_{row['id']}"):
                admin_toggle_code(row["id"], False)
                st.success(f"Revoked {row['name']}")
                st.rerun()
        else:
            if st.button(f"Reinstate", key=f"rei_{row['id']}"):
                admin_toggle_code(row["id"], True)
                st.success(f"Reinstated {row['name']}")
                st.rerun()
        st.markdown("---")

    st.markdown("### ➕ Create New Access Code")
    n1, n2, n3, n4 = st.columns([2, 2, 1, 1])
    new_name = n1.text_input("Name", placeholder="e.g. John Smith")
    _default_code = "BETA-" + gen_code()[4:] if True else gen_code()  # start with BETA- suggested
    new_code = n2.text_input("Code (auto-generated, editable)", value=gen_code())
    trial_days = n3.number_input("Trial days", min_value=0, max_value=365, value=7,
                                  help="0 = no expiry (full member). 7 = 7-day beta trial.")
    if n4.button("Create", type="primary", use_container_width=True):
        if new_name and new_code:
            days = int(trial_days) if trial_days > 0 else None
            if admin_insert_code(new_code.strip().upper(), new_name.strip(), trial_days=days):
                exp_note = f" · expires in {days} days" if days else " · no expiry"
                st.success(f"✅ Created code for **{new_name}**: `{new_code.upper()}`{exp_note}")
                st.rerun()
            else:
                st.error("Failed — code may already exist.")
        else:
            st.warning("Enter a name and code.")

    st.stop()

# ─── Access code gate ─────────────────────────────────────────────────────────
def validate_code(code: str):
    """Returns (name, code_id, error_key) — error_key is None on success, 'expired' or 'invalid' otherwise."""
    if not SUPABASE_URL:
        return None, False, "invalid"
    import datetime as _dt
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    url = (f"{SUPABASE_URL}/rest/v1/access_codes"
           f"?code=eq.{urllib.parse.quote(code.strip().upper())}&active=eq.true&select=id,name,usage_count,expires_at,daily_limit")
    req = urllib.request.Request(url, headers={
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    })
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=10) as r:
            rows = json.loads(r.read().decode())
        if rows:
            row = rows[0]
            # Check expiry if set
            expires_at = row.get("expires_at")
            if expires_at:
                try:
                    exp = _dt.datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
                    if _dt.datetime.now(_dt.timezone.utc) > exp:
                        return None, False, "expired"
                except Exception:
                    pass
            return row["name"], row["id"], None, row.get("daily_limit")
        return None, False, "invalid", None
    except Exception:
        return None, False, "invalid"

def record_code_use(code_id: int):
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    import datetime
    data = json.dumps({
        "usage_count": f"usage_count + 1",
        "last_used": datetime.datetime.utcnow().isoformat() + "Z",
    }).encode()
    # Use RPC-style raw SQL via PostgREST
    rpc_data = json.dumps({"p_id": code_id}).encode()
    # Simpler: just PATCH with a fresh read-modify-write
    url = f"{SUPABASE_URL}/rest/v1/access_codes?id=eq.{code_id}"
    # Fetch current count first
    req = urllib.request.Request(url + "&select=usage_count", headers={
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    })
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=10) as r:
            current = json.loads(r.read().decode())[0]["usage_count"]
        import datetime as dt
        patch = json.dumps({
            "usage_count": current + 1,
            "last_used": dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        }).encode()
        req2 = urllib.request.Request(url, data=patch, headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        }, method="PATCH")
        with urllib.request.urlopen(req2, context=ctx, timeout=10):
            pass
    except Exception:
        pass

if not st.session_state.get("access_granted"):
    st.markdown(
        f"""
        <div style="max-width:420px; margin:60px auto; padding:36px 28px;
                    background:#1e2130; border-radius:12px;
                    border:1px solid #2e3250; text-align:center;">
            <div style="font-size:2.2rem; margin-bottom:8px;">💎</div>
            <h2 style="margin-bottom:4px;">{APP_NAME}</h2>
            <p style="color:#aaa; font-size:0.85rem; margin-bottom:24px;">
                Enter your access code to continue.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    # Pre-fill from query param ?k=CODE (set on successful login)
    _saved_code = st.query_params.get("k", "")
    # JS: set autocomplete=new-password (browsers ignore "off") and re-apply saved code after autofill fires
    st.components.v1.html(f"""
    <script>
    (function() {{
        var code = {repr(_saved_code)};
        function fix() {{
            var inputs = window.parent.document.querySelectorAll('input[type="text"], input:not([type])');
            inputs.forEach(function(inp) {{
                inp.setAttribute('autocomplete', 'new-password');
                if (code && inp.value !== code) {{
                    inp.value = code;
                    inp.dispatchEvent(new Event('input', {{bubbles: true}}));
                    inp.dispatchEvent(new Event('change', {{bubbles: true}}));
                }}
            }});
        }}
        setTimeout(fix, 100);
        setTimeout(fix, 500);
        setTimeout(fix, 1200);
    }})();
    </script>
    """, height=0)
    gc1, gc2, gc3 = st.columns([1, 2, 1])
    with gc2:
        entered_code = st.text_input("Access Code", value=_saved_code, placeholder="XXXX-XXXX", label_visibility="collapsed", autocomplete="off")
        if st.button("Enter", use_container_width=True, type="primary"):
            clean_code = entered_code.strip().upper()
            # Owner bypass — never touches Supabase
            if clean_code == "DFS-MASTER":
                st.query_params["k"] = clean_code
                st.session_state.access_granted = True
                st.session_state.access_name = "Duane"
                st.session_state.access_code_id = 1
                st.session_state.agreed = True
                st.session_state.is_beta = False
                st.rerun()
            else:
                name, code_id, err, daily_limit = validate_code(entered_code)
                if name and code_id:
                    st.query_params["k"] = clean_code
                    st.session_state.access_granted = True
                    st.session_state.access_name = name
                    st.session_state.access_code_id = code_id
                    st.session_state.access_daily_limit = daily_limit
                    # BETA- prefix = limited preview access
                    st.session_state.is_beta = clean_code.startswith("BETA-")
                    # Store expiry label for sidebar display
                    if st.session_state.get("_login_expires_at"):
                        import datetime as _dt2
                        try:
                            exp2 = _dt2.datetime.fromisoformat(
                                st.session_state["_login_expires_at"].replace("Z", "+00:00"))
                            dl = (exp2 - _dt2.datetime.now(_dt2.timezone.utc)).days
                            st.session_state.trial_expires_label = f"Trial ends in {dl} day{'s' if dl != 1 else ''}."
                        except Exception:
                            pass
                    record_code_use(code_id)
                    st.rerun()
                elif err == "expired":
                    st.error("⏰ Your trial has ended. Contact us to upgrade to full access.")
                else:
                    st.error("Invalid or inactive access code.")
    st.stop()

# ─── Disclaimer gate ──────────────────────────────────────────────────────────
if not st.session_state.get("agreed"):
    st.markdown(
        f"""
        <div style="max-width:680px; margin:40px auto; padding:32px 24px;
                    background:#1e2130; border-radius:12px;
                    border:1px solid #2e3250; text-align:center;">
            <div style="font-size:2.5rem; margin-bottom:8px;">💎</div>
            <h2 style="margin-bottom:4px;">{APP_NAME}</h2>
            <p style="color:#aaa; font-size:0.85rem; margin-bottom:24px;">
                Please read and accept the disclaimer before continuing.
            </p>
            <div style="text-align:left; background:#0f1117; border-radius:8px;
                        padding:20px; margin-bottom:24px;
                        font-size:0.88rem; color:#ccc; line-height:1.7;">
                <strong style="color:#fafafa;">Disclaimer</strong><br><br>
                {APP_NAME} is a research and decision-support tool only.
                All pricing data (GemRate, eBay) is pulled from third-party sources
                and may be incomplete, delayed, or inaccurate. Gem rates and market
                values fluctuate — always verify data independently before submitting
                cards for grading.<br><br>
                All grading decisions and associated costs are <strong style="color:#fafafa;">
                solely your responsibility</strong>. {APP_NAME} assumes no liability
                for financial outcomes resulting from use of this tool.<br><br>
                <span style="font-size:0.8rem; color:#888;">
                ©️ 2026 {APP_NAME}. All rights reserved.
                </span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    col_l, col_btn, col_r = st.columns([2, 2, 2])
    with col_btn:
        if st.button("✅ I Understand & Agree", use_container_width=True, type="primary"):
            st.session_state.agreed = True
            st.rerun()
    st.stop()

# ─── SSL ──────────────────────────────────────────────────────────────────────
def ssl_ctx():
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx

# ─── Supabase tracker ─────────────────────────────────────────────────────────
TRACKER_COLS = [
    "id", "date_added", "card_description", "year", "set_name", "parallel",
    "raw_buy_price", "psa_tier", "psa_fee", "psa10_avg_price", "target_price",
    "gem_rate", "go_no_go", "est_net", "est_roi",
    "date_submitted", "status", "grade_returned",
    "actual_sell_price", "actual_net", "actual_roi", "notes",
]

def sb_headers():
    return {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }

def sb_get():
    if not SUPABASE_URL:
        return []
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/grading_tracker?order=id.asc",
        headers=sb_headers(),
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
            return json.loads(r.read().decode())
    except Exception:
        return []

def sb_insert(row: dict):
    if not SUPABASE_URL:
        return
    data = json.dumps(row).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/grading_tracker",
        data=data,
        headers=sb_headers(),
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
            return json.loads(r.read().decode())
    except Exception:
        pass

def sb_update(row_id: int, updates: dict):
    if not SUPABASE_URL:
        return
    data = json.dumps(updates).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/grading_tracker?id=eq.{row_id}",
        data=data,
        headers={**sb_headers(), "Prefer": "return=minimal"},
        method="PATCH",
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10):
            pass
    except Exception:
        pass

def sb_delete(row_id: int):
    if not SUPABASE_URL:
        return
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/grading_tracker?id=eq.{row_id}",
        headers=sb_headers(),
        method="DELETE",
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10):
            pass
    except Exception:
        pass

# ─── Scan Stacks Supabase helpers ────────────────────────────────────────────
def _sb_req(method, path, data=None, extra_headers=None):
    """Generic Supabase REST helper."""
    if not SUPABASE_URL:
        return None
    h = {**sb_headers(), **(extra_headers or {})}
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/{path}",
        data=json.dumps(data).encode() if data else None,
        headers=h, method=method,
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
            body = r.read().decode()
            return json.loads(body) if body.strip() else []
    except Exception:
        return None

def stacks_list():
    return _sb_req("GET", "scan_stacks?order=id.desc&select=*") or []

def stack_create(name, notes=""):
    rows = _sb_req("POST", "scan_stacks", {"name": name, "notes": notes, "status": "open"})
    return rows[0] if rows else None

def stack_update(stack_id, updates):
    _sb_req("PATCH", f"scan_stacks?id=eq.{stack_id}", updates,
            extra_headers={"Prefer": "return=minimal"})

def stack_delete(stack_id):
    _sb_req("DELETE", f"scan_stacks?id=eq.{stack_id}", extra_headers={"Prefer": "return=minimal"})

def stack_cards_get(stack_id):
    return _sb_req("GET", f"scan_cards?stack_id=eq.{stack_id}&order=idx.asc&select=*") or []

def stack_card_upsert(card: dict):
    """Insert or replace a card row (matched on stack_id + idx)."""
    _sb_req("POST", "scan_cards", card,
            extra_headers={"Prefer": "resolution=merge-duplicates,return=minimal",
                           "on_conflict": "stack_id,idx"})

def stack_cards_delete(stack_id):
    _sb_req("DELETE", f"scan_cards?stack_id=eq.{stack_id}",
            extra_headers={"Prefer": "return=minimal"})

def stack_card_update(card_id, updates):
    _sb_req("PATCH", f"scan_cards?id=eq.{card_id}", updates,
            extra_headers={"Prefer": "return=minimal"})

# ─── Shipment Intake Supabase helpers ─────────────────────────────────────────
def sb_intake_get():
    """Fetch all shipment intake records, newest first. Returns (rows, error_str)."""
    if not SUPABASE_URL:
        return [], "Supabase not configured"
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/shipment_intake?order=id.desc",
        headers=sb_headers(),
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
            return json.loads(r.read().decode()), None
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors="replace")
        return [], f"HTTP {e.code}: {body}"
    except Exception as e:
        return [], str(e)

def sb_intake_insert(row: dict):
    """Insert a new shipment intake record. Returns (result, error_str)."""
    if not SUPABASE_URL:
        return None, "Supabase not configured"
    data = json.dumps(row).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/shipment_intake",
        data=data,
        headers=sb_headers(),
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
            return json.loads(r.read().decode()), None
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors="replace")
        return None, f"HTTP {e.code}: {body}"
    except Exception as e:
        return None, str(e)

def _json_safe(obj):
    """Recursively convert non-JSON-serializable types (date, numpy, NaT) to safe values."""
    import datetime, math
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    try:
        import numpy as np
        if isinstance(obj, (np.integer,)):  return int(obj)
        if isinstance(obj, (np.floating,)): return None if math.isnan(float(obj)) else float(obj)
        if isinstance(obj, np.bool_):       return bool(obj)
    except ImportError:
        pass
    if obj is pd.NaT:
        return None
    try:
        if isinstance(obj, float) and math.isnan(obj):
            return None
    except Exception:
        pass
    return obj

def sb_intake_update(row_id: int, updates: dict):
    """Update a shipment intake record by ID."""
    if not SUPABASE_URL:
        return
    data = json.dumps(_json_safe(updates)).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/shipment_intake?id=eq.{row_id}",
        data=data,
        headers={**sb_headers(), "Prefer": "return=minimal"},
        method="PATCH",
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10):
            pass
    except Exception:
        pass

def sb_intake_delete(row_id: int):
    """Delete a shipment intake record by ID."""
    if not SUPABASE_URL:
        return
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/shipment_intake?id=eq.{row_id}",
        headers=sb_headers(),
        method="DELETE",
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10):
            pass
    except Exception:
        pass

# ─── Live-pricing daily usage cap (protects the CardHedger API budget) ────────
# Supabase table `pricing_usage` (code text, day date, used int) persists the
# count across sessions/refreshes; falls back to session_state if it's missing.
def _today_iso():
    return date.today().isoformat()

def _pricing_key():
    return str(st.session_state.get("access_code_id") or st.session_state.get("access_name") or "anon")

def pricing_unlimited():
    """Owner + Robert (marketing partner) are never capped."""
    return st.session_state.get("access_name", "") in ("Duane", "Robert Bass")

def pricing_used_today():
    """Live look-ups already used today (max of Supabase + this session)."""
    day = _today_iso()
    sess_n = st.session_state.get("_pricing_used", {}).get(day, 0)
    if not SUPABASE_URL:
        return sess_n
    try:
        code = urllib.parse.quote(_pricing_key())
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/pricing_usage?code=eq.{code}&day=eq.{day}&select=used",
            headers=sb_headers())
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=8) as r:
            rows = json.loads(r.read().decode())
        return max(sess_n, rows[0]["used"] if rows else 0)
    except Exception:
        return sess_n

def pricing_remaining():
    if pricing_unlimited():
        return 10 ** 9
    cap = st.session_state.get("access_daily_limit") or DAILY_PRICING_CAP
    return max(0, cap - pricing_used_today())

def pricing_bump(n):
    """Record n live look-ups against today's budget (session + Supabase)."""
    if n <= 0 or pricing_unlimited():
        return
    day = _today_iso()
    sess = st.session_state.setdefault("_pricing_used", {})
    sess[day] = sess.get(day, 0) + n
    if not SUPABASE_URL:
        return
    try:
        code = _pricing_key()
        q = urllib.parse.quote(code)
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/pricing_usage?code=eq.{q}&day=eq.{day}&select=id,used",
            headers=sb_headers())
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=8) as r:
            rows = json.loads(r.read().decode())
        if rows:
            req2 = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/pricing_usage?id=eq.{rows[0]['id']}",
                data=json.dumps({"used": rows[0]["used"] + n}).encode(),
                headers={**sb_headers(), "Prefer": "return=minimal"}, method="PATCH")
        else:
            req2 = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/pricing_usage",
                data=json.dumps({"code": code, "day": day, "used": n}).encode(),
                headers={**sb_headers(), "Prefer": "return=minimal"}, method="POST")
        urllib.request.urlopen(req2, context=ssl_ctx(), timeout=8)
    except Exception:
        pass

# ─── GemRate API ──────────────────────────────────────────────────────────────
def _gemrate_single(query: str):
    payload = json.dumps({"query": query, "limit": 10}).encode()
    browser_hdrs = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    }

    # Route through WordPress proxy first (avoids Cloudflare IP block on Streamlit Cloud)
    try:
        req = urllib.request.Request(WP_PROXY_URL, data=payload, headers=browser_hdrs, method="POST")
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=15) as r:
            result = json.loads(r.read().decode())
            if isinstance(result, list) and result:
                return result
    except Exception:
        pass

    # Direct fallback (works locally, may be blocked on cloud)
    try:
        req = urllib.request.Request(
            "https://www.gemrate.com/universal-search-query",
            data=payload, headers=browser_hdrs, method="POST",
        )
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=15) as r:
            return json.loads(r.read().decode())
    except Exception:
        return []

def _build_queries(query: str):
    """Expand user query into multiple GemRate-friendly variants."""
    q = query.strip()
    queries = [q]
    ql = q.lower()

    # auto / autograph expansion
    if "auto" in ql and "autograph" not in ql:
        queries.append(q.lower().replace("auto", "autograph").strip())
        queries.append(q + " rookie autograph")

    # rc → rookie
    if " rc" in ql or ql.endswith(" rc"):
        queries.append(q.lower().replace(" rc", " rookie").strip())

    # rookie without auto → add autograph variant
    if "rookie" in ql and "auto" not in ql and "autograph" not in ql:
        queries.append(q + " autograph")

    # Prizm → Panini Prizm (GemRate indexes with manufacturer name)
    if "prizm" in ql and "panini" not in ql:
        queries.append(q.replace("Prizm", "Panini Prizm").replace("prizm", "Panini Prizm"))

    # Optic → Panini Optic
    if "optic" in ql and "panini" not in ql:
        queries.append(q.replace("Optic", "Panini Optic").replace("optic", "Panini Optic"))

    # Chrome → Topps Chrome (if not Bowman Chrome already)
    if "chrome" in ql and "topps" not in ql and "bowman" not in ql:
        queries.append(q.replace("Chrome", "Topps Chrome").replace("chrome", "Topps Chrome"))

    # Bowman Chrome prospect autograph variant
    if "bowman" in ql and "chrome" in ql and "autograph" not in ql:
        queries.append(q + " autograph")

    # Draft Picks → Prizm Draft Picks (Panini)
    if "draft picks" in ql and "panini" not in ql:
        queries.append("Panini " + q)

    return list(dict.fromkeys(queries))  # dedupe while preserving order

@st.cache_data(ttl=300, show_spinner=False)
def search_gemrate(query: str):
    queries = _build_queries(query)
    seen_ids = set()
    combined = []
    for q in queries:
        for r in _gemrate_single(q):
            gid = r.get("gemrate_id") or r.get("id") or str(r)
            if gid not in seen_ids:
                seen_ids.add(gid)
                combined.append(r)
    # Sort: prioritise results where set_name or parallel contains user keywords
    ql = query.lower()
    keywords = [w for w in ql.split() if len(w) > 2]
    def relevance(r):
        text = f"{r.get('set_name','')} {r.get('parallel','')} {r.get('name','')}".lower()
        return sum(1 for kw in keywords if kw in text)
    combined.sort(key=relevance, reverse=True)
    return combined

# ─── eBay scraper fallback (no API key required) ──────────────────────────────
import re as _re

@st.cache_data(ttl=1800, show_spinner=False)
def _scrape_ebay_sold(query: str, max_results: int = 15):
    q = urllib.parse.quote_plus(query)
    url = (f"https://www.ebay.com/sch/i.html?_nkw={q}"
           f"&LH_Sold=1&LH_Complete=1&_sop=13&_ipg=25")
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept-Language": "en-US,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    }
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=15) as r:
            html = r.read().decode("utf-8", errors="ignore")
    except Exception:
        return []

    # Each listing lives in a <li class="s-item ..."> block
    items = _re.split(r'class="s-item["\s]', html)
    results = []
    for chunk in items[1:]:
        # Title
        tm = _re.search(r's-item__title[^>]*>([^<]{5,})<', chunk)
        title = tm.group(1).strip() if tm else ""
        if not title or title.lower() in ("shop on ebay", "results matching fewer words"):
            continue
        # Price — handle ranges like "$12.00 to $15.00" by taking the first
        pm = _re.search(r'\$\s*([\d,]+\.?\d*)', chunk)
        if not pm:
            continue
        try:
            price = float(pm.group(1).replace(",", ""))
        except ValueError:
            continue
        if price <= 0:
            continue
        # Image
        im = _re.search(r's-item__image-img[^>]+src="([^"]+)"', chunk)
        image = im.group(1) if im else ""
        # Link
        lm = _re.search(r'href="(https://www\.ebay\.com/itm/[^"]+)"', chunk)
        item_url = lm.group(1).split("?")[0] if lm else ""
        results.append({"title": title, "price": price, "date": "", "image": image, "url": item_url})
        if len(results) >= max_results:
            break
    return results

# ─── eBay Finding API ─────────────────────────────────────────────────────────
@st.cache_data(ttl=1800, show_spinner=False)
def fetch_ebay_sold(query: str, app_id: str, max_results: int = 15):
    if not app_id:
        return _scrape_ebay_sold(query, max_results)
    params = {
        "OPERATION-NAME": "findCompletedItems",
        "SERVICE-VERSION": "1.0.0",
        "SECURITY-APPNAME": app_id,
        "RESPONSE-DATA-FORMAT": "JSON",
        "keywords": query,
        "itemFilter(0).name": "SoldItemsOnly",
        "itemFilter(0).value": "true",
        "sortOrder": "EndTimeSoonest",
        "paginationInput.entriesPerPage": str(max_results),
    }
    url = "https://svcs.ebay.com/services/search/FindingService/v1?" + urllib.parse.urlencode(params)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=12) as r:
            data = json.loads(r.read().decode())
        items = (data.get("findCompletedItemsResponse", [{}])[0]
                     .get("searchResult", [{}])[0]
                     .get("item", []))
        results = []
        for item in items:
            price_val = (item.get("sellingStatus", [{}])[0]
                             .get("currentPrice", [{}])[0]
                             .get("__value__", ""))
            end_time = (item.get("listingInfo", [{}])[0]
                            .get("endTime", [None])[0] or "")
            title = item.get("title", [""])[0]
            try:
                price = float(price_val)
            except Exception:
                continue
            image = item.get("galleryURL", [""])[0]
            item_url = item.get("viewItemURL", [""])[0]
            results.append({"title": title, "price": price, "date": end_time[:10], "image": image, "url": item_url})
        return results
    except Exception:
        return []

def ebay_avg(sold_items):
    if not sold_items:
        return None
    prices = sorted(x["price"] for x in sold_items)
    if len(prices) >= 6:
        cut = max(1, len(prices) // 10)
        prices = prices[cut:-cut]
    return round(sum(prices) / len(prices), 2)

# ─── CardHedger API ───────────────────────────────────────────────────────────
def _ch_post(endpoint: str, payload: dict):
    if not CARDHEDGER_KEY:
        return None
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        f"{CARDHEDGER_BASE}{endpoint}", data=data,
        headers={"X-API-Key": CARDHEDGER_KEY, "Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=30) as r:
            return json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        _body = ""
        try:
            _body = e.read().decode("utf-8", errors="replace")[:500]
        except Exception:
            pass
        return {"_ch_error": f"HTTP {e.code} {e.reason}", "_ch_body": _body}
    except Exception as _ex:
        return {"_ch_error": str(_ex)}

@st.cache_data(ttl=1800, show_spinner=False)
def ch_search(query: str):
    result = _ch_post("/v1/cards/card-search", {"search": query, "page": 1, "page_size": 5})
    if not result:
        return []
    for key in ("cards", "data", "results", "items"):
        if key in result and isinstance(result[key], list):
            return result[key]
    return result if isinstance(result, list) else []

@st.cache_data(ttl=600, show_spinner=False)
def ch_card_match(query: str):
    """AI-powered best-match search — returns card with prices by grade."""
    result = _ch_post("/v1/cards/card-match", {"query": query, "page": 1, "page_size": 5})
    if not result or "match" not in result:
        return None
    return result["match"]

@st.cache_data(ttl=1800, show_spinner=False)
def ch_comps(card_id, grade: str):
    return _ch_post("/v1/cards/comps", {
        "card_id": card_id, "grade": grade,
        "count": 20, "include_raw_prices": True,
    }) or {}

@st.cache_data(ttl=1800, show_spinner=False)
def ch_fmv(card_id, grade: str):
    """Fair Market Value — statistically cleaned price (Winsorized median) with a
    confidence grade and a low–high band. More reliable than the raw comp average.
    Returns {} if unavailable."""
    return _ch_post("/v1/cards/card-fmv", {
        "card_id": card_id, "grade": grade,
    }) or {}

def fmv_price(d):
    """Pull a usable price out of an FMV dict, else None."""
    try:
        v = float(d.get("price"))
        return v if v > 0 else None
    except Exception:
        return None

def fmv_caption(d):
    """One-line FMV display: '🎯 FMV $X  ($low–$high, conf A)' or '' if no price."""
    p = fmv_price(d)
    if not p:
        return ""
    lo, hi = d.get("price_low"), d.get("price_high")
    grade = d.get("confidence_grade") or ""
    band = ""
    try:
        if lo and hi:
            band = f" (${float(lo):,.0f}–${float(hi):,.0f}"
            band += f", conf {grade})" if grade else ")"
    except Exception:
        band = f" (conf {grade})" if grade else ""
    return f"🎯 **FMV ${p:,.2f}**{band}"

def fmv_band_conf(d):
    """'🎯 FMV · $lo–$hi · conf A' for use under a metric already showing the FMV price."""
    if not fmv_price(d):
        return ""
    lo, hi = d.get("price_low"), d.get("price_high")
    grade = d.get("confidence_grade") or ""
    parts = []
    try:
        if lo and hi:
            parts.append(f"${float(lo):,.0f}–${float(hi):,.0f}")
    except Exception:
        pass
    if grade:
        parts.append(f"conf {grade}")
    return "🎯 FMV · " + " · ".join(parts) if parts else "🎯 FMV"

@st.cache_data(ttl=3600, show_spinner=False)
def ch_price_history(card_id, grade: str, days: int = 90):
    return _ch_post("/v1/cards/prices-by-card", {
        "card_id": card_id, "grade": grade, "days": days,
    }) or {}

@st.cache_data(ttl=3600, show_spinner=False)
def ch_card_image(card_id: str) -> str:
    """Fetch the card image URL from CardHedger card-details."""
    result = _ch_post("/v1/cards/card-details", {"card_id": card_id})
    if not result:
        return ""
    cards = result.get("cards") or []
    if cards and isinstance(cards, list):
        return cards[0].get("image", "") or ""
    return result.get("image", "") or ""

@st.cache_data(ttl=1800, show_spinner=False)
def ch_card_meta(card_id):
    """card-details for one card — includes '7 Day Sales' / '30 Day Sales' volume."""
    result = _ch_post("/v1/cards/card-details", {"card_id": card_id})
    cards = (result or {}).get("cards") or []
    return cards[0] if cards else (result or {})

def _ch_get(endpoint: str):
    """GET against CardHedger (top-movers is GET, not POST)."""
    if not CARDHEDGER_KEY:
        return None
    req = urllib.request.Request(f"{CARDHEDGER_BASE}{endpoint}",
                                 headers={"X-API-Key": CARDHEDGER_KEY})
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=15) as r:
            return json.loads(r.read().decode())
    except Exception:
        return None

@st.cache_data(ttl=900, show_spinner=False)
def ch_top_movers(category=None, count=25):
    """Weekly biggest price gainers (optionally filtered to a sport category)."""
    q = f"/v1/cards/top-movers?count={int(count)}"
    if category:
        q += f"&category={urllib.parse.quote(category)}"
    d = _ch_get(q) or {}
    return d.get("cards", []) if isinstance(d, dict) else []

@st.cache_data(ttl=1800, show_spinner=False)
def ch_all_prices(card_id):
    """Latest price for EVERY grade in one call (Raw..PSA 10). Cheaper than N comps calls."""
    d = _ch_post("/v1/cards/all-prices-by-card", {"card_id": card_id}) or {}
    return d.get("prices", []) if isinstance(d, dict) else []

@st.cache_data(ttl=1800, show_spinner=False)
def ch_price_est(card_id, grade: str):
    """Single-card price estimate with confidence band and freshness. Faster than comps."""
    return _ch_post("/v1/cards/price-estimate", {"card_id": card_id, "grade": grade}) or {}

def ch_batch_price_est(items: list):
    """Bulk price estimate — items=[{card_id, grade}, ...]. One call for many cards."""
    if not items:
        return []
    d = _ch_post("/v1/cards/batch-price-estimate", {"items": items}) or {}
    return d.get("results", []) if isinstance(d, dict) else []

@st.cache_data(ttl=1800, show_spinner=False)
def ch_sales_stats(player, interval="week", periods=8):
    """Player sales volume + $ per period → demand trend. Returns list of buckets."""
    d = _ch_post("/v1/cards/sales-stats-by-player",
                 {"players": [player], "interval": interval, "periods": periods}) or {}
    res = d.get("results", []) if isinstance(d, dict) else []
    return (res[0].get("buckets", []) if res else [])

@st.cache_data(ttl=900, show_spinner=False)
def ch_image_match(image_b64):
    """Identify a raw card from a photo (base64). Returns dict with candidates[]."""
    # Try both field names the CardHedger API has used across versions
    for _field in ("image_base64", "image"):
        _r = _ch_post("/v1/cards/image-match", {_field: image_b64})
        if _r and not _r.get("_ch_error"):
            # Check if this response has actual match data (not just an empty success)
            for _ck in ("candidates", "matches", "cards", "results", "data", "items"):
                if isinstance((_r or {}).get(_ck), list) and _r[_ck]:
                    return _r
            # Non-error response but no list found — return it anyway for inspection
            if _r:
                return _r
    # Return last error so caller can show it
    return _r or {}

# ─── Claude Vision Card Identification ────────────────────────────────────────

_CLAUDE_CARD_PROMPT = """You are a professional trading card expert. Identify this trading card from the image.

IMPORTANT: If the card is inside a graded slab (PSA, BGS, SGC, CGC, etc.), read the label carefully for the grader name, grade number, and certification/serial number.

Return ONLY a valid JSON object with these exact fields — no prose, no markdown fences:
{
  "player": "Full player/subject name exactly as printed on the card",
  "year": "4-digit release year",
  "brand": "Manufacturer (Topps, Panini, Bowman, Upper Deck, Fleer, etc.)",
  "set": "Full set name (e.g. Bowman Chrome, Topps Heritage, Panini Prizm, Topps Update)",
  "card_number": "Card number with any prefix (e.g. #187, BCP-53, TOG-14)",
  "parallel": "Parallel or refractor type (e.g. Silver Prizm, Gold Refractor, Wave Prizm). Empty string for base cards.",
  "sport": "Baseball, Basketball, Football, Soccer, or Hockey",
  "numbered": "Print run if visible on card (e.g. /99, /250). Empty string if not visible.",
  "rookie": true or false,
  "team": "Team name as shown on card",
  "notes": "Any other notable details: RC, 1st Bowman, autograph, relic, etc.",
  "graded": true or false,
  "grader": "PSA, BGS, SGC, CGC, or empty string if not in a graded slab",
  "grade": "Numeric grade from slab label (e.g. 10, 9.5, 9) — just the number, no grader prefix. Empty string if not graded.",
  "cert_number": "Certification/serial number printed on the slab label. Empty string if not graded or not visible."
}

Read the card carefully — player name, year, set name, card number are usually printed directly on the card.
For graded slabs: the grader logo (PSA/BGS/SGC/CGC) and grade number appear prominently on the label; the cert number is a long numeric code (e.g. 12345678).
Only report what you can actually see. Use empty string for anything you cannot read clearly."""

def claude_identify_card(image_b64: str, media_type: str = "image/jpeg") -> dict:
    """Identify a trading card using Claude Vision. Returns structured card data."""
    if not ANTHROPIC_KEY:
        return {"_error": "Anthropic API key not configured"}
    payload = {
        "model": "claude-sonnet-5",
        "max_tokens": 1024,
        "messages": [{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {"type": "base64", "media_type": media_type, "data": image_b64},
                },
                {"type": "text", "text": _CLAUDE_CARD_PROMPT},
            ],
        }],
    }
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=data,
        headers={
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=30) as r:
            resp = json.loads(r.read().decode())
        text = resp["content"][0]["text"].strip()
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if m:
            return json.loads(m.group(0))
        return {"_error": "No JSON in response", "_raw": text}
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        return {"_error": f"HTTP {e.code}", "_raw": body}
    except Exception as e:
        return {"_error": str(e)}

_CLAUDE_GRADE_PROMPT = """You are a PSA expert grader with 20+ years of experience. Analyze the trading card image(s) and return a grading assessment.

PSA GRADING CRITERIA:
- PSA 10 (Gem Mint): Centering ≤55/45 front AND ≤75/25 back. All four corners perfectly sharp. All four edges clean. Surface flawless.
- PSA 9 (Mint): Centering ≤65/35 front AND ≤85/15 back. Corners very sharp — one may have slight fraying. Edges very clean. Surface nearly flawless.
- PSA 8 (NM-MT): Centering ≤65/35. One or two corners/edges with slight wear. Minor surface issues.
- PSA 7 (NM): Centering ≤70/30. Noticeable corner wear or surface issue. No creases.
- PSA 6 (EX-MT) and below: Obvious wear on multiple features.

WHAT TO LOOK FOR:
Centering: Estimate the white border widths. If left border = 45% and right = 55% of total horizontal space, that is 45/55 (fails PSA 10 front threshold of 55/45).
Corners: Look at all four corners — NW (top-left), NE (top-right), SW (bottom-left), SE (bottom-right). Rate each: sharp / slight fraying / moderate wear / heavy wear.
Edges: Look at top, bottom, left, right edges. Rate each: clean / rough / minor chip / chip / nick.
Surface: Look for scratches, print lines, stains, dimples, fingerprints, creases.

Return ONLY valid JSON — no prose, no explanation outside the JSON:

{
  "predicted_grade": 10,
  "grade_label": "PSA 10",
  "confidence": "high",
  "centering": {
    "front_left_right": "52/48",
    "front_top_bottom": "51/49",
    "back_left_right": "55/45",
    "front_passes_10": true,
    "back_passes_10": true,
    "note": "Well-centered on both sides"
  },
  "corners": {
    "NW": "sharp",
    "NE": "sharp",
    "SW": "sharp",
    "SE": "slight fraying",
    "worst": "SE",
    "passes_10": false,
    "note": "SE corner has very slight fraying visible under close inspection"
  },
  "edges": {
    "top": "clean",
    "bottom": "clean",
    "left": "clean",
    "right": "clean",
    "passes_10": true,
    "note": "All edges clean"
  },
  "surface": {
    "front": "clean",
    "back": "clean",
    "front_defects": [],
    "back_defects": [],
    "passes_10": true,
    "note": "No surface issues detected"
  },
  "fatal_flaws": ["SE corner: slight fraying — this will likely hold the grade to PSA 9"],
  "positive_attributes": ["Excellent centering", "Three corners gem-sharp", "Clean edges", "Pristine surface"],
  "estimated_grade_range": "PSA 9 – PSA 9.5",
  "recommendation": "Strong PSA 9 candidate. Submit only if PSA 9 market value comfortably exceeds the grading fee.",
  "submit_recommended": true,
  "caveat": "AI grading is based on visible image quality only. Micro-defects under loupe, UV surface issues, and back defects not in the photo may affect the actual PSA grade."
}"""

def claude_grade_card(front_b64: str, back_b64: str | None,
                       front_mime: str = "image/jpeg", back_mime: str = "image/jpeg") -> dict:
    """Predict PSA grade from front (and optionally back) card images using Claude Vision."""
    if not ANTHROPIC_KEY:
        return {"_error": "Anthropic API key not configured"}
    content = [
        {"type": "text", "text": "FRONT OF CARD:"},
        {"type": "image", "source": {"type": "base64", "media_type": front_mime, "data": front_b64}},
    ]
    if back_b64:
        content += [
            {"type": "text", "text": "BACK OF CARD:"},
            {"type": "image", "source": {"type": "base64", "media_type": back_mime, "data": back_b64}},
        ]
    else:
        content.append({"type": "text", "text": "(No back image provided — centering/surface back assessment will be estimated.)"})
    content.append({"type": "text", "text": _CLAUDE_GRADE_PROMPT})
    payload = {
        "model": "claude-sonnet-5",
        "max_tokens": 1800,
        "messages": [{"role": "user", "content": content}],
    }
    data = json.dumps(payload).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=data,
        headers={
            "x-api-key": ANTHROPIC_KEY,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
    )
    try:
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=45) as r:
            resp = json.loads(r.read().decode())
        text = resp["content"][0]["text"].strip()
        m = re.search(r"\{.*\}", text, re.DOTALL)
        if m:
            return json.loads(m.group(0))
        return {"_error": "No JSON in response", "_raw": text}
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        return {"_error": f"HTTP {e.code}", "_raw": body}
    except Exception as e:
        return {"_error": str(e)}

def build_card_query(info: dict, include_parallel: bool = True) -> str:
    """Build a precise eBay search query from Claude-identified card fields."""
    parts = []
    year     = (info.get("year")        or "").strip()
    brand    = (info.get("brand")       or "").strip()
    set_name = (info.get("set")         or "").strip()
    player   = (info.get("player")      or "").strip()
    card_num = (info.get("card_number") or "").strip()
    parallel = (info.get("parallel")    or "").strip()
    numbered = (info.get("numbered")    or "").strip()

    if year:
        parts.append(year)
    if brand:
        parts.append(brand)

    # Strip brand from the front of the set name to avoid "Topps Topps Tier One…"
    if set_name:
        set_clean = re.sub(rf"^{re.escape(brand)}\s*", "", set_name, flags=re.IGNORECASE).strip()
        if set_clean:
            parts.append(set_clean)

    if player:
        parts.append(player)
    if card_num:
        parts.append(card_num)

    # Skip parallel if it's already embedded in the set name (e.g. "Silver Signatures" in "Tier One Silver Signatures")
    if include_parallel and parallel and parallel.lower() not in set_name.lower():
        parts.append(parallel)

    if numbered:
        parts.append(numbered)

    return " ".join(dict.fromkeys(parts))  # deduplicate preserving order

@st.cache_data(ttl=1800, show_spinner=False)
def ch_prices_by_cert(cert, grader="PSA", days=180):
    """Look up a graded card by its cert number → cert_info + card + price history."""
    return _ch_post("/v1/cards/prices-by-cert",
                    {"cert": str(cert), "grader": grader, "days": int(days)}) or {}

# ── Batch Scanner helpers ────────────────────────────────────────────────────

_RAW_CONDITIONS = {
    "Near Mint (NM)":    "2750",
    "Excellent (EX)":    "3000",
    "Very Good (VG)":    "4000",
    "Good (G)":          "5000",
    "Poor":              "6000",
}

def ch_cert_ocr(image_url: str):
    """details-by-cert-ocr: send a public image URL, get card identity + grade."""
    return _ch_post("/v1/cards/details-by-cert-ocr", {"image_url": image_url}) or {}

def ch_image_match_url(image_url: str):
    """image-match via URL: visual KNN search — best for raw cards with no cert label."""
    return _ch_post("/v1/cards/image-match", {"image_url": image_url, "k": 3}) or {}

def ch_image_match_raw(image_url: str, k: int = 5):
    """image-match with k candidates — primary identification for raw cards."""
    return _ch_post("/v1/cards/image-match", {"image_url": image_url, "k": k}) or {}

def imgbb_upload(b64_bytes: bytes, name: str = "card") -> str:
    """Upload image bytes to imgbb (free host). Returns permanent public URL or ''."""
    key = IMGBB_KEY or st.session_state.get("imgbb_key_input", "")
    if not key:
        return ""
    try:
        import urllib.parse as _up
        b64_str = base64.b64encode(b64_bytes).decode() if isinstance(b64_bytes, (bytes, bytearray)) else b64_bytes
        payload = _up.urlencode({"key": key, "image": b64_str, "name": name[:100]}).encode()
        req = urllib.request.Request("https://api.imgbb.com/1/upload", data=payload)
        with urllib.request.urlopen(req, timeout=30) as r:
            d = json.loads(r.read().decode())
        return d.get("data", {}).get("url", "") or ""
    except Exception:
        return ""

def ch_fmv_batch(items: list):
    """card-fmv-batch: price up to 100 card/grade combos in one call."""
    return _ch_post("/v1/cards/card-fmv-batch", {"items": items}) or {}

def ch_comps_raw(card_id: str, count: int = 10):
    """comps for a raw card — returns comp_price, high, low, raw_prices list."""
    return _ch_post("/v1/cards/comps", {
        "card_id": card_id,
        "count": count,
        "grade": "Raw",
        "time_weighted": True,
        "include_raw_prices": True,
    }) or {}

def _raw_title(player, set_name, number, variant=""):
    """eBay title for an ungraded raw card (no grade appended)."""
    player_up = (player or "").upper()
    year_m = re.search(r"\b(19|20)\d{2}\b", set_name or "")
    year = year_m.group() if year_m else ""
    set_clean = re.sub(r"^\d{4}\s*", "", set_name or "").strip()
    var = variant if variant and variant.lower() not in ("base", "base set", "") else ""
    parts = [p for p in [player_up, year, set_clean,
                          f"#{number}" if number else "", var] if p]
    return " ".join(parts)[:80]

def _scan_upload_to_supabase(image_bytes: bytes, filename: str):
    """Upload a scan to Supabase company-files/scanner/ and return a 1-year signed URL."""
    import time as _t
    path = f"scanner/{int(_t.time())}_{filename}"
    # Upload
    up_url = f"{SUPABASE_URL}/storage/v1/object/company-files/{path}"
    up_req = urllib.request.Request(up_url, data=image_bytes, method="POST")
    up_req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
    up_req.add_header("Content-Type", "image/jpeg")
    try:
        with urllib.request.urlopen(up_req, context=ssl_ctx(), timeout=30):
            pass
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"Upload failed {e.code}: {e.read().decode()[:200]}")
    # Signed URL (1 year = 31 536 000 s)
    sign_url = f"{SUPABASE_URL}/storage/v1/object/sign/company-files/{path}"
    sign_payload = json.dumps({"expiresIn": 31536000}).encode()
    sign_req = urllib.request.Request(sign_url, data=sign_payload, method="POST")
    sign_req.add_header("Authorization", f"Bearer {SUPABASE_KEY}")
    sign_req.add_header("Content-Type", "application/json")
    with urllib.request.urlopen(sign_req, context=ssl_ctx(), timeout=15) as r:
        data = json.loads(r.read().decode())
    signed = data.get("signedURL") or data.get("signedUrl") or ""
    if not signed:
        raise RuntimeError("No signed URL returned from Supabase")
    # Make absolute if relative
    if signed.startswith("/"):
        signed = SUPABASE_URL + signed
    return signed, path

def _scan_shipping(price: float) -> str:
    SHIP_FREE   = "Calculated: US_eBayStandardEnvelope free, 2 bu (315021080021)"
    SHIP_STD    = "Flat: US_eBayStandardEnvelope $.99, 2 busines (315934471021)"
    SHIP_GROUND = "$15 to $50 Ground"
    if abs(price - 2.49) < 0.005:
        return SHIP_FREE
    if price < 20:
        return SHIP_STD
    return SHIP_GROUND

def _scan_title(player, set_name, number, grade, variant=""):
    player_up = (player or "").upper()
    year_m = re.search(r"\b(19|20)\d{2}\b", set_name or "")
    year = year_m.group() if year_m else ""
    set_clean = re.sub(r"^\d{4}\s*", "", set_name or "").strip()
    var = variant if variant and variant.lower() not in ("base", "base set", "") else ""
    parts = [p for p in [player_up, year, set_clean,
                          f"#{number}" if number else "", var, grade] if p]
    return " ".join(parts)[:80]

def _scan_description(title: str, front_url: str = "") -> str:
    img_block = (
        f'<div style="text-align:center;margin:20px 0;">'
        f'<img src="{front_url}" alt="{title}" style="max-width:400px;width:100%;border:1px solid #ddd;border-radius:4px;">'
        f'</div>'
    ) if front_url else ""
    return (
        '<div style="font-family:Arial,sans-serif;color:#1f2937;line-height:1.6;max-width:600px;margin:0 auto;">'
        f'<h2 style="margin:0 0 12px;font-size:22px;text-transform:uppercase;text-align:center;">{title}</h2>'
        f'{img_block}'
        '<ol style="margin:0 0 20px;padding-left:20px;">'
        '<li style="margin-bottom:10px;">All cards are scanned. If you see any lines or want more pictures, just ask!</li>'
        '<li style="margin-bottom:10px;">Cards over $20 ship with tracking. Memorabilia/patch cards ship Ground Advantage to prevent damage.</li>'
        '<li style="margin-bottom:10px;">Cards $75–$199 ship with signature required — you\'re protected as the buyer.</li>'
        '</ol>'
        '<p style="margin:0;">My goal is to have you receive the card in tip-top shape. No Returns Accepted. Questions? Please ask!!!</p>'
        '</div>'
    )

# ── Cross-platform reconciliation (CollX ↔ eBay) ─────────────────────────────
def _rec_num(x):
    try:
        return float(str(x).replace("$", "").replace(",", "").strip() or 0) or None
    except Exception:
        return None

def _rec_norm(s):
    return re.sub(r"[^a-z0-9 ]", " ", (s or "").lower())

def _rec_numpart(n):
    m = re.search(r"\d+", n or "")
    return m.group() if m else None

def _rec_header_row(rows, needle):
    """eBay reports sometimes have a preamble row — find the real header."""
    for i, r in enumerate(rows[:8]):
        if any(needle in (c or "").lower() for c in r):
            return i
    return 0

def parse_collx_csv(text):
    return list(csv.DictReader(io.StringIO(text)))

def parse_ebay_sold_csv(text):
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return []
    hi = _rec_header_row(rows, "item title")
    col = {c: j for j, c in enumerate(rows[hi])}
    def g(r, name):
        j = col.get(name)
        return r[j] if (j is not None and j < len(r)) else ""
    out = []
    for r in rows[hi + 1:]:
        t = g(r, "Item Title").strip()
        if not t:
            continue
        out.append({"title": t, "sold_for": g(r, "Sold For"),
                    "date": g(r, "Sale Date"), "item_number": g(r, "Item Number")})
    return out

def parse_ebay_active_csv(text):
    rows = list(csv.reader(io.StringIO(text)))
    if not rows:
        return []
    hi = _rec_header_row(rows, "item number")
    col = {c: j for j, c in enumerate(rows[hi])}
    def g(r, name):
        j = col.get(name)
        return r[j] if (j is not None and j < len(r)) else ""
    out = []
    for r in rows[hi + 1:]:
        t = g(r, "Title").strip()
        itemno = g(r, "Item number").strip()
        if not t or not itemno:
            continue
        out.append({"title": t, "item_number": itemno,
                    "sku": g(r, "Custom label (SKU)"), "price": g(r, "Current price")})
    return out

def match_collx_to_ebay(collx_rows, ebay_items):
    """Join each CollX card to its best eBay item by name+year+number.
    Returns list of (collx_card, confidence, ebay_item)."""
    idx = collections.defaultdict(list)
    prepped = []
    for e in ebay_items:
        t = _rec_norm(e.get("title", ""))
        ts = set(t.split())
        prepped.append((t, ts, e))
        for w in ts:
            if len(w) > 2:
                idx[w].append(len(prepped) - 1)
    out = []
    for c in collx_rows:
        parts = _rec_norm(c.get("name", "")).split()
        if not parts:
            continue
        last = parts[-1]
        year = (c.get("year", "") or "").strip()
        yr2 = year[-2:] if len(year) >= 2 else year
        numv = _rec_numpart(c.get("number", ""))
        best = None
        for i in set(idx.get(last, [])):
            t, ts, e = prepped[i]
            if last not in ts:
                continue
            year_ok = (year and year in t) or (yr2 and yr2 in ts)
            num_ok = (numv in ts) if numv else False
            first_ok = parts[0] in ts if len(parts) > 1 else True
            if year_ok and num_ok and first_ok:
                conf = "High"
            elif year_ok and num_ok:
                conf = "Medium"
            elif year_ok and first_ok:
                conf = "Low"
            else:
                continue
            best = (conf, e)
            if conf == "High":
                break
        if best:
            out.append((c, best[0], best[1]))
    return out

def ebay_end_file(item_numbers):
    """eBay File Exchange 'End Listings' CSV text for bulk-ending listings."""
    lines = ["Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8),ItemID,EndCode"]
    for n in item_numbers:
        lines.append(f"End,{n},NotAvailable")
    return "\n".join(lines) + "\n"

def render_price_trend(card_id, key_prefix="pt", default_grade="PSA 10"):
    """Sold-price line chart with 7/30/60/90-day windows + sales-volume context.

    CardHedger only has SOLD data (no active-listing counts), so 'how many are
    for sale' is shown as n/a on purpose rather than faked.
    """
    if not card_id:
        return
    st.markdown("#### 📈 Sold Price Trend")
    gsel, wsel = st.columns([1, 2])
    grades = ["PSA 10", "PSA 9", "Raw"]
    gi = grades.index(default_grade) if default_grade in grades else 0
    grade_sel = gsel.selectbox("Grade", grades, index=gi, key=f"{key_prefix}_grade")
    window = wsel.radio("Window", ["7d", "30d", "60d", "90d"], index=3,
                        horizontal=True, key=f"{key_prefix}_win")
    days_map = {"7d": 7, "30d": 30, "60d": 60, "90d": 90}

    hist = ch_price_history(card_id, grade_sel, days=90)
    pts = hist.get("prices") if isinstance(hist, dict) else hist
    rows = []
    for p in (pts or []):
        try:
            dt = (p.get("closing_date") or p.get("date") or "")[:10]
            val = float(p.get("price"))
            if dt and val > 0:
                rows.append((dt, val))
        except Exception:
            pass
    if not rows:
        st.info(f"No sold-price history for {grade_sel} on this card.")
        return
    df = pd.DataFrame(rows, columns=["date", "price"]).drop_duplicates("date")
    df["date"] = pd.to_datetime(df["date"])
    df = df.sort_values("date")

    latest = df["price"].iloc[-1]

    def _pct(days):
        cutoff = df["date"].max() - pd.Timedelta(days=days)
        win = df[df["date"] >= cutoff]
        if len(win) < 2:
            return None
        a = win["price"].iloc[0]
        return ((win["price"].iloc[-1] - a) / a * 100) if a else None

    st.metric(f"{grade_sel} — latest sold", f"${latest:,.2f}")
    cc = st.columns(4)
    for col, w in zip(cc, ["7d", "30d", "60d", "90d"]):
        pct = _pct(days_map[w])
        if pct is None:
            col.metric(f"Δ {w}", "—")
        else:
            arrow = "📈" if pct > 1 else "📉" if pct < -1 else "➡️"
            col.metric(f"Δ {w}", f"{pct:+.1f}% {arrow}")

    cutoff = df["date"].max() - pd.Timedelta(days=days_map[window])
    dfw = df[df["date"] >= cutoff].set_index("date")
    st.line_chart(dfw["price"], height=240)
    st.caption(
        f"Each point = a daily sold price (CardHedger). Showing last {window}. "
        "Short windows show momentum; 90d shows the longer-term trend."
    )

    meta = ch_card_meta(card_id)
    s7, s30 = meta.get("7 Day Sales"), meta.get("30 Day Sales")
    vc1, vc2, vc3 = st.columns(3)
    vc1.metric("🔁 Sales · 7 days", f"{int(s7):,}" if isinstance(s7, (int, float)) else "—",
               help="How many of this card sold in the last 7 days. Higher = more liquid, easier to sell.")
    vc2.metric("🔁 Sales · 30 days", f"{int(s30):,}" if isinstance(s30, (int, float)) else "—",
               help="30-day sales volume — a fuller read on demand.")
    vc3.metric("🛒 Active for sale", "n/a",
               help="CardHedger tracks SOLD sales only, not live listings — 'how many are for sale right now' isn't available from this API.")

# ── ESPN free player news + injuries (no API key) ────────────────────────────
ESPN_SPORT = {
    "Baseball":   ("baseball", "mlb"),
    "Basketball": ("basketball", "nba"),
    "Football":   ("football", "nfl"),
    "Hockey":     ("hockey", "nhl"),
}
_ESPN_UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

def _espn_get(url):
    try:
        req = urllib.request.Request(url, headers={"User-Agent": _ESPN_UA, "Accept": "application/json"})
        with urllib.request.urlopen(req, context=ssl_ctx(), timeout=12) as r:
            return json.loads(r.read().decode())
    except Exception:
        return None

@st.cache_data(ttl=1800, show_spinner=False)
def espn_injuries_map(sport, league):
    """Whole-league injury report → {player_name_lower: {status,type,comment,team}}."""
    d = _espn_get(f"https://site.api.espn.com/apis/site/v2/sports/{sport}/{league}/injuries")
    out = {}
    for team in (d or {}).get("injuries", []):
        for inj in team.get("injuries", []):
            nm = ((inj.get("athlete") or {}).get("displayName") or "").strip().lower()
            if not nm:
                continue
            typ = inj.get("type")
            typ = typ.get("description") if isinstance(typ, dict) else typ
            out[nm] = {
                "status":  inj.get("status") or "",
                "type":    typ or "",
                "comment": inj.get("shortComment") or inj.get("longComment") or "",
                "team":    team.get("displayName", ""),
            }
    return out

@st.cache_data(ttl=900, show_spinner=False)
def espn_player_news(player, limit=4):
    """Recent ESPN articles matching a player name → [{title,url,date}]."""
    if not player:
        return []
    d = _espn_get(f"https://site.web.api.espn.com/apis/search/v2?query={urllib.parse.quote(player)}&limit={limit}")
    out = []
    for g in (d or {}).get("results", []):
        if g.get("type") == "article":
            for it in g.get("contents", []):
                title = it.get("displayName") or ""
                link = (it.get("link") or {}).get("web") if isinstance(it.get("link"), dict) else ""
                if title and link:
                    out.append({"title": title, "url": link, "date": (it.get("date") or "")[:10]})
            break
    # Prefer headlines that actually name the player (drops generic league noise).
    last = player.lower().split()[-1] if player else ""
    preferred = [a for a in out if last and last in a["title"].lower()]
    return (preferred or out)[:limit]

def render_player_watch(card_id, key_prefix="pw"):
    """News + injury status for the card's player, via the free ESPN feeds."""
    meta = ch_card_meta(card_id) if card_id else {}
    player = (meta.get("player") or "").strip()
    if not player:
        return
    category = meta.get("category") or ""
    st.markdown("#### 📰 Player Watch — news & injury")

    sport = ESPN_SPORT.get(category.title())
    injury = None
    if sport:
        inj_map = espn_injuries_map(*sport)
        injury = inj_map.get(player.lower())
        if not injury and " " in player:            # last-name fallback
            last = player.lower().split()[-1]
            for nm, v in inj_map.items():
                if nm.split()[-1] == last:
                    injury = v
                    break
    if injury:
        msg = f"🔴 **{player} — {injury['status']}**"
        if injury["type"]:
            msg += f" · {injury['type']}"
        if injury["comment"]:
            msg += f"  \n{injury['comment']}"
        st.error(msg)
    elif sport:
        st.success(f"🟢 {player} — not on the {sport[1].upper()} injury report right now.")

    news = espn_player_news(player)
    if news:
        for n in news:
            line = f"- [{n['title']}]({n['url']})"
            if n["date"]:
                line += f" · _{n['date']}_"
            st.markdown(line)
    else:
        srch = urllib.parse.quote(player)
        st.caption(f"No recent ESPN articles matched **{player}** (common for prospects). "
                   f"[Search ESPN ↗](https://www.espn.com/search/_/q/{srch})")
    st.caption("⚡ News & injuries move card prices before the comps do. Source: ESPN (free).")

def render_player_demand(card_id, key_prefix="pd"):
    """Weekly sold-volume trend across ALL of a player's cards — a demand signal."""
    meta = ch_card_meta(card_id) if card_id else {}
    player = (meta.get("player") or "").strip()
    if not player:
        return
    buckets = ch_sales_stats(player, interval="week", periods=8)
    rows = []
    for b in buckets:
        wk = (b.get("start") or "")[:10]
        try:
            rows.append({"week": wk, "sales": int(b.get("count") or 0),
                         "avg": float(b.get("average_sale") or 0)})
        except Exception:
            pass
    if len(rows) < 2:
        return
    df = pd.DataFrame(rows).sort_values("week")
    st.markdown(f"#### 📊 Player Demand — {player} (all cards)")
    st.bar_chart(df.set_index("week")["sales"], height=200)

    # Demand direction: recent half vs older half of the weekly sales counts.
    vals = df["sales"].tolist()
    mid = len(vals) // 2
    older = sum(vals[:mid]) / max(mid, 1)
    recent = sum(vals[mid:]) / max(len(vals) - mid, 1)
    pct = ((recent - older) / older * 100) if older else 0
    arrow = "📈 rising" if pct > 10 else "📉 cooling" if pct < -10 else "➡️ steady"
    m1, m2, m3 = st.columns(3)
    m1.metric("Latest week — sales", f"{vals[-1]:,}")
    m2.metric("Avg sale (latest wk)", f"${df['avg'].iloc[-1]:,.0f}")
    m3.metric("Demand trend", arrow, f"{pct:+.0f}%")
    st.caption(
        f"Weekly count of ALL {player} cards sold across marketplaces (CardHedger). "
        "Rising volume = growing demand/attention — a leading signal for where prices head."
    )

def safe_image_url(u):
    """Return a fetchable http(s) URL (normalizing protocol-relative), else None.

    st.image crashes the whole app if handed a non-URL string (it tries to open
    it as a local file → MediaFileStorageError), so anything that isn't a real
    web URL must be filtered out before it reaches st.image.
    """
    if not isinstance(u, str):
        return None
    u = u.strip()
    if u.startswith("//"):
        u = "https:" + u
    return u if u.lower().startswith(("http://", "https://")) else None

def render_card_image(url, placeholder=True, **kwargs):
    """Safely render a card image. Returns True if shown, False if not (bad/missing URL)."""
    url = safe_image_url(url)
    if url:
        try:
            st.image(url, **kwargs)
            return True
        except Exception:
            pass
    if placeholder:
        st.markdown(
            '<div style="background:#1e2130;border:1px solid #2e3250;border-radius:10px;'
            'height:180px;display:flex;align-items:center;justify-content:center;'
            'color:#4a5568;font-size:2rem;">🃏</div>',
            unsafe_allow_html=True,
        )
    return False

def _extract_price(p):
    for k in ("price", "sale_price", "sold_price", "value", "amount"):
        try:
            v = float(p[k])
            if v > 0:
                return v
        except Exception:
            pass
    return None

def calculate_trend(history_data):
    """Return (direction, pct_change) from CardHedger price history."""
    prices = []
    if isinstance(history_data, list):
        prices = history_data
    elif isinstance(history_data, dict):
        for key in ("prices", "data", "sales", "history", "results"):
            if key in history_data and isinstance(history_data[key], list):
                prices = history_data[key]
                break
    vals = [_extract_price(p) for p in prices]
    vals = [v for v in vals if v]
    if len(vals) < 4:
        return None, 0.0
    mid = len(vals) // 2
    older_avg = sum(vals[:mid]) / mid
    recent_avg = sum(vals[mid:]) / (len(vals) - mid)
    if older_avg == 0:
        return None, 0.0
    pct = ((recent_avg - older_avg) / older_avg) * 100
    direction = "up" if pct > 5 else "down" if pct < -5 else "flat"
    return direction, pct

def trend_badge(direction, pct):
    if direction == "up":
        return f"📈 Trending Up +{pct:.0f}%"
    elif direction == "down":
        return f"📉 Trending Down {pct:.0f}%"
    return f"➡️ Flat ({pct:+.0f}%)"

# ─── Quick search target list ─────────────────────────────────────────────────
QUICK_SEARCHES = [
    ("🏀", "Cooper Flagg Topps Chrome Rookie"),
    ("🏀", "Steph Curry Topps Chrome Paradox Refractor"),
    ("🏀", "Caitlin Clark Panini Prizm WNBA Rookie"),
    ("🏀", "Luka Doncic Panini Prizm Silver Rookie"),
    ("🏀", "Zion Williamson Panini Prizm Rookie"),
    ("🏈", "Cam Ward Panini Prizm Draft Picks Rookie"),
    ("🏈", "Ashton Jeanty Panini Prizm Draft Picks Rookie"),
    ("🏈", "Travis Hunter Panini Prizm Draft Picks Rookie"),
    ("🏈", "Drake Maye Panini Prizm Rookie"),
    ("🏈", "Shedeur Sanders Panini Prizm Draft Picks Rookie"),
    ("⚾", "Paul Skenes Topps Chrome Rookie"),
    ("⚾", "Aaron Judge Topps"),
    ("⚾", "Elly De La Cruz Bowman Chrome Rookie"),
]

# ─── Gem rate visual helpers ──────────────────────────────────────────────────
def gem_color_hex(g):
    """Return a hex color for a gem rate percentage."""
    if g is None:
        return "#94a3b8"
    if g >= 60:
        return "#22c55e"   # green
    if g >= 35:
        return "#f59e0b"   # yellow
    return "#ef4444"       # red

def gem_bar_html(g):
    """Return an HTML progress bar for embedding in st.markdown."""
    if g is None:
        return "—"
    color = gem_color_hex(g)
    pct = min(g, 100)
    return (
        f'<div style="display:flex;align-items:center;gap:8px">'
        f'<span style="color:{color};font-weight:700;font-size:15px;min-width:42px">{g:.1f}%</span>'
        f'<div style="flex:1;background:#22263a;border-radius:4px;height:8px;min-width:80px">'
        f'<div style="width:{pct}%;height:8px;border-radius:4px;background:{color}"></div>'
        f'</div></div>'
    )

# ─── URL builders ─────────────────────────────────────────────────────────────
def gemrate_url(gid):
    return f"https://www.gemrate.com/card/{gid}"

def ebay_raw_url(desc):
    q = urllib.parse.quote_plus(desc + " raw")
    return f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_Sold=1&LH_Complete=1&LH_BIN=1&_sop=13"

def ebay_raw_sold_all_url(desc):
    """Raw sold comps — all sale types (BIN + auction + offers)."""
    q = urllib.parse.quote_plus(desc)
    return f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_Sold=1&LH_Complete=1&_sop=13&_sacat=212"

def ebay_buy_bin_url(desc):
    """Buy raw — BIN listings, cheapest first."""
    q = urllib.parse.quote_plus(desc)
    return f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_BIN=1&_sop=15&_sacat=212"

def ebay_buy_auction_url(desc):
    """Buy raw — auctions ending soonest."""
    q = urllib.parse.quote_plus(desc)
    return f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_Auction=1&_sop=1&_sacat=212"

def ebay_graded_sold_url(desc):
    q = urllib.parse.quote_plus(desc + " PSA 10")
    return f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_Sold=1&LH_Complete=1&LH_BIN=1&_sop=13"

def ebay_psa9_sold_url(desc):
    q = urllib.parse.quote_plus(desc + " PSA 9")
    return f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_Sold=1&LH_Complete=1&_sop=13&_sacat=212"

def ebay_graded_buy_url(desc):
    q = urllib.parse.quote_plus(desc + " PSA 10")
    return f"https://www.ebay.com/sch/i.html?_nkw={q}&LH_BIN=1&_sop=15"

def psa_pop_url(desc):
    q = urllib.parse.quote_plus(desc)
    return f"https://www.psacard.com/pop/search?q={q}"

# ─── ROI logic ────────────────────────────────────────────────────────────────
def calc_opp_cost(raw_cost, psa_tier, opp_rate: float, ship_cost: float = 0.0) -> float:
    """Return the opportunity cost of having money locked up at PSA.

    All upfront cash (card + grading fee + shipping) is locked up during the wait,
    so all of it earns zero return for that period.
    opp_rate: annual return % you expect elsewhere (e.g. 12.0 = 12%/yr)
    """
    if opp_rate <= 0:
        return 0.0
    fee = PSA_FEES.get(psa_tier, 50)
    biz_days = PSA_DAYS.get(psa_tier, 60)
    cal_days = biz_days * 1.4
    capital_at_risk = raw_cost + fee + ship_cost   # everything paid upfront is locked
    return round(capital_at_risk * (opp_rate / 100) * (cal_days / 365), 2)

def target_price(raw_cost, psa_tier, roi=4.0, opp_rate=0.0, ship_cost=0.0):
    fee = PSA_FEES.get(psa_tier, 50)
    opp = calc_opp_cost(raw_cost, psa_tier, opp_rate, ship_cost)
    total_cost = raw_cost + fee + ship_cost + opp
    return (total_cost * roi) / (1 - EBAY_FEE)

def calc_net_roi(raw_cost, psa_tier, graded_price, opp_rate=0.0, ship_cost=0.0):
    fee = PSA_FEES.get(psa_tier, 50)
    opp = calc_opp_cost(raw_cost, psa_tier, opp_rate, ship_cost)
    total_cost = raw_cost + fee + ship_cost + opp
    net = graded_price * (1 - EBAY_FEE) - total_cost
    roi = (net / total_cost * 100) if total_cost > 0 else 0
    return round(net, 2), round(roi, 1)

def verdict(raw_cost, psa_tier, gem_rate, graded_price, min_gem, roi_target, opp_rate=0.0, ship_cost=0.0):
    if gem_rate is None or gem_rate < min_gem:
        return "❌ NO-GO", "red", f"Gem rate {gem_rate or 0:.1f}% below {min_gem:.0f}% floor"
    tgt = target_price(raw_cost, psa_tier, roi_target, opp_rate, ship_cost)
    net, roi = calc_net_roi(raw_cost, psa_tier, graded_price, opp_rate, ship_cost)
    if graded_price >= tgt:
        return "✅ GO", "green", f"Gem 10 avg ${graded_price:,.0f} clears ${tgt:,.0f} target | Net ~${net:,.0f} | ROI ~{roi:.0f}%"
    return "❌ NO-GO", "red", f"Gem 10 avg ${graded_price:,.0f} needs ${tgt:,.0f} for {roi_target:.0f}× | ROI only ~{roi:.0f}%"

# ─── Grade-vs-Flip decision logic (the holding-cost engine) ───────────────────
def hold_cost(capital, cal_days, opp_rate):
    """Opportunity cost of `capital` locked up for cal_days at opp_rate %/yr."""
    if opp_rate <= 0 or capital <= 0:
        return 0.0
    return capital * (opp_rate / 100.0) * (cal_days / 365.0)

def flip_raw_net(raw_comp, raw_buy):
    """Net from selling the card raw right now (after eBay fees)."""
    if not raw_comp:
        return None
    return raw_comp * (1 - EBAY_FEE) - raw_buy

def grade_sale_net(sale_price, raw_buy, fee, ship, hold):
    """Net from grading then selling at sale_price, after every cost incl. holding."""
    if not sale_price:
        return None
    return sale_price * (1 - EBAY_FEE) - (raw_buy + fee + ship + hold)

def expected_graded_sale(gem_rate, psa10, psa9):
    """Gem-rate-weighted expected graded sale: P(10)*PSA10 + (1-P(10))*PSA9."""
    if psa10 is None and psa9 is None:
        return None
    p10 = max(0.0, min(1.0, (gem_rate or 0) / 100.0))
    s10 = psa10 if psa10 is not None else psa9
    s9 = psa9 if psa9 is not None else psa10
    return p10 * s10 + (1 - p10) * s9

def grade_vs_flip(raw_buy, raw_comp, psa10, psa9, gem_rate, tier, ship, opp_rate):
    """Full grade-vs-flip economics for one card. Returns a dict of outcomes."""
    fee = PSA_FEES.get(tier, 50)
    cal_days = int(PSA_DAYS.get(tier, 60) * 1.4)
    capital = raw_buy + fee + ship
    hold = round(hold_cost(capital, cal_days, opp_rate), 2)
    raw_net = flip_raw_net(raw_comp, raw_buy)
    net10 = grade_sale_net(psa10, raw_buy, fee, ship, hold)
    net9 = grade_sale_net(psa9, raw_buy, fee, ship, hold)
    exp_sale = expected_graded_sale(gem_rate, psa10, psa9)
    net_exp = grade_sale_net(exp_sale, raw_buy, fee, ship, hold)
    premium = (net_exp - raw_net) if (net_exp is not None and raw_net is not None) else None
    return {
        "fee": fee, "cal_days": cal_days, "capital": capital, "hold": hold,
        "raw_net": raw_net, "net10": net10, "net9": net9,
        "exp_sale": exp_sale, "net_exp": net_exp, "premium": premium,
        "p10": max(0.0, min(1.0, (gem_rate or 0) / 100.0)),
    }

def grade_flip_verdict(d):
    """Return (label, color, message) recommending grade vs flip for one card."""
    net_exp, raw_net, net9 = d["net_exp"], d["raw_net"], d["net9"]
    if net_exp is None:
        return "ℹ️ Need comps", "info", "Not enough comp data to compare — enter prices manually."
    downside = ""
    if net9 is not None and net9 < 0:
        downside = f" ⚠️ Downside: a PSA 9 nets **−${abs(net9):,.0f}** (a loss)."
    if raw_net is None:
        if net_exp > 0:
            return "✅ GRADE", "green", f"Expected net **${net_exp:,.0f}** after the ${d['hold']:,.0f} holding cost.{downside}"
        return "❌ SKIP", "red", f"Expected net only **${net_exp:,.0f}** — not worth grading.{downside}"
    if net_exp > raw_net and net_exp > 0:
        return "✅ GRADE", "green", (
            f"Expected **${net_exp:,.0f}** vs **${raw_net:,.0f}** flipping raw — "
            f"**+${d['premium']:,.0f}** for the ~{d['cal_days']}-day wait (after ${d['hold']:,.0f} holding cost).{downside}"
        )
    # Neither path makes money at this buy price → don't buy it.
    if raw_net <= 0 and net_exp <= 0:
        return "❌ PASS", "red", (
            f"Both paths lose at this cost: flipping raw nets **${raw_net:,.0f}**, "
            f"grading's expected is **${net_exp:,.0f}**. Skip it (or pay less).{downside}"
        )
    return "💵 FLIP RAW", "amber", (
        f"Flipping raw nets **${raw_net:,.0f}** now; grading's expected **${net_exp:,.0f}** "
        f"isn't worth locking **${d['capital']:,.0f}** for ~{d['cal_days']} days.{downside}"
    )

def fmt_gem(g):
    if g is None:
        return "—"
    return "<0.1%" if g < 0.05 else f"{g:.1f}%"

def gem_signal(g):
    if g is None or g < 40:
        return "🔴"
    if g < 60:
        return "🟡"
    return "🟢"

# ─── Sidebar ──────────────────────────────────────────────────────────────────
is_beta = st.session_state.get("is_beta", False)

# Settings — read from session_state so widgets (defined below) persist across reruns
roi_target   = st.session_state.get("m_roi",      4.0)
min_gem      = st.session_state.get("m_mingem",  40.0)
default_tier = st.session_state.get("m_tier",    list(PSA_FEES.keys())[0])
opp_rate     = st.session_state.get("m_opp",     12.0)
ship_cost    = round(st.session_state.get("m_ship_to", 0.0) + st.session_state.get("m_ship_ret", 0.0), 2)

with st.sidebar:
    # ── Branding ──────────────────────────────────────────────────────────────
    st.markdown(
        f"""
        <div style="padding:14px 4px 10px 4px;">
          <div style="font-size:1.25rem;font-weight:800;letter-spacing:-0.5px;color:#e2e8f0;">
            {APP_NAME}
          </div>
          <div style="font-size:0.72rem;color:#94a3b8;margin-top:3px;font-style:italic;">
            {APP_TAGLINE}
          </div>
          <div style="font-size:0.65rem;color:#475569;margin-top:5px;">v{APP_VERSION}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    access_name = st.session_state.get("access_name", "")
    if access_name:
        if is_beta:
            st.markdown(f"👤 **{access_name}** &nbsp; `BETA`")
            _exp = st.session_state.get("trial_expires_label", "")
            st.info(f"🔓 Beta Preview unlocked.{' ' + _exp if _exp else ''}", icon="ℹ️")
        else:
            st.markdown(
                f'<div style="font-size:0.8rem;color:#94a3b8;padding:2px 0 10px 0;">👤 {access_name}</div>',
                unsafe_allow_html=True,
            )

    st.markdown("---")

    # ── Features ──────────────────────────────────────────────────────────────
    st.markdown(
        '<div style="font-size:0.7rem;font-weight:700;letter-spacing:0.08em;'
        'color:#64748b;text-transform:uppercase;margin-bottom:8px;">Features</div>',
        unsafe_allow_html=True,
    )

    # Tab index = position in the st.tabs() list defined below
    _FEATURES = [
        ("🔍", "Card Research",    "FMV · trend · buy signal · comps",         0),
        ("🔥", "Hot Movers",       "What the market is chasing right now",      1),
        ("📷", "Scan",             "Photo ID · AI batch · graded slab",         2),
        ("📦", "Batch → eBay",     "Drip-scheduled CSV export with photos",     2),
        ("📦", "Inventory Check",  "Grading ROI · gem rate · break-even",       3),
        ("🧰", "Operations",       "Reprice queue · Sunday workflow",            4),
        ("📬", "Submissions",      "PSA order tracker · cert lookup",            5),
        ("💰", "Sales & P&L",      "eBay / CollX revenue · monthly P&L",        10),
        ("🏷️", "Consignments",    "Track cards you're selling for others",      8),
        ("📸", "Image Prep",       "Corner crops · eBay-ready photo sets",       11),
    ]

    for _icon, _name, _desc, _tidx in _FEATURES:
        if st.button(f"{_icon}  {_name}", key=f"feat_{_name}", use_container_width=True, help=_desc):
            st.session_state["_goto_tab"] = _tidx

    st.markdown("---")

    # ── Guide button ──────────────────────────────────────────────────────────
    if st.button("📖 How to Maximize This App", use_container_width=True, key="open_guide_btn"):
        st.session_state["show_guide"] = True

    # ── Admin (Duane only) ────────────────────────────────────────────────────
    if st.session_state.get("access_name") == "Duane":
        st.markdown("---")
        if st.button("🔐 Admin Panel", use_container_width=True, key="open_admin_btn"):
            st.session_state["show_admin"] = True

    st.markdown("---")
    st.caption(f"v{APP_VERSION} · Settings ⚙️ in main area below")
# ─── App Guide dialog ─────────────────────────────────────────────────────────
@st.dialog(f"📖 How to Maximize {APP_NAME}", width="large")
def _show_guide():
    st.markdown("### Get the most out of every feature")
    _GUIDE_SECTIONS = [
        ("🔍", "Card Research — your first stop",
         [
             "Type any player, year, set, or parallel into the search bar — no image needed.",
             "Check the **Buy Signal**: 🔥 HOT means rising demand; 🛑 COOLING means the market is softening. Time your sells accordingly.",
             "Use the **Sell-Through %** (green/yellow/red) to see how fast copies are actually moving on eBay.",
             "Tap **📋 Copy Comps** to paste recent sold prices directly into a message or spreadsheet.",
         ]),
        ("📷", "Scanning & AI Batch — raw cards",
         [
             "Drop front + back images interleaved (front1, back1, front2, back2…) into the **📦 Batch → eBay** tab.",
             "Hit **Identify & Price All** — CardHedger visually matches each card and pulls FMV in one shot.",
             "Edit the **Price ($)** column directly in the table before exporting. FMV is a starting point, not a floor.",
             "Toggle **Drip Schedule** to stagger listings hourly — eBay rewards consistent daily activity with better placement.",
             "Export once → upload once to eBay Seller Hub. No manual listing needed.",
         ]),
        ("🎫", "Graded Slabs — PSA / BGS / SGC",
         [
             "Go to **📷 Scan → 🎫 Graded slab (cert #)** and enter your cert number.",
             "You get price history, trend direction, and a buy signal for that exact graded copy.",
             "Hit **➕ Add to eBay Batch** — the card lands in the export table with grade, grader, and cert # pre-filled. No scanning needed.",
             "The eBay CSV automatically populates the CD: grader, grade, and cert fields so eBay shows the slab details correctly.",
         ]),
        ("📦", "Inventory Check — grading decisions",
         [
             "Upload your card list or scan a card to run the **Grading ROI calculator**.",
             "Set your **ROI target** and **Min gem rate** in ⚙️ Settings (sidebar) to filter GO / NO-GO automatically.",
             "A card that pencils out at 4× ROI but has a 10% gem rate is a coin flip — the gem rate matters as much as the math.",
             "Use **Opportunity Cost** in Settings to see the true cost of cash locked at PSA for 30–90 days.",
         ]),
        ("🧰", "Operations — weekly workflow",
         [
             "Every Sunday: go to **🧰 Operations → 📅 Sunday Reprice**. Upload your eBay active listings CSV.",
             "Filter to 7–30 day listings that haven't sold. CardHedger pulls fresh FMV on each.",
             "Download the bulk-edit CSV → upload to eBay Seller Hub to reprice the whole batch in one shot.",
             "Consistent repricing = fewer stale listings and faster inventory turns.",
         ]),
        ("💰", "Sales & P&L — track what matters",
         [
             "Import your eBay sold CSV and CollX sales under **💰 Sales & P&L → Import**.",
             "The monthly P&L breaks down revenue by channel (eBay vs CollX vs Whatnot).",
             "Match SKU prefixes between your listings and sales to track per-lot profitability in **📦 Purchases**.",
         ]),
        ("⚙️", "Pro tips",
         [
             "Set a **SKU prefix** (e.g. your initials) in ⚙️ Settings so every eBay listing has a traceable ID.",
             "Use **Best Offer** with a min floor — it catches buyers who won't pay BIN but will negotiate.",
             "Drip 8–12 cards/day max. Flooding eBay with 40 at once suppresses your own listings.",
             "Check **🔥 Hot Movers** before scanning a new lot — if a player is trending up, price up and hold a day.",
         ]),
    ]
    for _gi, (_icon, _title, _points) in enumerate(_GUIDE_SECTIONS):
        st.markdown(
            f'<div style="background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,0.08);'
            f'border-radius:10px;padding:14px 18px;margin-bottom:12px;">'
            f'<div style="font-size:1rem;font-weight:700;color:#e2e8f0;margin-bottom:8px;">'
            f'{_icon} {_title}</div>'
            + "".join(
                f'<div style="display:flex;gap:8px;align-items:flex-start;margin-bottom:6px;">'
                f'<span style="color:#6366f1;font-size:0.75rem;margin-top:3px;">▸</span>'
                f'<span style="font-size:0.82rem;color:#94a3b8;line-height:1.5;">{p}</span></div>'
                for p in _points
            )
            + "</div>",
            unsafe_allow_html=True,
        )

# ─── Admin dialog (Duane only) ────────────────────────────────────────────────
@st.dialog(f"🔐 {APP_NAME} — Admin", width="large")
def _show_admin():
    import datetime as _dt_adm
    st.caption(f"v{APP_VERSION} · Access code management")

    with st.expander("🛠 First-time setup SQL (run once in Supabase if daily limits aren't saving)", expanded=False):
        st.code("alter table access_codes add column if not exists daily_limit integer;", language="sql")

    codes = admin_get_codes()
    now_utc = _dt_adm.datetime.utcnow().replace(tzinfo=_dt_adm.timezone.utc)

    st.markdown(f"### 👥 Access Codes ({len(codes)} total)")
    for row in codes:
        c1, c2, c3, c4, c5, c6 = st.columns([2, 2, 1, 1, 1, 2])
        c1.markdown(f"**{row['name']}**")
        c2.code(row["code"])
        c3.markdown("🟢" if row["active"] else "🔴")
        c4.markdown(f"Uses: **{row['usage_count']}**")
        c5.markdown(f"{(row.get('last_used') or 'Never')[:10]}")
        exp_raw = row.get("expires_at")
        if exp_raw:
            try:
                exp_dt   = _dt_adm.datetime.fromisoformat(exp_raw.replace("Z", "+00:00"))
                days_left = (exp_dt - now_utc).days
                if days_left < 0:
                    c6.markdown("🔴 **Expired**")
                elif days_left == 0:
                    c6.markdown("🟡 **Expires today**")
                else:
                    c6.markdown(f"⏳ {days_left}d left ({exp_dt.strftime('%b %d')})")
            except Exception:
                c6.markdown(exp_raw[:10])
        else:
            c6.markdown("♾️ No expiry")

        # Expiry edit row
        _exp_c1, _exp_c2, _exp_c3 = st.columns([2, 1, 1])
        _exp_c1.caption("Set / extend expiry:")
        _exp_days = _exp_c2.number_input(
            "Days from now", min_value=0, max_value=365, value=7,
            key=f"adm_exp_{row['id']}", label_visibility="collapsed",
            help="0 = remove expiry (no limit)"
        )
        if _exp_c3.button("⏳ Set", key=f"adm_exp_save_{row['id']}"):
            admin_set_expiry(row["id"], int(_exp_days))
            _exp_label = f"{_exp_days} days" if _exp_days else "removed"
            st.success(f"Expiry {_exp_label} for {row['name']}")
            st.rerun()

        # Daily limit row
        _dl_cur = row.get("daily_limit") or 0
        _dl_c1, _dl_c2, _dl_c3 = st.columns([2, 1, 1])
        _dl_c1.caption(f"Daily lookup limit: {'**' + str(_dl_cur) + '**/day**' if _dl_cur else 'Global default (' + str(DAILY_PRICING_CAP) + '/day)'}")
        _new_limit = _dl_c2.number_input(
            "Set limit", min_value=0, max_value=500, value=_dl_cur,
            key=f"adm_lim_{row['id']}", label_visibility="collapsed",
            help="0 = use global default. Set per-user daily lookup cap."
        )
        if _dl_c3.button("💾 Save", key=f"adm_lim_save_{row['id']}"):
            admin_set_daily_limit(row["id"], int(_new_limit))
            st.success(f"Limit set to {_new_limit}/day for {row['name']}")
            st.rerun()

        if row["active"]:
            if st.button("Revoke", key=f"adm_rev_{row['id']}"):
                admin_toggle_code(row["id"], False)
                st.rerun()
        else:
            if st.button("Reinstate", key=f"adm_rei_{row['id']}"):
                admin_toggle_code(row["id"], True)
                st.rerun()
        st.markdown("---")

    st.markdown("### ➕ Create New Code")
    n1, n2, n3, n4 = st.columns([2, 2, 1, 1])
    new_name  = n1.text_input("Name", placeholder="e.g. John Smith", key="adm_new_name")
    new_code  = n2.text_input("Code", value=gen_code(), key="adm_new_code")
    trial_days = n3.number_input("Trial days", min_value=0, max_value=365, value=7, key="adm_trial")
    if n4.button("Create", type="primary", use_container_width=True, key="adm_create"):
        if new_name and new_code:
            days = int(trial_days) if trial_days > 0 else None
            if admin_insert_code(new_code.strip().upper(), new_name.strip(), trial_days=days):
                exp_note = f" · expires in {days} days" if days else " · no expiry"
                st.success(f"✅ Created `{new_code.upper()}` for **{new_name}**{exp_note}")
                st.rerun()
            else:
                st.error("Failed — code may already exist.")
        else:
            st.warning("Enter a name and code.")

# ─── What's New dialog ────────────────────────────────────────────────────────
_WN_KEY = f"wn_seen_{APP_VERSION}"

@st.dialog(f"🎉 What's New in {APP_NAME}", width="large")
def _show_whats_new():
    # Mark as seen immediately — so closing via X doesn't re-trigger on next rerun
    st.session_state[_WN_KEY] = True
    notes = RELEASE_NOTES.get(APP_VERSION, {})
    st.markdown(
        f'<div style="display:inline-block;background:#1e2130;border:1px solid #2e3250;'
        f'border-radius:8px;padding:4px 12px;font-size:0.75rem;color:#7c8db5;'
        f'margin-bottom:12px">v{APP_VERSION}</div>',
        unsafe_allow_html=True,
    )
    st.markdown(f"### {notes.get('emoji','🎉')} {notes.get('title','Updates')}")
    st.markdown("")
    for icon, text in notes.get("items", []):
        st.markdown(
            f'<div style="display:flex;gap:10px;align-items:flex-start;'
            f'margin-bottom:10px;padding:10px 14px;background:#1e2130;'
            f'border-radius:8px;border:1px solid #2e3250">'
            f'<span style="font-size:1.2rem;line-height:1.4">{icon}</span>'
            f'<span style="font-size:0.9rem;color:#c8d3e8;line-height:1.5">{text}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
    st.markdown("")
    if st.button("Got it, let's go →", type="primary", use_container_width=True):
        st.session_state[_WN_KEY] = True
        st.rerun()

if not st.session_state.get(_WN_KEY) and APP_VERSION in RELEASE_NOTES:
    _show_whats_new()

if st.session_state.get("show_guide"):
    st.session_state["show_guide"] = False
    _show_guide()

if st.session_state.get("show_admin") and st.session_state.get("access_name") == "Duane":
    st.session_state["show_admin"] = False
    _show_admin()

# ─── Dashboard KPI tiles ──────────────────────────────────────────────────────
if not is_beta and SUPABASE_URL:
    _dash_rows = sb_get()
    _df_d = pd.DataFrame(_dash_rows) if _dash_rows else pd.DataFrame()
    _total   = len(_df_d)
    _go      = len(_df_d[_df_d["go_no_go"].str.startswith("✅", na=False)]) if not _df_d.empty else 0
    _pending = len(_df_d[_df_d["status"] == "Submitted"]) if not _df_d.empty else 0
    _returned= len(_df_d[_df_d["status"].isin(["Received","Sold"])]) if not _df_d.empty else 0
    try:
        _go_mask  = _df_d["go_no_go"].str.startswith("✅", na=False)
        _sub_cost = pd.to_numeric(_df_d.loc[_go_mask, "psa_fee"], errors="coerce").sum()
        _est_net  = pd.to_numeric(_df_d.loc[_go_mask, "est_net"], errors="coerce").sum()
    except Exception:
        _sub_cost = _est_net = 0

    def _kpi(label, value, sub, color):
        return (
            f'<div style="background:#1e2130;border:1px solid #2e3250;border-radius:10px;padding:16px 18px">'
            f'<div style="font-size:0.68rem;color:#7c8db5;text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px">{label}</div>'
            f'<div style="font-size:1.9rem;font-weight:700;color:{color};line-height:1.1">{value}</div>'
            f'<div style="font-size:0.7rem;color:#556;margin-top:5px">{sub}</div>'
            f'</div>'
        )

    st.markdown(
        f'<div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(155px,1fr));gap:12px;margin-bottom:20px">'
        + _kpi("Cards in Tracker",    _total,              "in submission tracker",   "#4f8ef7")
        + _kpi("Flagged: Submit",     _go,                 "pass your ROI threshold", "#22c55e")
        + _kpi("Est. Submission Cost",f"${_sub_cost:,.0f}","for GO cards",            "#f59e0b")
        + _kpi("Est. Net Profit",     f"${_est_net:,.0f}", "if all grade PSA 10",     "#22c55e")
        + _kpi("Pending at PSA",      _pending,            "cards submitted",         "#4f8ef7")
        + _kpi("Grades Returned",     _returned,           "received back",           "#22c55e")
        + _kpi("ROI Threshold",       f"{roi_target:.0f}×","minimum to flag submit",  "#f59e0b")
        + '</div>',
        unsafe_allow_html=True,
    )

# ─── Navigation ───────────────────────────────────────────────────────────────
if is_beta:
    st.info("🔓 **Beta Preview** — You have access to Card Research and Inventory Check. Submission Tracker and Downloads unlock with a full membership.", icon="💎")

# ── Settings (main content — always visible regardless of sidebar) ────────────
with st.expander("⚙️ Settings", expanded=False):
    _s1, _s2, _s3 = st.columns(3)
    with _s1:
        roi_target   = st.number_input("ROI target (×)", min_value=1.0, max_value=20.0, value=4.0, step=0.5, key="m_roi")
        min_gem      = st.number_input("Min gem rate (%)", min_value=0.0, max_value=100.0, value=40.0, step=5.0, key="m_mingem")
    with _s2:
        default_tier = st.selectbox("Default grading tier", list(PSA_FEES.keys()), index=0, key="m_tier")
        opp_rate     = st.slider("Opportunity cost (%/yr)", min_value=0.0, max_value=50.0, value=12.0, step=1.0, key="m_opp",
                                 help="e.g. 12%/yr: $649 locked at Express (~35 days) = $7.47 hidden cost")
    with _s3:
        st.markdown("**📦 Shipping ($/card)**")
        ship_to     = st.number_input("To PSA", min_value=0.0, value=0.0, step=0.50, key="m_ship_to")
        ship_return = st.number_input("Return",  min_value=0.0, value=0.0, step=0.50, key="m_ship_ret")
    ship_cost = round(ship_to + ship_return, 2)
    st.caption(f"PSA fees: " + " · ".join(f"{t.split('(')[0].strip()} ${f['fee']:.0f}" for t, f in PSA_FEES_ALL.items()))

_NAV_LABELS = [
    "🔍 Card Research", "🔥 Hot Movers", "📷 Scan", "📦 Inventory Check",
    "🧰 Operations", "📬 Submission Tracker", "📥 Downloads", "🚚 Shipment Intake",
    "🏷️ Consignments", "📦 Purchases", "💰 Sales & P&L", "📸 Image Prep",
]

# If a sidebar feature button was clicked, update both nav state AND the radio widget's own state key
if "_goto_tab" in st.session_state:
    _target = st.session_state.pop("_goto_tab")
    st.session_state["_nav"] = _target
    st.session_state["_nav_radio"] = _NAV_LABELS[_target]

_active_tab = st.session_state.get("_nav", 0)
_selected_label = st.radio(
    "Section", _NAV_LABELS, index=_active_tab,
    horizontal=True, label_visibility="collapsed", key="_nav_radio",
)
_active_tab = _NAV_LABELS.index(_selected_label)
st.session_state["_nav"] = _active_tab

st.markdown("---")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Card Research
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 0:
    st.markdown("## 🔍 Card Research")
    st.markdown("Search any card — owned or not. Get gem rate, graded value comps, and eBay links.")
    st.caption("Tip: include the full set name for best results — e.g. *Steph Curry Topps Chrome Paradox* not just *Steph Curry Paradox*")

    col_q, col_btn, col_clr = st.columns([5, 1, 1])
    with col_q:
        query = st.text_input("Search", placeholder="e.g.  Curry Topps Chrome Paradox  |  Luka Prizm RC auto  |  Wemby Optic", label_visibility="collapsed")
    with col_btn:
        do_search = st.button("Search", use_container_width=True, type="primary")
    with col_clr:
        if st.button("🔄 Clear", use_container_width=True, help="Clear cached results and retry"):
            st.cache_data.clear()
            st.session_state.pop("gr_results", None)
            st.session_state.pop("ch_match_result", None)
            st.session_state.pop("last_q", None)
            st.rerun()

    # ── Quick search buttons ──────────────────────────────────────────────────
    with st.expander("⚡ Quick Search — Target Card List", expanded=False):
        st.caption("One-click searches for your buy targets. Click any card to auto-search.")
        sports = ["🏀", "🏈", "⚾"]
        sport_names = {"🏀": "Basketball", "🏈": "Football", "⚾": "Baseball"}
        for sport_emoji in sports:
            sport_cards = [(e, q) for e, q in QUICK_SEARCHES if e == sport_emoji]
            if sport_cards:
                st.markdown(f"**{sport_emoji} {sport_names[sport_emoji]}**")
                cols = st.columns(min(len(sport_cards), 4))
                for idx, (_, qs) in enumerate(sport_cards):
                    short = qs[:35] + ("…" if len(qs) > 35 else "")
                    if cols[idx % 4].button(short, key=f"qs_{qs}", use_container_width=True):
                        st.session_state.quick_search_query = qs
                        st.rerun()

    # Pick up quick-search selection — always overrides whatever is in the text box
    if st.session_state.get("quick_search_query"):
        query = st.session_state.pop("quick_search_query")
        do_search = True

    if query and (do_search or st.session_state.get("last_q") != query):
        st.session_state.last_q = query
        with st.spinner("Searching card database..."):
            # CardHedger first — fast and reliable for prices
            if CARDHEDGER_KEY:
                st.session_state.ch_match_result = ch_card_match(query)
                # Also fetch all variants so user can switch if wrong one matched
                _ch_all_variants = ch_search(query)
                st.session_state.ch_search_results = _ch_all_variants[:15] if _ch_all_variants else []
            # Also try GemRate for population/gem-rate data (fails silently if offline)
            st.session_state.gr_results = search_gemrate(query)
        # Reset variant picker when new search runs
        st.session_state.pop("ch_variant_pick", None)
        # Only use GemRate-primary path if CardHedger is not configured
        if st.session_state.gr_results and not CARDHEDGER_KEY:
            st.session_state.pop("ch_match_result", None)

    results = st.session_state.get("gr_results", [])
    ch_match_data = st.session_state.get("ch_match_result")

    if results:
        st.markdown(f"**{len(results)} result(s) from GemRate**")
        rows = [{
            "Signal": gem_signal(r.get("gem_rate")),
            "Year": r.get("year", ""),
            "Set": r.get("set_name", ""),
            "Player": r.get("name", ""),
            "Parallel": r.get("parallel") or "Base",
            "#": r.get("card_number", ""),
            "Grader": r.get("population_type", ""),
            "Total Pop": f"{r.get('total_population', 0):,}",
            "Gem Copies": f"{r.get('gems', 0):,}",
            "Gem Rate": fmt_gem(r.get("gem_rate")),
        } for r in results]

        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True, height=280)

        gr_search_url = f"https://www.gemrate.com/search?q={urllib.parse.quote_plus(query)}"
        st.caption(f"Don't see the card? [Search directly on GemRate ↗]({gr_search_url}) — their database may use a different set name.")

        opts = [
            f"{r.get('year','')} {r.get('set_name','')} {r.get('name','')} "
            f"{r.get('parallel') or 'Base'} #{r.get('card_number','')}"
            for r in results
        ]
        selected = st.selectbox("Select card to analyze", opts)
        sel = results[opts.index(selected)]
        gem = sel.get("gem_rate")
        desc = f"{sel.get('year','')} {sel.get('set_name','')} {sel.get('name','')} {sel.get('parallel') or ''}".strip()

        st.markdown("---")
        st.text_input("📋 Card name (tap → select all → copy)", value=selected, key="card_name_copy")

        # ── Card image via CardHedger ────────────────────────────────────────
        gr_image_url = ""
        if CARDHEDGER_KEY:
            gr_ch_match = ch_card_match(desc)
            if gr_ch_match:
                gr_image_url = gr_ch_match.get("image", "") or ""
                if not gr_image_url and gr_ch_match.get("card_id"):
                    gr_image_url = ch_card_image(gr_ch_match["card_id"])
        gr_image_url = safe_image_url(gr_image_url)

        if gr_image_url:
            img_c, stats_c = st.columns([1, 3])
            with img_c:
                render_card_image(gr_image_url, use_container_width=True)
            with stats_c:
                m1, m2, m3 = st.columns(3)
                with m1:
                    st.markdown("**Gem Rate (PSA 10)**")
                    st.markdown(gem_bar_html(gem), unsafe_allow_html=True)
                m2.metric("Total Pop", f"{sel.get('total_population', 0):,}")
                m3.metric("Gem Copies", f"{sel.get('gems', 0):,}")
                st.markdown(f"[📊 Full pop report on GemRate]({gemrate_url(sel.get('gemrate_id',''))})")
        else:
            m1, m2, m3 = st.columns(3)
            with m1:
                st.markdown("**Gem Rate (PSA 10)**")
                st.markdown(gem_bar_html(gem), unsafe_allow_html=True)
            m2.metric("Total Pop", f"{sel.get('total_population', 0):,}")
            m3.metric("Gem Copies", f"{sel.get('gems', 0):,}")
            st.markdown(f"[📊 Full pop report on GemRate]({gemrate_url(sel.get('gemrate_id',''))})")

        st.markdown("#### ROI Analysis")

        # ── CardHedger: live sold comps + trend ───────────────────────────────
        ch_raw_avg = None
        ch_psa10_avg = None
        ch_psa9_avg = None
        ch_trend_dir = None
        ch_trend_pct = 0.0
        ch_raw_sales = []
        ch_psa10_sales = []
        ch_card_name = ""
        ch_matched_desc = ""
        ch_matched_variant = ""
        ch_match_confidence = None
        ch_id = None
        ch_raw_fmv = {}
        ch_psa10_fmv = {}
        ch_psa9_fmv = {}

        if CARDHEDGER_KEY:
            with st.spinner("Fetching live sold comps, FMV & trend data..."):
                # Reuse already-cached card_match result (already called above for image);
                # fall back to ch_search only if card_match gave nothing.
                _ch_reuse = st.session_state.get("ch_match_result")
                if _ch_reuse:
                    ch_matches = [_ch_reuse]
                else:
                    ch_matches = ch_search(desc)
                if ch_matches:
                    ch_card = ch_matches[0]
                    ch_id = ch_card.get("card_id") or ch_card.get("id")
                    ch_card_name = ch_card.get("name") or ch_card.get("title") or ""
                    ch_matched_desc = ch_card.get("description") or ch_card.get("name") or ch_card.get("title") or ""
                    ch_matched_variant = ch_card.get("variant") or ""
                    ch_match_confidence = ch_card.get("confidence")
                    if ch_id:
                        # Parallel: comps × 3 + price_history — all independent calls
                        def _fetch_raw():   return ch_comps(ch_id, "Raw")
                        def _fetch_psa10(): return ch_comps(ch_id, "PSA 10")
                        def _fetch_psa9():  return ch_comps(ch_id, "PSA 9")
                        def _fetch_hist():  return ch_price_history(ch_id, "PSA 10", days=90)
                        with ThreadPoolExecutor(max_workers=4) as _ex:
                            _f_raw  = _ex.submit(_fetch_raw)
                            _f_p10  = _ex.submit(_fetch_psa10)
                            _f_p9   = _ex.submit(_fetch_psa9)
                            _f_hist = _ex.submit(_fetch_hist)
                            raw_data  = _f_raw.result()
                            psa_data  = _f_p10.result()
                            psa9_data = _f_p9.result()
                            history   = _f_hist.result()
                        # Batch FMV — 1 call instead of 3
                        _fmv_batch = ch_fmv_batch([
                            {"card_id": ch_id, "grade": "Raw"},
                            {"card_id": ch_id, "grade": "PSA 10"},
                            {"card_id": ch_id, "grade": "PSA 9"},
                        ])
                        _fmv_items = _fmv_batch.get("items") or _fmv_batch.get("results") or []
                        _fmv_map = {}
                        for _fi in _fmv_items:
                            _fmv_map[(_fi.get("grade") or "").upper()] = _fi
                        ch_raw_fmv   = _fmv_map.get("RAW")   or ch_fmv(ch_id, "Raw")
                        ch_psa10_fmv = _fmv_map.get("PSA 10") or ch_fmv(ch_id, "PSA 10")
                        ch_psa9_fmv  = _fmv_map.get("PSA 9")  or ch_fmv(ch_id, "PSA 9")
                        ch_raw_avg   = raw_data.get("comp_price") or raw_data.get("average") or raw_data.get("mean")
                        ch_psa10_avg = psa_data.get("comp_price") or psa_data.get("average") or psa_data.get("mean")
                        ch_psa9_avg  = psa9_data.get("comp_price") or psa9_data.get("average") or psa9_data.get("mean")
                        for k in ("raw_prices", "sales", "comps", "data"):
                            if k in raw_data and isinstance(raw_data[k], list):
                                ch_raw_sales = raw_data[k]; break
                        for k in ("raw_prices", "sales", "comps", "data"):
                            if k in psa_data and isinstance(psa_data[k], list):
                                ch_psa10_sales = psa_data[k]; break
                        ch_trend_dir, ch_trend_pct = calculate_trend(history)

            # Show which card was matched (important for numbered parallels)
            if ch_matched_desc:
                import re as _re
                _conf_str = f" · {ch_match_confidence*100:.0f}% match" if ch_match_confidence else ""
                _variant_str = f" · **{ch_matched_variant}**" if ch_matched_variant else ""
                st.caption(f"🎯 Matched: {ch_matched_desc}{_variant_str}{_conf_str}")
                # Detect numbered parallel in user query (/150, /99, /50, /25, /10, /5, /1)
                _parallel_m = _re.search(r"/(\d+)", desc)
                if _parallel_m and ch_matched_variant and "base" in ch_matched_variant.lower():
                    st.warning(
                        f"⚠️ You searched for **/{_parallel_m.group(1)}** (a numbered parallel) but CardHedger matched the **Base** card — "
                        f"pricing below is for the **base** version, NOT your numbered parallel. "
                        f"Search for the specific parallel name (e.g. 'Gold Refractor /150') for accurate pricing."
                    )

            if ch_raw_avg or ch_psa10_avg or ch_raw_sales or ch_psa10_sales:
                def _fmt_sale_row(s):
                    p = _extract_price(s)
                    if not p:
                        return None
                    raw_date = s.get("sale_date", "")
                    try:
                        import datetime as _dt
                        d = _dt.datetime.fromisoformat(raw_date.replace("Z", "+00:00"))
                        date_str = d.strftime("%-m/%-d/%y")
                    except Exception:
                        date_str = raw_date[:10] if raw_date else ""
                    sale_type = s.get("sale_type", "") or ""
                    type_label = "🏷 BIN" if "bin" in sale_type.lower() or sale_type == "" else \
                                 "🔨 Auction" if "auction" in sale_type.lower() else \
                                 "💬 Offer" if "offer" in sale_type.lower() else sale_type
                    return {"Price": f"${p:,.2f}", "Date": date_str, "Type": type_label}

                fc1, fc2 = st.columns(2)
                with fc1:
                    st.markdown("**📦 Raw — recent sold comps**")
                    if ch_raw_sales:
                        rows_r = [r for s in ch_raw_sales[:15] if (r := _fmt_sale_row(s))]
                        if rows_r:
                            st.dataframe(pd.DataFrame(rows_r), use_container_width=True, hide_index=True, height=200)
                    if ch_raw_avg:
                        st.markdown(f"**Comp avg: ${ch_raw_avg:,.2f}**")
                    _rf = fmv_caption(ch_raw_fmv)
                    if _rf:
                        st.markdown(_rf)
                    elif not ch_raw_avg and not ch_raw_sales:
                        st.info("No raw comps found")
                with fc2:
                    st.markdown("**💎 PSA 10 — recent sold comps**")
                    if ch_psa10_sales:
                        rows_g = [r for s in ch_psa10_sales[:15] if (r := _fmt_sale_row(s))]
                        if rows_g:
                            st.dataframe(pd.DataFrame(rows_g), use_container_width=True, hide_index=True, height=200)
                    if ch_psa10_avg:
                        st.markdown(f"**Comp avg: ${ch_psa10_avg:,.2f}**")
                    _gf = fmv_caption(ch_psa10_fmv)
                    if _gf:
                        st.markdown(_gf)
                    elif not ch_psa10_avg and not ch_psa10_sales:
                        st.info("No PSA 10 comps found")
            if ch_raw_sales or ch_psa10_sales:
                st.caption("⚠️ Comp avg includes all sale types. 🏷 BIN = fixed price (most reliable). 💬 Offer = accepted below ask. 🔨 Auction = bidding (use with caution).")
                st.caption("🎯 **FMV** = Fair Market Value: a statistically cleaned price (Winsorized median, recent sales) with a confidence grade (A best). More reliable than the comp avg — drives the cost/sell pre-fills below.")
            elif CARDHEDGER_KEY:
                st.info("No CardHedger match found for this card — enter prices manually below.")
        else:
            st.info("📊 Live sold comps will appear here once the CardHedger API is connected.")

        # ── Quick pricing summary + GemRate link ─────────────────────────────
        _raw_show   = fmv_price(ch_raw_fmv)   or ch_raw_avg
        _p10_show   = fmv_price(ch_psa10_fmv) or ch_psa10_avg
        if _raw_show or _p10_show:
            _ps_c1, _ps_c2, _ps_c3, _ps_c4 = st.columns(4)
            _ps_c1.metric("📦 Raw Avg", f"${_raw_show:,.2f}" if _raw_show else "—")
            _ps_c2.metric("💎 PSA 10 Avg", f"${_p10_show:,.2f}" if _p10_show else "—")
            if _raw_show and _p10_show:
                _uplift = _p10_show - _raw_show
                _ps_c3.metric("💰 Grade Uplift", f"${_uplift:,.2f}", help="PSA 10 avg minus raw avg — what grading adds in dollar terms")
            _ps_c4.metric("🔵 Gem Rate", fmt_gem(gem) if gem is not None else "—")
            _gr_btn_url = gemrate_url(sel.get("gemrate_id", ""))
            if _gr_btn_url:
                st.link_button("📊 View Full Pop Report on GemRate →", _gr_btn_url)
            st.markdown("---")

        # Prefer cleaned FMV for pre-fills; fall back to the comp average.
        raw_auto    = fmv_price(ch_raw_fmv)   or ch_raw_avg
        graded_auto = fmv_price(ch_psa10_fmv) or ch_psa10_avg

        ra1, ra2, ra3 = st.columns(3)
        with ra1:
            raw_cost = st.number_input(
                "Your cost for the raw card ($)", min_value=0.0,
                value=float(raw_auto) if raw_auto else 50.0,
                step=5.0, key="t1_raw",
            )
            st.caption("What you paid (or plan to pay) for the ungraded card. Pre-filled from raw FMV (cleaned) — update to your actual price.")
        with ra2:
            tier = default_tier
            st.caption(f"**Grading tier:** {tier} (${PSA_FEES[tier]:.2f}) — change in sidebar ⚙️")
        with ra3:
            graded_price = st.number_input(
                "Expected PSA 10 sell price ($)", min_value=0.0,
                value=float(graded_auto) if graded_auto else 0.0,
                step=10.0, key="t1_graded",
            )
            st.caption("The price you expect to sell a PSA 10 for on eBay. Pre-filled from PSA 10 FMV (cleaned) — adjust up or down based on your read of the market.")

        if raw_cost > 0:
            fee      = PSA_FEES[tier]
            opp      = calc_opp_cost(raw_cost, tier, opp_rate, ship_cost)
            tgt      = target_price(raw_cost, tier, roi_target, opp_rate, ship_cost)
            cal_days = int(PSA_DAYS.get(tier, 60) * 1.4)
            total_in = raw_cost + fee + ship_cost + opp
            bd1, bd2, bd3, bd4, bd5, bd6 = st.columns(6)
            bd1.metric("Raw card",                             f"${raw_cost:,.2f}")
            bd2.metric("Grading fee",                         f"${fee:.2f}")
            bd3.metric("Shipping (to + return)",              f"${ship_cost:.2f}",
                       help="Per-card shipping to PSA + return. Set in sidebar ⚙️")
            bd4.metric(f"⏳ Time cost ({cal_days} days)",     f"${opp:,.2f}",
                       help=f"At {opp_rate:.0f}%/yr, ${raw_cost+fee+ship_cost:,.0f} locked for {cal_days} days = ${opp:,.2f} you can't redeploy")
            bd5.metric(f"Target PSA 10 ({roi_target:.0f}×)",  f"${tgt:,.0f}")
            bd6.metric("eBay fees",                           f"${graded_price * EBAY_FEE:,.2f}" if graded_price else "—")

            st.caption(
                f"💡 **True total cost: ${total_in:,.2f}** — raw ${raw_cost:,.2f} + grading ${fee:.2f} + "
                f"shipping ${ship_cost:.2f} + time cost ${opp:,.2f}. "
                f"Most apps only count the grading fee."
            )

            if graded_price > 0:
                v, color, msg = verdict(raw_cost, tier, gem, graded_price, min_gem, roi_target, opp_rate, ship_cost)
                net, roi = calc_net_roi(raw_cost, tier, graded_price, opp_rate, ship_cost)
                if color == "green":
                    st.success(f"{v} — {msg}")
                else:
                    st.error(f"{v} — {msg}")
                r1, r2, r3 = st.columns(3)
                r1.metric("Est. Net Profit", f"${net:,.0f}", help="After ALL costs: raw, grading, shipping, time cost, and eBay fees")
                r2.metric("Est. ROI",        f"{roi:.0f}%")
                if ch_trend_dir:
                    r3.metric("90-Day Trend", trend_badge(ch_trend_dir, ch_trend_pct))

                summary = f"""{query}
Gem Rate: {fmt_gem(gem)} | Raw: ${raw_cost:,.2f} | Gem 10 Avg: ${graded_price:,.2f}
Grading fee: ${fee:.2f} | Shipping: ${ship_cost:.2f} | Time cost ({cal_days} days @ {opp_rate:.0f}%/yr): ${opp:,.2f}
True total cost: ${total_in:,.2f} | Target: ${tgt:,.0f} | Net: ${net:,.0f} | ROI: {roi:.0f}%
{v}"""
                with st.expander("📋 Copy Analysis"):
                    st.code(summary, language=None)
            else:
                st.info("Enter a Gem 10 avg price above to get a GO/NO-GO decision")

        # ── ⚖️ Grade vs Flip — the holding-cost decision ──────────────────────
        if CARDHEDGER_KEY and raw_cost > 0 and (ch_raw_avg or ch_psa10_avg or ch_psa9_avg):
            st.markdown("#### ⚖️ Grade vs Flip — the real decision")
            st.caption(
                "Grading locks up your cash for months. Here's what each path actually nets — "
                "weighted by this card's gem rate — so you decide with eyes open, not just \"I can 4× it.\""
            )
            # Use cleaned FMV where available, fall back to the comp average.
            gvf_raw   = fmv_price(ch_raw_fmv)   or ch_raw_avg
            gvf_psa10 = fmv_price(ch_psa10_fmv) or ch_psa10_avg
            gvf_psa9  = fmv_price(ch_psa9_fmv)  or ch_psa9_avg
            gvf = grade_vs_flip(raw_cost, gvf_raw, gvf_psa10, gvf_psa9, gem, tier, ship_cost, opp_rate)

            d1, d2, d3 = st.columns(3)
            with d1:
                st.markdown("**💵 Flip Raw Now**")
                st.metric("Sell ~", f"${gvf_raw:,.0f}" if gvf_raw else "—", help="Raw FMV (cleaned) — or sold-comp average if FMV unavailable")
                st.metric("Net profit", f"${gvf['raw_net']:,.0f}" if gvf['raw_net'] is not None else "—",
                          help="After eBay fees + your buy cost. Cash back in ~3-7 days, nothing locked up.")
            with d2:
                st.markdown("**💎 Grade → PSA 10**")
                st.metric("Sell ~", f"${gvf_psa10:,.0f}" if gvf_psa10 else "—")
                st.metric("Net profit", f"${gvf['net10']:,.0f}" if gvf['net10'] is not None else "—",
                          help=f"After grading, shipping, eBay fees, and ${gvf['hold']:,.0f} holding cost.")
            with d3:
                st.markdown("**🥈 Grade → PSA 9**")
                st.metric("Sell ~", f"${gvf_psa9:,.0f}" if gvf_psa9 else "—")
                st.metric("Net profit", f"${gvf['net9']:,.0f}" if gvf['net9'] is not None else "—",
                          help="The downside if it doesn't gem — same costs, lower sale.")

            ev1, ev2, ev3 = st.columns(3)
            ev1.metric(f"🎯 Expected net (gem {fmt_gem(gem)})",
                       f"${gvf['net_exp']:,.0f}" if gvf['net_exp'] is not None else "—",
                       help=f"Gem-rate-weighted: {gvf['p10']*100:.0f}% chance of a 10, {(1-gvf['p10'])*100:.0f}% a 9. After all costs incl. holding.")
            ev2.metric("💸 Holding cost", f"${gvf['hold']:,.0f}",
                       help=f"${gvf['capital']:,.0f} of your cash locked ~{gvf['cal_days']} days at {opp_rate:.0f}%/yr")
            if gvf['premium'] is not None:
                ev3.metric("Grade premium vs raw",
                           f"{'+' if gvf['premium'] >= 0 else '−'}${abs(gvf['premium']):,.0f}",
                           help="Expected graded net minus flip-raw net — your reward for the wait.")

            _lbl, _clr, _msg = grade_flip_verdict(gvf)
            (st.success if _clr == "green" else st.warning if _clr == "amber"
             else st.error if _clr == "red" else st.info)(f"**{_lbl}** — {_msg}")

            st.caption(
                f"⏳ Grading this card ties up **${gvf['capital']:,.0f}** "
                f"(buy ${raw_cost:,.0f} + fee ${gvf['fee']:.0f}"
                + (f" + ship ${ship_cost:.0f}" if ship_cost else "")
                + f") for ~**{gvf['cal_days']} days** (~{gvf['cal_days']/30:.1f} months). Flipping raw frees that cash now."
            )

        if CARDHEDGER_KEY and ch_id:
            render_price_trend(ch_id, key_prefix="main")
            render_player_watch(ch_id, key_prefix="main")
            render_player_demand(ch_id, key_prefix="main")

        st.markdown("#### 🔍 Card Finder Links")
        st.caption("All the links you need — buy raw, check sold comps, verify gem rate, or look up the PSA pop report.")

        _link_style = (  # noqa: F841  (also used in CardHedger fallback block below)
            "display:inline-flex;align-items:center;gap:6px;padding:8px 14px;"
            "border-radius:6px;font-size:12px;font-weight:600;text-decoration:none;"
            "border:1px solid;margin:3px;"
        )
        st.markdown(
            f"""
            <div style="display:flex;flex-wrap:wrap;gap:2px;margin-bottom:8px">
              <a href="{ebay_buy_bin_url(desc)}" target="_blank"
                 style="{_link_style}color:#e5a310;border-color:rgba(229,163,16,.4);background:rgba(229,163,16,.08)">
                 🛒 Buy Raw (BIN)</a>
              <a href="{ebay_buy_auction_url(desc)}" target="_blank"
                 style="{_link_style}color:#e5a310;border-color:rgba(229,163,16,.4);background:rgba(229,163,16,.08)">
                 ⚡ Buy Raw (Auction)</a>
              <a href="{ebay_raw_sold_all_url(desc)}" target="_blank"
                 style="{_link_style}color:#94a3b8;border-color:rgba(148,163,184,.3);background:rgba(148,163,184,.06)">
                 📊 Raw Sold Comps</a>
              <a href="{ebay_graded_sold_url(desc)}" target="_blank"
                 style="{_link_style}color:#22c55e;border-color:rgba(34,197,94,.4);background:rgba(34,197,94,.08)">
                 🔴 PSA 10 Sold Comps</a>
              <a href="{ebay_psa9_sold_url(desc)}" target="_blank"
                 style="{_link_style}color:#f59e0b;border-color:rgba(245,158,11,.4);background:rgba(245,158,11,.08)">
                 🟡 PSA 9 Sold Comps</a>
              <a href="{gemrate_url(sel.get('gemrate_id',''))}" target="_blank"
                 style="{_link_style}color:#4f8ef7;border-color:rgba(79,142,247,.4);background:rgba(79,142,247,.08)">
                 💎 GemRate.com</a>
              <a href="{psa_pop_url(desc)}" target="_blank"
                 style="{_link_style}color:#7c5cfc;border-color:rgba(124,92,252,.4);background:rgba(124,92,252,.08)">
                 🏆 PSA Pop Report</a>
            </div>
            """,
            unsafe_allow_html=True,
        )

        st.markdown("---")
        if st.button("➕ Add to Submission Tracker", type="secondary"):
            if raw_cost <= 0:
                st.warning("Enter a raw buy price first")
            else:
                fee = PSA_FEES[tier]
                tgt = target_price(raw_cost, tier, roi_target, opp_rate, ship_cost)
                net, roi = calc_net_roi(raw_cost, tier, graded_price, opp_rate, ship_cost) if graded_price > 0 else (None, None)
                v, _, _ = verdict(raw_cost, tier, gem, graded_price, min_gem, roi_target, opp_rate, ship_cost) if graded_price > 0 else ("Pending", "", "")
                sb_insert({
                    "date_added": date.today().isoformat(),
                    "card_description": selected,
                    "year": sel.get("year", ""),
                    "set_name": sel.get("set_name", ""),
                    "parallel": sel.get("parallel") or "Base",
                    "raw_buy_price": raw_cost,
                    "psa_tier": tier,
                    "psa_fee": fee,
                    "psa10_avg_price": graded_price or None,
                    "target_price": round(tgt, 2),
                    "gem_rate": round(gem, 2) if gem else None,
                    "go_no_go": v,
                    "est_net": net,
                    "est_roi": f"{roi:.0f}%" if roi is not None else None,
                    "status": "Queued",
                })
                st.success("Added to Submission Tracker ✓")

    elif ch_match_data:
        # ── CardHedger AI-match fallback (GemRate unavailable) ───────────────

        # Variant picker — let user switch if CardHedger matched wrong parallel
        _ch_all = st.session_state.get("ch_search_results", [])
        _active_match = ch_match_data
        if _ch_all and len(_ch_all) > 1:
            _var_labels = [
                f"{c.get('variant','Base') or 'Base'} — {c.get('description','')}"
                for c in _ch_all
            ]
            _default_var = 0
            _matched_id = ch_match_data.get("card_id", "")
            for _vi, _vc in enumerate(_ch_all):
                if _vc.get("card_id") == _matched_id:
                    _default_var = _vi
                    break
            _picked_var = st.selectbox(
                "🎨 Variant (CardHedger matched this — change if wrong)",
                _var_labels,
                index=_default_var,
                key="ch_variant_pick",
            )
            _picked_idx = _var_labels.index(_picked_var)
            if _picked_idx != _default_var:
                # User switched variant — build a synthetic match object from search result
                _picked_card = _ch_all[_picked_idx]
                _active_match = {
                    "card_id":    _picked_card.get("card_id"),
                    "description": _picked_card.get("description", ""),
                    "player":     _picked_card.get("player", ""),
                    "set":        _picked_card.get("set", ""),
                    "variant":    _picked_card.get("variant", ""),
                    "image":      _picked_card.get("image", ""),
                    "confidence": None,
                    "prices":     _picked_card.get("prices", []),
                }
        elif _ch_all:
            _active_match = ch_match_data

        confidence  = _active_match.get("confidence", 0) or 0
        reasoning   = _active_match.get("reasoning", "")
        desc        = _active_match.get("description", query)
        player      = _active_match.get("player", "")
        set_name    = _active_match.get("set", "")
        variant     = _active_match.get("variant", "")
        prices_list = _active_match.get("prices", [])

        # Build price map
        price_map = {}
        for p in prices_list:
            try:
                price_map[p["grade"]] = float(p["price"])
            except Exception:
                pass
        ch_psa10 = price_map.get("PSA 10")
        ch_psa9  = price_map.get("PSA 9")
        ch_raw   = price_map.get("Raw")

        # Cleaned FMV per grade — batch call instead of 3 sequential calls
        _cid = _active_match.get("card_id")
        if _cid:
            _fb_batch = ch_fmv_batch([
                {"card_id": _cid, "grade": "Raw"},
                {"card_id": _cid, "grade": "PSA 10"},
                {"card_id": _cid, "grade": "PSA 9"},
            ])
            _fb_items = _fb_batch.get("items") or _fb_batch.get("results") or []
            _fb_map = {(fi.get("grade") or "").upper(): fi for fi in _fb_items}
            fb_raw_fmv   = _fb_map.get("RAW")    or ch_fmv(_cid, "Raw")
            fb_psa10_fmv = _fb_map.get("PSA 10") or ch_fmv(_cid, "PSA 10")
            fb_psa9_fmv  = _fb_map.get("PSA 9")  or ch_fmv(_cid, "PSA 9")
        else:
            fb_raw_fmv = fb_psa10_fmv = fb_psa9_fmv = {}
        raw_val   = fmv_price(fb_raw_fmv)   or ch_raw
        psa10_val = fmv_price(fb_psa10_fmv) or ch_psa10
        psa9_val  = fmv_price(fb_psa9_fmv)  or ch_psa9

        _link_style_fb = (
            "display:inline-flex;align-items:center;gap:6px;padding:8px 14px;"
            "border-radius:6px;font-size:12px;font-weight:600;text-decoration:none;"
            "border:1px solid;margin:3px;"
        )

        st.info(
            "💎 **Showing live CardHedger prices** — GemRate is currently unavailable. "
            "Gem rate (PSA pop) cannot be checked right now; ROI calculation still works.",
            icon="ℹ️",
        )

        # ── Card image ────────────────────────────────────────────────────────
        image_url = _active_match.get("image", "") or ""
        if not image_url and _active_match.get("card_id"):
            with st.spinner("Loading card image..."):
                image_url = ch_card_image(_active_match["card_id"])

        img_col, info_col = st.columns([1, 3])
        with img_col:
            render_card_image(image_url, use_container_width=True)
        with info_col:
            st.markdown(f"### {desc}")
            ci1, ci2, ci3 = st.columns(3)
            ci1.markdown(f"**Set:** {set_name}")
            ci2.markdown(f"**Variant:** {variant}")
            if confidence:
                ci3.markdown(f"**AI Match confidence:** {confidence * 100:.0f}%")
            else:
                ci3.markdown("**Source:** CardHedger")

        st.markdown("#### 💰 Market Prices (CardHedger)")
        if prices_list:
            price_rows_fb = [{"Grade": p.get("grade", ""), "Avg Market Price": f"${float(p.get('price', 0)):,.2f}"} for p in prices_list]
            st.dataframe(pd.DataFrame(price_rows_fb), use_container_width=True, hide_index=True)

        st.markdown("---")
        st.text_input("📋 Card name (tap → select all → copy)", value=desc, key="card_name_copy")

        fb1, fb2, fb3 = st.columns(3)
        with fb1:
            st.markdown("**Gem Rate (PSA 10)**")
            gem_fb = st.slider(
                "Your gem rate estimate (%)", 0, 100, 50, 5, key="fb_gem_top",
                help="How often this card grades PSA 10. Check the PSA Pop Report link below for historical rates. 50% is a neutral starting point.",
            )
            st.caption(f"[Look up PSA pop →]({psa_pop_url(desc)})")
        if psa10_val:
            fb2.metric("PSA 10", f"${psa10_val:,.2f}")
            _c10 = fmv_band_conf(fb_psa10_fmv)
            if _c10:
                fb2.caption(_c10)
        if psa9_val:
            fb3.metric("PSA 9", f"${psa9_val:,.2f}")
            _c9 = fmv_band_conf(fb_psa9_fmv)
            if _c9:
                fb3.caption(_c9)
        if fmv_price(fb_raw_fmv) or fmv_price(fb_psa10_fmv):
            st.caption("🎯 Showing FMV (cleaned value) where available, else the CardHedger average. Confidence A = best.")

        st.markdown("#### ROI Analysis")

        ra1, ra2, ra3 = st.columns(3)
        with ra1:
            raw_cost = st.number_input(
                "Your cost for the raw card ($)", min_value=0.0,
                value=float(raw_val) if raw_val else 50.0,
                step=5.0, key="t1_raw",
            )
            st.caption("Pre-filled from raw FMV (cleaned) — update to your actual price.")
        with ra2:
            tier = default_tier
            st.caption(f"**Grading tier:** {tier} (${PSA_FEES[tier]:.2f}) — change in sidebar ⚙️")
        with ra3:
            graded_price = st.number_input(
                "Expected PSA 10 sell price ($)", min_value=0.0,
                value=float(psa10_val) if psa10_val else 0.0,
                step=10.0, key="t1_graded",
            )
            st.caption("Pre-filled from PSA 10 FMV (cleaned) — adjust as needed.")

        if raw_cost > 0:
            fee      = PSA_FEES[tier]
            opp      = calc_opp_cost(raw_cost, tier, opp_rate, ship_cost)
            tgt      = target_price(raw_cost, tier, roi_target, opp_rate, ship_cost)
            cal_days = int(PSA_DAYS.get(tier, 60) * 1.4)
            total_in = raw_cost + fee + ship_cost + opp
            bd1, bd2, bd3, bd4, bd5, bd6 = st.columns(6)
            bd1.metric("Raw card",                            f"${raw_cost:,.2f}")
            bd2.metric("Grading fee",                        f"${fee:.2f}")
            bd3.metric("Shipping (to + return)",             f"${ship_cost:.2f}",
                       help="Per-card shipping to PSA + return. Set in sidebar ⚙️")
            bd4.metric(f"⏳ Time cost ({cal_days} days)",    f"${opp:,.2f}",
                       help=f"At {opp_rate:.0f}%/yr, ${raw_cost+fee+ship_cost:,.0f} locked for {cal_days} days")
            bd5.metric(f"Target PSA 10 ({roi_target:.0f}×)", f"${tgt:,.0f}")
            bd6.metric("eBay fees",                          f"${graded_price * EBAY_FEE:,.2f}" if graded_price else "—")

            st.caption(
                f"💡 **True total cost: ${total_in:,.2f}** — raw ${raw_cost:,.2f} + grading ${fee:.2f} + "
                f"shipping ${ship_cost:.2f} + time cost ${opp:,.2f}. "
                f"Most apps only count the grading fee."
            )

            if graded_price > 0:
                v_fb, color_fb, msg_fb = verdict(raw_cost, tier, gem_fb, graded_price, min_gem, roi_target, opp_rate, ship_cost)
                net, roi = calc_net_roi(raw_cost, tier, graded_price, opp_rate, ship_cost)
                if color_fb == "green":
                    st.success(f"{v_fb} — {msg_fb}")
                else:
                    st.error(f"{v_fb} — {msg_fb}")
                r1, r2, r3 = st.columns(3)
                r1.metric("Est. Net Profit", f"${net:,.0f}", help="After ALL costs: raw, grading, shipping, time cost, and eBay fees")
                r2.metric("Est. ROI",        f"{roi:.0f}%")
                r3.metric("Gem Rate Est.",   f"{gem_fb}%", help="Your estimate set above. Check PSA pop report to calibrate.")

                summary_fb = f"""{desc}
Source: CardHedger | Gem Rate estimate: {gem_fb}% (set manually — verify on PSA pop)
Raw: ${raw_cost:,.2f} | PSA 10 FMV: ${graded_price:,.2f}
Grading fee: ${fee:.2f} | Shipping: ${ship_cost:.2f} | Time cost ({cal_days} days @ {opp_rate:.0f}%/yr): ${opp:,.2f}
True total cost: ${total_in:,.2f} | Target: ${tgt:,.0f} | Net: ${net:,.0f} | ROI: {roi:.0f}%
Verdict: {v_fb} — {msg_fb}"""
                with st.expander("📋 Copy Analysis"):
                    st.code(summary_fb, language=None)
            else:
                st.info("Enter a PSA 10 price above to get a GO/NO-GO decision")

        # ── ⚖️ Grade vs Flip ──────────────────────────────────────────────────
        if raw_cost > 0 and (raw_val or psa10_val or psa9_val):
            st.markdown("#### ⚖️ Grade vs Flip — the real decision")
            st.caption(
                f"Grading locks up your cash for months. Using your {gem_fb}% gem rate estimate "
                f"(set above) to weight the PSA 10 vs PSA 9 outcome."
            )
            gvf = grade_vs_flip(raw_cost, raw_val, psa10_val, psa9_val, gem_fb, tier, ship_cost, opp_rate)
            d1, d2, d3 = st.columns(3)
            with d1:
                st.markdown("**💵 Flip Raw Now**")
                st.metric("Sell ~", f"${raw_val:,.0f}" if raw_val else "—", help="Raw FMV (cleaned)")
                st.metric("Net profit", f"${gvf['raw_net']:,.0f}" if gvf['raw_net'] is not None else "—",
                          help="After eBay fees + your buy cost. Cash back in days, nothing locked up.")
            with d2:
                st.markdown("**💎 Grade → PSA 10**")
                st.metric("Sell ~", f"${psa10_val:,.0f}" if psa10_val else "—")
                st.metric("Net profit", f"${gvf['net10']:,.0f}" if gvf['net10'] is not None else "—",
                          help=f"After grading, shipping, eBay fees, and ${gvf['hold']:,.0f} holding cost.")
            with d3:
                st.markdown("**🥈 Grade → PSA 9**")
                st.metric("Sell ~", f"${psa9_val:,.0f}" if psa9_val else "—")
                st.metric("Net profit", f"${gvf['net9']:,.0f}" if gvf['net9'] is not None else "—",
                          help="The downside if it doesn't gem — same costs, lower sale.")
            ev1, ev2, ev3 = st.columns(3)
            ev1.metric(f"🎯 Expected net (gem {gem_fb}%)",
                       f"${gvf['net_exp']:,.0f}" if gvf['net_exp'] is not None else "—",
                       help="Gem-rate-weighted expected profit after all costs incl. holding.")
            ev2.metric("💸 Holding cost", f"${gvf['hold']:,.0f}",
                       help=f"${gvf['capital']:,.0f} of cash locked ~{gvf['cal_days']} days at {opp_rate:.0f}%/yr")
            if gvf['premium'] is not None:
                ev3.metric("Grade premium vs raw",
                           f"{'+' if gvf['premium'] >= 0 else '−'}${abs(gvf['premium']):,.0f}",
                           help="Expected graded net minus flip-raw net — your reward for the wait.")
            _lbl, _clr, _msg = grade_flip_verdict(gvf)
            (st.success if _clr == "green" else st.warning if _clr == "amber"
             else st.error if _clr == "red" else st.info)(f"**{_lbl}** — {_msg}")
            st.caption(
                f"⏳ Grading ties up **${gvf['capital']:,.0f}** for ~**{gvf['cal_days']} days** "
                f"(~{gvf['cal_days']/30:.1f} months). Flipping raw frees that cash now."
            )

        render_price_trend(_cid, key_prefix="fb")
        render_player_watch(_cid, key_prefix="fb")
        render_player_demand(_cid, key_prefix="fb")

        st.markdown("#### 🔍 Card Finder Links")
        st.markdown(
            f"""<div style="display:flex;flex-wrap:wrap;gap:2px;margin-bottom:8px">
              <a href="{ebay_buy_bin_url(desc)}" target="_blank"
                 style="{_link_style_fb}color:#e5a310;border-color:rgba(229,163,16,.4);background:rgba(229,163,16,.08)">
                 🛒 Buy Raw (BIN)</a>
              <a href="{ebay_buy_auction_url(desc)}" target="_blank"
                 style="{_link_style_fb}color:#e5a310;border-color:rgba(229,163,16,.4);background:rgba(229,163,16,.08)">
                 ⚡ Buy Raw (Auction)</a>
              <a href="{ebay_raw_sold_all_url(desc)}" target="_blank"
                 style="{_link_style_fb}color:#94a3b8;border-color:rgba(148,163,184,.3);background:rgba(148,163,184,.06)">
                 📊 Raw Sold Comps</a>
              <a href="{ebay_graded_sold_url(desc)}" target="_blank"
                 style="{_link_style_fb}color:#22c55e;border-color:rgba(34,197,94,.4);background:rgba(34,197,94,.08)">
                 🔴 PSA 10 Sold Comps</a>
              <a href="{ebay_psa9_sold_url(desc)}" target="_blank"
                 style="{_link_style_fb}color:#f59e0b;border-color:rgba(245,158,11,.4);background:rgba(245,158,11,.08)">
                 🟡 PSA 9 Sold Comps</a>
              <a href="{psa_pop_url(desc)}" target="_blank"
                 style="{_link_style_fb}color:#7c5cfc;border-color:rgba(124,92,252,.4);background:rgba(124,92,252,.08)">
                 🏆 PSA Pop Report</a>
            </div>""",
            unsafe_allow_html=True,
        )

        if reasoning:
            with st.expander("🤖 AI Match Reasoning"):
                st.caption(reasoning)

        st.markdown("---")
        if st.button("➕ Add to Submission Tracker", type="secondary"):
            if raw_cost <= 0:
                st.warning("Enter a raw buy price first")
            else:
                fee = PSA_FEES[tier]
                tgt = target_price(raw_cost, tier, roi_target, opp_rate, ship_cost)
                net, roi_val = calc_net_roi(raw_cost, tier, graded_price, opp_rate, ship_cost) if graded_price > 0 else (None, None)
                v_fb = "✅ GO" if graded_price > 0 and graded_price >= tgt else "❌ NO-GO"
                sb_insert({
                    "date_added":       date.today().isoformat(),
                    "card_description": desc,
                    "year":             "",
                    "set_name":         set_name or "",
                    "parallel":         variant or "Base",
                    "raw_buy_price":    raw_cost,
                    "psa_tier":         tier,
                    "psa_fee":          fee,
                    "psa10_avg_price":  graded_price or None,
                    "target_price":     round(tgt, 2),
                    "gem_rate":         gem_fb,
                    "go_no_go":         v_fb,
                    "est_net":          net,
                    "est_roi":          f"{roi_val:.0f}%" if roi_val is not None else None,
                    "status":           "Queued",
                })
                st.success("Added to Submission Tracker ✓")

    elif query:
        gr_search_url = f"https://www.gemrate.com/search?q={urllib.parse.quote_plus(query)}"
        if CARDHEDGER_KEY:
            st.warning(
                f"No results found on GemRate or CardHedger for **{query}**. "
                "Try adjusting the search — include player name, set, year, and parallel "
                "*(e.g. Cooper Flagg 2025 Topps Chrome Rookie)*."
            )
        else:
            st.warning(f"No results found — GemRate may not have this set indexed yet. Try a different search term or [search directly on GemRate ↗]({gr_search_url})")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Inventory Check
# ══════════════════════════════════════════════════════════════════════════════

TEMPLATE_COLS = [
    "Card Description", "Player", "Year", "Set", "Parallel",
    "Card Number", "Category", "Cost Basis ($)", "Listed Price ($)",
    "Source", "Notes"
]

TEMPLATE_EXAMPLE = [
    ["2021 Topps Chrome Julio Rodriguez Auto RC", "Julio Rodriguez", "2021", "Topps Chrome", "Base", "RA-JR", "Baseball", 68, 155, "Whatnot", ""],
    ["2024 Panini Prizm Luka Doncic Silver", "Luka Doncic", "2024", "Panini Prizm", "Silver", "10", "Basketball", 120, 280, "eBay Lot", ""],
    ["2023 Bowman Chrome Patrick Bailey Auto", "Patrick Bailey", "2023", "Bowman Chrome", "Base", "CDA-PB", "Baseball", 25, 55, "Card Show", ""],
]

def make_template_csv():
    import io
    buf = io.StringIO()
    df = pd.DataFrame(TEMPLATE_EXAMPLE, columns=TEMPLATE_COLS)
    df.to_csv(buf, index=False)
    return buf.getvalue().encode()

def load_inventory(uploaded_file):
    """Load inventory from either the DFS workbook or the standard template CSV/XLSX."""
    name = uploaded_file.name.lower()
    try:
        if name.endswith(".csv"):
            df = pd.read_csv(uploaded_file)
            df.columns = [c.strip() for c in df.columns]
            return df, "template"
        else:
            xl = pd.ExcelFile(uploaded_file)
            # Detect file type by sheet names
            if "Inventory & Aging" in xl.sheet_names:
                # DFS Operations Workbook
                df = pd.read_excel(uploaded_file, sheet_name="Inventory & Aging", header=2)
                df.columns = [str(c).strip() for c in df.columns]
                if "Card Description" not in df.columns:
                    df.columns = ["Acquired", "Listed Date", "Source", "Card Description",
                                  "Category", "Cost Basis ($)", "Listed Price ($)",
                                  "Status", "Days Listed", "Aging Flag", "Next Action"]
                df = df[df["Card Description"].notna()]
                df = df[~df["Card Description"].astype(str).str.startswith("💡")]
                df = df[~df["Acquired"].astype(str).str.startswith("SUMMARY", na=True)]
                return df.reset_index(drop=True), "workbook"
            else:
                # Standard template XLSX (first sheet)
                df = pd.read_excel(uploaded_file, sheet_name=0)
                df.columns = [str(c).strip() for c in df.columns]
                return df.reset_index(drop=True), "template"
    except Exception as e:
        return None, str(e)

# ─── Reprice Assistant helpers ────────────────────────────────────────────────
def clean_title_for_ch(title: str) -> str:
    """Strip eBay-specific cruft from a listing title before sending to CardHedger.

    eBay titles often include pack types, raw condition markers, and other noise
    that confuse the AI matcher and drop match rates.
    """
    t = (title or "").strip()
    # Remove trailing raw-condition marker " R" that eBay sellers append
    t = re.sub(r'\s+R\s*$', '', t)
    # Remove eBay pack-type descriptors that aren't part of the card identity
    t = re.sub(r'\b(HTA CHOICE|HTA|CHOICE REFRACTOR|JUMBO PACK|JUMBO|HOBBY|RETAIL)\b', '', t, flags=re.IGNORECASE)
    # Season-year "2025-26" → "2025" (CardHedger uses release year, not season span)
    t = re.sub(r'\b(20\d{2})-\d{2}\b', r'\1', t)
    # Collapse extra whitespace left by removals
    t = re.sub(r'\s{2,}', ' ', t).strip()
    return t


def sane_price(suggested, current, floor_pct=0.20, ceil_mult=4.0):
    """Return suggested price only if it passes a sanity check vs the current list price.

    Drops suggestions that are <20% or >4× the current price — almost always a bad match.
    """
    if not suggested or not current or current <= 0:
        return suggested
    ratio = suggested / current
    if ratio < floor_pct or ratio > ceil_mult:
        return None
    return suggested


def detect_grade(desc):
    """Infer the grade to price against from the card description. Defaults to Raw."""
    d = (desc or "").upper()
    if "PSA 10" in d or "GEM MINT 10" in d or "GEM MT 10" in d: return "PSA 10"
    if "PSA 9.5" in d: return "PSA 9.5"
    if "PSA 9"  in d: return "PSA 9"
    if "PSA 8.5" in d: return "PSA 8.5"
    if "PSA 8"  in d: return "PSA 8"
    if "PSA 7"  in d: return "PSA 7"
    if "PSA 6"  in d: return "PSA 6"
    if "PSA 5"  in d: return "PSA 5"
    if "BGS 10" in d or "BGS PRISTINE" in d: return "BGS 10"
    if "BGS 9.5" in d: return "BGS 9.5"
    if "BGS 9"  in d: return "BGS 9"
    if "SGC 10" in d: return "SGC 10"
    if "SGC 9.5" in d: return "SGC 9.5"
    if "SGC 9"  in d: return "SGC 9"
    return "Raw"

def fetch_market(desc, grade):
    """Look up a card via CardHedger card-match (AI) and return comp avg + 90-day trend."""
    out = {"grade": grade, "comp_avg": None, "trend_dir": None, "trend_pct": 0.0, "matched": False}
    if not CARDHEDGER_KEY:
        return out

    # Use card-match (AI-powered) for better accuracy than card-search
    match = ch_card_match(desc)
    if not match:
        return out

    cid = match.get("card_id") or match.get("id")
    if not cid:
        return out

    out["matched"] = True

    # First try: grade-specific comp from the match's prices array
    prices = match.get("prices") or []
    for p in prices:
        if str(p.get("grade","")).upper() == grade.upper():
            try:
                out["comp_avg"] = float(p["price"])
            except Exception:
                pass
            break

    # Second try: all-prices-by-card — ONE cached call covers every grade, so the
    # Reprice loop stops making a separate comps call per grade/card (cost saver).
    if not out["comp_avg"]:
        for p in ch_all_prices(cid):
            if str(p.get("grade", "")).upper() == grade.upper():
                try:
                    out["comp_avg"] = float(p.get("price"))
                except Exception:
                    pass
                break

    # Third try: dedicated comps endpoint (recent-sales average) as a final fallback
    if not out["comp_avg"]:
        cdata = ch_comps(cid, grade) or {}
        out["comp_avg"] = cdata.get("comp_price") or cdata.get("average") or cdata.get("mean")

    # Fall back to Raw if still nothing and the card is graded
    if not out["comp_avg"] and grade != "Raw":
        for p in prices:
            if str(p.get("grade","")).upper() == "RAW":
                try:
                    out["comp_avg"] = float(p["price"])
                    out["grade"] = "Raw (fallback)"
                except Exception:
                    pass
                break

    out["trend_dir"], out["trend_pct"] = calculate_trend(ch_price_history(cid, grade, days=90))
    return out

def suggest_reprice(comp_avg, trend_pct, strategy, adj_pct):
    """Suggested list price from the market comp, shaped by the chosen strategy."""
    if not comp_avg or comp_avg <= 0:
        return None
    if strategy == "Match market":
        return comp_avg * (1 + adj_pct / 100.0)
    if strategy == "Undercut to sell faster":
        return comp_avg * (1 - adj_pct / 100.0)
    if strategy == "List high for offers":
        return comp_avg * (1 + adj_pct / 100.0)
    # Trend-following (default): lean into momentum, capped so a wild swing can't run away
    t = max(-25.0, min(25.0, trend_pct or 0.0))
    move = t * (adj_pct / 100.0)  # adj_pct is sensitivity 0..100
    return comp_avg * (1 + move / 100.0)

def trend_label(direction, pct):
    if not direction:
        return "—"
    if direction == "up":
        return f"↑ Up +{pct:.0f}%"
    if direction == "down":
        return f"↓ Down {pct:.0f}%"
    return f"→ Flat ({pct:+.0f}%)"

# ─── Sport / rookie detection + Supabase listings helpers ────────────────────
_SPORT_KEYS = {
    "Soccer": [
        'premier league','bundesliga','la liga','champions league','serie a',
        'liga mx','fifa','messi','pulisic','yamal','ronaldo','neymar','beckham',
        'vlahovic','gavi','pedri','mbappe','haaland','lewandowski','di maria',
        'topps ucc','panini select fifa','topps now uefa','soccer','futbol',
    ],
    "Basketball": [
        'nba',' hoops ','prizm nba','mosaic nba','select nba','basketball','wnba',
        'hailey van lith','player of the day','timberwolves','nuggets','heat ',
        'bucks ','suns ','clippers','maverick','thunder','pelicans','grizzlies',
        'knicks','lakers','celtics','warriors','strawther','beringer','knueppel',
        'joan beringer','julian strawther',
    ],
    "Football": [
        'nfl','gridiron','panini contenders season ticket','panini prestige',
        'panini absolute football','ohio state university nil',
        'lagway','hartman','pierce','jalen mcmillan','dallas turner',
        'jayden daniels','jordan addison','tuli tuipulotu','tyreek hill','julian sayin',
        'chiefs','eagles','cowboys','patriots','seahawks','49ers','ravens','bengals',
        'bills','commanders','vikings','jaguars','titans','giants','bears ','packers',
        'lions ','browns ','saints','falcons','panthers','cardinals','rams ','chargers',
        'raiders','broncos','dolphins','donruss optic rated rookie',
    ],
    "Baseball": [
        'bowman','topps','upper deck','fleer','stadium club','heritage',
        '1st bowman','minor league','mlb','yankees','braves','cubs','dodgers',
        'red sox','mets','padres','astros','phillies','winfield','palmeiro',
        'topps five star',"bowman's best",'bowman draft',
    ],
}

def detect_sport(title):
    t = (title or "").lower()
    for sport in ("Soccer", "Basketball", "Football", "Baseball"):
        if any(k in t for k in _SPORT_KEYS[sport]):
            return sport
    return "Unknown"

def is_rookie_card(title):
    t = (title or "").lower()
    return any(k in t for k in [' rc ', ' rc/', '(rc)', 'rookie', '1st bowman',
                                  'prospect', 'rated rookie', 'bowman draft'])

def price_freq_for(sport):
    return {"Baseball": "daily", "Soccer": "triweekly"}.get(sport, "weekly")

def needs_pricing_today(last_priced_at_str, freq):
    if not last_priced_at_str:
        return True
    try:
        from datetime import timezone as _tz
        lp = datetime.fromisoformat(last_priced_at_str.replace("Z", "+00:00"))
        now = datetime.now(_tz.utc)
        days = (now - lp).days
        return days >= {"daily": 1, "triweekly": 3}.get(freq, 7)
    except Exception:
        return True

def parse_ebay_csv_to_listings(df):
    col_map = {str(c).lower(): c for c in df.columns}
    def col(name):
        return col_map.get(name.lower(), name)
    rows = []
    for _, r in df.iterrows():
        item_num = str(r.get(col("item number"), "") or "").strip()
        if not item_num:
            continue
        title = str(r.get(col("title"), "") or "").strip()
        try:
            price = float(r.get(col("current price"), 0) or 0)
        except Exception:
            price = 0.0
        sport = detect_sport(title)
        rookie = is_rookie_card(title)
        freq = price_freq_for(sport)
        def _parse_ebay_date(s):
            s = str(s or "").strip()
            if not s or s == "nan":
                return None
            for tz_suffix in [" PDT", " PST", " EDT", " EST"]:
                s = s.replace(tz_suffix, "")
            try:
                return datetime.strptime(s, "%b-%d-%y %H:%M:%S").isoformat()
            except Exception:
                return None
        def _int(v):
            try: return int(v or 0)
            except Exception: return 0
        rows.append({
            "item_number":   item_num,
            "title":         title,
            "sku":           str(r.get(col("custom label (sku)"), "") or "").strip() or None,
            "current_price": price,
            "start_date":    _parse_ebay_date(r.get(col("start date"))),
            "end_date":      _parse_ebay_date(r.get(col("end date"))),
            "watchers":      _int(r.get(col("watchers"))),
            "sold_qty":      _int(r.get(col("sold quantity"))),
            "condition":     str(r.get(col("condition"), "") or "").strip() or None,
            "grader":        str(r.get(col("cd:professional grader - (id: 27501)"), "") or "").strip() or None,
            "grade":         str(r.get(col("cd:grade - (id: 27502)"), "") or "").strip() or None,
            "cert_number":   str(r.get(col("cda:certification number - (id: 27503)"), "") or "").strip() or None,
            "sport":         sport,
            "is_rookie":     rookie,
            "price_freq":    freq,
            "updated_at":    datetime.utcnow().isoformat() + "Z",
        })
    return rows

def _sb_ctx():
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    return ctx

def _sb_base_headers():
    return {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json"}

def upsert_listings(rows):
    if not SUPABASE_URL or not rows:
        return 0
    ctx = _sb_ctx()
    total = 0
    for i in range(0, len(rows), 500):
        chunk = rows[i:i+500]
        body = json.dumps(chunk).encode()
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/listings",
            data=body, method="POST",
            headers={**_sb_base_headers(),
                     "Prefer": "resolution=merge-duplicates,return=minimal"},
        )
        try:
            with urllib.request.urlopen(req, context=ctx, timeout=30):
                total += len(chunk)
        except Exception:
            pass
    return total

def load_listings(min_price=20, limit=1000):
    if not SUPABASE_URL:
        return []
    ctx = _sb_ctx()
    url = (f"{SUPABASE_URL}/rest/v1/listings"
           f"?current_price=gte.{min_price}&order=current_price.desc&limit={limit}")
    req = urllib.request.Request(url, headers=_sb_base_headers())
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
            data = json.loads(r.read().decode())
            return data if isinstance(data, list) else []
    except Exception:
        return []

def update_listing(item_number, updates):
    if not SUPABASE_URL:
        return False
    ctx = _sb_ctx()
    body = json.dumps(updates).encode()
    req = urllib.request.Request(
        f"{SUPABASE_URL}/rest/v1/listings?item_number=eq.{urllib.parse.quote(item_number)}",
        data=body, method="PATCH",
        headers={**_sb_base_headers(), "Prefer": "return=minimal"},
    )
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=10):
            return True
    except Exception:
        return False

def save_listing_pricing(item_number, comp_avg, trend_dir, trend_pct, suggested):
    return update_listing(item_number, {
        "comp_avg": comp_avg, "trend_dir": trend_dir,
        "trend_pct": trend_pct, "suggested_price": suggested,
        "last_priced_at": datetime.utcnow().isoformat() + "Z",
        "updated_at":     datetime.utcnow().isoformat() + "Z",
    })

# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — Hot Movers (top-movers)
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 1:
    st.markdown("## 🔥 Hot Movers")
    st.markdown("The week's biggest price gainers — what's heating up right now. A buy-radar, not a card you already own.")
    if not CARDHEDGER_KEY:
        st.info("📊 Connect the CardHedger API to load hot movers.")
    else:
        hm1, hm2 = st.columns([2, 1])
        cat = hm1.selectbox("Category", ["All", "Basketball", "Baseball", "Football", "Hockey", "Pokemon", "Soccer"], key="hm_cat")
        count = hm2.slider("How many", 10, 100, 25, 5, key="hm_count")
        if st.button("🔥 Load hot movers", key="hm_go"):
            st.session_state["hm_data"] = ch_top_movers(None if cat == "All" else cat, count)
        movers = st.session_state.get("hm_data")
        if movers is not None:
            def _gp(card, grade):
                for p in card.get("prices", []):
                    if p.get("grade") == grade:
                        try:
                            return float(p.get("price"))
                        except Exception:
                            return None
                return None
            rows = []
            for c in movers:
                g = c.get("gain")
                rows.append({
                    "Card": c.get("description", ""),
                    "Gain %": round(float(g), 1) if g is not None else None,
                    "7d Sales": c.get("7 Day Sales"),
                    "30d Sales": c.get("30 Day Sales"),
                    "PSA 10": _gp(c, "PSA 10"),
                    "Raw": _gp(c, "Raw"),
                })
            if rows:
                dfm = pd.DataFrame(rows).sort_values("Gain %", ascending=False)
                st.dataframe(
                    dfm, use_container_width=True, hide_index=True,
                    column_config={
                        "Gain %": st.column_config.NumberColumn(format="%.1f%%"),
                        "PSA 10": st.column_config.NumberColumn(format="$%.2f"),
                        "Raw": st.column_config.NumberColumn(format="$%.2f"),
                    },
                )
                st.caption(f"{len(rows)} movers · {cat}. Gain = recent weekly price change (CardHedger). Cross-check volume — a spike on thin sales is fragile.")
            else:
                st.info("No movers returned for that category — try another.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 8 — Scan (image-match raw card + cert lookup)
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 2:
    st.markdown("## 📷 Scan")
    st.markdown("Identify a card fast — snap a raw card, or look up a graded slab by its cert number.")
    if not CARDHEDGER_KEY:
        st.info("📊 Connect the CardHedger API to use scanning.")
    else:
        # ── TCP helpers shared with AI Batch → eBay export ────────────────
        _AB_TCP_NFL = {'Cardinals','Falcons','Ravens','Bills','Panthers','Bears','Bengals','Browns','Cowboys','Broncos','Lions','Packers','Texans','Colts','Jaguars','Chiefs','Raiders','Chargers','Rams','Dolphins','Vikings','Patriots','Saints','Giants','Jets','Eagles','Steelers','Seahawks','49ers',"49's",'Buccaneers','Titans','Commanders','Niners'}
        _AB_TCP_NBA = {'Hawks','Celtics','Nets','Hornets','Bulls','Cavaliers','Mavericks','Nuggets','Pistons','Warriors','Rockets','Pacers','Clippers','Lakers','Grizzlies','Heat','Bucks','Timberwolves','Pelicans','Knicks','Thunder','Magic','Sixers','76ers','Suns','Blazers','Kings','Spurs','Raptors','Jazz'}
        _AB_TCP_MLB = {'Orioles','Red Sox','Yankees','Rays','Blue Jays','White Sox','Guardians','Tigers','Royals','Twins','Astros','Angels','Athletics',"A's",'Mariners','Rangers','Braves','Marlins','Mets','Phillies','Nationals','Cubs','Reds','Brewers','Pirates','Cardinals','Diamondbacks','Rockies','Dodgers','Padres','Giants','Indians'}
        _AB_TCP_MFRS = ['Panini','Topps','Bowman','Upper Deck','Donruss','Score','Leaf','Fleer',"Collector's Edge",'Pro Set','Hoops','SkyBox','Playoff','Pacific','Wild Card']
        _AB_TCP_SETS = ['Prizm','Mosaic','Select','Phoenix','Illusions','Absolute','Donruss','Chronicles','National Treasures','Flawless','Immaculate','Revolution','Optic','Contenders','Rookies & Stars','Score','Origins','Elements','Obsidian','Hoops','Court Kings','Certified','Noir','Spectra','Gold Standard','Chrome','Series 1','Series 2','Update','Allen & Ginter','Stadium Club','Heritage','Finest','Now','Gypsy Queen','Archives','Opening Day','Holiday','Big League','Draft','Platinum','Sapphire']
        _AB_TCP_PARALLELS = ['LogoFractor','Superfractor','X-Fractor','Refractor','Sandglitter','Elevate','Reactive Blue','Reactive Purple','Reactive','Cracked Ice','Mojo','Disco','Laser','Neon','Pulsar','Holo','Wave','Scope','Atomic','Shimmer','Gold','Silver','Blue','Red','Green','Orange','Purple','Pink','Rainbow','Prizm']
        _AB_TCP_HEADER = ['*Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8)','CustomLabel','*Category','StoreCategory','*Title','Subtitle','Relationship','*ConditionID','*C:Graded','*C:Sport','*C:Player/Athlete','*C:Parallel/Variety','*C:Manufacturer','C:Season','*C:Features','*C:Set','CD:Grade - (ID: 27502)','*C:League','CD:Professional Grader - (ID: 27501)','*C:Team','*C:Autographed','CD:Card Condition - (ID: 40001)','*C:Card Name','*C:Card Number','CDA:Certification Number - (ID: 27503)','*C:Type','C:Signed By','C:Autograph Authentication','C:Year Manufactured','C:Card Size','C:Country/Region of Manufacturer','C:Material','C:Autograph Format','C:Vintage','C:Original/Licensed Reprint','C:Event/Tournament','C:Language','C:Autograph Authentication Number','C:Bundle Description','C:California Prop 65 Warning','C:Card Thickness','C:Custom Bundle','C:Insert Set','C:Print Run','PicURL','GalleryType','*Description','*Format','*Duration','*StartPrice','BuyItNowPrice','*Quantity','PayPalAccepted','PayPalEmailAddress','ImmediatePayRequired','PaymentInstructions','*Location','PostalCode','WeightMajor','WeightMinor','ShippingType','ShippingService-1:Option','ShippingService-1:FreeShipping','ShippingService-1:Cost','ShippingService-1:AdditionalCost','ShippingService-2:Option','ShippingService-2:Cost','*DispatchTimeMax','PromotionalShippingDiscount','ShippingDiscountProfileID','*ReturnsAcceptedOption','ReturnsWithinOption','RefundOption','ShippingCostPaidByOption','AdditionalDetails','ShippingProfileName','ReturnProfileName','PaymentProfileName','TakeBackPolicyID','ProductCompliancePolicyID','ScheduleTime','BestOfferEnabled','MinimumBestOfferPrice','BestOfferAutoAcceptPrice','*C:Rookie','*C:Memorabilia','ActiveListings','SoldListings','Confidence','PricingPulledFrom']
        def _ab_build_desc(title, img_url=""):
            img_tag = (f'<img src="{img_url}" alt="{title}" style="width:100%;max-width:180px;border-radius:12px;border:1px solid #e5e7eb;" />'
                       if img_url else '<div style="width:180px;height:240px;background:#f3f4f6;border-radius:12px;"></div>')
            return (
                '<div style="font-family:Arial,sans-serif;color:#1f2937;line-height:1.6;">'
                '<table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;"><tr>'
                f'<td valign="top" style="width:180px;padding-right:20px;">{img_tag}</td>'
                '<td valign="top">'
                f'<h2 style="margin:0 0 12px;font-size:24px;text-transform:uppercase;">{title}</h2>'
                '<ol style="margin:0 0 20px;padding-left:20px;">'
                '<li style="margin-bottom:10px;">All cards are scanned. If you see any lines, if you want more pictures, just ask!&nbsp;</li>'
                '<li style="margin-bottom:10px;">I send all cards over $20 with tracking, and all mem patches are sent ground advantage to prevent damage.&nbsp;</li>'
                '<li style="margin-bottom:10px;">Cards $75–$199 ship with signature required, which protects you as the buyer.&nbsp;</li>'
                '</ol>'
                '<p style="margin:0;">My goal is to have you receive the card in tip-top shape. No Returns Accepted. Questions, please ask!!!</p>'
                '</td></tr></table></div>'
            )

        def _ab_tcp_sport(sport_str):
            """Map Vision sport string → (sport_col, league_col)."""
            s = (sport_str or '').upper()
            if 'SOCCER' in s or 'FOOTBALL' in s and 'AMERICAN' not in s: pass
            mapping = {
                'BASKETBALL': ('BASKETBALL', 'NBA'),
                'FOOTBALL':   ('FOOTBALL',   'NFL'),
                'BASEBALL':   ('BASEBALL',   'MLB'),
                'SOCCER':     ('SOCCER',     'SOCCER'),
                'HOCKEY':     ('HOCKEY',     'NHL'),
            }
            for key, val in mapping.items():
                if key in s:
                    return val
            return ('BASEBALL', 'MLB')

        def _ab_package_dims(price_str):
            """Return (ShippingPackage, WeightMinor, L, W, D) for a given price.
            WeightMinor is oz; WeightMajor is always 0 (under 1 lb).
            Weights are for single card — buyer/seller adjusts for multiples."""
            try:
                p = float(str(price_str).replace(',', '').strip() or 0)
            except (ValueError, TypeError):
                p = 0
            if p < 20:
                return ('LargeEnvelope', '1', '6', '4', '0.25')     # envelope, 1 oz, 6x4x0.25
            elif p < 50:
                return ('PackageThickEnvelope', '4', '9', '6', '1') # $20–$49: 9x6x1 mailer, 4 oz
            else:
                return ('PackageThickEnvelope', '4', '6', '4', '2') # $50+: 6x4x2 box, 4 oz

        def _ab_shipping_policy(price_str):
            """Return the eBay ShippingProfileName for a given price."""
            try:
                p = float(str(price_str).replace(',', '').strip() or 0)
            except (ValueError, TypeError):
                p = 0
            if p < 20:
                return "Flat: US_eBayStandardEnvelope $.99, 2 busines (315934471021)"
            elif p <= 50:
                return "$15 to $50 Ground"
            elif p <= 100:
                return "$51 to $100 - box"
            elif p <= 200:
                return "$100-$200"
            elif p <= 300:
                return "$201-$300"
            elif p <= 400:
                return "$301-$400"
            elif p <= 500:
                return "$401-500"
            else:
                return "$501-600"

        def _ab_expand_title(base_title, r, target=78):
            """Pad eBay title toward target chars using card metadata."""
            t = base_title.strip()
            team     = str(r.get('Team', '') or '').strip()
            sport    = str(r.get('Sport', '') or '').strip()
            numbered = str(r.get('Numbered', '') or '').strip()
            is_rc    = bool(r.get('Rookie', False))
            notes    = str(r.get('Notes', '') or '').strip()
            extras = []
            if is_rc and 'Rookie' not in t:
                extras.append('Rookie Card')
            if team and team.upper() not in t.upper():
                extras.append(team)
            if numbered and numbered not in t:
                extras.append(numbered)
            if sport and sport.upper() not in t.upper():
                extras.append(sport)
            if notes and len(notes) <= 25 and notes.upper() not in t.upper():
                extras.append(notes)
            extras += ['Sports Trading Card', 'Mint']
            for extra in extras:
                candidate = t + ' ' + extra
                if len(candidate) <= target:
                    t = candidate
                elif len(t) >= target:
                    break
            return t[:80]

        def _ab_make_tcp_row(r, idx):
            """Build a TCP/eBay row dict from a Vision scan result dict."""
            player    = str(r.get('Player', '') or '')
            year      = str(r.get('Year', '') or '')
            set_      = str(r.get('Set', '') or '')
            card_num  = str(r.get('Card #', '') or '')
            parallel  = str(r.get('Parallel', '') or '')
            sport_s   = str(r.get('Sport', '') or '')
            pic_urls  = str(r.get('PicURLs', '') or '')   # real uploaded photos (pipe-sep)
            front_url = str(r.get('FrontURL', '') or '')   # front only, for description img
            card_img  = front_url or str(r.get('CH Image', '') or '')  # fallback to CH ref image
            parts = [p for p in [year, set_, player, parallel, f'#{card_num}' if card_num else ''] if p.strip()]
            title = ' '.join(parts)[:80]
            sport_col, league_col = _ab_tcp_sport(sport_s)
            sku = f'LOT{idx:03d}'
            mfr = next((m for m in _AB_TCP_MFRS if re.search(r'\b'+re.escape(m)+r'\b', set_, re.I)), '')
            set_name = next((s for s in _AB_TCP_SETS if re.search(r'\b'+re.escape(s)+r'\b', set_, re.I)), set_)
            para_col = next((p for p in _AB_TCP_PARALLELS if re.search(r'\b'+re.escape(p)+r'\b', parallel, re.I)), parallel)
            rookie = 'Yes' if re.search(r'\bRC\b|\bRookie\b|\(RC\)', title, re.I) else 'No'
            desc = _ab_build_desc(title.replace('"', '&quot;'), card_img)
            row = {k: '' for k in _AB_TCP_HEADER}
            row.update({
                '*Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8)': 'Add',
                'CustomLabel':                       sku,
                '*Category':                         '261328',
                'StoreCategory':                     '0',
                '*Title':                            title,
                '*ConditionID':                      '4000',
                '*C:Graded':                         'No',
                '*C:Sport':                          sport_col,
                '*C:Player/Athlete':                 player[:50],
                '*C:Parallel/Variety':               para_col,
                '*C:Manufacturer':                   mfr,
                'C:Season':                          year,
                '*C:Features':                       '',
                '*C:Set':                            set_name,
                'CD:Grade - (ID: 27502)':            '',
                '*C:League':                         league_col,
                'CD:Professional Grader - (ID: 27501)': '',
                '*C:Team':                           '',
                '*C:Autographed':                    'No',
                'CD:Card Condition - (ID: 40001)':   EBAY_CONDITION_DEFAULT,
                '*C:Card Name':                      player[:50],
                '*C:Card Number':                    card_num,
                'CDA:Certification Number - (ID: 27503)': '',
                '*C:Type':                           'Sports Trading Card',
                'C:Year Manufactured':               year,
                'C:Print Run':                       '',
                'PicURL':                            pic_urls or card_img,
                'GalleryType':                       'Gallery' if (pic_urls or card_img) else '',
                '*Description':                      desc,
                '*Format':                           'FixedPrice',
                '*Duration':                         'GTC',
                '*StartPrice':                       str(r.get('FMV_raw', '') or ''),
                '*Quantity':                         '1',
                'PayPalAccepted':                    '1',
                'ImmediatePayRequired':              '1',
                '*Location':                         'Scottsdale, AZ',
                'PostalCode':                        '85255',
                'WeightMajor':                       '0',
                'WeightMinor':                       '4',
                'ShippingType':                      '',
                'ShippingService-1:Option':          '',
                'ShippingService-1:FreeShipping':    '',
                'ShippingService-1:Cost':            '',
                'ShippingService-1:AdditionalCost':  '',
                '*DispatchTimeMax':                  '1',
                '*ReturnsAcceptedOption':            'ReturnsNotAccepted',
                'ShippingProfileName':               '',
                'ReturnProfileName':                 'Returns',
                'PaymentProfileName':                'BIN',
                'BestOfferEnabled':                  '1',
                '*C:Rookie':                         rookie,
                '*C:Memorabilia':                    'No',
            })
            return row

        def _ab_make_tcp_csv(rows):
            buf = io.StringIO()
            info = ['Info', 'Version=1.0.0', 'Template=fx_category_template_EBAY_US'] + [''] * (len(_AB_TCP_HEADER) - 3)
            buf.write(','.join(info) + '\n')
            w = csv.DictWriter(buf, fieldnames=_AB_TCP_HEADER)
            w.writeheader()
            w.writerows(rows)
            return buf.getvalue().encode('utf-8')

        scan_search, scan_raw, scan_ai_batch, scan_cert, scan_batch, scan_grade = st.tabs(["🔍 Title Search", "🃏 Raw card photo", "🤖 AI Batch", "🎫 Graded slab (cert #)", "📦 Batch → eBay", "🔬 Grade Predictor"])

        # ── TITLE SEARCH ──────────────────────────────────────────────────────
        with scan_search:
            st.caption("Type any card title — player, year, set, parallel — and get live FMV + trend instantly. No image needed.")

            _sq = st.text_input(
                "Card title",
                placeholder="e.g. 2026 Bowman Chrome Cam Schlittler #53 RC Red",
                key="scan_text_q",
                label_visibility="collapsed",
            )
            _sg = st.radio("Grade", ["Raw", "PSA 10", "PSA 9", "PSA 8"], horizontal=True, key="scan_text_grade")

            if _sq:
                with st.spinner("Looking up…"):
                    _tm = ch_card_match(_sq)
                if not _tm:
                    st.error("No card found — try adding year, set name, or card number.")
                else:
                    _tid  = _tm.get("card_id") or _tm.get("id") or ""
                    _tdsc = _tm.get("description") or _tm.get("name") or _tm.get("title") or _sq
                    _tset = _tm.get("set") or _tm.get("set_name") or ""
                    _tvar = _tm.get("variant") or _tm.get("parallel") or ""

                    st.markdown(f"### {_tdsc}")
                    if _tset or _tvar:
                        st.caption(f"{_tset}{' · ' + _tvar if _tvar else ''}")

                    if _tid:
                        with st.spinner("Loading market data…"):
                            _tc   = ch_comps(_tid, _sg)
                            _tfmv = ch_fmv(_tid, _sg)
                            _th   = ch_price_history(_tid, _sg, 90)
                            _tmeta = ch_card_meta(_tid)

                        _tavg  = _tc.get("comp_price") or _tc.get("average") or _tc.get("mean")
                        _tfmvv = fmv_price(_tfmv)
                        _tshow = _tfmvv or _tavg
                        _tdir, _tpct = calculate_trend(_th)
                        _ts7  = _tmeta.get("7 Day Sales")
                        _ts30 = _tmeta.get("30 Day Sales")

                        # Trend signal — big and clear
                        if _tdir == "up" and _tpct > 5:
                            st.success(f"🔥 **UP {_tpct:+.0f}%** over 90 days — rising demand.")
                        elif _tdir == "down" and _tpct < -5:
                            st.error(f"🛑 **DOWN {abs(_tpct):.0f}%** over 90 days — market softening.")
                        elif _tdir is not None:
                            st.info(f"➡️ **FLAT** ({_tpct:+.0f}% in 90d) — stable market.")
                        else:
                            st.warning("⚠️ Not enough sales history to call trend.")

                        # Price metrics
                        _pc1, _pc2, _pc3, _pc4 = st.columns(4)
                        _pc1.metric(f"{_sg} FMV", f"${_tshow:,.2f}" if _tshow else "—")
                        try:
                            _lo = float(_tfmv.get("price_low") or 0)
                            _hi = float(_tfmv.get("price_high") or 0)
                            _pc2.metric("Range", f"${_lo:,.0f}–${_hi:,.0f}" if _lo and _hi else "—")
                        except Exception:
                            _pc2.metric("Range", "—")
                        _pc3.metric("7-day sales",  str(_ts7)  if _ts7  else "—")
                        _pc4.metric("30-day sales", str(_ts30) if _ts30 else "—")

                        # Recent sales
                        _tsales = []
                        for _k in ("raw_prices", "sales", "comps", "data"):
                            if _k in _tc and isinstance(_tc[_k], list):
                                _tsales = _tc[_k]; break
                        if _tsales:
                            with st.expander("Recent sales"):
                                _srows = []
                                for _s in _tsales[:15]:
                                    try:
                                        _srows.append({
                                            "Date":    _s.get("date") or _s.get("sold_date") or "",
                                            "Price":   f"${float(_s.get('price') or _s.get('sale_price') or 0):,.2f}",
                                            "Grade":   _s.get("grade") or _sg,
                                            "Platform":_s.get("platform") or _s.get("source") or "eBay",
                                        })
                                    except Exception:
                                        pass
                                if _srows:
                                    st.dataframe(pd.DataFrame(_srows), hide_index=True, use_container_width=True)
                    else:
                        st.warning("Card matched but no ID returned — can't pull live pricing.")

        with scan_raw:
            st.caption("Upload a high-res photo of the card front. Claude AI reads the card and pulls live eBay sold comps instantly.")

            if not ANTHROPIC_KEY:
                st.warning("⚠️ Add your Anthropic API key to `.streamlit/secrets.toml` under `[anthropic] api_key = \"sk-ant-...\"` to enable AI card recognition.")

            _up = st.file_uploader("Card photo (front)", type=["jpg", "jpeg", "png", "webp", "heic"], key="scan_up")

            if _up is not None:
                import io as _io
                _up_bytes = _up.getvalue()
                _mime = "image/jpeg"
                if _up.name.lower().endswith(".png"):
                    _mime = "image/png"
                elif _up.name.lower().endswith(".webp"):
                    _mime = "image/webp"

                # Resize to keep API payload reasonable (~1.5MB max)
                try:
                    from PIL import Image as _PIL
                    _img = _PIL.open(_io.BytesIO(_up_bytes))
                    _img.thumbnail((1600, 1600), _PIL.LANCZOS)
                    _buf = _io.BytesIO()
                    _img.save(_buf, format="JPEG", quality=90)
                    _b64 = base64.b64encode(_buf.getvalue()).decode()
                    _mime = "image/jpeg"
                except Exception:
                    _b64 = base64.b64encode(_up_bytes).decode()

                # ── Run Claude Vision identification ──────────────────────
                _scan_file_key = _up.name + str(_up.size)
                if st.session_state.get("_cv_file_key") != _scan_file_key:
                    st.session_state["_cv_file_key"] = _scan_file_key
                    st.session_state.pop("_cv_result", None)
                    st.session_state.pop("_cv_query", None)
                    st.session_state.pop("_cv_ebay_sold", None)

                if "_cv_result" not in st.session_state:
                    with st.spinner("🔍 AI reading card…"):
                        _cv = claude_identify_card(_b64, _mime)
                    st.session_state["_cv_result"] = _cv
                else:
                    _cv = st.session_state["_cv_result"]

                # ── Split: image left, fields right ───────────────────────
                _col_img, _col_fields = st.columns([1, 2])
                with _col_img:
                    st.image(_up_bytes, use_container_width=True)

                with _col_fields:
                    if _cv.get("_error"):
                        st.error(f"Recognition error: {_cv['_error']}")
                        if _cv.get("_raw"):
                            with st.expander("Raw response"):
                                st.text(_cv["_raw"])
                    else:
                        st.markdown("#### 🃏 Identified Card")
                        _rc1, _rc2 = st.columns(2)
                        _cv_player  = _rc1.text_input("Player", value=_cv.get("player", ""), key="cv_player")
                        _cv_year    = _rc2.text_input("Year",   value=_cv.get("year", ""),   key="cv_year")
                        _cv_brand   = _rc1.text_input("Brand",  value=_cv.get("brand", ""),  key="cv_brand")
                        _cv_set     = _rc2.text_input("Set",    value=_cv.get("set", ""),    key="cv_set")
                        _cv_num     = _rc1.text_input("Card #", value=_cv.get("card_number", ""), key="cv_num")
                        _cv_par     = _rc2.text_input("Parallel / Variety", value=_cv.get("parallel", ""), key="cv_par")
                        _cv_sport   = _rc1.text_input("Sport",  value=_cv.get("sport", ""),  key="cv_sport")
                        _cv_team    = _rc2.text_input("Team",   value=_cv.get("team", ""),   key="cv_team")
                        _cv_num_pr  = _rc1.text_input("Print run", value=_cv.get("numbered", ""), key="cv_numbered")
                        _cv_notes   = _rc2.text_input("Notes",  value=_cv.get("notes", ""),  key="cv_notes")

                        # Build the eBay search query from (possibly edited) fields
                        _cv_info_live = {
                            "player": _cv_player, "year": _cv_year, "brand": _cv_brand,
                            "set": _cv_set, "card_number": _cv_num, "parallel": _cv_par,
                            "numbered": _cv_num_pr,
                        }
                        _cv_query = build_card_query(_cv_info_live)
                        st.caption(f"**Search query:** `{_cv_query}`")

                # ── eBay sold comps ────────────────────────────────────────
                if not _cv.get("_error") and _cv_query:
                    st.divider()
                    _comp_col1, _comp_col2 = st.columns([3, 1])
                    with _comp_col1:
                        st.markdown("### 📊 eBay Sold Comps")
                    with _comp_col2:
                        _cv_grade = st.selectbox("Grade", ["Raw", "PSA 10", "PSA 9", "PSA 8"], key="cv_grade_sel")

                    # Append grade to query if not Raw
                    _ebay_q = _cv_query if _cv_grade == "Raw" else f"{_cv_query} {_cv_grade}"

                    if st.session_state.get("_cv_query") != _ebay_q:
                        st.session_state["_cv_query"] = _ebay_q
                        st.session_state.pop("_cv_ebay_sold", None)

                    if "_cv_ebay_sold" not in st.session_state:
                        with st.spinner("Pulling eBay sold listings…"):
                            _cv_sold = fetch_ebay_sold(_ebay_q, DEFAULT_EBAY_KEY, max_results=20)
                        st.session_state["_cv_ebay_sold"] = _cv_sold
                    else:
                        _cv_sold = st.session_state["_cv_ebay_sold"]

                    if _cv_sold:
                        _cv_prices = [x["price"] for x in _cv_sold if x.get("price", 0) > 0]
                        _cv_avg    = ebay_avg(_cv_sold)
                        _cv_lo     = min(_cv_prices) if _cv_prices else None
                        _cv_hi     = max(_cv_prices) if _cv_prices else None

                        _m1, _m2, _m3, _m4 = st.columns(4)
                        _m1.metric("Avg sold", f"${_cv_avg:,.2f}" if _cv_avg else "—")
                        _m2.metric("Low",  f"${_cv_lo:,.2f}"  if _cv_lo  else "—")
                        _m3.metric("High", f"${_cv_hi:,.2f}"  if _cv_hi  else "—")
                        _m4.metric("Sales found", str(len(_cv_sold)))

                        _cv_rows = []
                        for _s in _cv_sold:
                            _cv_rows.append({
                                "Title": _s.get("title", "")[:60],
                                "Sold": f"${_s['price']:,.2f}",
                                "Date": _s.get("date", "")[:10],
                                "Link": _s.get("url", ""),
                            })
                        _cv_df = pd.DataFrame(_cv_rows)
                        st.dataframe(
                            _cv_df,
                            use_container_width=True,
                            hide_index=True,
                            height=320,
                            column_config={
                                "Title": st.column_config.TextColumn("Title", width="large"),
                                "Sold":  st.column_config.TextColumn("Sold",  width="small"),
                                "Date":  st.column_config.TextColumn("Date",  width="small"),
                                "Link":  st.column_config.LinkColumn("eBay",  width="small"),
                            },
                        )
                        st.caption(f"eBay completed sales · query: `{_ebay_q}` · avg trims top/bottom 10%")

                        # Narrow query option (strip parallel for base card comparison)
                        if _cv_par:
                            with st.expander("🔍 Compare without parallel (base card comps)"):
                                _base_q = build_card_query(_cv_info_live, include_parallel=False)
                                _base_q_g = _base_q if _cv_grade == "Raw" else f"{_base_q} {_cv_grade}"
                                _base_sold = fetch_ebay_sold(_base_q_g, DEFAULT_EBAY_KEY, max_results=15)
                                if _base_sold:
                                    _base_avg = ebay_avg(_base_sold)
                                    st.metric("Base card avg", f"${_base_avg:,.2f}" if _base_avg else "—")
                                    _brows = [{"Title": s.get("title","")[:60], "Sold": f"${s['price']:,.2f}", "Date": s.get("date","")[:10]} for s in _base_sold]
                                    st.dataframe(pd.DataFrame(_brows), hide_index=True, use_container_width=True)
                                else:
                                    st.info("No base card comps found.")
                    else:
                        st.warning("No eBay sold results — try editing the fields above to narrow or broaden the search.")
                        st.markdown(f"🔗 [Search eBay sold manually](https://www.ebay.com/sch/i.html?_nkw={urllib.parse.quote_plus(_ebay_q)}&LH_Sold=1&LH_Complete=1)")

                    # ── CardHedger FMV (secondary source) ─────────────────
                    if CARDHEDGER_KEY and _cv_query:
                        with st.expander("📈 CardHedger FMV (cross-check)"):
                            with st.spinner("CardHedger lookup…"):
                                _ch_match = ch_card_match(_cv_query)
                            if _ch_match:
                                _ch_id  = _ch_match.get("card_id") or _ch_match.get("id") or ""
                                _ch_dsc = _ch_match.get("description") or _ch_match.get("name") or _ch_match.get("title") or ""
                                if _ch_dsc:
                                    st.caption(f"Matched: **{_ch_dsc}**")
                                if _ch_id:
                                    _ch_fmv_r = ch_fmv(_ch_id, _cv_grade)
                                    _ch_fmv_v = fmv_price(_ch_fmv_r)
                                    _ch_comp_r = ch_comps(_ch_id, _cv_grade)
                                    _ch_avg = _ch_comp_r.get("comp_price") or _ch_comp_r.get("average")
                                    _ch_show = _ch_fmv_v or _ch_avg
                                    if _ch_show:
                                        st.metric(f"CardHedger FMV ({_cv_grade})", f"${_ch_show:,.2f}")
                                    else:
                                        st.info("No CardHedger FMV for this card/grade.")
                            else:
                                st.info("No CardHedger match for this query.")

        # ── AI BATCH SCANNER ──────────────────────────────────────────────────
        with scan_ai_batch:
            st.caption("Drop all card photos into one box — interleaved front/back (front1, back1, front2, back2…). Claude reads the fronts, backs become corner crops. ~1¢ per card.")

            if not ANTHROPIC_KEY:
                st.warning("⚠️ Add your Anthropic API key to `.streamlit/secrets.toml` under `[anthropic] api_key = \"sk-ant-...\"` to enable AI card recognition.")

            _ab_all_files = st.file_uploader(
                "Drop all card photos here — front1, back1, front2, back2… (or fronts only)",
                type=["jpg", "jpeg", "png", "webp"],
                accept_multiple_files=True,
                key="ai_batch_files",
            )

            # Split interleaved into fronts + backs
            if _ab_all_files and len(_ab_all_files) > 1 and len(_ab_all_files) % 2 == 0:
                _ab_files      = _ab_all_files[0::2]   # even indices = fronts
                _ab_back_files = _ab_all_files[1::2]   # odd indices  = backs
                st.caption(f"✅ {len(_ab_files)} fronts + {len(_ab_back_files)} backs detected (interleaved)")
            else:
                _ab_files      = _ab_all_files or []
                _ab_back_files = []
                if _ab_all_files and len(_ab_all_files) % 2 != 0:
                    st.caption(f"📷 {len(_ab_files)} front(s) only — add a back for every front to enable corner crops")

            # imgbb key — stored in secrets.toml or entered here
            if not IMGBB_KEY:
                with st.expander("📸 imgbb auto-upload (free — required for automatic eBay photo URLs)"):
                    st.markdown("Get a free API key at [imgbb.com/api](https://api.imgbb.com/) — takes 30 seconds.")
                    st.text_input("imgbb API key", key="imgbb_key_input", type="password",
                                  placeholder="Paste key here — or add to secrets.toml under [imgbb] api_key")
            else:
                st.success("📸 imgbb connected — all 8 photos per card will auto-upload to imgbb and land in the eBay CSV", icon="✅")

            _ab_grade = st.radio("Grade for comps", ["Raw", "PSA 10", "PSA 9", "PSA 8"], horizontal=True, key="ai_batch_grade")
            _ab_run_comps = st.checkbox("Pull CardHedger comps + trend after scan", value=True, key="ai_batch_comps_chk")

            if _ab_files:
                _ab_n = len(_ab_files)
                _ab_n_back = len(_ab_back_files) if _ab_back_files else 0
                _ab_cost_lo = _ab_n * 0.007
                _ab_cost_hi = _ab_n * 0.014
                _ab_has_imgbb = bool(IMGBB_KEY or st.session_state.get("imgbb_key_input", ""))
                _ab_img_note = (f" · {_ab_n_back} back{'s' if _ab_n_back != 1 else ''} → 8 photos/card auto-uploaded to imgbb + ZIP" if (_ab_n_back and _ab_has_imgbb)
                                else f" · {_ab_n_back} back{'s' if _ab_n_back != 1 else ''} → ZIP only (add imgbb key to auto-upload)" if _ab_n_back
                                else " · No backs uploaded — add backs for corner crops + imgbb upload")
                st.info(
                    f"**{_ab_n} front{'s' if _ab_n != 1 else ''}** · Claude Vision cost: ~${_ab_cost_lo:.2f}–${_ab_cost_hi:.2f}{_ab_img_note}"
                )
                if _ab_n_back and _ab_n_back != _ab_n:
                    st.warning(f"⚠️ {_ab_n} fronts but {_ab_n_back} backs — they must match in count and order. Extra backs will be ignored.", icon="⚠️")

                if ANTHROPIC_KEY and st.button("🤖 Scan All Cards", type="primary", key="ai_batch_go"):
                    import io as _aio
                    import zipfile as _ab_zipmod
                    from PIL import Image as _PILB

                    # Helper: crop one corner from a PIL image
                    _AB_CROP_PCT = 0.28
                    def _ab_corner_bytes(img, side):
                        w, h = img.size
                        cx, cy = int(w * _AB_CROP_PCT), int(h * _AB_CROP_PCT)
                        regions = [("tl",(0,0,cx,cy)), ("tr",(w-cx,0,w,cy)),
                                   ("bl",(0,h-cy,cx,h)), ("br",(w-cx,h-cy,w,h))]
                        out = {}
                        for name, box in regions:
                            buf = _aio.BytesIO()
                            img.crop(box).convert("RGB").save(buf, format="JPEG", quality=92)
                            out[f"{side}_{name}"] = buf.getvalue()
                        return out

                    def _ab_img_to_bytes(img):
                        buf = _aio.BytesIO()
                        img.convert("RGB").save(buf, format="JPEG", quality=92)
                        return buf.getvalue()

                    _ab_results = []
                    _ab_img_zips = {}   # card_index → zip bytes (only when back provided)
                    _ab_bar    = st.progress(0, text="Starting…")
                    _ab_status = st.empty()

                    for _abi, _abf in enumerate(_ab_files):
                        _ab_status.markdown(f"Scanning **{_abi + 1}/{_ab_n}**: `{_abf.name}`")

                        # Load + resize front (for Vision)
                        try:
                            _abimg_front = _PILB.open(_aio.BytesIO(_abf.getvalue()))
                            _abimg_front.thumbnail((1600, 1600), _PILB.LANCZOS)
                            _abbuf = _aio.BytesIO()
                            _abimg_front.save(_abbuf, format="JPEG", quality=90)
                            _abb64 = base64.b64encode(_abbuf.getvalue()).decode()
                            _abmime = "image/jpeg"
                        except Exception:
                            _abb64  = base64.b64encode(_abf.getvalue()).decode()
                            _abmime = "image/jpeg"
                            _abimg_front = None

                        # Step 1: Claude Vision — front only, no tokens wasted on back
                        _abcv = claude_identify_card(_abb64, _abmime)
                        _ab_query = build_card_query(_abcv) if not _abcv.get("_error") else ""

                        # Step 2: CardHedger comps + trend
                        _ab_ch_match = ""
                        _ab_ch_card_num = ""   # card number from CH match — fallback when vision misses it
                        _ab_fmv = ""
                        _ab_fmv_raw = ""
                        _ab_comp_avg = ""
                        _ab_low = ""
                        _ab_high = ""
                        _ab_trend = ""

                        if not _abcv.get("_error") and _ab_run_comps and _ab_query and CARDHEDGER_KEY:
                            _ab_status.markdown(f"Pricing **{_abi + 1}/{_ab_n}**: `{_abf.name}`")
                            _abm = ch_card_match(_ab_query)
                            if _abm:
                                _ab_ch_id  = _abm.get("card_id") or _abm.get("id") or ""
                                _ab_ch_match = (_abm.get("description") or _abm.get("name") or _abm.get("title") or "")[:50]
                                # Card number from CH — fallback when not visible on front scan
                                _ab_ch_card_num = str(_abm.get("number") or _abm.get("card_number") or "").strip()
                                if not _ab_ch_card_num:
                                    # Try extracting from CH match title via regex
                                    _cn_m = re.search(r'\b([A-Z]{1,4}-?\d{1,4}|\d{1,4})\b', _ab_ch_match)
                                    if _cn_m:
                                        _ab_ch_card_num = _cn_m.group(1)
                                if _ab_ch_id:
                                    _ab_fmv_r  = ch_fmv(_ab_ch_id, _ab_grade)
                                    _ab_comp_r = ch_comps(_ab_ch_id, _ab_grade)
                                    _ab_hist_r = ch_price_history(_ab_ch_id, _ab_grade, 90)

                                    _ab_fmv_v = fmv_price(_ab_fmv_r)
                                    _ab_fmv   = f"${_ab_fmv_v:,.2f}" if _ab_fmv_v else ""
                                    _ab_fmv_raw = f"{_ab_fmv_v:.2f}" if _ab_fmv_v else ""

                                    _ab_avg_v = _ab_comp_r.get("comp_price") or _ab_comp_r.get("average")
                                    _ab_comp_avg = f"${float(_ab_avg_v):,.2f}" if _ab_avg_v else ""

                                    _ab_prices = [x.get("price") or x.get("sale_price") for x in (_ab_comp_r.get("raw_prices") or _ab_comp_r.get("sales") or [])]
                                    _ab_prices = [float(p) for p in _ab_prices if p]
                                    _ab_low  = f"${min(_ab_prices):,.2f}" if _ab_prices else ""
                                    _ab_high = f"${max(_ab_prices):,.2f}" if _ab_prices else ""

                                    _ab_dir, _ab_pct = calculate_trend(_ab_hist_r)
                                    if _ab_dir == "up":
                                        _ab_trend = f"↑ {_ab_pct:+.0f}%"
                                    elif _ab_dir == "down":
                                        _ab_trend = f"↓ {_ab_pct:.0f}%"
                                    elif _ab_dir is not None:
                                        _ab_trend = f"→ {_ab_pct:+.0f}%"

                        # Step 3: eBay image pack (front + back corners) + imgbb auto-upload
                        _ab_has_images = False
                        _ab_pic_urls   = ""   # pipe-separated for eBay PicURL column
                        _ab_front_url  = ""   # just the front, for description <img>
                        _ab_use_imgbb  = bool(IMGBB_KEY or st.session_state.get("imgbb_key_input", ""))
                        if _abimg_front and _ab_back_files and _abi < len(_ab_back_files):
                            try:
                                _ab_status.markdown(f"Generating images **{_abi + 1}/{_ab_n}**: `{_abf.name}`")
                                _abimg_back = _PILB.open(_aio.BytesIO(_ab_back_files[_abi].getvalue()))
                                _ab_slug = re.sub(r"[^\w]+", "_", (
                                    _abcv.get("player") or _abf.name.rsplit(".",1)[0]
                                ).strip()).strip("_")[:30]
                                _ab_prefix = f"card_{_abi+1:02d}_{_ab_slug}"

                                _ab_front_corners = _ab_corner_bytes(_abimg_front, "front")
                                _ab_back_corners  = _ab_corner_bytes(_abimg_back,  "back")

                                # Build ordered list of (filename, bytes) — same order = eBay display order
                                _ab_img_files = [
                                    (f"{_ab_prefix}_1_front.jpg",    _ab_img_to_bytes(_abimg_front)),
                                    (f"{_ab_prefix}_2_back.jpg",     _ab_img_to_bytes(_abimg_back)),
                                    (f"{_ab_prefix}_3_front_tl.jpg", _ab_front_corners["front_tl"]),
                                    (f"{_ab_prefix}_4_front_tr.jpg", _ab_front_corners["front_tr"]),
                                    (f"{_ab_prefix}_5_front_bl.jpg", _ab_front_corners["front_bl"]),
                                    (f"{_ab_prefix}_6_front_br.jpg", _ab_front_corners["front_br"]),
                                    (f"{_ab_prefix}_7_back_tl.jpg",  _ab_back_corners["back_tl"]),
                                    (f"{_ab_prefix}_8_back_tr.jpg",  _ab_back_corners["back_tr"]),
                                ]

                                _ab_zip_buf = _aio.BytesIO()
                                with _ab_zipmod.ZipFile(_ab_zip_buf, "w", _ab_zipmod.ZIP_DEFLATED) as _abzf:
                                    for _fn, _fb in _ab_img_files:
                                        _abzf.writestr(_fn, _fb)
                                _ab_zip_buf.seek(0)
                                _ab_img_zips[_abi] = (_ab_prefix, _ab_zip_buf.getvalue())
                                _ab_has_images = True

                                # Auto-upload all 8 to imgbb → public URLs for eBay CSV
                                if _ab_use_imgbb:
                                    _ab_status.markdown(f"Uploading images **{_abi + 1}/{_ab_n}** to imgbb…")
                                    _uploaded_urls = []
                                    for _fn, _fb in _ab_img_files:
                                        _u = imgbb_upload(_fb, name=_fn.rsplit(".",1)[0])
                                        if _u:
                                            _uploaded_urls.append(_u)
                                    if _uploaded_urls:
                                        _ab_front_url = _uploaded_urls[0]
                                        _ab_pic_urls  = "|".join(_uploaded_urls)
                            except Exception:
                                pass

                        _ab_ebay_q   = urllib.parse.quote_plus(_ab_query)
                        _ab_sold_url = f"https://www.ebay.com/sch/i.html?_nkw={_ab_ebay_q}&_sacat=261328&LH_Sold=1&LH_Complete=1"
                        _ab_results.append({
                            "#":           _abi + 1,
                            "Player":      _abcv.get("player", ""),
                            "Year":        _abcv.get("year", ""),
                            "Set":         _abcv.get("set", ""),
                            "Card #":      _abcv.get("card_number", "") or _ab_ch_card_num,
                            "Parallel":    _abcv.get("parallel", ""),
                            "Sport":       _abcv.get("sport", ""),
                            "Team":        _abcv.get("team", ""),
                            "Rookie":      bool(_abcv.get("rookie", False)),
                            "Numbered":    str(_abcv.get("numbered", "") or ""),
                            "Notes":       str(_abcv.get("notes", "") or ""),
                            "CH Match":    _ab_ch_match,
                            "CH Image":    (_abm.get("image", "") if _abm else ""),
                            "PicURLs":     _ab_pic_urls,
                            "FrontURL":    _ab_front_url,
                            "FMV":         _ab_fmv,
                            "FMV_raw":     _ab_fmv_raw,
                            "Comp Avg":    _ab_comp_avg,
                            "Low":         _ab_low,
                            "High":        _ab_high,
                            "Trend (90d)": _ab_trend,
                            "Images":      "✅ 8 images" if _ab_has_images else ("—" if not _ab_back_files else "⚠️ error"),
                            "🔍 Sold":     _ab_sold_url,
                            "Query":       _ab_query,
                            "Status":      ("Error: " + _abcv["_error"]) if _abcv.get("_error") else "OK",
                        })

                        _ab_bar.progress((_abi + 1) / _ab_n, text=f"Done {_abi + 1}/{_ab_n}")

                    _ab_status.empty()
                    _ab_bar.empty()
                    st.session_state["ai_batch_results"] = _ab_results
                    st.session_state["ai_batch_img_zips"] = _ab_img_zips
                    st.success(f"✅ Scanned {len(_ab_results)} cards!" + (f" · {len(_ab_img_zips)} eBay image packs ready." if _ab_img_zips else ""))
                    st.rerun()

            if st.session_state.get("ai_batch_results"):
                _ab_res      = st.session_state["ai_batch_results"]
                _ab_img_zips = st.session_state.get("ai_batch_img_zips", {})
                _ab_df  = pd.DataFrame(_ab_res)
                st.markdown(f"### Results — {len(_ab_res)} cards · grade: {_ab_grade}")

                # ── Card Image Preview ──────────────────────────────────────────
                _ab_prev_opts = [
                    f"#{r['#']} — {r.get('Player','?')}  {r.get('Year','')}  {r.get('Set','')}"
                    for r in _ab_res
                ]
                _ab_prev_idx = st.selectbox(
                    "📸 Card Preview", _ab_prev_opts, index=0,
                    key="ab_preview_sel",
                ).split("—")[0].strip().lstrip("#")
                try:
                    _ab_prev_idx = int(_ab_prev_idx) - 1
                except Exception:
                    _ab_prev_idx = 0
                _ab_prev_r     = _ab_res[_ab_prev_idx]
                _ab_prev_pics  = [u.strip() for u in str(_ab_prev_r.get("PicURLs","")).split("|") if u.strip()]
                _ab_prev_front = _ab_prev_r.get("FrontURL","") or (_ab_prev_pics[0] if _ab_prev_pics else "")
                _ab_prev_back  = _ab_prev_pics[1] if len(_ab_prev_pics) > 1 else ""
                _ab_prev_ch    = _ab_prev_r.get("CH Image","")
                _ab_pcols = st.columns(3 if (_ab_prev_front and _ab_prev_back and _ab_prev_ch) else 2 if (_ab_prev_front and _ab_prev_back) else 1)
                if _ab_prev_front:
                    _ab_pcols[0].image(_ab_prev_front, caption="Front ↕ click to expand", use_container_width=True)
                if _ab_prev_back and len(_ab_pcols) > 1:
                    _ab_pcols[1].image(_ab_prev_back, caption="Back ↕ click to expand", use_container_width=True)
                if _ab_prev_ch and len(_ab_pcols) > 2:
                    _ab_pcols[2].image(_ab_prev_ch, caption=f"📚 CH: {_ab_prev_r.get('CH Match','Reference')[:40]}", use_container_width=True)
                elif _ab_prev_ch and not _ab_prev_back:
                    _ab_pcols[1].image(_ab_prev_ch, caption=f"📚 CH: {_ab_prev_r.get('CH Match','Reference')[:40]}", use_container_width=True)
                st.markdown("---")

                # ── Results table ───────────────────────────────────────────────
                _ab_table_cols = ["#","FrontURL","Player","Year","Set","Card #","Parallel","Sport","FMV","Comp Avg","Low","High","Trend (90d)","Images","🔍 Sold","Status"]
                _ab_df_disp = _ab_df[[c for c in _ab_table_cols if c in _ab_df.columns]]
                st.dataframe(
                    _ab_df_disp,
                    hide_index=True,
                    use_container_width=True,
                    height=min(600, 45 + len(_ab_res) * 36),
                    column_config={
                        "#":           st.column_config.NumberColumn("#",           width="small"),
                        "FrontURL":    st.column_config.ImageColumn("📸",           width="small"),
                        "Player":      st.column_config.TextColumn("Player",        width="medium"),
                        "Year":        st.column_config.TextColumn("Year",          width="small"),
                        "Set":         st.column_config.TextColumn("Set",           width="medium"),
                        "Card #":      st.column_config.TextColumn("Card #",        width="small"),
                        "Parallel":    st.column_config.TextColumn("Parallel",      width="medium"),
                        "Sport":       st.column_config.TextColumn("Sport",         width="small"),
                        "FMV":         st.column_config.TextColumn("FMV",           width="small"),
                        "Comp Avg":    st.column_config.TextColumn("Comp Avg",      width="small"),
                        "Low":         st.column_config.TextColumn("Low",           width="small"),
                        "High":        st.column_config.TextColumn("High",          width="small"),
                        "Trend (90d)": st.column_config.TextColumn("Trend (90d)",   width="small"),
                        "Images":      st.column_config.TextColumn("Images",        width="small"),
                        "🔍 Sold":     st.column_config.LinkColumn("🔍 Sold", display_text="eBay ↗", width="small"),
                        "Status":      st.column_config.TextColumn("Status",        width="small"),
                    },
                )

                # ── Download row ───────────────────────────────────────────
                _ab_btn_cols = st.columns([1, 1, 3])
                _ab_csv = _ab_df.to_csv(index=False)
                _ab_btn_cols[0].download_button("📥 Export CSV", _ab_csv, "ai_batch_scan.csv", "text/csv", key="ai_batch_dl")

                if _ab_img_zips:
                    # Bundle all per-card ZIPs into one master ZIP
                    import io as _abio2
                    import zipfile as _abzm2
                    _ab_master_buf = _abio2.BytesIO()
                    with _abzm2.ZipFile(_ab_master_buf, "w", _abzm2.ZIP_DEFLATED) as _ab_mzf:
                        for _ab_idx, (_ab_pfx, _ab_zbytes) in _ab_img_zips.items():
                            # Each card's 8 images go into their own sub-folder
                            import zipfile as _abzm3
                            _ab_inner = _abio2.BytesIO(_ab_zbytes)
                            with _abzm3.ZipFile(_ab_inner, "r") as _ab_inner_zf:
                                for _ab_inner_name in _ab_inner_zf.namelist():
                                    _ab_mzf.writestr(
                                        f"{_ab_pfx}/{_ab_inner_name}",
                                        _ab_inner_zf.read(_ab_inner_name)
                                    )
                    _ab_master_buf.seek(0)
                    _ab_btn_cols[1].download_button(
                        f"📸 Download {len(_ab_img_zips)} eBay image packs",
                        _ab_master_buf.getvalue(),
                        "ebay_images_all_cards.zip",
                        "application/zip",
                        key="ai_batch_img_dl",
                        type="primary",
                    )
                    st.caption("ZIP contains one folder per card, each with 8 images: front, back, 4 front corners, 2 back corners. Unzip → open card folder → drag all 8 into eBay.")

                if _ab_btn_cols[2].button("🗑 Clear results", key="ai_batch_clear"):
                    st.session_state.pop("ai_batch_results", None)
                    st.session_state.pop("ai_batch_img_zips", None)
                    st.rerun()

                # ── Step 2 — Pricing ──────────────────────────────────────────
                st.markdown("---")
                st.markdown("#### Step 2 — Pricing")
                _ab_tcp_rows   = [_ab_make_tcp_row(r, i + 1) for i, r in enumerate(_ab_res)]
                # Auto-expand all titles to ~80 chars
                for _i2, (_r2, _row2) in enumerate(zip(_ab_res, _ab_tcp_rows)):
                    _row2['*Title'] = _ab_expand_title(_row2.get('*Title', ''), _r2)

                # ── Editable title + price review ─────────────────────────────
                st.markdown("**Review titles & prices before downloading** (click any cell to edit)")
                _ab_edit_df = pd.DataFrame([{
                    'SKU':   row.get('CustomLabel', f'LOT{i+1:03d}'),
                    'Title': row.get('*Title', ''),
                    'Chars': len(row.get('*Title', '')),
                    'Price': row.get('*StartPrice', ''),
                    'Source': 'CH ✅' if _ab_res[i].get('FMV_raw') else '— needs TCP',
                } for i, row in enumerate(_ab_tcp_rows)])
                _ab_edited_df = st.data_editor(
                    _ab_edit_df,
                    hide_index=True,
                    use_container_width=True,
                    height=min(500, 55 + len(_ab_tcp_rows) * 40),
                    column_config={
                        'SKU':    st.column_config.TextColumn('SKU',     width='small',  disabled=True),
                        'Title':  st.column_config.TextColumn('Title (80 max)', width='large'),
                        'Chars':  st.column_config.NumberColumn('Chars', width='small',  disabled=True),
                        'Price':  st.column_config.NumberColumn('$ Price', width='small', format='$%.2f'),
                        'Source': st.column_config.TextColumn('Priced By', width='small', disabled=True),
                    },
                    key='ab_title_editor',
                )
                # Apply edited titles and prices back to tcp_rows
                for _ei, _erow in _ab_edited_df.iterrows():
                    if _ei < len(_ab_tcp_rows):
                        _new_title = str(_erow.get('Title') or _ab_tcp_rows[_ei].get('*Title', ''))[:80]
                        _ab_tcp_rows[_ei]['*Title'] = _new_title
                        _new_price = _erow.get('Price')
                        if _new_price is not None and str(_new_price).strip() not in ('', 'nan', 'None'):
                            _ab_tcp_rows[_ei]['*StartPrice'] = f"{float(_new_price):.2f}"

                _ab_ch_priced  = sum(1 for row in _ab_tcp_rows if str(row.get('*StartPrice') or '').strip())
                _ab_ch_missing = len(_ab_tcp_rows) - _ab_ch_priced

                # ── Option A: CardHedger prices → direct eBay CSV ─────────────
                if _ab_ch_priced > 0:
                    if _ab_ch_priced == len(_ab_res):
                        st.success(f"✅ All {len(_ab_res)} cards priced by CardHedger — ready to list directly!")
                    else:
                        st.success(f"✅ {_ab_ch_priced}/{len(_ab_res)} cards priced by CardHedger")
                        st.warning(f"⚠️ {_ab_ch_missing} card(s) not found in CardHedger — use Option B below for those")
                    # Build eBay-ready CSV using CH prices for all CH-priced cards
                    _ab_ch_ebay_rows = [row for row in _ab_tcp_rows if row.get('*StartPrice')]
                    _ab_ch_ebay_csv  = _ab_make_tcp_csv(_ab_ch_ebay_rows)
                    st.download_button(
                        f"⚡ Download eBay CSV — {_ab_ch_priced} cards (CardHedger prices)",
                        _ab_ch_ebay_csv,
                        "lot_scan_ebay_ch_prices.csv",
                        "text/csv",
                        key="ab_ch_ebay_dl",
                        type="primary",
                    )
                    st.caption("Prices from CardHedger FMV · shipping policy auto-set by price tier · upload directly to eBay File Exchange")
                else:
                    st.warning("CardHedger couldn't price any cards in this batch — use TradingCardPricer.com below")

                # ── Option B: TradingCardPricer fallback ──────────────────────
                if _ab_ch_missing > 0 or _ab_ch_priced == 0:
                    st.markdown(f"**{'Option B' if _ab_ch_priced > 0 else 'Option A'} — TradingCardPricer.com** ({_ab_ch_missing if _ab_ch_priced > 0 else len(_ab_res)} cards need pricing)")
                    # Export only unpriced cards to TCP
                    _ab_tcp_unpriced = [row for row in _ab_tcp_rows if not row.get('*StartPrice')]
                    _ab_tcp_csv = _ab_make_tcp_csv(_ab_tcp_unpriced if _ab_tcp_unpriced else _ab_tcp_rows)
                    st.download_button(
                        "📊 Download TCP Upload CSV",
                        _ab_tcp_csv,
                        "lot_scan_tcp_upload.csv",
                        "text/csv",
                        key="ab_tcp_dl",
                    )
                    st.caption(f"{len(_ab_tcp_unpriced or _ab_tcp_rows)} cards · upload to [tradingcardpricer.com](https://tradingcardpricer.com) · TCP fills in *StartPrice and returns a priced CSV")

                # ── Step 3 — Import TCP Results → eBay Listing CSV ────────────
                st.markdown("---")
                st.markdown("#### Step 3 — Import TCP Results → eBay Listing CSV")
                st.caption("Upload the priced CSV you downloaded from TradingCardPricer.com. The app builds a ready-to-upload eBay Add CSV with prices filled in.")
                _ab_tcp_result_file = st.file_uploader(
                    "TCP results CSV (from TradingCardPricer.com)",
                    type=["csv"],
                    key="ab_tcp_result_upload",
                    label_visibility="visible",
                )
                if _ab_tcp_result_file:
                    _ab_tcp_result_bytes = _ab_tcp_result_file.read()
                    _ab_tcp_result_text  = _ab_tcp_result_bytes.decode('utf-8-sig')
                    # TCP result CSVs have an Info header row — skip it
                    _ab_tcp_result_lines = _ab_tcp_result_text.splitlines()
                    _ab_tcp_skip = 0
                    if _ab_tcp_result_lines and _ab_tcp_result_lines[0].startswith('Info,'):
                        _ab_tcp_skip = 1
                    _ab_tcp_result_rows = list(csv.DictReader(_ab_tcp_result_lines[_ab_tcp_skip:]))
                    if _ab_tcp_result_rows:
                        # Build preview table
                        _ab_prev_cols = ['CustomLabel', '*Title', '*StartPrice', 'Confidence', 'PricingPulledFrom']
                        _ab_prev_data = [
                            {c: row.get(c, '') for c in _ab_prev_cols}
                            for row in _ab_tcp_result_rows
                        ]
                        st.dataframe(
                            pd.DataFrame(_ab_prev_data),
                            hide_index=True,
                            use_container_width=True,
                            height=min(400, 45 + len(_ab_prev_data) * 36),
                            column_config={
                                'CustomLabel':      st.column_config.TextColumn('SKU',          width='small'),
                                '*Title':           st.column_config.TextColumn('Title',         width='large'),
                                '*StartPrice':      st.column_config.TextColumn('Price',         width='small'),
                                'Confidence':       st.column_config.TextColumn('Confidence',    width='small'),
                                'PricingPulledFrom':st.column_config.TextColumn('Priced From',   width='medium'),
                            },
                        )
                        # Build eBay Add CSV — pull price from TCP, rebuild our description (TCP overwrites it)
                        # Build SKU → scan result lookup so we can restore PicURL + FrontURL
                        _ab_sku_map = {f"LOT{i+1:03d}": r for i, r in enumerate(_ab_res)}
                        _ab_ebay_rows = []
                        for _tcp_r in _ab_tcp_result_rows:
                            _ebay_r = {k: '' for k in _AB_TCP_HEADER}
                            for k in _AB_TCP_HEADER:
                                if k in _tcp_r:
                                    _ebay_r[k] = _tcp_r[k]
                            # Restore our branded description — TCP replaces it with their own
                            _ab_sku       = str(_tcp_r.get('CustomLabel', '') or '').strip().upper()
                            _ab_orig      = _ab_sku_map.get(_ab_sku, {})
                            _ab_title_r   = str(_tcp_r.get('*Title', '') or '').strip()
                            _ab_front_r   = str(_ab_orig.get('FrontURL', '') or _tcp_r.get('PicURL', '') or '').split('|')[0]
                            _ab_pic_r     = str(_ab_orig.get('PicURLs', '') or _tcp_r.get('PicURL', '') or '')
                            _ebay_r['*Description']  = _ab_build_desc(_ab_title_r.replace('"', '&quot;'), _ab_front_r)
                            _ebay_r['PicURL']        = _ab_pic_r or _ebay_r.get('PicURL', '')
                            _ebay_r['GalleryType']   = 'Gallery' if _ebay_r['PicURL'] else ''
                            _ebay_r['*Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8)'] = 'Add'
                            _ebay_r['BestOfferEnabled'] = '1'
                            # Business policies — shipping tier based on TCP price
                            _ab_price_r = str(_tcp_r.get('*StartPrice', '') or '')
                            _ebay_r['ShippingProfileName']              = _ab_shipping_policy(_ab_price_r)
                            _ebay_r['ReturnProfileName']                = 'Returns'
                            _ebay_r['PaymentProfileName']               = 'BIN'
                            _ebay_r['*Location']                        = 'Scottsdale, AZ'
                            _ebay_r['PostalCode']                       = '85255'
                            _ebay_r['WeightMajor']                      = '0'
                            _ebay_r['WeightMinor']                      = '4'
                            _ebay_r['CD:Card Condition - (ID: 40001)']  = EBAY_CONDITION_DEFAULT
                            _ebay_r['ShippingType']                     = ''
                            _ebay_r['ShippingService-1:Option']         = ''
                            _ebay_r['ShippingService-1:FreeShipping']   = ''
                            _ebay_r['ShippingService-1:Cost']           = ''
                            _ebay_r['ShippingService-1:AdditionalCost'] = ''
                            _ab_ebay_rows.append(_ebay_r)
                        _ab_ebay_csv = _ab_make_tcp_csv(_ab_ebay_rows)
                        st.download_button(
                            "📥 Download eBay Listing CSV",
                            _ab_ebay_csv,
                            "lot_scan_ebay_add.csv",
                            "text/csv",
                            key="ab_ebay_add_dl",
                            type="primary",
                        )
                        st.caption(f"{len(_ab_ebay_rows)} listings ready · upload to eBay File Exchange to list all at once")
                    else:
                        st.warning("No rows found in TCP results CSV — make sure you uploaded the file returned by TradingCardPricer.com.")

        with scan_cert:
            st.caption("Enter the cert number printed on the slab label to pull the card + recent sold prices and buy signal.")
            cc1, cc2, cc3 = st.columns([2, 1, 1])
            cert = cc1.text_input("Cert number", key="cert_num")
            grader = cc2.selectbox("Grader", ["PSA", "BGS", "SGC", "CGC"], key="cert_grader")
            days = cc3.selectbox("History", [90, 180, 365], index=1, key="cert_days")
            if st.button("🎫 Look up cert", key="cert_go") and cert.strip():
                with st.spinner("Looking up cert…"):
                    cres = ch_prices_by_cert(cert.strip(), grader, days)
                # Persist results so Add to Batch button works on next rerun
                st.session_state["cert_lookup_result"] = cres
                st.session_state["cert_lookup_cert"]   = cert.strip()
                st.session_state["cert_lookup_grader"] = grader

            # Read persisted results (survives reruns)
            _clr   = st.session_state.get("cert_lookup_result")
            info   = (_clr or {}).get("cert_info") or {}
            _cl_cert   = st.session_state.get("cert_lookup_cert", "")
            _cl_grader = st.session_state.get("cert_lookup_grader", "PSA")
            # Clear if user changed cert number
            if cert.strip() and cert.strip() != _cl_cert:
                info = {}

            if info.get("description"):
                    st.markdown(f"**{info.get('description', '')}**")
                    st.caption(f"{info.get('grader','').upper()} {info.get('grade','')} · cert {info.get('cert','')}")

                    pr = (_clr or {}).get("prices") or []
                    vals = []
                    for p in pr:
                        try:
                            vals.append({"date": (p.get("closing_date") or "")[:10], "price": float(p.get("price"))})
                        except Exception:
                            pass

                    if vals:
                        dfp = pd.DataFrame(vals).drop_duplicates("date").sort_values("date")
                        _cert_dir, _cert_pct = calculate_trend({"prices": pr})

                        # Buy signal
                        if _cert_dir == "up" and _cert_pct > 5:
                            st.success(f"🔥 **HOT** — Up **{_cert_pct:+.0f}%** over {days}d. Rising demand.")
                        elif _cert_dir == "down" and _cert_pct < -5:
                            st.error(f"🛑 **COOLING** — Down **{abs(_cert_pct):.0f}%** over {days}d. Softening market.")
                        elif _cert_dir is not None:
                            st.info(f"➡️ **STABLE** — Flat ({_cert_pct:+.0f}% over {days}d).")
                        else:
                            st.warning("⚠️ Not enough history to call direction.")

                        cl1, cl2, cl3 = st.columns(3)
                        cl1.metric("Latest sold",  f"${dfp['price'].iloc[-1]:,.2f}")
                        cl2.metric("90d high",     f"${dfp['price'].max():,.2f}")
                        cl3.metric("90d low",      f"${dfp['price'].min():,.2f}")

                        # Individual sales table
                        st.markdown("**📋 Individual sales**")
                        cert_sale_rows = []
                        for p in pr[:10]:
                            try:
                                cert_sale_rows.append({
                                    "Date":  (p.get("closing_date") or "")[:10],
                                    "Price": f"${float(p.get('price')):,.2f}",
                                })
                            except Exception:
                                pass
                        if cert_sale_rows:
                            st.dataframe(pd.DataFrame(cert_sale_rows), use_container_width=True, hide_index=True, height=240)

                        st.markdown("**📈 Price trend**")
                        st.line_chart(dfp.set_index("date")["price"], height=240)
                        st.caption(f"Each point = one sold transaction (CardHedger). Showing last {days} days.")
                    else:
                        vals = []
                        st.info("Cert matched, but no recent sold-price history for this exact card.")
            elif _cl_cert:
                    st.warning("No card found for that cert / grader. Double-check the number and grader.")

            # ── Add to eBay Batch ─────────────────────────────────────────────
            if info.get("description"):
                st.markdown("---")
                if st.button("➕ Add to eBay Batch", key="cert_add_to_batch", type="primary",
                             help="Adds this graded card to the Batch → eBay tab for export, pulling its image from PSA"):
                    import urllib.request as _ur_cb, json as _json_cb
                    _cb_cert   = cert.strip()
                    _cb_player = info.get("description", "")
                    _cb_grader = grader
                    _raw_grade = str(info.get("grade", "") or "").strip()
                    # Strip grader prefix if API returns e.g. "PSA 10" instead of "10"
                    for _pfx in ("PSA ", "BGS ", "SGC ", "CGC "):
                        if _raw_grade.upper().startswith(_pfx):
                            _raw_grade = _raw_grade[len(_pfx):]
                            break
                    _cb_grade  = _raw_grade
                    # Build set name from description if possible (CardHedger gives full card description)
                    _cb_set    = ""
                    _cb_number = ""
                    # Try to fetch PSA image
                    _cb_image  = ""
                    if grader == "PSA":
                        try:
                            _cb_req = _ur_cb.Request(
                                f"https://api.psacard.com/publicapi/cert/GetByCertNumber/{_cb_cert}",
                                headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
                            )
                            with _ur_cb.urlopen(_cb_req, timeout=8) as _cb_resp:
                                _cb_psa = _json_cb.loads(_cb_resp.read().decode())
                            _cb_ci = _cb_psa.get("PSACert") or {}
                            _cb_image   = (_cb_ci.get("ImageURL") or "").strip()
                            _cb_player  = (_cb_ci.get("Subject") or _cb_player).strip()
                            _cb_year    = str(_cb_ci.get("Year") or "").strip()
                            _cb_brand   = (_cb_ci.get("Brand") or "").strip()
                            _cb_set     = f"{_cb_year} {_cb_brand}".strip()
                            _cb_number  = str(_cb_ci.get("CardNumber") or "").strip()
                            _cb_variety = (_cb_ci.get("Variety") or "").strip()
                            _cb_team    = (_cb_ci.get("Team") or "").strip()
                            _cb_grade_raw = str(_cb_ci.get("PSAGrade") or _cb_grade).strip()
                            for _pfx2 in ("PSA ", "BGS ", "SGC ", "CGC "):
                                if _cb_grade_raw.upper().startswith(_pfx2):
                                    _cb_grade_raw = _cb_grade_raw[len(_pfx2):]
                                    break
                            _cb_grade = _cb_grade_raw
                        except Exception as _cb_err:
                            _cb_variety = ""
                            _cb_team    = ""
                            st.caption(f"PSA image fetch failed: {_cb_err} — will use placeholder image")
                    else:
                        _cb_variety = ""
                        _cb_team    = ""
                    # Cache PSA image
                    if _cb_image:
                        if "psa_image_cache" not in st.session_state:
                            st.session_state["psa_image_cache"] = {}
                        st.session_state["psa_image_cache"][_cb_cert] = _cb_image
                    # Build batch entry
                    _eb2 = st.session_state.get("raw_batch", [])
                    _ni2 = max((x.get("idx", -1) for x in _eb2), default=-1) + 1
                    _cb_fmv = None
                    _cb_vals = [{"date": (p.get("closing_date") or "")[:10], "price": float(p.get("price"))}
                                for p in ((_clr or {}).get("prices") or []) if p.get("price")]
                    if _cb_vals:
                        import statistics as _stats
                        _cb_fmv = round(_stats.median([v["price"] for v in _cb_vals[:5]]), 2)
                    _new_c2 = {
                        "idx":         _ni2,
                        "front_file":  "",
                        "front_url":   _cb_image,
                        "back_url":    "",
                        "card_id":     "",
                        "player":      _cb_player,
                        "set_name":    _cb_set,
                        "number":      _cb_number,
                        "variant":     _cb_variety,
                        "team":        _cb_team,
                        "similarity":  100,
                        "candidates":  [],
                        "status":      "identified",
                        "low_conf":    False,
                        "title":       _raw_title(_cb_player, _cb_set, _cb_number, _cb_variety),
                        "fmv":         _cb_fmv,
                        "fmv_low":     None,
                        "fmv_high":    None,
                        "fmv_conf":    "",
                        "price":       _cb_fmv or 2.49,
                        "graded":      True,
                        "grader":      _cb_grader,
                        "grade":       _cb_grade,
                        "cert_number": _cb_cert,
                    }
                    _eb2.append(_new_c2)
                    st.session_state.raw_batch = _eb2
                    st.success(f"✅ Added to batch: {_cb_player} — {_cb_grader} {_cb_grade}. Switch to **📦 Batch → eBay** to export.")

        # ── BATCH → EBAY ──────────────────────────────────────────────────────
        with scan_batch:
            import subprocess as _sp
            import socket as _sock

            # ── eBay Export Settings (persists in session) ─────────────────
            _BATCH_DEFAULTS = {
                "sku_prefix":            "DFS",
                "default_condition":     "Near Mint (NM)",
                "best_offer":            True,
                "bo_min_price":          "",
                "bo_auto_accept_price":  "",
                "sport":                 "BASEBALL",
                "store_cat_baseball":    "44411116016",
                "store_cat_basketball":  "44411138016",
                "store_cat_football":    "44411117016",
                "store_cat_soccer":      "44411118016",
                "store_cat_other":       "0",
                "desc_template": (
                    '<div style="font-family:Arial,sans-serif;color:#1f2937;line-height:1.6;max-width:600px;margin:0 auto;">'
                    '<table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;">'
                    '<tr>'
                    '<td valign="top" style="width:180px;padding-right:20px;">'
                    '<img src="[FRONT_IMAGE_URL]" alt="[LISTING_TITLE]" style="width:100%;max-width:180px;border-radius:8px;border:1px solid #e5e7eb;"/>'
                    '</td>'
                    '<td valign="top">'
                    '<h2 style="margin:0 0 12px;font-size:22px;text-transform:uppercase;">[LISTING_TITLE]</h2>'
                    '<ol style="margin:0 0 16px;padding-left:20px;">'
                    '<li style="margin-bottom:8px;">All cards are scanned. If you see any lines or want more pictures, just ask!</li>'
                    '<li style="margin-bottom:8px;">Cards over $20 ship with tracking. Memorabilia/patch cards ship Ground Advantage to prevent damage.</li>'
                    '<li style="margin-bottom:8px;">Cards $75–$199 ship with signature required — you\'re protected as the buyer.</li>'
                    '</ol>'
                    '<p style="margin:0;">My goal is to have you receive the card in tip-top shape. No Returns Accepted. Questions? Please ask!!!</p>'
                    '</td>'
                    '</tr>'
                    '</table>'
                    '</div>'
                ),
            }
            for _k, _v in _BATCH_DEFAULTS.items():
                if f"bx_{_k}" not in st.session_state:
                    st.session_state[f"bx_{_k}"] = _v

            with st.expander("⚙️ eBay Export Settings", expanded=False):
                _sc1, _sc2 = st.columns(2)
                with _sc1:
                    st.markdown("**Listing Basics**")
                    st.session_state["bx_sku_prefix"] = st.text_input(
                        "SKU Prefix", value=st.session_state["bx_sku_prefix"], key="bxw_sku")
                    st.session_state["bx_default_condition"] = st.selectbox(
                        "Default Condition",
                        list(_RAW_CONDITIONS.keys()),
                        index=list(_RAW_CONDITIONS.keys()).index(st.session_state["bx_default_condition"]),
                        key="bxw_cond")
                    st.session_state["bx_sport"] = st.selectbox(
                        "Sport", ["BASEBALL","BASKETBALL","FOOTBALL","SOCCER","OTHER"],
                        index=["BASEBALL","BASKETBALL","FOOTBALL","SOCCER","OTHER"].index(st.session_state["bx_sport"]),
                        key="bxw_sport")
                    st.session_state["bx_best_offer"] = st.checkbox(
                        "Enable Best Offer", value=st.session_state["bx_best_offer"], key="bxw_bo")
                    if st.session_state["bx_best_offer"]:
                        _boc1, _boc2 = st.columns(2)
                        st.session_state["bx_bo_min_price"] = _boc1.text_input(
                            "Min Best Offer ($)", value=st.session_state.get("bx_bo_min_price",""),
                            placeholder="e.g. 1.50", key="bxw_bo_min",
                            help="eBay auto-declines offers below this amount. Leave blank to allow all offers.")
                        st.session_state["bx_bo_auto_accept_price"] = _boc2.text_input(
                            "Auto-Accept ($)", value=st.session_state.get("bx_bo_auto_accept_price",""),
                            placeholder="e.g. 3.00", key="bxw_bo_aa",
                            help="eBay auto-accepts offers at or above this amount.")
                with _sc2:
                    st.markdown("**eBay Store Category IDs**")
                    st.caption("Find your category IDs in eBay Seller Hub → Store → Manage Categories")
                    st.session_state["bx_store_cat_baseball"]   = st.text_input("Baseball",    value=st.session_state["bx_store_cat_baseball"],   key="bxw_sb")
                    st.session_state["bx_store_cat_basketball"] = st.text_input("Basketball",  value=st.session_state["bx_store_cat_basketball"], key="bxw_sk")
                    st.session_state["bx_store_cat_football"]   = st.text_input("Football",    value=st.session_state["bx_store_cat_football"],   key="bxw_sf")
                    st.session_state["bx_store_cat_soccer"]     = st.text_input("Soccer",      value=st.session_state["bx_store_cat_soccer"],     key="bxw_sc")
                    st.session_state["bx_store_cat_other"]      = st.text_input("Other/Default", value=st.session_state["bx_store_cat_other"],    key="bxw_so")

                st.markdown("---")
                st.markdown("**Business Policies** *(optional — leave blank to use manual shipping/return/payment settings)*")
                st.caption("Use eBay Seller Hub → Account → Business Policies to find your policy names exactly as shown.")
                _bp1, _bp2, _bp3 = st.columns(3)
                st.session_state["bx_shipping_profile"] = _bp1.text_input(
                    "Shipping policy name", value=st.session_state.get("bx_shipping_profile", ""),
                    placeholder="e.g. Standard Shipping", key="bxw_ship_prof")
                st.session_state["bx_return_profile"] = _bp2.text_input(
                    "Return policy name", value=st.session_state.get("bx_return_profile", ""),
                    placeholder="e.g. No Returns", key="bxw_ret_prof")
                st.session_state["bx_payment_profile"] = _bp3.text_input(
                    "Payment policy name", value=st.session_state.get("bx_payment_profile", ""),
                    placeholder="e.g. Immediate Pay", key="bxw_pay_prof")

                st.markdown("---")
                st.markdown("**⏱ Drip Schedule** *(stagger listings hourly to boost eBay visibility)*")
                st.caption("eBay rewards consistent activity. Upload once — eBay lists each card at its scheduled time automatically.")
                _drip_enabled = st.checkbox("Enable drip posting", value=st.session_state.get("bx_drip_enabled", False), key="bxw_drip")
                st.session_state["bx_drip_enabled"] = _drip_enabled
                if _drip_enabled:
                    _dsc1, _dsc2 = st.columns(2)
                    st.session_state["bx_drip_spread_hours"] = _dsc1.number_input(
                        "Spread over (hours)", min_value=1, max_value=48,
                        value=st.session_state.get("bx_drip_spread_hours", 8),
                        key="bxw_drip_hrs",
                        help="App auto-calculates cards per hour. 40 cards over 8 hrs = 5/hr.")
                    st.session_state["bx_drip_start"] = _dsc2.text_input(
                        "Start time (YYYY-MM-DD HH:MM, 24h local)",
                        value=st.session_state.get("bx_drip_start", ""),
                        placeholder="e.g. 2026-08-10 09:00",
                        key="bxw_drip_start",
                        help="Leave blank to start from the next full hour.")
                    # Live schedule preview based on current batch
                    import datetime as _dtp
                    _n_est = sum(1 for _r in st.session_state.get("raw_batch", []) if _r.get("status") != "listed")
                    _hrs = int(st.session_state.get("bx_drip_spread_hours", 8))
                    if _n_est > 0 and _hrs > 0:
                        _cph_est = max(1, -(-_n_est // _hrs))
                        _slots_est = -(-_n_est // _cph_est)
                        _drip_start_str = st.session_state.get("bx_drip_start", "").strip()
                        if _drip_start_str:
                            try:
                                _drip_base_p = _dtp.datetime.strptime(_drip_start_str, "%Y-%m-%d %H:%M")
                            except ValueError:
                                _drip_base_p = (_dtp.datetime.now() + _dtp.timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
                        else:
                            _drip_base_p = (_dtp.datetime.now() + _dtp.timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
                        _drip_end_p = _drip_base_p + _dtp.timedelta(hours=_slots_est - 1)
                        st.info(f"📅 **{_n_est} cards · {_cph_est}/hr · {_slots_est} slots** — {_drip_base_p.strftime('%-I:%M %p')} → {_drip_end_p.strftime('%-I:%M %p on %b %-d')}")

                st.markdown("---")
                st.markdown("**Description Template**")
                st.caption("Use `[LISTING_TITLE]` and `[FRONT_IMAGE_URL]` as placeholders — they'll be replaced per card on export.")
                new_tmpl = st.text_area(
                    "HTML Template", value=st.session_state["bx_desc_template"],
                    height=200, key="bxw_tmpl", label_visibility="collapsed")
                st.session_state["bx_desc_template"] = new_tmpl

                # Live preview
                _prev_title = "JAZZ CHISHOLM 2019 Bowman Chrome #BTP-63 Green Refractor"
                _prev_html  = new_tmpl.replace("[LISTING_TITLE]", _prev_title).replace("[FRONT_IMAGE_URL]", "")
                with st.expander("👁 Preview description"):
                    st.markdown(_prev_html, unsafe_allow_html=True)

            # ── Stacks session state ────────────────────────────────────────
            if "raw_batch" not in st.session_state:
                st.session_state.raw_batch = []
            if "raw_batch_comps" not in st.session_state:
                st.session_state.raw_batch_comps = {}
            if "bx_active_stack_id" not in st.session_state:
                st.session_state["bx_active_stack_id"] = None
            if "bx_active_stack_name" not in st.session_state:
                st.session_state["bx_active_stack_name"] = ""
            if "bx_stacks_cache" not in st.session_state:
                st.session_state["bx_stacks_cache"] = None

            _STACKS_SQL = """create table if not exists scan_stacks (
  id bigint primary key generated always as identity,
  name text not null,
  notes text,
  status text default 'open',
  total_cards int default 0,
  created_at timestamptz default now()
);
create table if not exists scan_cards (
  id bigint primary key generated always as identity,
  stack_id bigint not null references scan_stacks(id) on delete cascade,
  idx int not null,
  front_file text,
  front_url text,
  back_url text,
  player text,
  title text,
  similarity numeric,
  fmv numeric,
  price numeric,
  status text default 'identified',
  card_data_json text,
  created_at timestamptz default now(),
  unique(stack_id, idx)
);
alter table scan_stacks disable row level security;
alter table scan_cards disable row level security;"""

            st.markdown("---")

            # ══════════════════════════════════════════════════════════════════
            # STACKS LIST VIEW  (no active stack)
            # ══════════════════════════════════════════════════════════════════
            if st.session_state["bx_active_stack_id"] is None:
                # Show warning if cards are pending in raw_batch (e.g. added from cert tab)
                _pending_cards = [c for c in st.session_state.get("raw_batch", []) if c.get("status") == "identified"]
                if _pending_cards:
                    st.warning(f"📦 **{len(_pending_cards)} card{'s' if len(_pending_cards) != 1 else ''} pending** (added from cert lookup). Create a new lot below — they'll be included automatically.")

                _slh1, _slh2 = st.columns([5, 1])
                _slh1.markdown("### 📦 Lots")
                if _slh2.button("➕ New Lot", type="primary", key="bx_new_stack_btn"):
                    st.session_state["bx_show_new_form"] = True

                if st.session_state.get("bx_show_new_form"):
                    with st.form("bx_new_stack_form", clear_on_submit=True):
                        _nsc1, _nsc2 = st.columns([2, 3])
                        _ns_name  = _nsc1.text_input("Lot name *", placeholder="RBLOT-08-06-26")
                        _ns_notes = _nsc2.text_input("Notes", placeholder="Baseball box, 131 cards")
                        if st.form_submit_button("✅ Create Lot & Open"):
                            _nm = (_ns_name or "").strip()
                            if _nm:
                                _ns = stack_create(_nm, (_ns_notes or "").strip())
                                if _ns:
                                    _px = _nm.split("-")[0].upper() if "-" in _nm else _nm[:8].upper()
                                    st.session_state["bx_active_stack_id"]   = _ns["id"]
                                    st.session_state["bx_active_stack_name"] = _ns["name"]
                                    st.session_state["bx_sku_prefix"]        = _px
                                    # Keep any cards already in raw_batch (e.g. from cert tab add)
                                    if not st.session_state.raw_batch:
                                        st.session_state.raw_batch = []
                                    st.session_state.raw_batch_comps         = {}
                                    st.session_state["bx_show_new_form"]     = False
                                    st.session_state["bx_stacks_cache"]      = None
                                    st.rerun()
                                else:
                                    st.error("Failed to create lot — run the SQL setup first:")
                                    st.code(_STACKS_SQL, language="sql")

                    with st.expander("📋 First-time setup SQL (run once in Supabase)", expanded=False):
                        st.code(_STACKS_SQL, language="sql")

                # Load + display stacks list
                if st.session_state["bx_stacks_cache"] is None:
                    _raw_sl = stacks_list()
                    for _s in _raw_sl:
                        _scs = stack_cards_get(_s["id"])
                        _s["_n"]       = len(_scs)
                        _s["_matched"] = sum(1 for c in _scs if c.get("status") not in ("error",))
                        _s["_over80"]  = sum(1 for c in _scs if float(c.get("similarity") or 0) >= 80)
                        _s["_listed"]  = sum(1 for c in _scs if c.get("status") == "listed")
                    st.session_state["bx_stacks_cache"] = _raw_sl

                _stacks = st.session_state["bx_stacks_cache"]
                if not _stacks:
                    if not st.session_state.get("bx_show_new_form"):
                        st.info("No stacks yet — click **➕ New Lot** to create your first batch.")
                else:
                    _hc = st.columns([3,1,1,1,1,1,1,1])
                    for _hl, _hv in zip(_hc, ["**Name**","**Cards**","**Matched**","**>80%**","**Listed**","**Created**","",""]):
                        _hl.caption(_hv)
                    st.markdown("---")
                    for _si, _s in enumerate(_stacks):
                        _rc = st.columns([3,1,1,1,1,1,1,1])
                        _rc[0].write(_s["name"])
                        _rc[1].write(str(_s["_n"]))
                        _rc[2].write(str(_s["_matched"]))
                        _rc[3].write(str(_s["_over80"]))
                        _rc[4].write(str(_s["_listed"]))
                        _rc[5].write((_s.get("created_at") or "")[:10])
                        if _rc[6].button("▶ Open", key=f"stk_open_{_s['id']}"):
                            _saved = stack_cards_get(_s["id"])
                            _loaded = []
                            for _sc in _saved:
                                try:
                                    _cd = json.loads(_sc.get("card_data_json") or "{}")
                                    if _cd:
                                        # Always trust the DB status column — card_data_json
                                        # may still say "identified" even after export marked it "listed"
                                        _cd["status"] = _sc.get("status") or _cd.get("status", "identified")
                                        _cd["idx"]    = _sc.get("idx", _cd.get("idx", 0))
                                        _loaded.append(_cd)
                                except Exception:
                                    pass
                            _px2 = _s["name"].split("-")[0].upper() if "-" in _s["name"] else _s["name"][:8].upper()
                            st.session_state["bx_active_stack_id"]   = _s["id"]
                            st.session_state["bx_active_stack_name"] = _s["name"]
                            st.session_state["bx_sku_prefix"]        = _px2
                            st.session_state.raw_batch               = _loaded
                            st.session_state.raw_batch_comps         = {}
                            st.rerun()
                        if _rc[7].button("🗑", key=f"stk_del_{_s['id']}", help="Delete lot and all its cards"):
                            stack_cards_delete(_s["id"])
                            stack_delete(_s["id"])
                            st.session_state["bx_stacks_cache"] = None
                            st.rerun()

            # ══════════════════════════════════════════════════════════════════
            # ACTIVE STACK VIEW
            # ══════════════════════════════════════════════════════════════════
            else:
                _ash1, _ash2 = st.columns([5, 1])
                _ash1.markdown(f"### 📦 {st.session_state['bx_active_stack_name']}")
                if _ash2.button("← Lots", key="bx_back_stacks"):
                    st.session_state["bx_active_stack_id"]   = None
                    st.session_state["bx_active_stack_name"] = ""
                    st.session_state["bx_stacks_cache"]      = None
                    st.session_state.raw_batch               = []
                    st.session_state.raw_batch_comps         = {}
                    st.rerun()

                import subprocess as _sp
                import socket as _sock

                def _scanner_running():
                    try:
                        s = _sock.create_connection(("localhost", 5100), timeout=1)
                        s.close()
                        return True
                    except OSError:
                        return False

                _SCANNER_PY = Path(__file__).parent / "card-scanner" / "app.py"
                _PYTHON     = "/Library/Frameworks/Python.framework/Versions/3.14/bin/python3"

                if _SCANNER_PY.exists():
                    if not _scanner_running():
                        st.markdown("### 📷 Batch Scanner")
                        st.info("Keyboard-driven bulk matching — upload images, hit 1–8 to pick the right parallel, Enter to confirm, auto-advances to the next card. Comps + trend fetch in the background.")
                        if st.button("🚀 Launch Batch Scanner", type="primary", key="launch_scanner"):
                            _sp.Popen(
                                [_PYTHON, str(_SCANNER_PY)],
                                cwd=str(_SCANNER_PY.parent),
                                stdout=_sp.DEVNULL, stderr=_sp.DEVNULL,
                            )
                            import time as _t; _t.sleep(2)
                            st.rerun()
                    else:
                        st.components.v1.iframe("http://localhost:5100", height=860, scrolling=True)

                # ── Batch ready banner ─────────────────────────────────────
                _pre_batch = st.session_state.raw_batch
                if _pre_batch:
                    _pre_identified = [c for c in _pre_batch if c.get("status") == "identified"]
                    if _pre_identified:
                        st.info(f"📦 **{len(_pre_identified)} card{'s' if len(_pre_identified) != 1 else ''} in your batch** — scroll down past the uploader to review, price, and export.")

                # ── Step 1: Select cards to scan ───────────────────────────
                st.markdown("### Select cards to scan")

                # Card Type selection
                st.markdown("**Card Type**")
                _batch_card_type = st.session_state.get("batch_card_type", "raw")
                _ct_col1, _ct_col2 = st.columns(2)
                with _ct_col1:
                    _ct_raw_border = "2px solid #1a73e8" if _batch_card_type == "raw" else "2px solid #ddd"
                    _ct_raw_bg     = "#f0f4ff" if _batch_card_type == "raw" else "transparent"
                    st.markdown(f'<div style="border:{_ct_raw_border};border-radius:8px;padding:14px 16px;background:{_ct_raw_bg};margin-bottom:8px;"><strong>● RAW (Not Graded)</strong></div>', unsafe_allow_html=True)
                    if st.button("Select RAW", key="ct_raw_btn", use_container_width=True, type="primary" if _batch_card_type == "raw" else "secondary"):
                        st.session_state["batch_card_type"] = "raw"
                        st.rerun()
                with _ct_col2:
                    _ct_gr_border = "2px solid #1a73e8" if _batch_card_type == "graded" else "2px solid #ddd"
                    _ct_gr_bg     = "#f0f4ff" if _batch_card_type == "graded" else "transparent"
                    st.markdown(f'<div style="border:{_ct_gr_border};border-radius:8px;padding:14px 16px;background:{_ct_gr_bg};margin-bottom:8px;"><strong>○ Graded</strong></div>', unsafe_allow_html=True)
                    if st.button("Select Graded", key="ct_graded_btn", use_container_width=True, type="primary" if _batch_card_type == "graded" else "secondary"):
                        st.session_state["batch_card_type"] = "graded"
                        st.rerun()
                _batch_card_type = st.session_state.get("batch_card_type", "raw")

                st.markdown("---")

                # ── Step 2: Upload files ────────────────────────────────────
                st.markdown("### Upload your cards")
                st.markdown("""
                <div style="border:1px solid #ddd;border-radius:10px;padding:14px 16px;margin-bottom:12px;background:#fafafa;">
                <p style="font-size:0.9em;margin:0 0 6px 0;font-weight:600;">📋 How to upload</p>
                <p style="font-size:0.85em;color:#555;margin:0 0 4px 0;">Select all images at once — fronts and backs interleaved in order:</p>
                <code style="font-size:0.8em;background:#f0f0f0;padding:2px 6px;border-radius:4px;margin-right:4px;">card-01-front.jpg</code>
                <code style="font-size:0.8em;background:#f0f0f0;padding:2px 6px;border-radius:4px;margin-right:4px;">card-01-back.jpg</code>
                <code style="font-size:0.8em;background:#f0f0f0;padding:2px 6px;border-radius:4px;margin-right:4px;">card-02-front.jpg</code>
                <code style="font-size:0.8em;background:#f0f0f0;padding:2px 6px;border-radius:4px;">card-02-back.jpg</code>
                <p style="font-size:0.8em;color:#888;margin:8px 0 0 0;">Up to 50 cards (100 images) recommended per batch. Front &amp; back required for eBay listings.</p>
                </div>""", unsafe_allow_html=True)

                all_files = st.file_uploader(
                    "📷 Drop all card images here",
                    type=["jpg", "jpeg", "png", "tif", "tiff"],
                    accept_multiple_files=True,
                    key="raw_all_files",
                    help="Select all images at once in front/back order. Cmd+A in Finder or drag the whole folder.",
                )

                n_files = len(all_files) if all_files else 0

                if n_files > 0:
                    if n_files % 2 != 0:
                        st.warning(f"⚠️ {n_files} images selected — need an even number (one front + one back per card). Got {n_files // 2} complete pairs + 1 leftover.")

                    n_pairs     = n_files // 2
                    front_files = all_files[0::2]
                    back_files  = all_files[1::2]

                    if n_pairs > 50:
                        st.warning(f"⚠️ {n_pairs} cards selected — 50 per batch recommended for best results.")

                    st.success(f"✅ {n_pairs} card{'s' if n_pairs != 1 else ''} ready (front + back)")
    
                    # Paired thumbnail preview — front and back side by side, compact
                    from PIL import Image as _PIL_Image
                    import io as _io
    
                    if "raw_batch_rots" not in st.session_state:
                        st.session_state.raw_batch_rots = {}
    
                    def _rotated_bytes(file_bytes, deg):
                        if deg == 0:
                            return file_bytes
                        img = _PIL_Image.open(_io.BytesIO(file_bytes))
                        img = img.rotate(-deg, expand=True)
                        buf = _io.BytesIO()
                        img.save(buf, format="JPEG")
                        return buf.getvalue()
    
                    preview_n = min(n_pairs, 6)
                    for ci in range(preview_n):
                        fkey = f"rot_f_{ci}"
                        bkey = f"rot_b_{ci}"
                        if fkey not in st.session_state.raw_batch_rots:
                            st.session_state.raw_batch_rots[fkey] = 0
                        if bkey not in st.session_state.raw_batch_rots:
                            st.session_state.raw_batch_rots[bkey] = 0

                        fb = _rotated_bytes(front_files[ci].getvalue(), st.session_state.raw_batch_rots[fkey])
                        bb = _rotated_bytes(back_files[ci].getvalue(), st.session_state.raw_batch_rots[bkey])

                        col_f, col_b, col_rf, col_rb, _sp = st.columns([3, 3, 1, 1, 4])
                        with col_f:
                            st.image(fb, caption=f"#{ci+1} Front", use_container_width=True)
                        with col_b:
                            st.image(bb, caption=f"#{ci+1} Back", use_container_width=True)
                        with col_rf:
                            st.write("")
                            st.write("")
                            if st.button("↻", key=f"rfbtn_{ci}", help="Rotate front"):
                                st.session_state.raw_batch_rots[fkey] = (st.session_state.raw_batch_rots[fkey] + 90) % 360
                                st.rerun()
                        with col_rb:
                            st.write("")
                            st.write("")
                            if st.button("↻", key=f"rbbtn_{ci}", help="Rotate back"):
                                st.session_state.raw_batch_rots[bkey] = (st.session_state.raw_batch_rots[bkey] + 90) % 360
                                st.rerun()

                    if n_pairs > 6:
                        st.caption(f"+ {n_pairs - 6} more pairs not shown")

                    st.markdown("---")
                    _scan_btn_label = f"🔍 Identify & Price All ({n_pairs} card{'s' if n_pairs != 1 else ''})"
                    if st.button(_scan_btn_label, type="primary", key="raw_batch_go", disabled=n_pairs == 0):
                            st.session_state.raw_batch = []
                            st.session_state.raw_batch_comps = {}
                            prog    = st.progress(0.0)
                            status  = st.empty()
                            cards   = []
                            _graded_upfront = _batch_card_type == "graded"

                            # Phase 1: upload + image-match (one card at a time)
                            for i, (ff, bf) in enumerate(zip(front_files, back_files)):
                                status.text(f"Identifying card {i+1}/{n_pairs}: {ff.name}…")
                                c = {
                                    "idx":          i,
                                    "front_file":   ff.name,
                                    "back_file":    bf.name,
                                    "status":       "error",
                                    "error":        "",
                                    "include":      True,
                                    "condition_id": "2750",
                                    "graded":       _graded_upfront,
                                }
                                try:
                                    _frot = st.session_state.raw_batch_rots.get(f"rot_f_{i}", 0)
                                    _brot = st.session_state.raw_batch_rots.get(f"rot_b_{i}", 0)
                                    _front_bytes = _rotated_bytes(ff.getvalue(), _frot)
                                    _back_bytes  = _rotated_bytes(bf.getvalue(), _brot)

                                    front_url, front_path = _scan_upload_to_supabase(_front_bytes, f"f_{i}_{ff.name}")
                                    back_url,  back_path  = _scan_upload_to_supabase(_back_bytes,  f"b_{i}_{bf.name}")
                                    c["front_url"]  = front_url
                                    c["front_path"] = front_path
                                    c["back_url"]   = back_url
                                    c["back_path"]  = back_path
    
                                    # Send as base64 to avoid URL accessibility issues
                                    _front_b64 = base64.b64encode(_front_bytes).decode()

                                    # Step 1: cert OCR — auto-detects graded slabs.
                                    # If the image has a PSA/BGS/SGC/CGC label, returns
                                    # grader, grade, cert number, and card identity.
                                    _cert_ocr  = ch_cert_ocr(front_url)
                                    _cert_info = _cert_ocr.get("cert_info") or {}
                                    if _cert_info.get("cert") or _cert_info.get("grade"):
                                        _ocr_grade = str(_cert_info.get("grade", "") or "").strip()
                                        for _pfx in ("PSA ", "BGS ", "SGC ", "CGC "):
                                            if _ocr_grade.upper().startswith(_pfx):
                                                _ocr_grade = _ocr_grade[len(_pfx):]
                                                break
                                        _ocr_grader = str(_cert_info.get("grader", "") or "").strip().upper() or "PSA"
                                        c.update({
                                            "graded":      True,
                                            "grader":      _ocr_grader,
                                            "grade":       _ocr_grade,
                                            "cert_number": str(_cert_info.get("cert", "") or "").strip(),
                                        })
                                        _ocr_card = _cert_ocr.get("card") or {}
                                        if _ocr_card.get("player"):
                                            c.update({
                                                "card_id":    _ocr_card.get("card_id", ""),
                                                "player":     _ocr_card.get("player", ""),
                                                "set_name":   _ocr_card.get("set", ""),
                                                "number":     str(_ocr_card.get("number", "") or ""),
                                                "variant":    _ocr_card.get("variant", ""),
                                                "similarity": 95.0,
                                                "candidates": [],
                                                "status":     "identified",
                                                "low_conf":   False,
                                            })
                                            c["title"] = _raw_title(c["player"], c["set_name"], c["number"], c.get("variant", ""))

                                    # Step 2: image-match (AI-powered) for card identity
                                    match_res  = _ch_post("/v1/cards/image-match", {"image_base64": _front_b64, "k": 5}) or {}
                                    best       = match_res.get("best_match") or {}
                                    candidates = match_res.get("candidates") or []
    
                                    # Fallback: image-search (broader KNN, no AI filter)
                                    if not best.get("card_id") and not candidates:
                                        search_res = _ch_post("/v1/cards/image-search", {"image_base64": _front_b64, "k": 5}) or {}
                                        candidates = search_res.get("results") or search_res.get("candidates") or []
                                        match_res  = search_res
    
                                    if best and best.get("card_id"):
                                        c.update({
                                            "card_id":    best.get("card_id", ""),
                                            "player":     best.get("player", ""),
                                            "set_name":   best.get("set", ""),
                                            "number":     best.get("number", ""),
                                            "variant":    best.get("variant", ""),
                                            "similarity": float(best.get("similarity") or 0),
                                            "candidates": candidates,
                                            "status":     "identified",
                                            "low_conf":   float(best.get("similarity") or 0) < 80,
                                        })
                                    elif candidates:
                                        top = candidates[0]
                                        c.update({
                                            "card_id":    top.get("card_id", ""),
                                            "player":     top.get("player", ""),
                                            "set_name":   top.get("set", ""),
                                            "number":     top.get("number", ""),
                                            "variant":    top.get("variant", ""),
                                            "similarity": float(top.get("similarity") or 0),
                                            "candidates": candidates,
                                            "status":     "identified",
                                            "low_conf":   True,
                                        })
                                    else:
                                        api_msg = match_res.get("message") or match_res.get("error") or "No candidates returned"
                                        c["error"] = f"No visual match — {api_msg}"
    
                                    if c.get("status") == "identified":
                                        c["title"] = _raw_title(
                                            c["player"], c["set_name"], c["number"], c.get("variant", "")
                                        )
                                except Exception as exc:
                                    c["error"] = str(exc)
    
                                cards.append(c)
                                prog.progress((i + 1) / n_pairs * 0.7)
    
                            # Phase 2: batch FMV for all identified cards (1 API call)
                            identified_cards = [c for c in cards if c.get("card_id")]
                            if identified_cards:
                                status.text(f"Fetching FMV for {len(identified_cards)} cards (batch)…")
                                fmv_items = [{"card_id": c["card_id"], "grade": "Raw"} for c in identified_cards]
                                fmv_res   = ch_fmv_batch(fmv_items)
                                fmv_by_id = {}
                                for r in (fmv_res.get("results") or []):
                                    cid = r.get("card_id")
                                    if cid:
                                        fmv_by_id[cid] = r
                                for c in identified_cards:
                                    fmv = fmv_by_id.get(c["card_id"], {})
                                    raw_p = fmv.get("price")
                                    c["fmv"]      = round(float(raw_p), 2) if raw_p else None
                                    c["fmv_low"]  = fmv.get("price_low")
                                    c["fmv_high"] = fmv.get("price_high")
                                    c["fmv_conf"] = fmv.get("confidence_grade", "")
                                    c["price"]    = c["fmv"] if c["fmv"] else 2.49
    
                            prog.progress(1.0)
                            n_ok  = sum(1 for c in cards if c["status"] == "identified")
                            n_err = len(cards) - n_ok
                            status.text(f"Done — {n_ok} identified, {n_err} failed.")
                            # Append to existing batch — don't replace (lot may already have cards)
                            existing_batch = st.session_state.raw_batch or []
                            _idx_offset = len(existing_batch)
                            for _c in cards:
                                _c["idx"] = _idx_offset + _c["idx"]
                            st.session_state.raw_batch = existing_batch + cards
                            # Persist to Supabase under active stack (upsert on stack_id+idx — safe)
                            if st.session_state.get("bx_active_stack_id"):
                                for _psc in cards:
                                    stack_card_upsert({
                                        "stack_id":       st.session_state["bx_active_stack_id"],
                                        "idx":            _psc["idx"],
                                        "front_file":     _psc.get("front_file", ""),
                                        "front_url":      _psc.get("front_url", ""),
                                        "back_url":       _psc.get("back_url", ""),
                                        "player":         _psc.get("player", ""),
                                        "title":          _psc.get("title", ""),
                                        "similarity":     _psc.get("similarity") or 0,
                                        "fmv":            _psc.get("fmv"),
                                        "price":          _psc.get("price") or 2.49,
                                        "status":         _psc.get("status", "error"),
                                        "card_data_json": json.dumps(_psc),
                                    })
                                stack_update(st.session_state["bx_active_stack_id"],
                                             {"total_cards": len(st.session_state.raw_batch)})
                                st.session_state["bx_stacks_cache"] = None
                            st.rerun()
    
                # ── Step 2: Review & Price ─────────────────────────────────────
                raw_batch = st.session_state.raw_batch
                if raw_batch:
                    identified = [c for c in raw_batch if c.get("status") == "identified"]
                    failed     = [c for c in raw_batch if c.get("status") != "identified"]
    
                    if failed:
                        st.markdown(f"#### ⚠️ {len(failed)} card(s) not matched — search to identify manually")
                        for _fc in failed:
                            _fi = _fc["idx"]
                            with st.container(border=True):
                                _fc1, _fc2 = st.columns([1, 3])
                                with _fc1:
                                    if _fc.get("front_url"):
                                        st.image(_fc["front_url"], caption=f"Card {_fi+1}", use_container_width=True)
                                    else:
                                        st.caption(f"Card {_fi+1} — {_fc.get('front_file','')}")
                                with _fc2:
                                    st.caption(f"Error: {_fc.get('error','No match found')}")
                                    _sq_key  = f"batch_manual_q_{_fi}"
                                    _sr_key  = f"batch_manual_r_{_fi}"
                                    _sq = st.text_input("Search by player / set / year", key=_sq_key, placeholder="e.g. Mike Trout 2011 Topps Update")
                                    if _sq and st.button("🔍 Search", key=f"batch_manual_go_{_fi}"):
                                        _sr = _ch_post("/v1/cards/card-search", {"search": _sq, "page": 1, "page_size": 6}) or {}
                                        st.session_state[_sr_key] = _sr.get("cards") or _sr.get("results") or []
                                    _candidates = st.session_state.get(_sr_key, [])
                                    if _candidates:
                                        st.markdown("**Pick the correct card:**")
                                        for _cand in _candidates:
                                            _clabel = f"{_cand.get('player','')} — {_cand.get('set','')} #{_cand.get('number','')} {_cand.get('variant','')}"
                                            if st.button(_clabel.strip(" —"), key=f"batch_pick_{_fi}_{_cand.get('card_id','')}"):
                                                # Update this card in raw_batch to identified
                                                for _rb in st.session_state.raw_batch:
                                                    if _rb.get("idx") == _fi:
                                                        _rb.update({
                                                            "card_id":    _cand.get("card_id",""),
                                                            "player":     _cand.get("player",""),
                                                            "set_name":   _cand.get("set",""),
                                                            "number":     str(_cand.get("number","") or ""),
                                                            "variant":    _cand.get("variant",""),
                                                            "similarity": 0.0,
                                                            "candidates": [],
                                                            "status":     "identified",
                                                            "low_conf":   False,
                                                            "title":      _raw_title(_cand.get("player",""), _cand.get("set",""), str(_cand.get("number","") or ""), _cand.get("variant","")),
                                                        })
                                                        if st.session_state.get("bx_active_stack_id"):
                                                            _sc_rows = stack_cards_get(st.session_state["bx_active_stack_id"])
                                                            for _scr in _sc_rows:
                                                                if _scr.get("idx") == _fi:
                                                                    stack_card_update(_scr["id"], {
                                                                        "player":  _rb["player"],
                                                                        "title":   _rb["title"],
                                                                        "status":  "identified",
                                                                        "card_data_json": json.dumps(_rb),
                                                                    })
                                                        break
                                                st.session_state.pop(_sr_key, None)
                                                st.rerun()
    
                    if identified:
                        n_low = sum(1 for c in identified if c.get("low_conf"))
                        if n_low:
                            st.warning(f"⚠️ {n_low} card(s) have low match confidence — review those rows carefully before exporting.")
                        else:
                            st.success(f"✅ {len(identified)} cards identified")
    
                        st.markdown("**Step 2 — Review & Price**")
    
                        # Build editable table
                        import urllib.parse as _uparse
                        import datetime as _dt_prev
                        _prev_date_str = _dt_prev.date.today().strftime("%m%d%y")
                        _prev_sku_pfx  = st.session_state.get("bx_sku_prefix", "DFS")
                        # Use only the name token (first segment before any "-") so the date
                        # and sequence we append don't double-up when the prefix includes dates.
                        _prev_sku_name = _prev_sku_pfx.split("-")[0] if _prev_sku_pfx else "DFS"
                        edit_rows = []
                        for _ei, c in enumerate(identified):
                            sim = c.get("similarity", 0)
                            fmv_display = f"${c['fmv']:.2f}" if c.get("fmv") else "—"
                            cond_label  = next((k for k, v in _RAW_CONDITIONS.items() if v == c.get("condition_id", "2750")), "Near Mint (NM)")
                            ebay_q      = _uparse.quote_plus(c.get("title", c.get("player", "")))
                            ebay_url    = f"https://www.ebay.com/sch/i.html?_nkw={ebay_q}&_sacat=261328&LH_Sold=1&LH_Complete=1"
                            raw_par = c.get("variant", "")
                            par_display = raw_par if raw_par and raw_par.lower() not in ("base","") else ""
                            auto_sku = f"{_prev_sku_name}-{_prev_date_str}-{_ei+1:04d}"
                            edit_rows.append({
                                "✓":          c.get("include", True),
                                "Player":     c.get("player", ""),
                                "Price ($)":  float(c.get("price") or 2.49),
                                "FMV":        fmv_display,
                                "eBay Title": c.get("title", ""),
                                "Graded":     c.get("graded", False),
                                "Grader":     c.get("grader", "PSA"),
                                "Grade":      c.get("grade", ""),
                                "Cert #":     c.get("cert_number", ""),
                                "Parallel":   par_display,
                                "Condition":  cond_label,
                                "Sport":      c.get("sport", st.session_state.get("bx_sport","BASEBALL")),
                                "Conf %":     f"{sim:.0f}" if sim else "?",
                                "FMV Conf":   c.get("fmv_conf", ""),
                                "Custom SKU": c.get("custom_sku", auto_sku),
                                "#":          c["idx"] + 1,
                                "🔍 eBay":    ebay_url,
                            })
    
                        edited = st.data_editor(
                            pd.DataFrame(edit_rows),
                            use_container_width=True,
                            hide_index=True,
                            key="raw_batch_editor",
                            column_config={
                                "✓":          st.column_config.CheckboxColumn("Include", width="small"),
                                "#":          st.column_config.NumberColumn("#", width="small", disabled=True),
                                "Custom SKU": st.column_config.TextColumn("SKU ✏️", width="medium",
                                                  help="Auto-generated from your SKU prefix + date. Edit to set a custom SKU per card."),
                                "Player":     st.column_config.TextColumn("Player", width="medium", disabled=True),
                                "Parallel":   st.column_config.TextColumn("Parallel ✏️", width="medium",
                                                  help="Edit if CardHedger got the parallel wrong — updates the title and CSV export"),
                                "eBay Title": st.column_config.TextColumn("eBay Title (editable)", max_chars=80, width="large"),
                                "Sport":      st.column_config.SelectboxColumn(
                                                  "Sport", options=["BASEBALL","BASKETBALL","FOOTBALL","SOCCER","OTHER"], width="small",
                                                  help="Override sport per card for mixed-sport batches"),
                                "Conf %":     st.column_config.TextColumn("Conf%", width="small", disabled=True,
                                                  help="Visual match confidence from CardHedger — below 80% means double-check"),
                                "FMV":        st.column_config.TextColumn("FMV", width="small", disabled=True),
                                "FMV Conf":   st.column_config.TextColumn("FMV Grade", width="small", disabled=True,
                                                  help="A=high data confidence, B=medium, C=low"),
                                "Price ($)":  st.column_config.NumberColumn("Your Price", min_value=0.01, format="$%.2f"),
                                "Condition":  st.column_config.SelectboxColumn(
                                                  "Condition", options=list(_RAW_CONDITIONS.keys()), width="medium"),
                                "Graded":     st.column_config.CheckboxColumn("Graded?", width="small",
                                                  help="Check if this card is in a graded slab (PSA/BGS/SGC/CGC)"),
                                "Grader":     st.column_config.SelectboxColumn(
                                                  "Grader", options=["PSA","BGS","SGC","CGC"], width="small"),
                                "Grade":      st.column_config.TextColumn("Grade ✏️", width="small",
                                                  help="e.g. 10, 9.5, 9 — added to title and CD: fields"),
                                "Cert #":     st.column_config.TextColumn("Cert # ✏️", width="medium",
                                                  help="PSA/BGS/SGC cert number from the label"),
                                "🔍 eBay":    st.column_config.LinkColumn("🔍 eBay", display_text="View Sold", width="small"),
                            },
                        )

                        # ── Upload card images ────────────────────────────────────
                        _missing_img = [c for c in identified if not c.get("front_url")]
                        _has_img     = [c for c in identified if c.get("front_url")]
                        if _missing_img or True:  # always show so user can replace images too
                            with st.expander(
                                f"📷 Upload card images ({len(_has_img)}/{len(identified)} have photos)",
                                expanded=bool(_missing_img),
                            ):
                                _img_card_opts = {
                                    f"Card {c['idx']+1} — {c.get('player','?')} {('✅' if c.get('front_url') else '⚠️ no image')}": c
                                    for c in identified
                                }
                                _img_sel_label = st.selectbox(
                                    "Which card?", list(_img_card_opts.keys()), key="img_upload_card_sel"
                                )
                                _img_target = _img_card_opts[_img_sel_label]
                                _iu_c1, _iu_c2 = st.columns(2)
                                _iu_front = _iu_c1.file_uploader(
                                    "Front image", type=["jpg","jpeg","png"], key="iu_front"
                                )
                                _iu_back = _iu_c2.file_uploader(
                                    "Back image (optional)", type=["jpg","jpeg","png"], key="iu_back"
                                )
                                if _iu_front and st.button("⬆️ Upload & attach", key="iu_upload_btn", type="primary"):
                                    with st.spinner("Uploading…"):
                                        try:
                                            _f_url, _ = _scan_upload_to_supabase(
                                                _iu_front.getvalue(), f"front_{_img_target['idx']}_{_iu_front.name}"
                                            )
                                            _b_url = ""
                                            if _iu_back:
                                                _b_url, _ = _scan_upload_to_supabase(
                                                    _iu_back.getvalue(), f"back_{_img_target['idx']}_{_iu_back.name}"
                                                )
                                            for _bc in st.session_state.raw_batch:
                                                if _bc.get("idx") == _img_target["idx"]:
                                                    _bc["front_url"] = _f_url
                                                    if _b_url:
                                                        _bc["back_url"] = _b_url
                                                    break
                                            st.success(f"✅ Images attached to Card {_img_target['idx']+1} — {_img_target.get('player','')}")
                                            st.rerun()
                                        except Exception as _iu_e:
                                            st.error(f"Upload failed: {_iu_e}")

                        # ── Re-identify low-confidence cards ──────────────────────
                        low_conf_cards = [c for c in identified if c.get("low_conf")]
                        if low_conf_cards:
                            st.markdown("---")
                            with st.expander("🔍 Re-identify a card (search by name / set)", expanded=True):
                                reid_options = {f"Card {c['idx']+1} — {c.get('player','')} {c.get('set_name','')} (Conf: {c.get('similarity',0):.0f}%)": c for c in low_conf_cards}
                                reid_label = st.selectbox("Which card to re-identify:", list(reid_options.keys()), key="reid_select")
                                reid_card  = reid_options[reid_label]
                                reid_q = st.text_input("Search query (player name, year, set):",
                                    value=f"{reid_card.get('player','')} {reid_card.get('set_name','')}".strip(),
                                    key="reid_query",
                                    placeholder="e.g. Ceddanne Rafaela 2024 Topps Chrome")
                                if st.button("🔍 Search", key="reid_go") and reid_q:
                                    with st.spinner("Searching catalog…"):
                                        sr = _ch_post("/v1/cards/card-search", {"search": reid_q, "page": 1, "page_size": 10}) or {}
                                        for _rk in ("cards", "data", "results", "items"):
                                            if _rk in sr and isinstance(sr[_rk], list):
                                                reid_results = sr[_rk]; break
                                        else:
                                            reid_results = sr if isinstance(sr, list) else []
                                        st.session_state["reid_results"] = reid_results
                                        st.session_state["reid_card_idx"] = reid_card["idx"]
    
                                reid_results = st.session_state.get("reid_results", [])
                                if reid_results and st.session_state.get("reid_card_idx") == reid_card["idx"]:
                                    st.caption(f"{len(reid_results)} results — pick the correct card:")
                                    for ri, r in enumerate(reid_results):
                                        # card-search returns card_name or player; set_name; card_number; year; parallel
                                        player = r.get("card_name") or r.get("player","")
                                        year   = str(r.get("year",""))
                                        set_n  = r.get("set_name") or r.get("set","")
                                        num    = r.get("card_number") or r.get("number","")
                                        par    = r.get("parallel") or r.get("variant","")
                                        par    = par if par and par.lower() not in ("base","") else ""
                                        rc_title = f"{year} {set_n} {player.upper()} #{num}".strip()
                                        if st.button(f"✓  {rc_title}{' — '+par if par else ''}", key=f"reid_pick_{ri}"):
                                            for bc in st.session_state.raw_batch:
                                                if bc["idx"] == reid_card["idx"]:
                                                    parts  = [year, set_n, player.upper(), f"#{num}" if num else "", par]
                                                    bc["title"]      = " ".join(p for p in parts if p)[:80]
                                                    bc["player"]     = player
                                                    bc["set_name"]   = set_n
                                                    bc["number"]     = num
                                                    bc["variant"]    = par
                                                    bc["card_id"]    = r.get("card_id") or r.get("id","")
                                                    bc["similarity"] = 90
                                                    bc["low_conf"]   = False
                                                    st.session_state["reid_results"] = []
                                                    st.rerun()
    
                        # ── Comps viewer ──────────────────────────────────────────
                        st.markdown("---")
                        st.markdown("**Comparable Sales**")
                        comp_options = {
                            f"Card {c['idx']+1} — {c.get('player','')} {c.get('set_name','')}": c
                            for c in identified if c.get("card_id")
                        }
                        if comp_options:
                            sel_label = st.selectbox("View comps for:", list(comp_options.keys()), key="raw_comp_select")
                            sel_card  = comp_options[sel_label]
                            cid       = sel_card["card_id"]
                            if cid not in st.session_state.raw_batch_comps:
                                with st.spinner("Loading comparable sales…"):
                                    st.session_state.raw_batch_comps[cid] = ch_comps_raw(cid)
                            comp = st.session_state.raw_batch_comps.get(cid, {})
                            sales = comp.get("raw_prices") or []
                            if sales:
                                mc1, mc2, mc3, mc4 = st.columns(4)
                                mc1.metric("Comp Price", f"${comp.get('comp_price', 0):.2f}")
                                mc2.metric("High",       f"${comp.get('high', 0):.2f}")
                                mc3.metric("Low",        f"${comp.get('low', 0):.2f}")
                                mc4.metric("Sales used", comp.get("count_used", "—"))
                                comp_df = pd.DataFrame([{
                                    "Date":   (s.get("sale_date") or "")[:10],
                                    "Price":  f"${s.get('price', 0):.2f}",
                                    "Source": s.get("price_source", ""),
                                } for s in sales[:10]])
                                st.dataframe(comp_df, use_container_width=True, hide_index=True)
                            else:
                                fmv_val = sel_card.get("fmv")
                                if fmv_val:
                                    st.info(f"No raw comps found in CardHedger — FMV estimate: ${fmv_val:.2f} ({sel_card.get('fmv_conf','')})")
                                else:
                                    st.info("No comparable raw sales found for this card.")
    
                        # ── Export ────────────────────────────────────────────────
                        st.markdown("---")
                        st.markdown("**Step 3 — Export to eBay**")
                        st.caption("✏️ Edit the **Your Price** column above to set your sell price per card before exporting.")
                        edited_records = edited.to_dict("records")
                        n_included = sum(1 for r in edited_records if r.get("✓"))

                        # ── Auto-fill grading info from cert numbers ──────────────
                        # Finds rows where Graded=True + Cert # filled + Grade missing,
                        # looks up the grade via CardHedger/PSA API, and updates raw_batch.
                        _needs_grade = [(i, r) for i, r in enumerate(edited_records)
                                        if r.get("Graded") and (r.get("Cert #") or "").strip()
                                        and not (r.get("Grade") or "").strip()]
                        if _needs_grade:
                            if st.button(f"🎫 Auto-fill grade for {len(_needs_grade)} card(s) from cert #",
                                         key="autofill_grade_btn", type="secondary"):
                                _fill_prog = st.progress(0.0)
                                for _fi, (_gi, _gr) in enumerate(_needs_grade):
                                    _fill_prog.progress((_fi + 1) / len(_needs_grade))
                                    _fc = (_gr.get("Cert #") or "").strip()
                                    _fg = (_gr.get("Grader") or "PSA").strip()
                                    try:
                                        # Try CardHedger cert lookup first
                                        _fcr  = ch_prices_by_cert(_fc, _fg, 90)
                                        _fci  = (_fcr or {}).get("cert_info") or {}
                                        _fgrade = str(_fci.get("grade", "") or "").strip()
                                        for _pfx in ("PSA ", "BGS ", "SGC ", "CGC "):
                                            if _fgrade.upper().startswith(_pfx):
                                                _fgrade = _fgrade[len(_pfx):]
                                                break
                                        # Fall back to PSA API if CardHedger didn't return grade
                                        if not _fgrade and _fg.upper() == "PSA":
                                            import urllib.request as _ur_fg, json as _json_fg
                                            _psa_req = _ur_fg.Request(
                                                f"https://api.psacard.com/publicapi/cert/GetByCertNumber/{_fc}",
                                                headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
                                            )
                                            with _ur_fg.urlopen(_psa_req, timeout=8) as _pr:
                                                _pci = _json_fg.loads(_pr.read().decode()).get("PSACert") or {}
                                            _raw_pg = str(_pci.get("PSAGrade", "") or "").strip()
                                            for _pfx in ("PSA ", "BGS ", "SGC ", "CGC "):
                                                if _raw_pg.upper().startswith(_pfx):
                                                    _raw_pg = _raw_pg[len(_pfx):]
                                                    break
                                            _fgrade = _raw_pg
                                        if _fgrade:
                                            # Update the matching raw_batch card
                                            _card_idx = _gr.get("#", _gi + 1) - 1
                                            for _bc in st.session_state.raw_batch:
                                                if _bc.get("idx") == _card_idx or _bc.get("cert_number") == _fc:
                                                    _bc["grade"]  = _fgrade
                                                    _bc["graded"] = True
                                                    _bc["grader"] = _fg
                                                    break
                                    except Exception:
                                        pass
                                _fill_prog.empty()
                                st.rerun()

                        # PSA image fetch — only show when graded cards are in the batch
                        _graded_rows = [(i, r) for i, r in enumerate(edited_records) if r.get("✓") and r.get("Graded") and (r.get("Cert #") or "").strip()]
                        if _graded_rows:
                            _psa_img_cache = st.session_state.get("psa_image_cache", {})
                            _psa_fetch_col1, _psa_fetch_col2 = st.columns([2, 2])
                            if _psa_fetch_col1.button("📸 Fetch PSA Images", key="fetch_psa_images_btn",
                                                      help="Pulls card images from PSA by cert number — uses them as eBay photo URLs"):
                                import urllib.request as _ur
                                import json as _json_psa
                                _fetched = 0
                                _failed  = []
                                for _gi, _gr in _graded_rows:
                                    _cert = (_gr.get("Cert #") or "").strip()
                                    if not _cert or _cert in _psa_img_cache:
                                        continue
                                    try:
                                        _req = _ur.Request(
                                            f"https://api.psacard.com/publicapi/cert/GetByCertNumber/{_cert}",
                                            headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0"}
                                        )
                                        with _ur.urlopen(_req, timeout=8) as _resp:
                                            _data = _json_psa.loads(_resp.read().decode())
                                        _img = (_data.get("PSACert") or {}).get("ImageURL") or \
                                               (_data.get("PSACert") or {}).get("FrontImageURL") or ""
                                        if _img:
                                            _psa_img_cache[_cert] = _img
                                            _fetched += 1
                                        else:
                                            _failed.append(_cert)
                                    except Exception as _e:
                                        _failed.append(f"{_cert} ({_e})")
                                st.session_state["psa_image_cache"] = _psa_img_cache
                                if _fetched:
                                    st.success(f"✅ Fetched images for {_fetched} cert(s)")
                                if _failed:
                                    st.warning(f"Could not fetch: {', '.join(_failed)}")
                            _cached_certs = [r.get("Cert #","").strip() for _,r in _graded_rows if (r.get("Cert #","").strip() in st.session_state.get("psa_image_cache",{}))]
                            if _cached_certs:
                                _psa_fetch_col2.success(f"📸 PSA images ready for {len(_cached_certs)} cert(s)")
    
                        # Helper: parse manufacturer and year from set_name
                        def _mfr(set_name):
                            s = (set_name or "").lower()
                            for m in ["Topps","Bowman","Panini","Donruss","Upper Deck","Score","Fleer","Pacific","Select","Prizm","Mosaic","Optic","Stadium Club"]:
                                if m.lower() in s:
                                    return m
                            return "Topps"
    
                        def _year(set_name):
                            m = re.search(r"\b(19|20)\d{2}\b", set_name or "")
                            return m.group() if m else ""
    
                        # Condition ID → eBay item specific value
                        # CD:Card Condition (40001) option IDs for raw (ungraded) cards.
                        # 4000→400011 confirmed from Card Dealer Pro export Aug 2026.
                        # Others estimated sequentially — update if eBay rejects them.
                        _COND_SPECIFIC = {
                            "2750": "400010",  # NM or Better (estimated)
                            "3000": "400010",  # Excellent (estimated — same tier as NM)
                            "4000": "400011",  # Very Good — CONFIRMED
                            "5000": "400011",  # Good (estimated)
                            "6000": "400011",  # Poor (estimated)
                        }
    
                        ACTION_COL = "*Action(SiteID=US|Country=US|Currency=USD|Version=1193)"
                        # Exact 82-column eBay File Exchange template (category 261328 — Sports Trading Cards)
                        COLS = [
                            ACTION_COL, "Custom label (SKU)", "Category ID", "Category name",
                            "Title", "Relationship", "Relationship details", "Schedule Time",
                            "P:UPC", "P:EPID", "Start price", "Quantity", "Item photo URL",
                            "VideoID", "Condition ID",
                            "Description", "Format", "Duration", "Buy It Now price",
                            "Best Offer Enabled", "Best Offer Auto Accept Price", "Minimum Best Offer Price",
                            "Immediate pay required", "Location",
                            "Shipping service 1 option", "Shipping service 1 cost", "Shipping service 1 priority",
                            "Shipping service 2 option", "Shipping service 2 cost", "Shipping service 2 priority",
                            "Max dispatch time", "Returns accepted option", "Returns within option",
                            "Refund option", "Return shipping cost paid by",
                            "Shipping profile name", "Return profile name", "Payment profile name",
                            "ProductCompliancePolicyID", "Regional ProductCompliancePolicies",
                            "C:Sport", "C:Player/Athlete", "C:Manufacturer", "C:Season",
                            "C:Parallel/Variety", "C:Features", "C:Set", "C:Team", "C:League",
                            "C:Autographed", "C:Card Name", "C:Card Number", "C:Type",
                            "CD:Professional Grader - (ID: 27501)", "CD:Grade - (ID: 27502)",
                            "CDA:Certification Number - (ID: 27503)", "CD:Card Condition - (ID: 40001)",
                            "C:Graded",
                            "Product Safety Pictograms", "Product Safety Statements", "Product Safety Component",
                            "Regulatory Document Ids",
                            "Manufacturer Name", "Manufacturer AddressLine1", "Manufacturer AddressLine2",
                            "Manufacturer City", "Manufacturer Country", "Manufacturer PostalCode",
                            "Manufacturer StateOrProvince", "Manufacturer Phone", "Manufacturer Email",
                            "Manufacturer ContactURL",
                            "Responsible Person 1", "Responsible Person 1 Type",
                            "Responsible Person 1 AddressLine1", "Responsible Person 1 AddressLine2",
                            "Responsible Person 1 City", "Responsible Person 1 Country",
                            "Responsible Person 1 PostalCode", "Responsible Person 1 StateOrProvince",
                            "Responsible Person 1 Phone", "Responsible Person 1 Email",
                            "Responsible Person 1 ContactURL",
                        ]
    
                        # Block export if Duane has graded cards without Grade filled
                        _is_duane_export = st.session_state.get("access_name") == "Duane"
                        _grade_warnings = [r.get("Player","Card") for r in edited_records
                                           if r.get("✓") and r.get("Graded") and not (r.get("Grade") or "").strip()] if _is_duane_export else []
                        if _grade_warnings:
                            st.error(f"⛔ Cannot export — {len(_grade_warnings)} graded card(s) are missing a Grade: **{', '.join(_grade_warnings)}**. Fill in the Grade column or use the 🎫 Auto-fill button above.")

                        ex_c1, ex_c2 = st.columns([2, 1])
                        if ex_c1.button(f"⬇️ Export eBay CSV ({n_included} cards)", type="primary", key="raw_export_btn",
                                        disabled=bool(_grade_warnings)):
                            import datetime as _dt_exp
                            date_str   = _dt_exp.date.today().strftime("%m%d%y")
                            export_idx = 1
    
                            # Pull settings from session state
                            _sx_sku    = st.session_state.get("bx_sku_prefix", "DFS")
                            _sx_sport  = st.session_state.get("bx_sport", "BASEBALL")
                            _sx_bo     = "1" if st.session_state.get("bx_best_offer", True) else "0"
                            _sx_tmpl   = st.session_state.get("bx_desc_template", "")

                            # Drip schedule setup — compute once before loop
                            import datetime as _dt_drip
                            _drip_on   = st.session_state.get("bx_drip_enabled", False)
                            _drip_hrs  = int(st.session_state.get("bx_drip_spread_hours", 8))
                            _drip_start_str = st.session_state.get("bx_drip_start", "").strip()
                            if _drip_start_str:
                                try:
                                    _drip_base = _dt_drip.datetime.strptime(_drip_start_str, "%Y-%m-%d %H:%M")
                                except ValueError:
                                    _drip_base = (_dt_drip.datetime.now() + _dt_drip.timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
                            else:
                                _drip_base = (_dt_drip.datetime.now() + _dt_drip.timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
                            # Count how many cards will actually be included
                            _drip_total = sum(1 for _r, _o in zip(edited_records, identified) if _r.get("✓"))
                            # Cards per hour = ceil(total / spread_hours), minimum 1
                            _drip_cph = max(1, -(-_drip_total // max(1, _drip_hrs)))
                            _sport_cats = {
                                "BASEBALL":   st.session_state.get("bx_store_cat_baseball",  "44411116016"),
                                "BASKETBALL": st.session_state.get("bx_store_cat_basketball","44411138016"),
                                "FOOTBALL":   st.session_state.get("bx_store_cat_football",  "44411117016"),
                                "SOCCER":     st.session_state.get("bx_store_cat_soccer",    "44411118016"),
                                "OTHER":      st.session_state.get("bx_store_cat_other",     "0"),
                            }
                            _league_map = {"BASEBALL":"MLB","BASKETBALL":"NBA","FOOTBALL":"NFL","SOCCER":"MLS","OTHER":""}
    
                            buf = io.StringIO()
                            # 3 INFO rows required by eBay File Exchange before the header
                            blank = [""] * (len(COLS) - 3)
                            buf.write(",".join(["#INFO", f"Created={int(__import__('time').time()*1000)}", " Indicates missing required fields"] + blank) + "\n")
                            buf.write(",".join(["#INFO", "Version=1.0", "Template=fx_category_template_EBAY_US", " Indicates missing recommended field"] + [""] * (len(COLS) - 4)) + "\n")
                            buf.write(",".join(["#INFO"] + [""] * (len(COLS) - 1)) + "\n")
                            # Row 4: Column headers
                            writer = csv.DictWriter(buf, fieldnames=COLS, extrasaction="ignore")
                            writer.writeheader()
    
                            _sx_bo_min = st.session_state.get("bx_bo_min_price", "").strip()
                            _sx_bo_aa  = st.session_state.get("bx_bo_auto_accept_price", "").strip()

                            for row, orig in zip(edited_records, identified):
                                if not row.get("✓"):
                                    continue
                                title    = (row.get("eBay Title") or orig.get("title", ""))[:80]
                                price    = float(row.get("Price ($)") or 2.49)
                                cond_lbl = row.get("Condition", st.session_state.get("bx_default_condition","Near Mint (NM)"))
                                cond_id  = _RAW_CONDITIONS.get(cond_lbl, "2750")
                                cond_sp  = _COND_SPECIFIC.get(cond_id, "400010")
                                # Graded card fields
                                _is_graded  = bool(row.get("Graded", False))
                                _grader_key = (row.get("Grader") or "PSA").strip()
                                _grade_val  = (row.get("Grade") or "").strip()
                                _cert_num   = (row.get("Cert #") or "").strip()
                                # CD: graded-card fields (27501/27502) only active for Duane
                                # until the eBay option ID mapping is fully verified.
                                _graded_export_enabled = st.session_state.get("access_name") == "Duane"
                                _has_grade  = bool(_is_graded and _grade_val and _graded_export_enabled)
                                _cd_grader  = EBAY_GRADER_VALUES.get(_grader_key, _grader_key) if _has_grade else ""
                                _cd_grade   = EBAY_GRADE_VALUES.get(str(_grade_val), f"{_grade_val} - (ID: 275020)") if _has_grade else ""
                                _cd_cert    = _cert_num if _has_grade else ""
                                _c_graded   = "Yes" if _is_graded else ""
                                # Custom SKU: use what's in the table (user may have edited it)
                                _sx_sku_name = _sx_sku.split("-")[0] if _sx_sku else "DFS"
                                sku      = (row.get("Custom SKU") or "").strip() or f"{_sx_sku_name}-{date_str}-{export_idx:04d}"
                                front_u  = orig.get("front_url", "")
                                back_u   = orig.get("back_url", "")
                                # Graded: prefer PSA image if fetched
                                _psa_img_cache = st.session_state.get("psa_image_cache", {})
                                _psa_img = _psa_img_cache.get(_cert_num, "") if _is_graded and _cert_num else ""
                                if _psa_img:
                                    pic_url = _psa_img
                                else:
                                    pic_url = f"{front_u}|{back_u}" if back_u else front_u
                                player   = orig.get("player", "")
                                set_n    = orig.get("set_name", "")
                                number   = orig.get("number", "")
                                team     = orig.get("team", "")
                                # Parallel: prefer what user typed in the table
                                _tbl_par = (row.get("Parallel") or "").strip()
                                _api_par = orig.get("variant", "")
                                par      = _tbl_par if _tbl_par else (_api_par if _api_par and _api_par.lower() not in ("base","") else "")
                                # Rebuild title if parallel was corrected
                                _base_title = row.get("eBay Title") or orig.get("title", "")
                                if _tbl_par and _tbl_par.lower() not in _base_title.lower():
                                    _parts = [_year(set_n), set_n, player.upper(), f"#{number}" if number else "", _tbl_par]
                                    title = " ".join(p for p in _parts if p)[:80]
                                else:
                                    title = _base_title[:80]
                                year      = _year(set_n)
                                mfr       = _mfr(set_n)
                                sim       = orig.get("similarity", 0)
                                # Per-row sport (editable in table for mixed-sport batches)
                                row_sport = (row.get("Sport") or _sx_sport).upper()
                                store_cat = _sport_cats.get(row_sport, _sport_cats.get(_sx_sport, "0"))
                                league    = _league_map.get(row_sport, _league_map.get(_sx_sport, ""))
                                features  = "Parallel" if par else ""

                                if price < 1.00:
                                    _exp_ship_cost = "0.00"; _exp_ship_free = "1"
                                elif price < 20:
                                    _exp_ship_cost = "0.74"; _exp_ship_free = "0"
                                else:
                                    _exp_ship_cost = "0.00"; _exp_ship_free = "1"

                                if _sx_tmpl:
                                    desc = _sx_tmpl.replace("[LISTING_TITLE]", title).replace("[FRONT_IMAGE_URL]", front_u)
                                else:
                                    desc = _scan_description(title, front_u)

                                # Drip: assign each card its scheduled slot (export_idx is 1-based count of included cards)
                                _schedule_time = ""
                                if _drip_on and _drip_total > 0:
                                    _slot = (export_idx - 1) // _drip_cph
                                    _card_dt = _drip_base + _dt_drip.timedelta(hours=_slot)
                                    # eBay ScheduleTime: ISO 8601 format (seller account timezone, not UTC-stamped)
                                    _schedule_time = _card_dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")

                                # SEO title: YEAR SET PLAYER #NUMBER PARALLEL — player always in CAPS
                                _seo_parts = [p for p in [year, set_n, player.upper() if player else "",
                                                           f"#{number}" if number else "", par] if p]
                                seo_title = " ".join(_seo_parts)[:80]
                                # Use pre-built title from table if set, otherwise use SEO format
                                final_title = title if title else seo_title
                                # Graded: append "PSA 10" / "BGS 9.5" etc. to title
                                if _is_graded and _grader_key and _grade_val:
                                    _grade_suffix = f"{_grader_key} {_grade_val}"
                                    final_title = f"{final_title[:80 - len(_grade_suffix) - 1]} {_grade_suffix}"[:80]

                                # Graded slabs: force condition 2750 (NM or Better) ONLY when
                                # both grader and grade are populated — eBay requires 27501+27502
                                # whenever condition 2750 is used in category 261328.
                                _export_cond_id = "2750" if _has_grade else cond_id

                                writer.writerow({
                                    ACTION_COL:                               "Add",
                                    "Custom label (SKU)":                     sku,
                                    "Category ID":                            "261328",
                                    "Category name":                          "Sports Trading Cards",
                                    "Title":                                  final_title,
                                    "Schedule Time":                          _schedule_time,
                                    "Condition ID":                           _export_cond_id,
                                    "Item photo URL":                         pic_url,
                                    "Description":                            desc,
                                    "Format":                                 "FixedPrice",
                                    "Duration":                               "GTC",
                                    "Start price":                            f"{price:.2f}",
                                    "Quantity":                               "1",
                                    "Immediate pay required":                 "1",
                                    "Location":                               "Scottsdale,AZ",
                                    "Shipping service 1 option":              "US_eBayStandardEnvelope",
                                    "Shipping service 1 cost":                _exp_ship_cost,
                                    "Shipping service 1 priority":            "1",
                                    "Max dispatch time":                      "2",
                                    "Returns accepted option":                "ReturnsNotAccepted",
                                    "Best Offer Enabled":                     _sx_bo,
                                    "Minimum Best Offer Price":               _sx_bo_min if _sx_bo == "1" else "",
                                    "Best Offer Auto Accept Price":           _sx_bo_aa  if _sx_bo == "1" else "",
                                    "Shipping profile name":                  st.session_state.get("bx_shipping_profile", ""),
                                    "Return profile name":                    st.session_state.get("bx_return_profile", ""),
                                    "Payment profile name":                   st.session_state.get("bx_payment_profile", ""),
                                    "C:Sport":                                row_sport.capitalize(),
                                    "C:Player/Athlete":                       player,
                                    "C:Manufacturer":                         mfr,
                                    "C:Season":                               year,
                                    "C:Parallel/Variety":                     par,
                                    "C:Features":                             features,
                                    "C:Set":                                  set_n,
                                    "C:Team":                                 team,
                                    "C:League":                               league,
                                    "C:Autographed":                          "No",
                                    "C:Card Name":                            player,
                                    "C:Card Number":                          number,
                                    "C:Type":                                 "Sports Trading Card",
                                    "CD:Professional Grader - (ID: 27501)":   _cd_grader,
                                    "CD:Grade - (ID: 27502)":                 _cd_grade,
                                    "CDA:Certification Number - (ID: 27503)": _cd_cert,
                                    "CD:Card Condition - (ID: 40001)":        "" if _has_grade else cond_sp,
                                    "C:Graded":                               _c_graded,
                                })
                                export_idx += 1
                            csv_bytes = buf.getvalue().encode("utf-8")
                            filename  = f"DFS_RawListings_{date_str}.csv"
                            # Mark exported cards as listed in Supabase
                            if st.session_state.get("bx_active_stack_id"):
                                _exp_idxs = {orig["idx"] for row, orig in zip(edited_records, identified) if row.get("✓")}
                                _sc_rows  = stack_cards_get(st.session_state["bx_active_stack_id"])
                                for _scr in _sc_rows:
                                    if _scr.get("idx") in _exp_idxs:
                                        stack_card_update(_scr["id"], {"status": "listed"})
                                        for _bc in st.session_state.raw_batch:
                                            if _bc.get("idx") == _scr.get("idx"):
                                                _bc["status"] = "listed"
                                st.session_state["bx_stacks_cache"] = None
                            st.download_button(
                                label=f"📥 Download {filename}",
                                data=csv_bytes,
                                file_name=filename,
                                mime="text/csv",
                                key="raw_dl_btn",
                            )
    
                        # ── Step 4: eBay Upload Error Reader ──────────────────
                        st.markdown("---")
                        st.markdown("**Step 4 — Review eBay Upload Results** *(optional)*")
                        st.caption("After uploading to eBay File Exchange, paste the response CSV here to see plain-English results.")
                        _ebay_resp_file = st.file_uploader(
                            "📋 Drop eBay response CSV here", type=["csv"],
                            key="ebay_resp_upload",
                            help="The file eBay emails/shows after a File Exchange upload — contains success/failure per row.",
                        )
                        if _ebay_resp_file:
                            import csv as _csv_resp
                            import io as _io_resp

                            _EBAY_ERR = {
                                "21920351": "Card Condition descriptor not valid for this condition — graded cards must use Condition ID 3000",
                                "21920352": "Condition descriptor value not accepted — check grader name or grade value",
                                "21916626": "Missing required item specific for this category",
                                "21916635": "Item specific value not valid for this field",
                                "240":      "Title is missing or too short",
                                "291":      "Listing already exists with this SKU",
                                "21916587": "Picture URL is invalid or unreachable",
                                "21916580": "Picture URL is empty",
                                "21919456": "Business policy warning — use policy IDs instead of legacy shipping fields",
                                "23015":    "Immediate payment warning with Best Offer (safe to ignore)",
                                "21920218": "Photo URL could not be fetched by eBay",
                                "17":       "Category ID not found",
                                "21916884": "Condition ID not valid for category 261328 — graded cards must use 2750 (NM or Better)",
                            }

                            try:
                                _resp_text = _ebay_resp_file.read().decode("utf-8", errors="replace")
                                _resp_reader = _csv_resp.DictReader(_io_resp.StringIO(_resp_text))
                                _resp_rows = list(_resp_reader)

                                _successes = [r for r in _resp_rows if (r.get("Status") or "").strip().lower() == "success"]
                                _failures  = [r for r in _resp_rows if (r.get("Status") or "").strip().lower() == "failure"]
                                _warnings  = [r for r in _resp_rows if (r.get("Status") or "").strip().lower() == "warning"]

                                if _successes:
                                    st.success(f"✅ {len(_successes)} listing{'s' if len(_successes)!=1 else ''} uploaded successfully" +
                                               (f" — Item IDs: {', '.join(r.get('ItemID','') for r in _successes if r.get('ItemID'))}" if any(r.get('ItemID') for r in _successes) else ""))

                                for _fr in _failures:
                                    _sku    = _fr.get("CustomLabel") or _fr.get("Custom label (SKU)") or "?"
                                    _codes  = [c.strip() for c in (_fr.get("ErrorCode") or "").split("|") if c.strip() and c.strip().isdigit()]
                                    _msgs   = [m.strip() for m in (_fr.get("ErrorMessage") or "").split("|") if m.strip() and not m.strip().startswith("Error -")]
                                    # Deduplicate messages
                                    _seen_msgs = set()
                                    _clean_msgs = []
                                    for _m in _msgs:
                                        _mk = _m.lower()[:60]
                                        if _mk not in _seen_msgs:
                                            _seen_msgs.add(_mk)
                                            _clean_msgs.append(_m)

                                    with st.container():
                                        st.error(f"❌ **{_sku}** — Upload failed")
                                        for _code in _codes:
                                            _plain = _EBAY_ERR.get(_code)
                                            if _plain:
                                                st.markdown(f"&nbsp;&nbsp;• **[{_code}]** {_plain}")
                                            else:
                                                # Fall back to eBay's raw message (trimmed)
                                                _raw = next((_m for _m in _clean_msgs if _m), "")
                                                if _raw:
                                                    st.markdown(f"&nbsp;&nbsp;• **[{_code}]** {_raw[:200]}")
                                        if _clean_msgs and not _codes:
                                            for _m in _clean_msgs[:3]:
                                                st.markdown(f"&nbsp;&nbsp;• {_m[:200]}")

                                if _warnings and not _failures:
                                    _warn_skus = set(r.get("CustomLabel","") for r in _warnings)
                                    _w_codes = set()
                                    for _wr in _warnings:
                                        for _wc in (_wr.get("WarningCode") or "").split("|"):
                                            _wc = _wc.strip()
                                            if _wc: _w_codes.add(_wc)
                                    _safe_warns = {"21919456", "23015"}
                                    if _w_codes <= _safe_warns:
                                        st.success(f"✅ Uploaded with minor warnings (safe to ignore) — {', '.join(_warn_skus)}")
                                    else:
                                        st.warning(f"⚠️ {len(_warnings)} row(s) with warnings — check eBay Seller Hub to confirm they listed.")

                                if not _successes and not _failures and not _warnings:
                                    st.info("No result rows found — make sure you're uploading the eBay *response* file, not your original upload CSV.")

                            except Exception as _re:
                                st.error(f"Could not parse response file: {_re}")

                        if ex_c2.button("🗑️ Clear Batch", key="raw_clear_btn"):
                            st.session_state.raw_batch = []
                            st.session_state.raw_batch_comps = {}
                            st.rerun()
    
                elif not all_files and not st.session_state.raw_batch:
                    st.info("💡 **How it works:** Select all scans at once — front and back interleaved (front1, back1, front2, back2…). CardHedger visually matches each card, shows recent sold comps, and exports an eBay Add CSV with both images on every listing.")

        # ── GRADE PREDICTOR ───────────────────────────────────────────────────
        with scan_grade:
            st.markdown("### 🔬 PSA Grade Predictor")
            st.caption("Upload a high-res front (and optionally back) scan. Claude Vision analyzes centering, corners, edges, and surface — then predicts your PSA grade.")

            # ── Step 1: Quick GemRate check ───────────────────────────────────
            st.markdown("#### Step 1 — Is it worth grading?")
            st.caption("Check gem rate and pricing before you scan. Skip if you already know.")
            _gq_col, _gq_btn = st.columns([5, 1])
            with _gq_col:
                _gq = st.text_input("Card name", placeholder="e.g. Luka Doncic Prizm RC PSA — or leave blank to skip", key="gp_quick_query", label_visibility="collapsed")
            with _gq_btn:
                _gq_go = st.button("Check", key="gp_quick_go", use_container_width=True)

            if _gq and (_gq_go or st.session_state.get("gp_last_q") != _gq):
                st.session_state["gp_last_q"] = _gq
                with st.spinner("Fetching gem rate + pricing…"):
                    _gq_gr = search_gemrate(_gq)
                    _gq_ch = ch_card_match(_gq) if CARDHEDGER_KEY else None
                st.session_state["gp_quick_results"] = (_gq_gr, _gq_ch)

            _gq_data = st.session_state.get("gp_quick_results")
            if _gq_data:
                _gq_gr_res, _gq_ch_res = _gq_data
                if _gq_gr_res:
                    _gq_sel = _gq_gr_res[0]
                    _gq_gem = _gq_sel.get("gem_rate")
                    _gq_pop = _gq_sel.get("total_population", 0)
                    _gq_gems = _gq_sel.get("gems", 0)
                    _gq_ch_raw, _gq_ch_p10 = None, None
                    if _gq_ch_res and CARDHEDGER_KEY:
                        _gq_id = _gq_ch_res.get("card_id") or _gq_ch_res.get("id")
                        if _gq_id:
                            _gq_fmv = ch_fmv_batch([{"card_id": _gq_id, "grade": "Raw"}, {"card_id": _gq_id, "grade": "PSA 10"}])
                            _gq_items = {(i.get("grade") or "").upper(): i for i in (_gq_fmv.get("items") or _gq_fmv.get("results") or [])}
                            _gq_ch_raw = fmv_price(_gq_items.get("RAW") or {})
                            _gq_ch_p10 = fmv_price(_gq_items.get("PSA 10") or {})

                    _gq_m1, _gq_m2, _gq_m3, _gq_m4 = st.columns(4)
                    with _gq_m1:
                        st.markdown("**Gem Rate**")
                        st.markdown(gem_bar_html(_gq_gem), unsafe_allow_html=True)
                    _gq_m2.metric("Total Pop", f"{_gq_pop:,}")
                    _gq_m3.metric("Gem Copies", f"{_gq_gems:,}")
                    if _gq_ch_raw or _gq_ch_p10:
                        _gq_m4.metric("💰 Grade Uplift", f"${(_gq_ch_p10 or 0) - (_gq_ch_raw or 0):,.2f}" if _gq_ch_raw and _gq_ch_p10 else "—")
                    _pr1, _pr2, _pr3 = st.columns(3)
                    if _gq_ch_raw:  _pr1.metric("📦 Raw Avg",    f"${_gq_ch_raw:,.2f}")
                    if _gq_ch_p10:  _pr2.metric("💎 PSA 10 Avg", f"${_gq_ch_p10:,.2f}")
                    _gq_url = gemrate_url(_gq_sel.get("gemrate_id", ""))
                    if _gq_url:
                        _pr3.markdown(f"[📊 Full pop report on GemRate ↗]({_gq_url})")
                else:
                    st.info("No GemRate results — try a different card name, or skip and proceed to scan.")

            st.markdown("#### Step 2 — Scan the card")
            if not ANTHROPIC_KEY:
                st.warning("⚠️ Add your Anthropic API key to `.streamlit/secrets.toml` to enable AI grading.")
            else:
                _gp_c1, _gp_c2 = st.columns(2)
                with _gp_c1:
                    st.markdown("**Front of card** *(required)*")
                    _gp_front = st.file_uploader("Front scan", type=["jpg","jpeg","png","webp"], key="gp_front", label_visibility="collapsed")
                with _gp_c2:
                    st.markdown("**Back of card** *(recommended)*")
                    _gp_back  = st.file_uploader("Back scan",  type=["jpg","jpeg","png","webp"], key="gp_back",  label_visibility="collapsed")

                if _gp_front:
                    _gp_img_c1, _gp_img_c2 = st.columns(2)
                    with _gp_img_c1:
                        st.image(_gp_front, caption="Front", use_container_width=True)
                    with _gp_img_c2:
                        if _gp_back:
                            st.image(_gp_back, caption="Back", use_container_width=True)
                        else:
                            st.markdown("*No back image — centering/surface back will be estimated*")

                    if st.button("🔬 Analyze Grade", type="primary", key="gp_run"):
                        with st.spinner("Analyzing with Claude Vision… (~10 seconds)"):
                            from PIL import Image as _PILG
                            import io as _iog

                            def _prep_gp(f):
                                _b = f.read()
                                _mime = "image/jpeg" if f.type in ("image/jpeg","image/jpg") else "image/png"
                                try:
                                    _img = _PILG.open(_iog.BytesIO(_b))
                                    _img.thumbnail((2000, 2000), _PILG.LANCZOS)
                                    _buf = _iog.BytesIO()
                                    _img.save(_buf, format="JPEG" if _mime == "image/jpeg" else "PNG", quality=92)
                                    _b = _buf.getvalue()
                                    _mime = "image/jpeg" if _mime == "image/jpeg" else "image/png"
                                except Exception:
                                    pass
                                return base64.b64encode(_b).decode(), _mime

                            _front_b64, _front_mime = _prep_gp(_gp_front)
                            _back_b64, _back_mime = None, "image/jpeg"
                            if _gp_back:
                                _back_b64, _back_mime = _prep_gp(_gp_back)

                            _gp_result = claude_grade_card(_front_b64, _back_b64, _front_mime, _back_mime)
                            st.session_state["gp_result"] = _gp_result

                    # ── Results ───────────────────────────────────────────────
                    _gpr = st.session_state.get("gp_result")
                    if _gpr:
                        if "_error" in _gpr:
                            st.error(f"Analysis failed: {_gpr['_error']}")
                            if "_raw" in _gpr:
                                with st.expander("Raw response"):
                                    st.text(_gpr["_raw"])
                        else:
                            _grade_num   = _gpr.get("predicted_grade", "?")
                            _grade_label = _gpr.get("grade_label", f"PSA {_grade_num}")
                            _confidence  = _gpr.get("confidence", "medium").capitalize()
                            _grade_range = _gpr.get("estimated_grade_range", "")
                            _submit      = _gpr.get("submit_recommended", False)
                            _rec         = _gpr.get("recommendation", "")
                            _caveat      = _gpr.get("caveat", "")

                            # Grade badge
                            _g_color = "#22c55e" if _grade_num == 10 else "#3b82f6" if _grade_num >= 9 else "#f59e0b" if _grade_num >= 7 else "#ef4444"
                            st.markdown(f"""
<div style="background:{_g_color}22;border:2px solid {_g_color};border-radius:12px;padding:16px 20px;margin:12px 0;display:flex;align-items:center;gap:20px;">
  <div style="font-size:3rem;font-weight:900;color:{_g_color};line-height:1;">{_grade_label}</div>
  <div>
    <div style="font-size:1rem;color:#94a3b8;">Predicted grade · Confidence: <strong>{_confidence}</strong></div>
    <div style="font-size:0.9rem;color:#64748b;">Range: {_grade_range}</div>
  </div>
</div>""", unsafe_allow_html=True)

                            # Scorecard
                            st.markdown("#### Scorecard")
                            def _pill(passes, label):
                                col = "#22c55e" if passes else "#ef4444"
                                txt = "PASS" if passes else "FAIL"
                                return f'<span style="background:{col}22;color:{col};border:1px solid {col};border-radius:6px;padding:2px 8px;font-size:0.75rem;font-weight:700;">{txt}</span> {label}'

                            _cen  = _gpr.get("centering", {})
                            _cor  = _gpr.get("corners", {})
                            _edg  = _gpr.get("edges", {})
                            _sur  = _gpr.get("surface", {})

                            _rows = [
                                ("Centering (front)", _cen.get("front_passes_10", False),
                                 f"L/R {_cen.get('front_left_right','?')} · T/B {_cen.get('front_top_bottom','?')}",
                                 _cen.get("note","")),
                                ("Centering (back)",  _cen.get("back_passes_10", False),
                                 f"L/R {_cen.get('back_left_right','?')}",
                                 ""),
                                ("Corners", _cor.get("passes_10", False),
                                 f"NW:{_cor.get('NW','?')} · NE:{_cor.get('NE','?')} · SW:{_cor.get('SW','?')} · SE:{_cor.get('SE','?')}",
                                 _cor.get("note","")),
                                ("Edges", _edg.get("passes_10", False),
                                 f"Top:{_edg.get('top','?')} · Bot:{_edg.get('bottom','?')} · L:{_edg.get('left','?')} · R:{_edg.get('right','?')}",
                                 _edg.get("note","")),
                                ("Surface (front)", _sur.get("passes_10", False),
                                 ", ".join(_sur.get("front_defects",[]) or ["Clean"]),
                                 _sur.get("note","")),
                                ("Surface (back)", _sur.get("passes_10", False),
                                 ", ".join(_sur.get("back_defects",[]) or ["Clean"]),
                                 ""),
                            ]
                            _sc_html = '<table style="width:100%;border-collapse:collapse;font-size:0.88rem;">'
                            for _rname, _passes, _detail, _note in _rows:
                                _sc_html += f'<tr style="border-bottom:1px solid #1e293b;"><td style="padding:7px 4px;width:140px;font-weight:600;">{_rname}</td><td style="padding:7px 8px;">{_pill(_passes, "")}</td><td style="padding:7px 4px;color:#94a3b8;">{_detail}</td><td style="padding:7px 4px;color:#64748b;font-style:italic;">{_note}</td></tr>'
                            _sc_html += "</table>"
                            st.markdown(_sc_html, unsafe_allow_html=True)

                            # Fatal flaws & positives
                            _flaws = _gpr.get("fatal_flaws", [])
                            _pos   = _gpr.get("positive_attributes", [])
                            _fl_c, _po_c = st.columns(2)
                            with _fl_c:
                                if _flaws:
                                    st.markdown("**⚠️ Issues found**")
                                    for _f in _flaws:
                                        st.markdown(f"- {_f}")
                                else:
                                    st.markdown("**✅ No issues detected**")
                            with _po_c:
                                if _pos:
                                    st.markdown("**💪 Strengths**")
                                    for _p in _pos:
                                        st.markdown(f"- {_p}")

                            # Recommendation
                            _rec_color = "#22c55e" if _submit else "#f59e0b"
                            _rec_icon  = "✅" if _submit else "⏸️"
                            st.markdown(f"""
<div style="background:{_rec_color}18;border-left:4px solid {_rec_color};border-radius:8px;padding:12px 16px;margin:12px 0;">
  <strong>{_rec_icon} Recommendation</strong><br/><span style="color:#cbd5e1;">{_rec}</span>
</div>""", unsafe_allow_html=True)

                            if _caveat:
                                st.caption(f"ℹ️ {_caveat}")

                else:
                    st.info("Upload the front scan to get started. High-res scans (600 DPI+) give the most accurate results.")

if _active_tab == 3:
    st.markdown("## 📦 Inventory Check")

    _is_owner = st.session_state.get("access_name", "") == "Duane"
    _wb_label = "DFS Operations Workbook" if _is_owner else "your Operations Workbook"

    st.markdown(
        f"Analyze cards from your intake log for grading potential. "
        f"Or upload a file directly here — use the **🚚 Shipment Intake** tab to bulk-import a shipment."
    )

    st.markdown("""
**How it works:**
1. Upload a CSV or Operations Workbook (.xlsx) → select a card → search GemRate → get GO/NO-GO
2. Add grading candidates directly to the Submission Tracker
""")

    uploaded = st.file_uploader(
        f"Upload inventory (CSV template or {_wb_label} .xlsx)",
        type=["csv", "xlsx"],
        label_visibility="visible",
    )

    if uploaded:
        inv, source = load_inventory(uploaded)

        if inv is None:
            st.error(f"Could not read file: {source}")
        else:
            _source_label = (_wb_label if source == "workbook" else "inventory template")
            st.success(f"Loaded **{len(inv)} cards** from {_source_label}")

            show_cols = [c for c in ["Card Description", "Player", "Year", "Set", "Parallel",
                                     "Category", "Cost Basis ($)", "Listed Price ($)", "Source"] if c in inv.columns]
            st.dataframe(inv[show_cols], use_container_width=True, hide_index=True)

            st.markdown("---")
            st.markdown("### Analyze a card for grading")

            descs = inv["Card Description"].dropna().tolist()
            sel_card = st.selectbox("Select card from your inventory", descs)
            sel_row = inv[inv["Card Description"] == sel_card].iloc[0]

            try:
                cost_basis = float(sel_row.get("Cost Basis ($)", 0) or 0)
            except Exception:
                cost_basis = 0.0

            ic1, ic2 = st.columns(2)
            inv_raw  = ic1.number_input("Cost basis / buy price ($)", value=cost_basis, min_value=0.0, step=5.0, key="inv_raw")
            inv_tier = ic2.selectbox("Grading tier", list(PSA_FEES.keys()), key="inv_tier")

            if st.button("🔎 Search GemRate for this card", key="inv_search"):
                with st.spinner("Searching GemRate..."):
                    st.session_state.inv_results = search_gemrate(sel_card)
                    st.session_state.inv_card = sel_card

            inv_results = st.session_state.get("inv_results", [])
            if inv_results and st.session_state.get("inv_card") == sel_card:
                inv_opts = [
                    f"{r.get('year','')} {r.get('set_name','')} {r.get('name','')} {r.get('parallel') or 'Base'} #{r.get('card_number','')}"
                    for r in inv_results
                ]
                inv_match = st.selectbox("Best GemRate match", inv_opts, key="inv_match")
                inv_sel = inv_results[inv_opts.index(inv_match)]
                gem_i = inv_sel.get("gem_rate")
                desc_i = f"{inv_sel.get('year','')} {inv_sel.get('set_name','')} {inv_sel.get('name','')} {inv_sel.get('parallel') or ''}".strip()

                g1, g2, g3 = st.columns(3)
                with g1:
                    st.markdown("**Gem Rate (PSA 10)**")
                    st.markdown(gem_bar_html(gem_i), unsafe_allow_html=True)
                g2.metric("Total Pop", f"{inv_sel.get('total_population', 0):,}")
                g3.metric("Gem Copies", f"{inv_sel.get('gems', 0):,}")
                st.markdown(f"[📊 GemRate pop report]({gemrate_url(inv_sel.get('gemrate_id',''))})")

                inv_graded_auto = None

                inv_graded = st.number_input(
                    "Gem 10 avg price ($)", min_value=0.0,
                    value=float(inv_graded_auto) if inv_graded_auto else 0.0,
                    step=10.0, key="inv_graded",
                )
                st.markdown(f"[📈 eBay Gem 10 sold comps]({ebay_graded_sold_url(desc_i)})")

                if inv_graded > 0 and inv_raw > 0:
                    v_i, color_i, msg_i = verdict(inv_raw, inv_tier, gem_i, inv_graded, min_gem, roi_target)
                    net_i, roi_i = calc_net_roi(inv_raw, inv_tier, inv_graded)
                    if color_i == "green":
                        st.success(f"{v_i} — {msg_i}")
                    else:
                        st.error(f"{v_i} — {msg_i}")

                    if st.button("➕ Add to Submission Tracker", key="inv_add"):
                        fee_i = PSA_FEES[inv_tier]
                        tgt_i = target_price(inv_raw, inv_tier, roi_target)
                        sb_insert({
                            "date_added": date.today().isoformat(),
                            "card_description": sel_card,
                            "year": inv_sel.get("year", ""),
                            "set_name": inv_sel.get("set_name", ""),
                            "parallel": inv_sel.get("parallel") or "Base",
                            "raw_buy_price": inv_raw,
                            "psa_tier": inv_tier,
                            "psa_fee": fee_i,
                            "psa10_avg_price": inv_graded,
                            "target_price": round(tgt_i, 2),
                            "gem_rate": round(gem_i, 2) if gem_i else None,
                            "go_no_go": v_i,
                            "est_net": net_i,
                            "est_roi": f"{roi_i:.0f}%",
                            "status": "Queued",
                        })
                        st.success("Added to Tracker ✓")

            # ── 💰 Reprice Assistant (bulk comps + trend) ──────────────────────
            st.markdown("---")
            st.markdown("### 💰 Reprice Assistant")
            st.caption(
                "Pull current sold comps + 90-day trend for every card, then get a suggested "
                "new list price. Built for repricing a full inventory at once instead of one eBay listing at a time."
            )

            if not CARDHEDGER_KEY:
                st.info("📊 Connect the CardHedger API to enable live comps and repricing.")
            else:
                rp1, rp2, rp3 = st.columns(3)
                strategy = rp1.selectbox(
                    "Pricing strategy",
                    ["Trend-following", "Match market", "Undercut to sell faster", "List high for offers"],
                    index=0, key="rp_strategy",
                    help="Trend-following raises the suggestion when a card is trending up and cuts it when trending down.",
                )
                if strategy == "Trend-following":
                    adj_pct = rp2.slider("Trend sensitivity (%)", 0, 100, 50, 5, key="rp_sens",
                                         help="How hard to lean into the trend. 50% moves the price half as much as the 90-day trend.")
                elif strategy == "Undercut to sell faster":
                    adj_pct = rp2.slider("Undercut below comp (%)", 0, 30, 7, 1, key="rp_under")
                elif strategy == "List high for offers":
                    adj_pct = rp2.slider("Premium above comp (%)", 0, 30, 10, 1, key="rp_prem")
                else:
                    adj_pct = rp2.slider("Adjust vs comp (%)", -20, 20, 0, 1, key="rp_match")
                misprice_thresh = rp3.slider("Mispricing flag threshold (%)", 5, 30, 10, 1, key="rp_thresh",
                                             help="How far your listed price can sit from market before it's flagged over/underpriced.")

                max_n = max(1, len(inv))
                proc_n = st.slider("How many cards to refresh this run", 1, max_n, min(25, max_n), key="rp_count",
                                   help="Each card makes a few CardHedger calls. Process in batches to stay fast and within your API quota.")

                if st.button("🔄 Refresh comps & trend", key="rp_run"):
                    rows_market = []
                    sub = inv.head(proc_n)
                    total = max(1, len(sub))
                    prog = st.progress(0.0, text="Fetching comps…")
                    for i, (_, row) in enumerate(sub.iterrows()):
                        d = str(row.get("Card Description", "") or "").strip()
                        if not d:
                            continue
                        grade = detect_grade(d)
                        mkt = fetch_market(d, grade)
                        try:
                            listed = float(row.get("Listed Price ($)", 0) or 0)
                        except Exception:
                            listed = 0.0
                        try:
                            cost = float(row.get("Cost Basis ($)", 0) or 0)
                        except Exception:
                            cost = 0.0
                        rows_market.append({
                            "Card": d, "Grade": grade, "Cost": cost, "Listed": listed,
                            "Comp": mkt["comp_avg"], "TrendDir": mkt["trend_dir"],
                            "TrendPct": mkt["trend_pct"], "Matched": mkt["matched"],
                        })
                        prog.progress((i + 1) / total, text=f"Fetching comps… {i + 1}/{total}")
                    prog.empty()
                    st.session_state["reprice_data"] = rows_market
                    st.session_state["reprice_when"] = date.today().isoformat()

                rows_market = st.session_state.get("reprice_data")
                if rows_market:
                    st.caption(
                        f"Last refreshed {st.session_state.get('reprice_when','')} · {len(rows_market)} cards. "
                        f"Adjust the strategy/sliders above to recompute instantly — no new API calls."
                    )
                    computed = []
                    for r in rows_market:
                        comp, listed = r["Comp"], r["Listed"]
                        sugg = suggest_reprice(comp, r["TrendPct"], strategy, adj_pct)
                        if not comp:
                            flag = "no comp"
                        elif listed <= 0:
                            flag = "🟡 No list price"
                        elif (listed - comp) / comp > misprice_thresh / 100:
                            flag = "🔴 Overpriced"
                        elif (listed - comp) / comp < -misprice_thresh / 100:
                            flag = "🟢 Underpriced"
                        else:
                            flag = "⚪ On market"
                        if sugg and r["Cost"] and sugg < r["Cost"]:
                            flag += " ⚠ under cost"
                        delta = (sugg - listed) if (sugg and listed) else None
                        delta_pct = (delta / listed * 100) if (delta is not None and listed) else None
                        computed.append({**r, "Sugg": sugg, "Flag": flag, "Delta": delta, "DeltaPct": delta_pct})

                    n_over = sum(1 for c in computed if c["Flag"].startswith("🔴"))
                    n_under = sum(1 for c in computed if c["Flag"].startswith("🟢"))
                    n_nomatch = sum(1 for c in computed if not c["Matched"])
                    mc1, mc2, mc3 = st.columns(3)
                    mc1.metric("🔴 Overpriced", n_over, help="Listed above market — lower to sell")
                    mc2.metric("🟢 Underpriced", n_under, help="Listed below market — room to raise")
                    mc3.metric("No match", n_nomatch, help="CardHedger couldn't price these — check the card name")

                    tdf = pd.DataFrame([{
                        "Card": c["Card"], "Grade": c["Grade"],
                        "Cost": f"${c['Cost']:,.0f}" if c["Cost"] else "—",
                        "Listed": f"${c['Listed']:,.0f}" if c["Listed"] else "—",
                        "Comp (Market)": f"${c['Comp']:,.0f}" if c["Comp"] else ("—" if c["Matched"] else "no match"),
                        "Trend": trend_label(c["TrendDir"], c["TrendPct"]),
                        "Suggested": f"${c['Sugg']:,.0f}" if c["Sugg"] else "—",
                        "Δ vs Listed": (f"{'+' if c['Delta'] >= 0 else ''}${c['Delta']:,.0f} ({c['DeltaPct']:+.0f}%)"
                                        if c["Delta"] is not None else "—"),
                        "Flag": c["Flag"],
                    } for c in computed])
                    st.dataframe(tdf, use_container_width=True, hide_index=True)

                    csv_buf = io.StringIO()
                    pd.DataFrame([{
                        "Card Description": c["Card"], "Grade": c["Grade"],
                        "Cost Basis": round(c["Cost"], 2) if c["Cost"] else "",
                        "Current Listed Price": round(c["Listed"], 2) if c["Listed"] else "",
                        "Comp Avg (Market)": round(c["Comp"], 2) if c["Comp"] else "",
                        "90-Day Trend": trend_label(c["TrendDir"], c["TrendPct"]),
                        "Suggested Price": round(c["Sugg"], 2) if c["Sugg"] else "",
                        "Flag": c["Flag"],
                    } for c in computed]).to_csv(csv_buf, index=False)
                    st.download_button(
                        "📥 Download reprice CSV",
                        data=csv_buf.getvalue().encode(),
                        file_name=f"reprice_{date.today().isoformat()}.csv",
                        mime="text/csv", key="rp_csv",
                    )
                    st.caption(
                        "Use this CSV with eBay's bulk-edit / File Exchange, or work down the list by hand. "
                        "There's no eBay API, so prices can't be pushed automatically."
                    )
    else:
        st.info("👆 Upload your inventory file above to get started, or download the template first.")

    # ── 📐 Submission Planner — Grade vs Flip a whole batch ────────────────────
    st.markdown("---")
    st.markdown("### 📐 Submission Planner — Grade vs Flip a whole batch")
    st.caption(
        "Model a full submission the way new sellers don't: how much cash gets locked up, for how long, "
        "and whether grading actually beats flipping the lot raw. Defaults to a 20-card order — change to your numbers."
    )
    with st.expander("📐 Open the batch planner", expanded=False):
        bp1, bp2, bp3 = st.columns(3)
        n_cards = bp1.number_input("Number of cards", min_value=1, max_value=500, value=20, step=1, key="bp_n")
        avg_buy = bp2.number_input("Avg buy cost / card ($)", min_value=0.0, value=40.0, step=5.0, key="bp_buy")
        bp_tier = bp3.selectbox("Grading tier", list(PSA_FEES.keys()), index=0, key="bp_tier")

        bc1, bc2, bc3, bc4 = st.columns(4)
        avg_raw = bc1.number_input("Avg RAW comp ($)", min_value=0.0, value=55.0, step=5.0, key="bp_raw",
                                   help="What each card sells for raw right now")
        avg_10 = bc2.number_input("Avg PSA 10 comp ($)", min_value=0.0, value=180.0, step=10.0, key="bp_10")
        avg_9 = bc3.number_input("Avg PSA 9 comp ($)", min_value=0.0, value=70.0, step=5.0, key="bp_9")
        avg_gem = bc4.number_input("Avg gem rate (%)", min_value=0.0, max_value=100.0, value=50.0, step=5.0, key="bp_gem")

        fee_b = PSA_FEES[bp_tier]
        cal_b = int(PSA_DAYS.get(bp_tier, 60) * 1.4)
        _bp_ship = float(ship_cost) if isinstance(ship_cost, (int, float)) else 14.0
        _bp_opp  = float(opp_rate)  if isinstance(opp_rate,  (int, float)) else 6.0
        gvf_b = grade_vs_flip(avg_buy, avg_raw, avg_10, avg_9, avg_gem, bp_tier, _bp_ship, _bp_opp)
        capital_total = gvf_b["capital"] * n_cards
        hold_total = gvf_b["hold"] * n_cards
        raw_total = (gvf_b["raw_net"] or 0) * n_cards
        exp_total = (gvf_b["net_exp"] or 0) * n_cards
        best_total = (gvf_b["net10"] or 0) * n_cards
        down_total = (gvf_b["net9"] or 0) * n_cards
        premium_total = exp_total - raw_total

        st.markdown("**If you GRADE the batch**")
        gm1, gm2, gm3 = st.columns(3)
        gm1.metric("💰 Cash locked up", f"${capital_total:,.0f}",
                   help=f"{n_cards} × (buy ${avg_buy:,.0f} + fee ${fee_b:.0f}"
                        + (f" + ship ${ship_cost:.0f}" if ship_cost else "") + ")")
        gm2.metric("⏳ Locked for", f"~{cal_b} days", help=f"~{cal_b/30:.1f} months at {bp_tier.split('(')[0].strip()}")
        gm3.metric("💸 Holding cost", f"${hold_total:,.0f}", help=f"At {opp_rate:.0f}%/yr opportunity cost")

        gn1, gn2, gn3 = st.columns(3)
        gn1.metric(f"🎯 Expected net (gem {avg_gem:.0f}%)", f"${exp_total:,.0f}",
                   help="Gem-rate-weighted across the batch, after every cost incl. holding")
        gn2.metric("💎 Best case (all 10s)", f"${best_total:,.0f}")
        gn3.metric("🥈 Downside (all 9s)", f"${down_total:,.0f}")

        st.markdown("**If you FLIP the batch RAW now**")
        fr1, fr2 = st.columns(2)
        fr1.metric("💵 Net now", f"${raw_total:,.0f}", help="Sell all raw today — cash in ~a week, nothing locked up")
        fr2.metric("Grade premium (expected)", f"{'+' if premium_total >= 0 else '−'}${abs(premium_total):,.0f}",
                   help="Expected graded net minus raw-flip net across the batch")

        if exp_total > raw_total and exp_total > 0:
            st.success(
                f"**✅ Grading wins (expected)** — across {n_cards} cards you'd net ~**${exp_total:,.0f}** vs "
                f"**${raw_total:,.0f}** flipping raw: **+${premium_total:,.0f}** for tying up **${capital_total:,.0f}** "
                f"for ~{cal_b/30:.1f} months. If none gem, downside is **${down_total:,.0f}**."
            )
        else:
            st.warning(
                f"**💵 Flipping raw may win** — grading's expected **${exp_total:,.0f}** doesn't beat the "
                f"**${raw_total:,.0f}** you'd net flipping raw now, and grading locks **${capital_total:,.0f}** "
                f"for ~{cal_b/30:.1f} months. Grade only the cards with the highest gem rates."
            )
        st.caption(
            "Rule of thumb: the lower the gem rate, the more the PSA 9 downside drags your expected return — "
            "and the longer your cash is locked, the more the holding cost eats the upside."
        )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — Submission Tracker
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 5:
    st.markdown("## 📬 Submission Tracker")
    if is_beta:
        st.warning("🔒 Submission Tracker is available with full membership. Your beta preview includes Card Research and Inventory Check.")
    elif not SUPABASE_URL:
        st.warning("Supabase not configured — tracker unavailable in this environment")
    else:
        rows = sb_get()
        if not rows:
            st.info("No submissions yet. Add cards from the Card Research or Inventory Check tabs.")
        else:
            df = pd.DataFrame(rows)

            # Summary
            total  = len(df)
            queued = len(df[df["status"] == "Queued"])
            sent   = len(df[df["status"] == "Submitted"])
            rcvd   = len(df[df["status"] == "Received"])
            goes   = len(df[df["go_no_go"].str.startswith("✅", na=False)])
            s1, s2, s3, s4, s5 = st.columns(5)
            s1.metric("Total", total)
            s2.metric("Queued", queued)
            s3.metric("Submitted", sent)
            s4.metric("Received", rcvd)
            s5.metric("GO Decisions", goes)

            st.markdown("---")
            display_cols = [c for c in [
                "id", "date_added", "card_description", "raw_buy_price", "psa_tier",
                "gem_rate", "go_no_go", "est_net", "est_roi",
                "status", "date_submitted", "grade_returned",
                "actual_sell_price", "actual_net", "actual_roi", "notes"
            ] if c in df.columns]

            edited = st.data_editor(
                df[display_cols],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "id": st.column_config.NumberColumn("ID", disabled=True),
                    "status": st.column_config.SelectboxColumn("Status", options=["Queued", "Submitted", "Received", "Sold"]),
                    "date_added": st.column_config.DateColumn("Added"),
                    "date_submitted": st.column_config.DateColumn("Submitted"),
                    "raw_buy_price": st.column_config.NumberColumn("Raw $", format="$%.2f"),
                    "psa10_avg_price": st.column_config.NumberColumn("Gem 10 Avg $", format="$%.2f"),
                    "target_price": st.column_config.NumberColumn("Target $", format="$%.2f"),
                    "est_net": st.column_config.NumberColumn("Est Net $", format="$%.2f"),
                    "actual_sell_price": st.column_config.NumberColumn("Actual Sell $", format="$%.2f"),
                    "actual_net": st.column_config.NumberColumn("Actual Net $", format="$%.2f"),
                    "go_no_go": st.column_config.TextColumn("Decision", disabled=True),
                },
                num_rows="dynamic",
            )

            if st.button("💾 Save changes", type="primary"):
                for _, row in edited.iterrows():
                    row_id = int(row.get("id", 0))
                    if row_id:
                        updates = {k: (None if pd.isna(v) else v) for k, v in row.items() if k != "id"}
                        sb_update(row_id, updates)
                st.success("Saved ✓")
                st.rerun()

            # ROI summary for received cards
            rcvd_df = df[df["status"].isin(["Received", "Sold"])].copy()
            if len(rcvd_df) > 0:
                st.markdown("---")
                st.markdown("### 📊 Actual returns")
                try:
                    invested = pd.to_numeric(rcvd_df["raw_buy_price"], errors="coerce").sum()
                    net_total = pd.to_numeric(rcvd_df["actual_net"], errors="coerce").sum()
                    avg_roi = (net_total / invested * 100) if invested > 0 else 0
                    r1, r2, r3 = st.columns(3)
                    r1.metric("Total Invested", f"${invested:,.0f}")
                    r2.metric("Total Net", f"${net_total:,.0f}")
                    r3.metric("Avg ROI", f"{avg_roi:.0f}%")
                except Exception:
                    pass

        # Manual add
        st.markdown("---")
        st.markdown("### ➕ Add manually")
        with st.expander("Add a card"):
            a1, a2 = st.columns(2)
            m_desc = a1.text_input("Card description", key="m_desc")
            m_raw  = a2.number_input("Raw buy price ($)", min_value=0.0, step=5.0, key="m_raw")
            a3, a4 = st.columns(2)
            m_tier = a3.selectbox("Grading tier", list(PSA_FEES.keys()), key="m_tier")
            m_gem  = a4.number_input("Gem rate (%)", min_value=0.0, max_value=100.0, key="m_gem")
            a5, a6 = st.columns(2)
            m_g10  = a5.number_input("Gem 10 avg price ($)", min_value=0.0, step=5.0, key="m_g10")
            m_notes = a6.text_input("Notes", key="m_notes")

            if st.button("Add", key="m_add"):
                if m_desc and m_raw > 0:
                    fee_m = PSA_FEES[m_tier]
                    tgt_m = target_price(m_raw, m_tier, roi_target)
                    net_m, roi_m = calc_net_roi(m_raw, m_tier, m_g10) if m_g10 > 0 else (None, None)
                    v_m, _, _ = verdict(m_raw, m_tier, m_gem or None, m_g10, min_gem, roi_target) if m_g10 > 0 else ("Pending", "", "")
                    sb_insert({
                        "date_added": date.today().isoformat(),
                        "card_description": m_desc,
                        "raw_buy_price": m_raw,
                        "psa_tier": m_tier,
                        "psa_fee": fee_m,
                        "psa10_avg_price": m_g10 or None,
                        "target_price": round(tgt_m, 2),
                        "gem_rate": m_gem or None,
                        "go_no_go": v_m,
                        "est_net": net_m,
                        "est_roi": f"{roi_m:.0f}%" if roi_m is not None else None,
                        "status": "Queued",
                        "notes": m_notes,
                    })
                    st.success("Added ✓")
                    st.rerun()
                else:
                    st.warning("Need a description and buy price")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — Downloads
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 6:
    st.markdown("## 📥 Downloads")
    if is_beta:
        st.warning("🔒 Downloads are available with full membership. Your beta preview includes Card Research and Inventory Check.")
    else:
        st.markdown("Tools to run your grading operation — built to work alongside the app.")
        st.markdown("---")

        # ── Operations Kit ──
        kit_path = Path(__file__).parent / "DFS_Card_Grader_Kit.xlsx"
        current_user = st.session_state.get("access_name", "")
        kit_unlocked = current_user == "Robert Bass"

        if kit_path.exists():
            with open(kit_path, "rb") as f:
                kit_bytes = f.read()

            c1, c2 = st.columns([2, 1])
            with c1:
                st.markdown(f"### 📊 {APP_NAME} — Operations Kit")
                st.markdown("""
A comprehensive Excel workbook built to run alongside this app. Includes:

- **PSA Fee Schedule** — all 7 tiers with eBay fee calculations
- **Grading ROI Calculator** — input a card, get instant GO/NO-GO with target price and net profit
- **Grading Tracker** — full 21-column submission log matching your tracker in this app
- **Grading Candidates** — shortlist cards from your inventory for PSA consideration
- **Inventory & Aging** — track every card with cost basis, days listed, and aging flags
- **Sales Log** — log every sale with fees, shipping, and net margin
- **Target Card List** — your buy list with max buy price and sell targets
- **Channel Fees Calculator** — live fee routing across all your sales channels
- **Consignment Tracker** — log DcSports, PWCC, Probstein submissions
- **Path to $40k** — monthly revenue projection model
- **+ 14 more sheets** covering capital velocity, lot evaluation, HeyStack priority scoring, and more
""")
            with c2:
                st.markdown("<div style='padding-top:48px'></div>", unsafe_allow_html=True)
                if kit_unlocked:
                    st.download_button(
                        label="⬇️ Download Operations Kit",
                        data=kit_bytes,
                        file_name=f"{APP_NAME.replace(' ', '_')}_Operations_Kit.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )
                    st.caption("Excel (.xlsx) · Works with Excel, Google Sheets, and Numbers")
                else:
                    st.button("🔒 Operations Kit", use_container_width=True, disabled=True)
                    st.caption("Coming soon — available as a premium add-on")

        st.markdown("---")

        # ── Inventory Template ──
        i1, i2 = st.columns([2, 1])
        with i1:
            st.markdown("### 📋 Inventory Template")
            st.markdown("""
A simple CSV template to get started tracking your inventory.
Upload it in the **Inventory Check** tab to search GemRate and get GO/NO-GO decisions on your cards.

- 11 columns: Card Description, Player, Year, Set, Parallel, Card Number, Category, Cost Basis, Listed Price, Source, Notes
- 3 example rows included to show the format
""")
        with i2:
            st.markdown("<div style='padding-top:48px'></div>", unsafe_allow_html=True)
            st.download_button(
                label="⬇️ Download Template",
                data=make_template_csv(),
                file_name="card_inventory_template.csv",
                mime="text/csv",
                use_container_width=True,
            )
            st.caption("CSV · Opens in Excel, Google Sheets, or Numbers")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — Shipment Intake
# ══════════════════════════════════════════════════════════════════════════════
INTAKE_PLATFORMS = ["eBay", "COMC", "Whatnot", "Card Show", "Private Sale", "MySlabs", "Facebook", "Other"]
INTAKE_STATUSES  = ["Received", "Sent to PSA", "In Collection", "Listed", "Sold"]

if _active_tab == 7:
    st.markdown("## 🚚 Shipment Intake")
    st.markdown("Log cards as they arrive. Build a queue to evaluate or send to PSA.")

    if not SUPABASE_URL:
        st.warning("Supabase not configured — intake log unavailable in this environment.")
    else:
        # ── Add Card form ─────────────────────────────────────────────────────
        st.markdown("### ➕ Add a Card")
        with st.form("intake_form", clear_on_submit=True):
            fi1, fi2, fi3, fi4 = st.columns([3, 1, 1, 1])
            with fi1:
                i_player = st.text_input("Player *", placeholder="e.g. Steph Curry")
            with fi2:
                i_year = st.text_input("Year", placeholder="2021")
            with fi3:
                i_card_num = st.text_input("Card #", placeholder="44")
            with fi4:
                i_date = st.date_input("Date Received", value=date.today())

            fi5, fi6 = st.columns([3, 2])
            with fi5:
                i_set = st.text_input("Set", placeholder="e.g. Topps Chrome")
            with fi6:
                i_parallel = st.text_input("Parallel / Variation", placeholder="e.g. Silver Prizm")

            fi7, fi8, fi9 = st.columns([2, 2, 2])
            with fi7:
                i_cost = st.number_input("Cost ($) *", min_value=0.0, step=1.0, format="%.2f",
                                         help="What you paid — include shipping if buying from a single seller")
            with fi8:
                i_platform = st.selectbox("Platform", INTAKE_PLATFORMS)
            with fi9:
                i_source = st.text_input("Source / Seller", placeholder="e.g. cardking88")

            i_notes = st.text_input("Notes", placeholder="e.g. Pack fresh, lot of 3, small corner wear")

            intake_submitted = st.form_submit_button("Log Card", type="primary", use_container_width=True)
            if intake_submitted:
                if not i_player:
                    st.session_state["intake_msg"] = ("error", "Player name is required.")
                else:
                    parts = [p for p in [i_year, i_set, i_player, i_parallel] if p]
                    auto_desc = " ".join(parts)
                    if i_card_num:
                        auto_desc += f" #{i_card_num}"
                    _res, _err = sb_intake_insert({
                        "date_received": i_date.isoformat(),
                        "player":        i_player.strip(),
                        "year":          i_year.strip(),
                        "set_name":      i_set.strip(),
                        "parallel":      i_parallel.strip(),
                        "card_number":   i_card_num.strip(),
                        "card_description": auto_desc.strip(),
                        "cost":          float(i_cost) if i_cost > 0 else None,
                        "source":        i_source.strip(),
                        "platform":      i_platform,
                        "notes":         i_notes.strip(),
                        "status":        "Received",
                    })
                    if _err:
                        st.session_state["intake_msg"] = ("error", f"Save failed: {_err}")
                    else:
                        st.session_state["intake_msg"] = ("success", f"✓ Logged: {auto_desc}")

        # ── Bulk Upload from Spreadsheet ──────────────────────────────────────
        with st.expander("📂 Bulk Upload from Spreadsheet", expanded=False):
            st.markdown(
                "Upload your inventory CSV or Operations Workbook (.xlsx) to log multiple cards at once. "
                "Each row becomes a card in the Received Cards log below."
            )
            bu1, bu2 = st.columns([3, 1])
            with bu2:
                st.download_button(
                    "⬇️ Download Template",
                    data=make_template_csv(),
                    file_name="card_inventory_template.csv",
                    mime="text/csv",
                    use_container_width=True,
                    help="Fill this out and upload it here",
                )
            with bu1:
                bulk_file = st.file_uploader(
                    "Upload CSV template or Operations Workbook (.xlsx)",
                    type=["csv", "xlsx"],
                    key="intake_bulk_upload",
                )

            if bulk_file:
                bulk_df, bulk_source = load_inventory(bulk_file)
                if bulk_df is None:
                    st.error(f"Could not read file: {bulk_source}")
                else:
                    st.success(f"Found **{len(bulk_df)} cards** — preview below. Click Import to save them all.")
                    preview_cols = [c for c in ["Card Description", "Player", "Year", "Set",
                                                "Parallel", "Card Number", "Cost Basis ($)", "Source", "Notes"]
                                    if c in bulk_df.columns]
                    st.dataframe(bulk_df[preview_cols].head(20), use_container_width=True, hide_index=True)

                    if st.button(f"⬆️ Import {len(bulk_df)} cards into Intake Log", type="primary", key="bulk_import_btn"):
                        saved, failed = 0, 0
                        for _, brow in bulk_df.iterrows():
                            def _str(col): return str(brow.get(col, "") or "").strip()
                            def _flt(col):
                                try: return float(brow.get(col, 0) or 0) or None
                                except: return None
                            parts = [p for p in [_str("Year"), _str("Set"), _str("Player"), _str("Parallel")] if p]
                            desc  = _str("Card Description") or " ".join(parts)
                            card_num = _str("Card Number")
                            if card_num and card_num not in desc:
                                desc += f" #{card_num}"
                            _, err = sb_intake_insert({
                                "date_received":   date.today().isoformat(),
                                "player":          _str("Player"),
                                "year":            _str("Year"),
                                "set_name":        _str("Set"),
                                "parallel":        _str("Parallel"),
                                "card_number":     card_num,
                                "card_description": desc,
                                "cost":            _flt("Cost Basis ($)"),
                                "source":          _str("Source"),
                                "platform":        "Other",
                                "notes":           _str("Notes"),
                                "status":          "Received",
                            })
                            if err:
                                failed += 1
                            else:
                                saved += 1
                        if failed:
                            st.error(f"Imported {saved} cards — {failed} failed. Error: {err}")
                        else:
                            st.success(f"✓ Imported {saved} cards into your Intake Log!")
                        st.rerun()

        # ── Feedback from last submit ─────────────────────────────────────────
        if "intake_msg" in st.session_state:
            _msg_type, _msg_text = st.session_state.pop("intake_msg")
            if _msg_type == "success":
                st.success(_msg_text)
            else:
                st.error(_msg_text)

        # ── Received cards log ────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### 📋 Received Cards")

        intake_rows, intake_err = sb_intake_get()
        if intake_err:
            st.error(f"Could not load cards: {intake_err}")
            st.stop()
        if not intake_rows:
            st.info("No cards logged yet — use the form above to start tracking incoming shipments.")
        else:
            df_i = pd.DataFrame(intake_rows)

            # Summary strip
            total_i    = len(df_i)
            cost_i     = pd.to_numeric(df_i.get("cost", pd.Series(dtype=float)), errors="coerce").sum()
            received_i = (df_i["status"] == "Received").sum() if "status" in df_i.columns else 0
            sent_i     = (df_i["status"] == "Sent to PSA").sum() if "status" in df_i.columns else 0

            mi1, mi2, mi3, mi4 = st.columns(4)
            mi1.metric("Total Cards",    total_i)
            mi2.metric("Total Cost",     f"${cost_i:,.2f}")
            mi3.metric("Awaiting Eval",  received_i)
            mi4.metric("Sent to PSA",    sent_i)

            st.markdown("")

            # Coerce types so data_editor doesn't crash on Supabase string returns
            if "date_received" in df_i.columns:
                df_i["date_received"] = pd.to_datetime(df_i["date_received"], errors="coerce").dt.date
            if "cost" in df_i.columns:
                df_i["cost"] = pd.to_numeric(df_i["cost"], errors="coerce")

            _show_cols = [c for c in [
                "id", "date_received", "card_description", "player", "year",
                "set_name", "card_number", "parallel", "cost",
                "platform", "source", "status", "notes",
            ] if c in df_i.columns]

            edited_i = st.data_editor(
                df_i[_show_cols],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "id":               st.column_config.NumberColumn("ID", disabled=True),
                    "date_received":    st.column_config.DateColumn("Date Received"),
                    "card_description": st.column_config.TextColumn("Card", disabled=True),
                    "player":           st.column_config.TextColumn("Player"),
                    "year":             st.column_config.TextColumn("Year"),
                    "set_name":         st.column_config.TextColumn("Set"),
                    "card_number":      st.column_config.TextColumn("Card #"),
                    "parallel":         st.column_config.TextColumn("Parallel"),
                    "cost":             st.column_config.NumberColumn("Cost $", format="$%.2f"),
                    "platform":         st.column_config.SelectboxColumn("Platform", options=INTAKE_PLATFORMS),
                    "source":           st.column_config.TextColumn("Seller / Source"),
                    "status":           st.column_config.SelectboxColumn("Status", options=INTAKE_STATUSES),
                    "notes":            st.column_config.TextColumn("Notes"),
                },
            )

            ci1, ci2 = st.columns([1, 1])
            with ci1:
                if st.button("💾 Save changes", type="primary", key="intake_save", use_container_width=True):
                    for _, row in edited_i.iterrows():
                        row_id = int(row.get("id", 0))
                        if row_id:
                            updates = {k: (None if (isinstance(v, float) and pd.isna(v)) else v)
                                       for k, v in row.items() if k != "id"}
                            sb_intake_update(row_id, updates)
                    st.success("Saved ✓")
                    st.rerun()
            with ci2:
                csv_bytes = df_i[_show_cols].to_csv(index=False).encode()
                st.download_button(
                    "⬇️ Export to CSV",
                    data=csv_bytes,
                    file_name=f"shipment_intake_{date.today().isoformat()}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — Operations & Inventory (paid, metered live pricing)
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 4:
    st.markdown("## 🧰 Operations & Inventory")

    if is_beta:
        st.warning("🔒 Operations is part of full membership. Your preview includes Card Research and Inventory Check.")
        st.caption(
            "Operations adds live real-time pricing across your whole inventory, an editable cost basis, "
            "and reprice + margin tools — metered with a daily look-up budget to keep pricing fast and fair."
        )
    elif not CARDHEDGER_KEY:
        st.info("📊 Connect the CardHedger API to enable live pricing.")
    else:
        from datetime import timezone as _optz

        op_tab_inv, op_tab_queue, op_tab_sunday, op_tab_promote, op_tab_tcp = st.tabs(["📦 Inventory & Aging", "🔄 Reprice Queue", "📅 Sunday Reprice", "📣 Promote Listings", "📊 TCP Reprice"])

        # ── helpers shared across both sub-tabs ───────────────────────────────
        def _days_since(dt_str):
            if not dt_str:
                return None
            try:
                d = datetime.fromisoformat(str(dt_str).replace("Z", "+00:00"))
                if d.tzinfo is None:
                    d = d.replace(tzinfo=_optz.utc)
                return (datetime.now(_optz.utc) - d).days
            except Exception:
                return None

        def _last_priced_label(ts):
            if not ts:
                return "Never"
            days = _days_since(ts)
            return f"{days}d ago" if days is not None else "?"

        # ── INVENTORY & AGING ─────────────────────────────────────────────────
        with op_tab_inv:
            st.markdown("### Sync from eBay")
            st.caption(
                "eBay Seller Hub → Reports → Active listings → Download CSV. "
                "Drop it here — all 4,000+ listings sync into Supabase in seconds."
            )
            op_sync_file = st.file_uploader("eBay active listings CSV", type=["csv"], key="op_sync")
            if op_sync_file:
                sync_df = pd.read_csv(op_sync_file, encoding="utf-8-sig")
                listing_rows = parse_ebay_csv_to_listings(sync_df)
                st.caption(f"Parsed {len(listing_rows):,} listings. Click below to sync.")
                if st.button(f"⬆️ Sync {len(listing_rows):,} listings to Supabase", type="primary", key="op_sync_btn"):
                    with st.spinner("Syncing… (batches of 500)"):
                        n_synced = upsert_listings(listing_rows)
                    if n_synced:
                        st.success(f"✅ Synced {n_synced:,} listings.")
                        st.session_state.pop("op_listings_cache", None)
                        st.rerun()
                    else:
                        st.error("Sync failed. Run the Supabase SQL to create the `listings` table first, then retry.")

            st.divider()

            # Load from Supabase
            if "op_listings_cache" not in st.session_state:
                with st.spinner("Loading from Supabase…"):
                    st.session_state["op_listings_cache"] = load_listings(min_price=20)

            all_ls = st.session_state.get("op_listings_cache", [])

            if not all_ls:
                st.info("No $20+ listings in Supabase yet. Sync your eBay CSV above to get started.")
            else:
                today_due = sum(1 for l in all_ls if needs_pricing_today(l.get("last_priced_at"), l.get("price_freq","weekly")))
                never_priced = sum(1 for l in all_ls if not l.get("last_priced_at"))
                total_value = sum(l.get("current_price") or 0 for l in all_ls)
                sc1, sc2, sc3, sc4 = st.columns(4)
                sc1.metric("$20+ Listings", f"{len(all_ls):,}")
                sc2.metric("Need Pricing Today", f"{today_due:,}")
                sc3.metric("Never Priced", f"{never_priced:,}")
                sc4.metric("Listed Value", f"${total_value:,.0f}")

                fc1, fc2, fc3 = st.columns(3)
                sports_avail = sorted(set(l.get("sport","Unknown") for l in all_ls))
                f_sport = fc1.selectbox("Sport", ["All"] + sports_avail, key="inv_f_sport")
                f_show  = fc2.selectbox("Show", ["All","Needs Pricing","Never Priced"], key="inv_f_show")
                if fc3.button("🔄 Refresh inventory", key="inv_refresh"):
                    st.session_state.pop("op_listings_cache", None)
                    st.rerun()

                filtered_ls = all_ls
                if f_sport != "All":
                    filtered_ls = [l for l in filtered_ls if l.get("sport") == f_sport]
                if f_show == "Needs Pricing":
                    filtered_ls = [l for l in filtered_ls if needs_pricing_today(l.get("last_priced_at"), l.get("price_freq","weekly"))]
                elif f_show == "Never Priced":
                    filtered_ls = [l for l in filtered_ls if not l.get("last_priced_at")]

                inv_rows = []
                for l in filtered_ls:
                    inv_rows.append({
                        "Title":        (l.get("title") or "")[:65],
                        "Sport":        l.get("sport","?"),
                        "RC":           "✓" if l.get("is_rookie") else "",
                        "Price ($)":    l.get("current_price"),
                        "Cost ($)":     l.get("cost_basis"),
                        "Comp ($)":     l.get("comp_avg"),
                        "Suggested ($)":l.get("suggested_price"),
                        "Trend":        trend_label(l.get("trend_dir"), l.get("trend_pct") or 0),
                        "Days Listed":  _days_since(l.get("start_date")),
                        "Last Priced":  _last_priced_label(l.get("last_priced_at")),
                        "Freq":         l.get("price_freq","weekly"),
                        "Item #":       l.get("item_number",""),
                    })

                st.dataframe(
                    pd.DataFrame(inv_rows), use_container_width=True, hide_index=True,
                    column_config={
                        "Price ($)":     st.column_config.NumberColumn(format="$%.2f"),
                        "Cost ($)":      st.column_config.NumberColumn(format="$%.2f"),
                        "Comp ($)":      st.column_config.NumberColumn(format="$%.2f"),
                        "Suggested ($)": st.column_config.NumberColumn(format="$%.2f"),
                    },
                )
                st.caption(f"Showing {len(filtered_ls):,} of {len(all_ls):,} listings.")

        # ── REPRICE QUEUE ─────────────────────────────────────────────────────
        with op_tab_queue:
            used = pricing_used_today()
            if pricing_unlimited():
                st.caption("⚡ Live pricing: **unlimited** on your account.")
            else:
                st.progress(min(1.0, used / DAILY_PRICING_CAP) if DAILY_PRICING_CAP else 1.0,
                            text=f"Live look-ups today: {used} / {DAILY_PRICING_CAP}")
                if pricing_remaining() == 0:
                    st.warning(f"You've used today's {DAILY_PRICING_CAP} live look-ups — resets tomorrow.")

            if "op_listings_cache" not in st.session_state:
                st.session_state["op_listings_cache"] = load_listings(min_price=20)

            all_ls2 = st.session_state.get("op_listings_cache", [])
            if not all_ls2:
                st.info("No listings in Supabase yet. Go to Inventory & Aging → sync your eBay CSV first.")
            else:
                force_all = st.toggle("Price all $20+ listings (ignore schedule)", key="q_force_all")

                due = all_ls2 if force_all else [l for l in all_ls2 if needs_pricing_today(l.get("last_priced_at"), l.get("price_freq","weekly"))]

                def _queue_priority(l):
                    sport, rc = l.get("sport","Unknown"), l.get("is_rookie", False)
                    if sport == "Baseball" and rc: return 0
                    if sport == "Baseball": return 1
                    if sport == "Soccer": return 2
                    return 3

                due.sort(key=_queue_priority)

                st.markdown(f"### Today's queue — {len(due)} cards")
                st.caption(
                    "Priority order: ⚾ Baseball rookies → ⚾ Baseball → ⚽ Soccer → 🏈 Football / 🏀 Basketball / Other. "
                    f"{len(all_ls2) - len(due)} cards already up to date."
                )

                if not due:
                    st.success("✅ All $20+ listings are current — toggle 'Price all' above to force a full run.")
                else:
                    qc1, qc2, qc3 = st.columns(3)
                    q_strat = qc1.selectbox("Strategy", ["Trend-following","Match market","Undercut to sell faster","List high for offers"], key="q_strat")
                    if q_strat == "Trend-following":
                        q_adj = qc2.slider("Trend sensitivity (%)", 0, 100, 50, 5, key="q_sens")
                    elif q_strat == "Undercut to sell faster":
                        q_adj = qc2.slider("Undercut below comp (%)", 0, 30, 7, 1, key="q_under")
                    elif q_strat == "List high for offers":
                        q_adj = qc2.slider("Premium above comp (%)", 0, 30, 10, 1, key="q_prem")
                    else:
                        q_adj = qc2.slider("Adjust vs comp (%)", -20, 20, 0, 1, key="q_match")

                    budget = pricing_remaining()
                    max_run = len(due) if pricing_unlimited() else max(1, min(len(due), budget))
                    run_n = qc3.number_input("Cards this run", 1, max(1, len(due)), min(min(100, len(due)), max_run), key="q_runn")

                    blocked = (budget <= 0 and not pricing_unlimited())
                    if st.button(f"🔄 Run queue ({int(run_n)} look-ups)", type="primary", key="q_run", disabled=blocked):
                        batch = due[:int(run_n)]
                        prog = st.progress(0.0, text="Pricing…")
                        q_res = []
                        for i, l in enumerate(batch):
                            title = l.get("title","")
                            grade = detect_grade(title)
                            mkt = fetch_market(title, grade)
                            sugg = suggest_reprice(mkt["comp_avg"], mkt["trend_pct"], q_strat, q_adj)
                            save_listing_pricing(l["item_number"], mkt["comp_avg"], mkt["trend_dir"], mkt["trend_pct"], sugg)
                            q_res.append({**l, "comp_avg": mkt["comp_avg"], "trend_dir": mkt["trend_dir"],
                                          "trend_pct": mkt["trend_pct"], "suggested_price": sugg})
                            prog.progress((i+1)/len(batch), text=f"Pricing… {i+1}/{len(batch)}")
                        pricing_bump(len(batch))
                        prog.empty()
                        st.session_state.pop("op_listings_cache", None)
                        st.session_state["q_results"] = q_res
                        st.success(f"✅ Priced {len(q_res)} cards. Results saved to Supabase.")
                        st.rerun()

                    display_list = st.session_state.get("q_results") or due[:100]
                    if display_list:
                        rtbl = []
                        for l in display_list:
                            cur  = l.get("current_price") or 0
                            comp = l.get("comp_avg")
                            sugg = l.get("suggested_price")
                            diff = ((sugg - cur) / cur * 100) if (sugg and cur) else None
                            rtbl.append({
                                "Title":         (l.get("title","") or "")[:65],
                                "Sport":         l.get("sport","?"),
                                "RC":            "✓" if l.get("is_rookie") else "",
                                "Current ($)":   cur,
                                "Comp ($)":      comp,
                                "Trend":         trend_label(l.get("trend_dir"), l.get("trend_pct") or 0),
                                "Suggested ($)": sugg,
                                "∆ vs Listed":   f"{diff:+.0f}%" if diff is not None else "—",
                                "Item #":        l.get("item_number",""),
                            })
                        st.dataframe(
                            pd.DataFrame(rtbl), use_container_width=True, hide_index=True,
                            column_config={
                                "Current ($)":   st.column_config.NumberColumn(format="$%.2f"),
                                "Comp ($)":      st.column_config.NumberColumn(format="$%.2f"),
                                "Suggested ($)": st.column_config.NumberColumn(format="$%.2f"),
                            },
                        )

                        ebay_rows = [l for l in display_list if l.get("suggested_price")]
                        if ebay_rows:
                            # eBay upload file — price-only, 2 columns
                            # Uploading any extra columns risks overwriting Heystack formatting
                            ebuf = io.StringIO()
                            pd.DataFrame([{
                                "Item number": l.get("item_number",""),
                                "Start price": round(l["suggested_price"], 2),
                            } for l in ebay_rows]).to_csv(ebuf, index=False)

                            # Reference sheet with full context (keep for your own records)
                            ref_buf = io.StringIO()
                            pd.DataFrame([{
                                "Item number":   l.get("item_number",""),
                                "Title":         l.get("title",""),
                                "Current Price": l.get("current_price",""),
                                "New Price":     round(l["suggested_price"], 2),
                                "Comp Avg":      round(l["comp_avg"], 2) if l.get("comp_avg") else "",
                                "Trend":         trend_label(l.get("trend_dir"), l.get("trend_pct") or 0),
                                "Sport":         l.get("sport",""),
                            } for l in ebay_rows]).to_csv(ref_buf, index=False)

                            ec1, ec2 = st.columns(2)
                            ec1.download_button(
                                "📤 Upload to eBay (price only)",
                                data=ebuf.getvalue().encode(),
                                file_name=f"ebay_reprice_{date.today().isoformat()}.csv",
                                mime="text/csv", key="q_export_ebay",
                                help="Safe to upload directly — only Item number + Start price. Nothing else touched.",
                            )
                            ec2.download_button(
                                "📋 Full reference sheet",
                                data=ref_buf.getvalue().encode(),
                                file_name=f"reprice_reference_{date.today().isoformat()}.csv",
                                mime="text/csv", key="q_export_ref",
                                help="Your records — current price, comp, trend, sport. Do NOT upload this to eBay.",
                            )

        # ── SUNDAY REPRICE ────────────────────────────────────────────────────
        with op_tab_sunday:
            st.markdown("### 📅 Sunday Reprice")
            st.caption(
                "Upload your eBay active listings CSV → filter to cards listed 7–30 days ago → "
                "run CardHedger FMV on every one → download the eBay revise CSV. "
                "Results are cached for the session so re-running the page burns zero extra API calls."
            )

            used_sun = pricing_used_today()
            if pricing_unlimited():
                st.caption("⚡ Live pricing: **unlimited** on your account.")
            else:
                st.progress(min(1.0, used_sun / DAILY_PRICING_CAP),
                            text=f"Live look-ups today: {used_sun} / {DAILY_PRICING_CAP}")
                if pricing_remaining() == 0:
                    st.warning(f"You've used today's {DAILY_PRICING_CAP} live look-ups.")

            uf1, uf2 = st.columns(2)
            sun_file = uf1.file_uploader(
                "eBay active listings CSV", type=["csv"], key="sun_csv",
                help="eBay Seller Hub → Reports → Active listings → Download CSV",
            )
            hay_file = uf2.file_uploader(
                "Haystack CSV (optional — improves match rate)",
                type=["csv"], key="sun_hay",
                help="Upload your Haystack listing file. Parallel/Variety, Player, Card Number fields give CardHedger a cleaner query than the eBay title alone.",
            )

            sc1, sc2, sc3 = st.columns(3)
            sun_min_days = sc1.number_input("Min days listed", 1, 60, 7, 1, key="sun_min_d")
            sun_max_days = sc2.number_input("Max days listed", 7, 180, 30, 1, key="sun_max_d")
            sun_floor    = sc3.number_input("Min price ($)", 0, 500, 10, 5, key="sun_floor",
                                             help="Skip cards listed below this price — not worth the API call")

            # Build SKU → Haystack query lookup if a Haystack file was uploaded
            hay_lookup = {}
            if hay_file:
                try:
                    hdf = pd.read_csv(hay_file, encoding="utf-8-sig", skiprows=1)
                    hcol = {str(c).lower(): c for c in hdf.columns}
                    def _hcol(n):
                        for k in hcol:
                            if n.lower() in k:
                                return hcol[k]
                        return None
                    sku_col      = _hcol("custom label")
                    player_col   = _hcol("player/athlete")
                    parallel_col = _hcol("parallel/variety")
                    season_col   = _hcol("season")
                    mfr_col      = _hcol("manufacturer")
                    num_col      = _hcol("card number")
                    title_col    = _hcol("title")
                    for _, hr in hdf.iterrows():
                        sku = str(hr.get(sku_col, "") or "").strip().upper() if sku_col else ""
                        if not sku:
                            continue
                        parts = []
                        season  = str(hr.get(season_col, "") or "").strip() if season_col else ""
                        mfr     = str(hr.get(mfr_col, "") or "").strip() if mfr_col else ""
                        player  = str(hr.get(player_col, "") or "").strip() if player_col else ""
                        cardnum = str(hr.get(num_col, "") or "").strip() if num_col else ""
                        para    = str(hr.get(parallel_col, "") or "").strip() if parallel_col else ""
                        htitle  = str(hr.get(title_col, "") or "").strip() if title_col else ""
                        if season:  parts.append(season)
                        if mfr:     parts.append(mfr)
                        if player:  parts.append(player)
                        if cardnum: parts.append(f"#{cardnum}")
                        if para and para.lower() not in ("base", "n/a", ""):
                            parts.append(para)
                        hay_lookup[sku] = " ".join(parts) if parts else htitle
                    st.caption(f"✅ Haystack loaded — {len(hay_lookup):,} SKUs mapped for cleaner queries.")
                except Exception as e:
                    st.warning(f"Haystack parse error: {e}")

            if sun_file:
                sun_df = pd.read_csv(sun_file, encoding="utf-8-sig")
                col_map_s = {str(c).lower(): c for c in sun_df.columns}
                def _scol(name):
                    return col_map_s.get(name.lower(), name)

                today_dt = datetime.utcnow()
                candidates = []
                for _, r in sun_df.iterrows():
                    item_num = str(r.get(_scol("item number"), "") or "").strip()
                    if not item_num:
                        continue
                    title = str(r.get(_scol("title"), "") or "").strip()
                    try:
                        price = float(r.get(_scol("current price"), 0) or 0)
                    except Exception:
                        price = 0.0
                    if price < sun_floor:
                        continue
                    raw_date = str(r.get(_scol("start date"), "") or "").strip()
                    for tz_s in [" PDT", " PST", " EDT", " EST"]:
                        raw_date = raw_date.replace(tz_s, "")
                    try:
                        listed_dt = datetime.strptime(raw_date, "%b-%d-%y %H:%M:%S")
                        days_listed = (today_dt - listed_dt).days
                    except Exception:
                        continue
                    if days_listed < sun_min_days or days_listed > sun_max_days:
                        continue
                    sku = str(r.get(_scol("custom label (sku)"), "") or "").strip().upper()
                    ch_query = hay_lookup.get(sku) or clean_title_for_ch(title)
                    candidates.append({
                        "item_number":  item_num,
                        "title":        title,
                        "ch_query":     ch_query,
                        "sku":          sku,
                        "current_price": price,
                        "days_listed":  days_listed,
                        "sport":        detect_sport(title),
                        "hay_matched":  sku in hay_lookup,
                    })

                # Sort: baseball rookies first, then baseball, then soccer, then other
                def _sun_priority(c):
                    s, t = c["sport"], c["title"].upper()
                    rc = is_rookie_card(t)
                    if s == "Baseball" and rc: return 0
                    if s == "Baseball": return 1
                    if s == "Soccer": return 2
                    return 3
                candidates.sort(key=_sun_priority)

                m1, m2, m3 = st.columns(3)
                m1.metric("In range", f"{len(candidates)}")
                m2.metric("Est. API calls", f"~{len(candidates) * 2}")
                m3.metric("Days filter", f"{sun_min_days}–{sun_max_days} days")

                if candidates:
                    # Cache key: date + file length + filter params so Sunday's run
                    # persists through reruns but a new file/filter clears it
                    cache_key = f"sun_results_{date.today().isoformat()}_{len(candidates)}_{sun_min_days}_{sun_max_days}_{sun_floor}_{len(hay_lookup)}"

                    already_run = st.session_state.get("sun_cache_key") == cache_key
                    sun_results = st.session_state.get("sun_results", []) if already_run else []

                    if already_run and sun_results:
                        matched  = sum(1 for r in sun_results if r.get("comp"))
                        no_match = sum(1 for r in sun_results if not r.get("comp"))
                        suspect  = sum(1 for r in sun_results if r.get("suspect"))
                        usable   = sum(1 for r in sun_results if r.get("suggested"))
                        parts = [f"✅ Cached — {usable} ready to upload"]
                        if no_match:
                            parts.append(f"{no_match} no match")
                        if suspect:
                            parts.append(f"{suspect} suspect (excluded)")
                        st.success("  ·  ".join(parts) + "  ·  Re-run below only to refresh.")
                    else:
                        budget = pricing_remaining()
                        can_run = pricing_unlimited() or budget >= len(candidates)
                        if not can_run:
                            st.warning(f"Only {budget} look-ups left today — {len(candidates)} needed. Run tomorrow or reduce the range.")

                        if st.button(
                            f"🔄 Run CardHedger on {len(candidates)} listings",
                            type="primary", key="sun_run",
                            disabled=(not can_run),
                        ):
                            prog = st.progress(0.0, text="Looking up prices…")
                            results_map = {}
                            completed_count = [0]

                            def _sun_lookup(idx_c):
                                idx, c = idx_c
                                grade = detect_grade(c["title"])
                                mkt   = fetch_market(c["ch_query"], grade)
                                raw_sugg = suggest_reprice(mkt["comp_avg"], mkt["trend_pct"], "Match market", 0)
                                sugg     = sane_price(raw_sugg, c["current_price"])
                                return idx, {
                                    **c,
                                    "grade":       grade,
                                    "comp":        mkt["comp_avg"],
                                    "trend":       trend_label(mkt["trend_dir"], mkt["trend_pct"] or 0),
                                    "matched":     mkt["matched"],
                                    "suggested":   sugg,
                                    "suspect":     (raw_sugg is not None and sugg is None),
                                    "hay_matched": c.get("hay_matched", False),
                                }

                            with ThreadPoolExecutor(max_workers=6) as _sun_ex:
                                _sun_futures = {
                                    _sun_ex.submit(_sun_lookup, (i, c)): i
                                    for i, c in enumerate(candidates)
                                }
                                for _fut in as_completed(_sun_futures):
                                    try:
                                        _idx, _row = _fut.result()
                                        results_map[_idx] = _row
                                    except Exception:
                                        pass
                                    completed_count[0] += 1
                                    prog.progress(
                                        completed_count[0] / len(candidates),
                                        text=f"Pricing… {completed_count[0]}/{len(candidates)}",
                                    )

                            results = [results_map[i] for i in sorted(results_map)]
                            pricing_bump(len(candidates))
                            prog.empty()
                            st.session_state["sun_results"]   = results
                            st.session_state["sun_cache_key"] = cache_key
                            st.rerun()

                    if sun_results:
                        has_price = [r for r in sun_results if r.get("suggested")]
                        no_price  = [r for r in sun_results if not r.get("suggested")]
                        suspect_rows = [r for r in sun_results if r.get("suspect")]

                        if suspect_rows:
                            st.warning(
                                f"⚠️ {len(suspect_rows)} listing(s) had a comp that was <20% or >4× "
                                "the current price — likely a bad card match, excluded from download. "
                                "Review and reprice manually."
                            )

                        hay_count = sum(1 for r in sun_results if r.get("hay_matched"))
                        if hay_count:
                            st.caption(f"🌾 {hay_count} listings used Haystack structured query · {len(sun_results)-hay_count} used eBay title")

                        tbl = []
                        for r in sun_results:
                            flag = " ⚠️" if r.get("suspect") else ""
                            tbl.append({
                                "Title":         r["title"][:55] + flag,
                                "Src":           "🌾" if r.get("hay_matched") else "eBay",
                                "Sport":         r["sport"],
                                "Days":          r["days_listed"],
                                "Current ($)":   r["current_price"],
                                "Comp ($)":      r.get("comp"),
                                "Suggested ($)": r.get("suggested"),
                                "Trend":         r.get("trend","—"),
                                "∆":             (
                                    f"{(r['suggested']-r['current_price'])/r['current_price']*100:+.0f}%"
                                    if r.get("suggested") and r["current_price"] else "—"
                                ),
                                "Item #":        r["item_number"],
                            })
                        st.dataframe(
                            pd.DataFrame(tbl), use_container_width=True, hide_index=True,
                            column_config={
                                "Current ($)":   st.column_config.NumberColumn(format="$%.2f"),
                                "Comp ($)":      st.column_config.NumberColumn(format="$%.2f"),
                                "Suggested ($)": st.column_config.NumberColumn(format="$%.2f"),
                            },
                        )
                        if no_price:
                            st.caption(f"⚠️ {len(no_price)} listings had no CardHedger match — excluded from download.")

                        if has_price:
                            # eBay Seller Hub bulk-edit upload format
                            ebay_buf = io.StringIO()
                            pd.DataFrame([{
                                "Item number": r["item_number"],
                                "Start price": round(r["suggested"], 2),
                            } for r in has_price]).to_csv(ebay_buf, index=False)

                            # Reference sheet
                            ref_buf2 = io.StringIO()
                            pd.DataFrame([{
                                "Item number":   r["item_number"],
                                "Title":         r["title"],
                                "Sport":         r["sport"],
                                "Days Listed":   r["days_listed"],
                                "Current Price": r["current_price"],
                                "Comp FMV":      round(r["comp"], 2) if r.get("comp") else "",
                                "Suggested":     round(r["suggested"], 2),
                                "Trend":         r.get("trend","—"),
                                "∆ vs Listed":   (
                                    f"{(r['suggested']-r['current_price'])/r['current_price']*100:+.1f}%"
                                    if r["current_price"] else ""
                                ),
                            } for r in has_price]).to_csv(ref_buf2, index=False)

                            dl1, dl2 = st.columns(2)
                            dl1.download_button(
                                f"📤 Upload to eBay ({len(has_price)} listings)",
                                data=ebay_buf.getvalue().encode(),
                                file_name=f"sunday_reprice_{date.today().isoformat()}.csv",
                                mime="text/csv", key="sun_dl_ebay",
                                help="eBay Seller Hub → Listings → Bulk Edit → Upload. Only Item number + Start price — nothing else touched.",
                            )
                            dl2.download_button(
                                "📋 Reference sheet",
                                data=ref_buf2.getvalue().encode(),
                                file_name=f"sunday_reprice_ref_{date.today().isoformat()}.csv",
                                mime="text/csv", key="sun_dl_ref",
                            )
                else:
                    st.info(f"No listings found in the {sun_min_days}–{sun_max_days} day range above ${sun_floor}. Try widening the filter.")
            else:
                st.info("Upload your eBay active listings CSV above to get started.")

        # ── PROMOTE LISTINGS TAB ──────────────────────────────────────────────
        with op_tab_promote:
            st.markdown("### 📣 Promote Listings")
            st.markdown("Upload your eBay active listings CSV to see how many listings fall into each ad-rate tier and get step-by-step instructions to promote them in Seller Hub.")

            st.info(
                "**How to get the file:** Seller Hub → Listings → Active → select all → Download → **Active listings report**\n\n"
                "eBay limits bulk promotion to **200 listings per pass** — the tool below tells you exactly how many passes each tier needs."
            )

            promote_file = st.file_uploader("Upload active listings CSV", type=["csv"], key="promote_upload")

            PROMOTE_TIERS = [
                {"label": "Tier 1 — $5.00 to $14.99",  "min": 5.00,  "max": 14.99, "rate": 5,  "emoji": "🟡"},
                {"label": "Tier 2 — $15.00 to $49.99", "min": 15.00, "max": 49.99, "rate": 8,  "emoji": "🟠"},
                {"label": "Tier 3 — $50.00 and up",    "min": 50.00, "max": None,  "rate": 10, "emoji": "🔴"},
            ]

            if promote_file:
                try:
                    import pandas as _pd_promo
                    _promo_df = _pd_promo.read_csv(promote_file)

                    # Find the price column — eBay calls it "Current price" or "Buy It Now price" or "Price"
                    _price_col = None
                    for _col in _promo_df.columns:
                        if "price" in _col.lower() or "buy it now" in _col.lower():
                            _price_col = _col
                            break

                    if _price_col is None:
                        st.error(f"Couldn't find a price column. Columns found: {list(_promo_df.columns)}")
                    else:
                        # Clean price column — strip $, commas
                        _promo_df["_price_clean"] = (
                            _promo_df[_price_col]
                            .astype(str)
                            .str.replace(r"[\$,]", "", regex=True)
                        )
                        _promo_df["_price_num"] = _pd_promo.to_numeric(_promo_df["_price_clean"], errors="coerce")

                        _total_eligible = _promo_df[_promo_df["_price_num"] >= 5.0]
                        _total_skip = len(_promo_df) - len(_total_eligible)

                        st.markdown(f"**{len(_promo_df):,} total listings** — {len(_total_eligible):,} eligible (≥$5) · {_total_skip:,} skipped (<$5)")
                        st.markdown("---")

                        for _tier in PROMOTE_TIERS:
                            if _tier["max"] is not None:
                                _mask = (_promo_df["_price_num"] >= _tier["min"]) & (_promo_df["_price_num"] <= _tier["max"])
                            else:
                                _mask = _promo_df["_price_num"] >= _tier["min"]

                            _tier_df = _promo_df[_mask]
                            _count = len(_tier_df)
                            _passes = (_count + 199) // 200  # ceil div

                            with st.expander(f"{_tier['emoji']} **{_tier['label']}** — {_count:,} listings · {_tier['rate']}% ad rate · {_passes} pass{'es' if _passes != 1 else ''}", expanded=True):
                                if _count == 0:
                                    st.caption("No listings in this range.")
                                    continue

                                st.markdown(f"**Step-by-step — Seller Hub Promoted Listings:**")
                                _steps = [
                                    "Go to **Seller Hub → Marketing → Promoted Listings Advanced**",
                                    "Click **Create campaign** → choose **General**",
                                    f"In **Listings**, filter by price: **Min ${_tier['min']:.2f}**" + (f" / Max ${_tier['max']:.2f}**" if _tier['max'] else " and up**"),
                                    "Click **Select all** (selects up to 200 at a time)",
                                    f"Set ad rate to **{_tier['rate']}%**",
                                    "Click **Add to campaign**",
                                ]
                                if _passes > 1:
                                    _steps.append(f"Repeat **Select all → Add** until all {_count:,} listings are added ({_passes} passes total — 200 per pass)")
                                _steps.append("Click **Launch campaign** and name it something like `DFS {label} {today}`.".replace("{label}", _tier["label"].split("—")[1].strip()).replace("{today}", str(date.today())))

                                for _i, _step in enumerate(_steps, 1):
                                    st.markdown(f"{_i}. {_step}")

                                st.caption(f"Estimated cost if all {_count:,} sell: {_tier['rate']}% of revenue")

                except Exception as _pe:
                    st.error(f"Error reading file: {_pe}")
            else:
                st.markdown("#### Ad Rate Tiers")
                _tc1, _tc2, _tc3 = st.columns(3)
                with _tc1:
                    st.metric("🟡 Tier 1", "$5 – $14.99", "5% ad rate")
                with _tc2:
                    st.metric("🟠 Tier 2", "$15 – $49.99", "8% ad rate")
                with _tc3:
                    st.metric("🔴 Tier 3", "$50+", "10% ad rate")
                st.caption("Upload your active listings CSV above to see counts and step-by-step instructions for each tier.")

        # ── TCP REPRICE TAB ───────────────────────────────────────────────────
        with op_tab_tcp:
            import io as _io
            import math as _math
            import re as _re
            from datetime import datetime as _dt

            st.markdown("### 📊 TCP Reprice")
            st.markdown("Two-step workflow: **Step 1** converts your eBay listings to TCP format for pricing. **Step 2** analyzes TCP results and produces your upload + review files.")

            # ── shared helpers ────────────────────────────────────────────────
            _TCP_NFL = {'Cardinals','Falcons','Ravens','Bills','Panthers','Bears','Bengals','Browns','Cowboys','Broncos','Lions','Packers','Texans','Colts','Jaguars','Chiefs','Raiders','Chargers','Rams','Dolphins','Vikings','Patriots','Saints','Giants','Jets','Eagles','Steelers','Seahawks','49ers',"49's",'Buccaneers','Titans','Commanders','Niners'}
            _TCP_NBA = {'Hawks','Celtics','Nets','Hornets','Bulls','Cavaliers','Mavericks','Nuggets','Pistons','Warriors','Rockets','Pacers','Clippers','Lakers','Grizzlies','Heat','Bucks','Timberwolves','Pelicans','Knicks','Thunder','Magic','Sixers','76ers','Suns','Blazers','Kings','Spurs','Raptors','Jazz'}
            _TCP_MLB = {'Orioles','Red Sox','Yankees','Rays','Blue Jays','White Sox','Guardians','Tigers','Royals','Twins','Astros','Angels','Athletics',"A's",'Mariners','Rangers','Braves','Marlins','Mets','Phillies','Nationals','Cubs','Reds','Brewers','Pirates','Cardinals','Diamondbacks','Rockies','Dodgers','Padres','Giants','Indians'}
            _TCP_MFRS = ['Panini','Topps','Bowman','Upper Deck','Donruss','Score','Leaf','Fleer',"Collector's Edge",'Pro Set','Hoops','SkyBox','Playoff','Pacific','Wild Card']
            _TCP_SETS = ['Prizm','Mosaic','Select','Phoenix','Illusions','Absolute','Donruss','Chronicles','National Treasures','Flawless','Immaculate','Revolution','Optic','Contenders','Rookies & Stars','Score','Origins','Elements','Obsidian','Hoops','Court Kings','Certified','Noir','Spectra','Gold Standard','Chrome','Series 1','Series 2','Update','Allen & Ginter','Stadium Club','Heritage','Finest','Now','Gypsy Queen','Archives','Opening Day','Holiday','Big League','Draft','Platinum','Sapphire']
            _TCP_PARALLELS = ['LogoFractor','Superfractor','X-Fractor','Refractor','Sandglitter','Elevate','Reactive Blue','Reactive Purple','Reactive','Cracked Ice','Mojo','Disco','Laser','Neon','Pulsar','Holo','Wave','Scope','Atomic','Shimmer','Gold','Silver','Blue','Red','Green','Orange','Purple','Pink','Rainbow','Prizm']
            _TCP_POKEMON = ['pokemon','pokémon','sudowoodo','psyduck','fuecoco','magnemite','quaxwell','scraggy','crocalor','drednaw','quaxly','reuniclus','floragato','pawmo','chansey','rockruff','pineco','sprigatito','pikachu','blitzle']
            _TCP_STARS = ['ohtani','messi','yamal','haaland']

            _TCP_HEADER = ['*Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8)','CustomLabel','*Category','StoreCategory','*Title','Subtitle','Relationship','*ConditionID','*C:Graded','*C:Sport','*C:Player/Athlete','*C:Parallel/Variety','*C:Manufacturer','C:Season','*C:Features','*C:Set','CD:Grade - (ID: 27502)','*C:League','CD:Professional Grader - (ID: 27501)','*C:Team','*C:Autographed','CD:Card Condition - (ID: 40001)','*C:Card Name','*C:Card Number','CDA:Certification Number - (ID: 27503)','*C:Type','C:Signed By','C:Autograph Authentication','C:Year Manufactured','C:Card Size','C:Country/Region of Manufacturer','C:Material','C:Autograph Format','C:Vintage','C:Original/Licensed Reprint','C:Event/Tournament','C:Language','C:Autograph Authentication Number','C:Bundle Description','C:California Prop 65 Warning','C:Card Thickness','C:Custom Bundle','C:Insert Set','C:Print Run','PicURL','GalleryType','*Description','*Format','*Duration','*StartPrice','BuyItNowPrice','*Quantity','PayPalAccepted','PayPalEmailAddress','ImmediatePayRequired','PaymentInstructions','*Location','PostalCode','WeightMajor','WeightMinor','ShippingType','ShippingService-1:Option','ShippingService-1:FreeShipping','ShippingService-1:Cost','ShippingService-1:AdditionalCost','ShippingService-2:Option','ShippingService-2:Cost','*DispatchTimeMax','PromotionalShippingDiscount','ShippingDiscountProfileID','*ReturnsAcceptedOption','ReturnsWithinOption','RefundOption','ShippingCostPaidByOption','AdditionalDetails','ShippingProfileName','ReturnProfileName','PaymentProfileName','TakeBackPolicyID','ProductCompliancePolicyID','ScheduleTime','BestOfferEnabled','MinimumBestOfferPrice','BestOfferAutoAcceptPrice','*C:Rookie','*C:Memorabilia','ActiveListings','SoldListings','Confidence','PricingPulledFrom']

            _TCP_DESC = ('<div style="background:#FDFEFE;border:1px solid #CBD4C2;color:#353535;padding:40px;line-height:1.6;font-family:Arial,sans-serif;font-size:16px;"><h1 style="text-align:center;">{t}</h1><table style="width:100%;margin-top:30px;border-spacing:0;"><tr><th align="left">Payment</th><td>Payment is due within 4 days. Unpaid items may be canceled and relisted.</td></tr><tr><th align="left">Shipping</th><td>Items ship via eBay Standard Envelope. Combined shipping may apply. Usually ships within 1 business day.</td></tr><tr><th align="left">Disclaimer</th><td>All cards are sold as-is. No refunds or returns. Contact us before leaving negative feedback.</td></tr></table><div style="text-align:center;margin:2.5rem 0;"><a href="https://www.tradingcardpricer.com" style="text-decoration:none;color:inherit;"><div style="display:inline-flex;align-items:center;gap:1rem;"><img src="https://s3.us-east-2.amazonaws.com/tcr.image.bucket/Logos/Priced+by+TCP.png" alt="EZ Price by TradingCardPricer" style="width:48%;max-width:300px;"><span style="font-size:2rem;font-weight:bold;">TradingCardPricer</span></div></a></div></div>')

            def _tcp_sport(title):
                t = title.upper()
                if any(k in t for k in ['SOCCER','FIFA','WORLD CUP','UEFA','MLS']): return 'SOCCER','SOCCER'
                if 'WWE' in t or 'WRESTLING' in t: return 'WRESTLING','WWE'
                if 'BASKETBALL' in t or 'NBA' in t: return 'BASKETBALL','NBA'
                if 'FOOTBALL' in t or 'NFL' in t: return 'FOOTBALL','NFL'
                if 'BASEBALL' in t or 'MLB' in t: return 'BASEBALL','MLB'
                if 'HOCKEY' in t or 'NHL' in t: return 'HOCKEY','NHL'
                for tm in _TCP_NFL:
                    if _re.search(r'\b'+_re.escape(tm)+r'\b', title, _re.I): return 'FOOTBALL','NFL'
                for tm in _TCP_NBA:
                    if _re.search(r'\b'+_re.escape(tm)+r'\b', title, _re.I): return 'BASKETBALL','NBA'
                for tm in _TCP_MLB:
                    if _re.search(r'\b'+_re.escape(tm)+r'\b', title, _re.I): return 'BASEBALL','MLB'
                if any(m in title for m in ['Topps','Bowman']): return 'BASEBALL','MLB'
                return 'FOOTBALL','NFL'

            def _tcp_meta(title):
                year = (_re.search(r'\b(19|20)\d{2}\b', title) or type('',(),{'group':lambda s,x:''})()).group(0) if _re.search(r'\b(19|20)\d{2}\b', title) else ''
                mfr  = next((m for m in _TCP_MFRS if _re.search(r'\b'+_re.escape(m)+r'\b', title, _re.I)), '')
                set_ = next((s for s in _TCP_SETS if _re.search(r'\b'+_re.escape(s)+r'\b', title, _re.I)), '')
                num  = (_re.search(r'#([A-Za-z0-9\-]+)', title) or type('',(),{'group':lambda s,x:''})()).group(1) if _re.search(r'#([A-Za-z0-9\-]+)', title) else ''
                para = next((p for p in _TCP_PARALLELS if _re.search(r'\b'+_re.escape(p)+r'\b', title, _re.I)), '')
                pr   = (_re.search(r'/(\d+)', title) or type('',(),{'group':lambda s,x:''})()).group(1) if _re.search(r'/(\d+)', title) else ''
                team = next((tm for tm_set in [_TCP_MLB,_TCP_NFL,_TCP_NBA] for tm in tm_set if _re.search(r'\b'+_re.escape(tm)+r'\b', title, _re.I)), '')
                rookie = 'Yes' if _re.search(r'\bRC\b|\bRookie\b|\(RC\)', title, _re.I) else 'No'
                graded_flag = 'Yes' if _re.search(r'\bPSA\b|\bBGS\b|\bSGC\b|\bCGC\b', title, _re.I) else 'No'
                # extract player
                t = title
                for sub in [year, mfr, set_]:
                    if sub: t = _re.sub(r'\b'+_re.escape(sub)+r'\b','',t,flags=_re.I,count=1)
                if num: t = _re.sub(r'#'+_re.escape(num),'',t)
                t = _re.sub(r'\(RC\)|\bRC\b|\bRookie\b','',t,flags=_re.I)
                t = _re.sub(r'/\d+','',t)
                for p in _TCP_PARALLELS: t = _re.sub(r'\b'+_re.escape(p)+r'\b','',t,flags=_re.I)
                t = _re.sub(r'\b(FOOTBALL|BASKETBALL|BASEBALL|HOCKEY|SOCCER|NFL|NBA|MLB|NHL|WWE|WNBA)\b','',t,flags=_re.I)
                for tm_set in [_TCP_NFL,_TCP_NBA,_TCP_MLB]:
                    for tm in tm_set: t = _re.sub(r'\b'+_re.escape(tm)+r'\b','',t,flags=_re.I)
                player = _re.sub(r'\s+',' ',t).strip(' ,#-')[:50]
                return year, mfr, set_, num, para, pr, team, rookie, player, graded_flag

            def _tcp_build_row(listing):
                title  = listing.get('Title','')
                sku    = listing.get('Custom label (SKU)','') or listing.get('Item number','')
                price  = listing.get('Current price', listing.get('Start price',''))
                grader = listing.get('CD:Professional Grader - (ID: 27501)','')
                grade  = listing.get('CD:Grade - (ID: 27502)','')
                cert   = listing.get('CDA:Certification Number - (ID: 27503)','')
                cond   = EBAY_CONDITION_DEFAULT
                year, mfr, set_, num, para, pr, team, rookie, player, graded_flag = _tcp_meta(title)
                if grader or grade: graded_flag = 'Yes'
                sport, league = _tcp_sport(title)
                desc = _TCP_DESC.replace('{t}', title.replace('"','&quot;'))
                return {
                    '*Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8)':'Add',
                    'CustomLabel':sku,'*Category':'261328','StoreCategory':'0','*Title':title,
                    'Subtitle':'','Relationship':'','*ConditionID':'4000','*C:Graded':graded_flag,
                    '*C:Sport':sport,'*C:Player/Athlete':player,'*C:Parallel/Variety':para,
                    '*C:Manufacturer':mfr,'C:Season':year,'*C:Features':'','*C:Set':set_,
                    'CD:Grade - (ID: 27502)':grade,'*C:League':league,
                    'CD:Professional Grader - (ID: 27501)':grader,'*C:Team':team,
                    '*C:Autographed':'No','CD:Card Condition - (ID: 40001)':cond,
                    '*C:Card Name':player,'*C:Card Number':num,
                    'CDA:Certification Number - (ID: 27503)':cert,'*C:Type':'Sports Trading Card',
                    'C:Signed By':'','C:Autograph Authentication':'','C:Year Manufactured':year,
                    'C:Card Size':'','C:Country/Region of Manufacturer':'','C:Material':'',
                    'C:Autograph Format':'','C:Vintage':'','C:Original/Licensed Reprint':'',
                    'C:Event/Tournament':'','C:Language':'','C:Autograph Authentication Number':'',
                    'C:Bundle Description':'','C:California Prop 65 Warning':'','C:Card Thickness':'',
                    'C:Custom Bundle':'','C:Insert Set':'','C:Print Run':pr,'PicURL':'','GalleryType':'',
                    '*Description':desc,'*Format':'FixedPrice','*Duration':'GTC','*StartPrice':price,
                    'BuyItNowPrice':'','*Quantity':'1','PayPalAccepted':'1','PayPalEmailAddress':'',
                    'ImmediatePayRequired':'1','PaymentInstructions':'','*Location':'Scottsdale, AZ',
                    'PostalCode':'85255','WeightMajor':'0','WeightMinor':'4',
                    'ShippingType':'','ShippingService-1:Option':'',
                    'ShippingService-1:FreeShipping':'','ShippingService-1:Cost':'',
                    'ShippingService-1:AdditionalCost':'','ShippingService-2:Option':'',
                    'ShippingService-2:Cost':'','*DispatchTimeMax':'1',
                    'PromotionalShippingDiscount':'','ShippingDiscountProfileID':'',
                    '*ReturnsAcceptedOption':'ReturnsNotAccepted','ReturnsWithinOption':'',
                    'RefundOption':'','ShippingCostPaidByOption':'','AdditionalDetails':'',
                    'ShippingProfileName':_ab_shipping_policy(price),'ReturnProfileName':'Returns','PaymentProfileName':'BIN',
                    'TakeBackPolicyID':'','ProductCompliancePolicyID':'','ScheduleTime':'',
                    'BestOfferEnabled':'','MinimumBestOfferPrice':'','BestOfferAutoAcceptPrice':'',
                    '*C:Rookie':rookie,'*C:Memorabilia':'No',
                    'ActiveListings':'','SoldListings':'','Confidence':'','PricingPulledFrom':'',
                }

            def _tcp_make_csv(rows):
                buf = _io.StringIO()
                info = ['Info','Version=1.0.0','Template=fx_category_template_EBAY_US'] + [''] * (len(_TCP_HEADER)-3)
                buf.write(','.join(info) + '\n')
                w = csv.DictWriter(buf, fieldnames=_TCP_HEADER)
                w.writeheader()
                w.writerows(rows)
                return buf.getvalue().encode('utf-8')

            def _tcp_make_revise_csv(approved_rows):
                buf = _io.StringIO()
                acol = '*Action(SiteID=US|Country=US|Currency=USD|Version=1193|CC=UTF-8)'
                buf.write('Info,Version=1.0.0,Template=fx_category_template_EBAY_US,,\n')
                w = csv.DictWriter(buf, fieldnames=[acol,'ItemID','CustomLabel','*StartPrice'])
                w.writeheader()
                for r in approved_rows:
                    w.writerow({acol:'Revise','ItemID':r['Item number'],'CustomLabel':r['SKU'],'*StartPrice':r['TCP price']})
                return buf.getvalue().encode('utf-8')

            def _tcp_make_analysis_csv(rows):
                fields = ['Item number','SKU','Title','Current price','TCP price','Change $','Change %','Direction','Confidence','Reason held','Matched to']
                buf = _io.StringIO()
                w = csv.DictWriter(buf, fieldnames=fields)
                w.writeheader()
                w.writerows(rows)
                return buf.getvalue().encode('utf-8')

            # ── STEP 1 ────────────────────────────────────────────────────────
            st.markdown("---")
            st.markdown("#### Step 1 — Convert eBay Listings → TCP Format")

            _tcp_ebay_upload = st.file_uploader("Upload eBay active listings report", type=["csv"], key="tcp_ebay_upload")

            if _tcp_ebay_upload:
                _tcp_ebay_bytes = _tcp_ebay_upload.read()
                _tcp_ebay_rows  = list(csv.DictReader(_io.StringIO(_tcp_ebay_bytes.decode('utf-8-sig'))))
                st.session_state['tcp_ebay_rows'] = _tcp_ebay_rows
                st.success(f"Loaded {len(_tcp_ebay_rows):,} listings")

            _tcp_ebay_rows_ss = st.session_state.get('tcp_ebay_rows', [])

            if _tcp_ebay_rows_ss:
                _tcpc1, _tcpc2 = st.columns(2)
                with _tcpc1:
                    _tcp_min_days = st.number_input("Min days listed", min_value=0, max_value=365, value=7, key="tcp_min_days")
                with _tcpc2:
                    _tcp_max_days = st.number_input("Max days listed", min_value=1, max_value=730, value=30, key="tcp_max_days")

                _tcp_today = _dt.today()

                # Case-insensitive column lookup for Start date
                try:
                    _tcp_col_map = {str(k).lower(): k for k in _tcp_ebay_rows_ss[0]}
                except Exception:
                    _tcp_col_map = {}
                _tcp_date_col = _tcp_col_map.get('start date', 'Start date')

                def _tcp_parse_age(row):
                    """Parse eBay start date → age in days. Handles timezone suffixes and single-digit days."""
                    sd = str(row.get(_tcp_date_col, '') or '').strip()
                    for tz in [' PDT', ' PST', ' EDT', ' EST', ' UTC']:
                        sd = sd.replace(tz, '')
                    sd = sd.strip()
                    for fmt in ('%b-%d-%y %H:%M:%S', '%b-%d-%y'):
                        try:
                            return (_tcp_today - _dt.strptime(sd, fmt)).days
                        except Exception:
                            pass
                    return None

                _tcp_filtered = []
                _tcp_age_buckets = {'0–7': 0, '8–30': 0, '31–60': 0, '61–90': 0, '91–180': 0, '180+': 0, 'unknown': 0}
                for _row in _tcp_ebay_rows_ss:
                    _age = _tcp_parse_age(_row)
                    if _age is None:
                        _tcp_age_buckets['unknown'] += 1
                        continue
                    if _age <= 7:       _tcp_age_buckets['0–7'] += 1
                    elif _age <= 30:    _tcp_age_buckets['8–30'] += 1
                    elif _age <= 60:    _tcp_age_buckets['31–60'] += 1
                    elif _age <= 90:    _tcp_age_buckets['61–90'] += 1
                    elif _age <= 180:   _tcp_age_buckets['91–180'] += 1
                    else:               _tcp_age_buckets['180+'] += 1
                    if _tcp_min_days <= _age <= _tcp_max_days:
                        _tcp_filtered.append(_row)

                # Show age distribution so user knows what's in the file
                _bucket_parts = [f"**{v}** {k}d" for k, v in _tcp_age_buckets.items() if v > 0 and k != 'unknown']
                if _tcp_age_buckets['unknown']:
                    _bucket_parts.append(f"{_tcp_age_buckets['unknown']} unparseable")
                if _bucket_parts:
                    st.caption("Age breakdown: " + " · ".join(_bucket_parts))

                _tcp_n_batches = _math.ceil(len(_tcp_filtered) / 250) if _tcp_filtered else 0
                st.info(f"**{len(_tcp_filtered):,} listings** in {_tcp_min_days}–{_tcp_max_days} day range → **{_tcp_n_batches} batch file(s)** of 250")

                if _tcp_filtered:
                    _tcp_built_rows = [_tcp_build_row(r) for r in _tcp_filtered]
                    _tcp_date_label = _tcp_today.strftime('%Y-%m-%d')
                    _dl_cols = st.columns(min(_tcp_n_batches, 4))
                    for _bi in range(_tcp_n_batches):
                        _batch = _tcp_built_rows[_bi*250:(_bi+1)*250]
                        _col = _dl_cols[_bi % len(_dl_cols)]
                        with _col:
                            st.download_button(
                                f"📥 Batch {_bi+1} ({len(_batch)} rows)",
                                data=_tcp_make_csv(_batch),
                                file_name=f"TCP_{_tcp_date_label}_{_tcp_min_days}to{_tcp_max_days}day_batch{str(_bi+1).zfill(2)}.csv",
                                mime="text/csv",
                                key=f"tcp_dl_batch_{_bi}",
                            )
                    st.caption("Upload these files to TradingCardPricer.com, then bring the completed files back to Step 2 below.")

            # ── STEP 2 ────────────────────────────────────────────────────────
            st.markdown("---")
            st.markdown("#### Step 2 — Analyze TCP Results & Build Upload Files")

            _tcp_completed_files = st.file_uploader(
                "Upload TCP completed CSV(s)", type=["csv"],
                accept_multiple_files=True, key="tcp_completed_upload"
            )

            _tcps2c1, _tcps2c2 = st.columns(2)
            with _tcps2c1:
                _tcp_conf_thresh = st.number_input("Min confidence to approve (%)", min_value=0, max_value=100, value=85, key="tcp_conf_thresh")
            with _tcps2c2:
                _tcp_chg_thresh = st.number_input("Max price change to approve (%)", min_value=1, max_value=100, value=20, key="tcp_chg_thresh")

            st.caption("⭐ Ohtani · Messi · Yamal · Haaland — never priced down (up only)")

            if _tcp_completed_files:
                _ebay_ref = st.session_state.get('tcp_ebay_rows', [])
                if not _ebay_ref:
                    _tcp_ebay2 = st.file_uploader("Also upload your original eBay listings CSV (needed for current prices)", type=["csv"], key="tcp_ebay_ref2")
                    if _tcp_ebay2:
                        _ebay_ref = list(csv.DictReader(_io.StringIO(_tcp_ebay2.read().decode('utf-8-sig'))))
                        st.session_state['tcp_ebay_rows'] = _ebay_ref

                if _ebay_ref:
                    _ebay_by_sku = {r.get('Custom label (SKU)','').strip(): r for r in _ebay_ref if r.get('Custom label (SKU)','')}

                    _approved_all, _review_all, _analysis_all = [], [], []
                    _batch_stats = []

                    for _cf in _tcp_completed_files:
                        _cf_lines = _cf.read().decode('utf-8-sig').splitlines()
                        _cf_start = 1 if _cf_lines[0].startswith('Info') else 0
                        _cf_rows  = list(csv.DictReader(_cf_lines[_cf_start:]))
                        for _r in _cf_rows:
                            for _k in list(_r.keys()):
                                if _k != _k.strip(): _r[_k.strip()] = _r.pop(_k)

                        _b_approved = _b_review = _b_skip = 0
                        for _r in _cf_rows:
                            _sku      = _r.get('CustomLabel','').strip()
                            _title    = _r.get('*Title','').strip()
                            _new_raw  = _r.get('*StartPrice','').strip()
                            _conf_raw = _r.get('Confidence','').strip().replace('%','')
                            _pulled   = _r.get('PricingPulledFrom','').strip()
                            _ebay_r   = _ebay_by_sku.get(_sku, {})
                            _item_num = _ebay_r.get('Item number','').strip()
                            _cur_raw  = _ebay_r.get('Current price', _ebay_r.get('Start price','')).strip()

                            try:
                                _new_p = float(_new_raw)
                                if _new_p <= 0: raise ValueError
                            except: _b_skip += 1; continue
                            if any(_k in _title.lower() for _k in _TCP_POKEMON): _b_skip += 1; continue
                            try: _cur_p = float(_cur_raw)
                            except: _b_skip += 1; continue
                            try: _conf = int(_conf_raw) if _conf_raw else 0
                            except: _conf = 0

                            _diff = _new_p - _cur_p
                            _pct  = (_diff / _cur_p * 100) if _cur_p else 0
                            _dir  = 'UP' if _diff > 0.005 else ('DOWN' if _diff < -0.005 else 'NO CHANGE')
                            _star = any(_s in _title.lower() for _s in _TCP_STARS)

                            if _star and _dir == 'DOWN': _b_skip += 1; continue
                            if _dir == 'NO CHANGE': _b_skip += 1; continue

                            _reasons = []
                            if _conf < _tcp_conf_thresh: _reasons.append(f"Conf {_conf}%<{_tcp_conf_thresh}%")
                            if abs(_pct) > _tcp_chg_thresh: _reasons.append(f"Chg {_pct:+.1f}%>{_tcp_chg_thresh}%")

                            _row_out = {
                                'Item number'  : _item_num,
                                'SKU'          : _sku,
                                'Title'        : _title,
                                'Current price': f"{_cur_p:.2f}",
                                'TCP price'    : f"{_new_p:.2f}",
                                'Change $'     : f"{_diff:+.2f}",
                                'Change %'     : f"{_pct:+.1f}%",
                                'Direction'    : _dir,
                                'Confidence'   : f"{_conf}%",
                                'Reason held'  : ' | '.join(_reasons),
                                'Matched to'   : _pulled[:100],
                            }
                            _analysis_all.append(_row_out)
                            if _reasons:
                                _review_all.append(_row_out); _b_review += 1
                            else:
                                _approved_all.append(_row_out); _b_approved += 1

                        _batch_stats.append((_cf.name, _b_approved, _b_review, _b_skip))

                    # Stats display
                    st.markdown("##### Results")
                    _sc1, _sc2, _sc3 = st.columns(3)
                    _sc1.metric("✅ Approved to Upload", len(_approved_all))
                    _sc2.metric("🔍 Held for Review", len(_review_all))
                    _sc3.metric("📊 Total Analyzed", len(_analysis_all))

                    if len(_batch_stats) > 1:
                        with st.expander("Per-file breakdown"):
                            for _fn, _a, _rv, _sk in _batch_stats:
                                st.caption(f"**{_fn}** — ✅{_a} approved / 🔍{_rv} review / ⏭{_sk} skipped")

                    if _approved_all:
                        _up_a   = sum(1 for r in _approved_all if r['Direction']=='UP')
                        _down_a = sum(1 for r in _approved_all if r['Direction']=='DOWN')
                        _up_r   = sum(1 for r in _review_all   if r['Direction']=='UP')
                        _down_r = sum(1 for r in _review_all   if r['Direction']=='DOWN')
                        st.caption(f"Approved: {_up_a} ↑ up / {_down_a} ↓ down  ·  Review: {_up_r} ↑ up / {_down_r} ↓ down")

                    _tcp_label = _dt.today().strftime('%Y-%m-%d')
                    _dl1, _dl2, _dl3 = st.columns(3)
                    with _dl1:
                        if _approved_all:
                            st.download_button(
                                f"📤 eBay Upload ({len(_approved_all)} listings)",
                                data=_tcp_make_revise_csv(_approved_all),
                                file_name=f"eBay_FileExchange_Revise_{_tcp_label}.csv",
                                mime="text/csv", key="tcp_dl_revise",
                                help="File Exchange format — upload to eBay Seller Hub"
                            )
                        else:
                            st.caption("No approved listings")
                    with _dl2:
                        if _analysis_all:
                            st.download_button(
                                f"📊 Price Analysis ({len(_analysis_all)} rows)",
                                data=_tcp_make_analysis_csv(_analysis_all),
                                file_name=f"eBay_PriceAnalysis_{_tcp_label}.csv",
                                mime="text/csv", key="tcp_dl_analysis",
                            )
                    with _dl3:
                        if _review_all:
                            st.download_button(
                                f"🔍 Review Sheet ({len(_review_all)} rows)",
                                data=_tcp_make_analysis_csv(_review_all),
                                file_name=f"eBay_PriceReview_{_tcp_label}.csv",
                                mime="text/csv", key="tcp_dl_review",
                            )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 9 — Consignments (DC Sports)
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 8:
    st.markdown("## 🏷️ Consignments")
    st.caption("DC Sports auction consignment tracking — import send history, track what sold and what was paid.")

    if not SUPABASE_URL:
        st.warning("Supabase not connected. Configure in sidebar to enable Consignments.")
    else:
        # ── Helpers ───────────────────────────────────────────────────────────
        def _csn_get(table, params=""):
            url = f"{SUPABASE_URL}/rest/v1/{table}{params}"
            req = urllib.request.Request(url, headers=sb_headers())
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
                    return json.loads(r.read().decode())
            except Exception:
                return []

        def _csn_post(table, payload, prefer="return=representation"):
            data = json.dumps(payload).encode()
            req = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/{table}",
                data=data,
                headers={**sb_headers(), "Prefer": prefer},
                method="POST",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
                    body = r.read()
                    return json.loads(body.decode()) if body else []
            except Exception:
                return None

        def _csn_patch(table, filt, updates):
            data = json.dumps(updates).encode()
            req = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/{table}?{filt}",
                data=data,
                headers={**sb_headers(), "Prefer": "return=minimal"},
                method="PATCH",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10):
                    return True
            except Exception:
                return False

        def _csn_delete_row(table, filt):
            req = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/{table}?{filt}",
                headers=sb_headers(),
                method="DELETE",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10):
                    return True
            except Exception:
                return False

        def _parse_dc_amount(val):
            if val is None or str(val).strip() in ("", "nan"):
                return 0.0
            try:
                return float(str(val).replace("$", "").replace(",", "").strip())
            except ValueError:
                return 0.0

        def _load_csn():
            ships = _csn_get("consignment_shipments", "?order=shipped_date.desc.nullslast,dc_package_id.desc")
            items = _csn_get("consignment_items", "?order=id.asc&limit=2000")
            return ships, items

        if "csn_data_v2" not in st.session_state:
            with st.spinner("Loading consignment data…"):
                st.session_state["csn_data_v2"] = _load_csn()

        ships_raw, items_raw = st.session_state["csn_data_v2"]
        csn_items = [dict(r) for r in items_raw]

        # Business snapshot
        paid_items   = [i for i in csn_items if (i.get("dc_status") or "").lower() == "paid"]
        active_items = [i for i in csn_items if (i.get("dc_status") or "").lower() not in ("paid", "unsold")]
        total_net    = sum(float(i.get("dc_net") or 0) for i in paid_items)

        sn1, sn2, sn3, sn4 = st.columns(4)
        sn1.metric("Cards Still Active", f"{len(active_items)}")
        sn2.metric("Total Cards (all)",  f"{len(csn_items)}")
        sn3.metric("Total Net (Paid)",   f"${total_net:,.2f}")
        sn4.metric("Lot P&L", "→ Purchases tab")

        st.divider()

        csn_t1, csn_t2, csn_t3, csn_t4 = st.tabs(["📦 Shipments", "🃏 Cards", "⬆️ Import", "🔍 Duplicates"])

        # ── SHIPMENTS ─────────────────────────────────────────────────────────
        with csn_t1:
            st.markdown("### DC Sports Shipments")
            if not ships_raw:
                st.info("No shipments yet. Import a DC Sports CSV to get started, or run the SQL below to create the tables first.")
            else:
                items_by_pkg = {}
                for item in csn_items:
                    pkg = item.get("dc_package_id") or "?"
                    items_by_pkg.setdefault(pkg, []).append(item)

                ship_rows = []
                for ship in ships_raw:
                    pkg = ship.get("dc_package_id") or "?"
                    pkg_items = items_by_pkg.get(pkg, [])
                    paid_c    = sum(1 for i in pkg_items if (i.get("dc_status") or "").lower() == "paid")
                    unsold_c  = sum(1 for i in pkg_items if (i.get("dc_status") or "").lower() == "unsold")
                    active_c  = len(pkg_items) - paid_c - unsold_c
                    paid_items = [i for i in pkg_items if (i.get("dc_status") or "").lower() == "paid"]
                    batch_gross = sum(float(i.get("dc_sale_price") or 0) for i in paid_items)
                    batch_fees  = sum(float(i.get("dc_fees") or 0) for i in paid_items)
                    batch_net   = sum(float(i.get("dc_net") or 0) for i in paid_items)
                    ship_rows.append({
                        "Package ID":   pkg,
                        "Shipped":      ship.get("shipped_date") or "—",
                        "Cards":        len(pkg_items),
                        "Paid":         paid_c,
                        "Unsold":       unsold_c,
                        "Active":       active_c,
                        "Gross ($)":    round(batch_gross, 2),
                        "Fees ($)":     round(batch_fees, 2),
                        "Net ($)":      round(batch_net, 2),
                        "Notes":        ship.get("notes") or "",
                    })

                st.dataframe(
                    pd.DataFrame(ship_rows),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Gross ($)": st.column_config.NumberColumn(format="$%.2f"),
                        "Fees ($)":  st.column_config.NumberColumn(format="$%.2f"),
                        "Net ($)":   st.column_config.NumberColumn(format="$%.2f"),
                    },
                )
                st.caption("💡 Cost basis and lot P&L are tracked in the **📦 Purchases** tab.")

            with st.expander("Edit shipment notes / shipped date"):
                if ships_raw:
                    pkg_opts = [s["dc_package_id"] for s in ships_raw if s.get("dc_package_id")]
                    sel_pkg = st.selectbox("Shipment", pkg_opts, key="csn_edit_pkg")
                    match = next((s for s in ships_raw if s["dc_package_id"] == sel_pkg), None)
                    if match:
                        raw_date = match.get("shipped_date")
                        default_date = pd.to_datetime(raw_date).date() if raw_date else None
                        new_date = st.date_input("Shipped date", value=default_date, key="csn_edit_date")
                        new_notes = st.text_input("Notes", value=match.get("notes") or "", key="csn_edit_notes")
                        if st.button("Save", key="csn_save_ship"):
                            _csn_patch("consignment_shipments", f"id=eq.{match['id']}", {
                                "shipped_date": str(new_date) if new_date else None,
                                "notes": new_notes.strip() or None,
                            })
                            st.success("Saved.")
                            st.session_state.pop("csn_data_v2", None)
                            st.rerun()
                else:
                    st.caption("No shipments to edit yet.")

        # ── CARDS ─────────────────────────────────────────────────────────────
        with csn_t2:
            st.markdown("### Card-Level Tracking")
            st.caption("Assign a Lot SKU to any card to connect its cost basis. Click **Save Lot SKU Assignments** after editing.")

            f1, f2 = st.columns(2)
            pkg_opts2 = ["All"] + sorted(set(i.get("dc_package_id") or "?" for i in csn_items))
            pkg_filt = f1.selectbox("Shipment", pkg_opts2, key="csn_cards_pkg")
            status_opts = ["All"] + sorted(set((i.get("dc_status") or "Unknown") for i in csn_items))
            status_filt = f2.selectbox("Status", status_opts, key="csn_cards_status")

            filtered = csn_items
            if pkg_filt != "All":
                filtered = [i for i in filtered if i.get("dc_package_id") == pkg_filt]
            if status_filt != "All":
                filtered = [i for i in filtered if (i.get("dc_status") or "Unknown") == status_filt]

            if not filtered:
                st.info("No cards match the current filters.")
            else:
                csn_card_ids = [i["id"] for i in filtered]
                card_rows = []
                for i in filtered:
                    card_rows.append({
                        "Title":      i.get("title") or "",
                        "Status":     i.get("dc_status") or "",
                        "Package":    i.get("dc_package_id") or "",
                        "Listed":     (i.get("dc_listing_date") or "")[:10],
                        "Ended":      (i.get("dc_ending_date") or "")[:10],
                        "Gross ($)":  float(i.get("dc_sale_price") or 0),
                        "Fees ($)":   float(i.get("dc_fees") or 0),
                        "Net ($)":    float(i.get("dc_net") or 0),
                        "🗑️":         False,
                    })
                edited_cards = st.data_editor(
                    pd.DataFrame(card_rows),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Gross ($)": st.column_config.NumberColumn(format="$%.2f"),
                        "Fees ($)":  st.column_config.NumberColumn(format="$%.2f"),
                        "Net ($)":   st.column_config.NumberColumn(format="$%.2f"),
                        "🗑️":        st.column_config.CheckboxColumn("🗑️", help="Check row(s) to delete, then click Delete below"),
                    },
                    disabled=["Title", "Status", "Package", "Listed", "Ended", "Gross ($)", "Fees ($)", "Net ($)"],
                    key="csn_cards_editor",
                )

                # Collect checked rows
                checked_mask  = edited_cards["🗑️"].tolist()
                del_ids       = [csn_card_ids[i] for i, v in enumerate(checked_mask) if v]
                del_titles    = edited_cards[edited_cards["🗑️"]]["Title"].tolist()

                if del_ids:
                    st.warning(
                        f"**{len(del_ids)} card(s) marked for deletion:**\n"
                        + "\n".join(f"- {t}" for t in del_titles)
                    )
                    _del_confirm_key = "csn_del_confirm_ids"
                    if st.session_state.get(_del_confirm_key) == del_ids:
                        c1, c2 = st.columns(2)
                        if c1.button("✅ Yes, permanently delete", type="primary", key="csn_del_yes"):
                            n_ok = 0
                            for rid in del_ids:
                                if _csn_delete_row("consignment_items", f"id=eq.{rid}"):
                                    n_ok += 1
                            st.session_state.pop(_del_confirm_key, None)
                            st.session_state.pop("csn_data_v2", None)
                            st.success(f"Deleted {n_ok} of {len(del_ids)} card(s).")
                            st.rerun()
                        if c2.button("✗ Cancel", key="csn_del_cancel"):
                            st.session_state.pop(_del_confirm_key, None)
                            st.rerun()
                    else:
                        if st.button("🗑️ Delete Selected", type="primary", key="csn_del_btn"):
                            st.session_state[_del_confirm_key] = del_ids
                            st.rerun()

        # ── IMPORT ────────────────────────────────────────────────────────────
        with csn_t3:
            st.markdown("### Import DC Sports Export")
            st.caption(
                "Accepts **two formats** — auto-detected on upload:\n\n"
                "• **Seller Dashboard CSV** — columns: Title, Status, ListingDate, EndingDate, BuyItNow, SalePrice, Fees, Net, FriendlyPackageId\n\n"
                "• **Financial Ledger** — Payments download with TransactionType, Description, GrossAmount, Fees, NetAmount columns. "
                "Only Sale rows are imported. These appear as settled items not tied to a shipment batch."
            )

            csn_file = st.file_uploader("DC Sports CSV", type=["csv"], key="csn_import_file")
            if csn_file:
                try:
                    raw_csn = csn_file.read().decode("utf-8-sig")
                    first_csn = raw_csn.split("\n")[0]
                    is_ledger_csn = "TransactionType" in first_csn and "GrossAmount" in first_csn

                    import_df = pd.read_csv(io.StringIO(raw_csn))
                    import_df.columns = [c.strip() for c in import_df.columns]

                    if is_ledger_csn:
                        import_df = import_df[import_df["TransactionType"].astype(str).str.strip() == "Sale"].copy()
                        st.info(f"📊 **Financial Ledger detected** — {len(import_df)} Sale rows will be imported as settled consignment results.")
                    else:
                        uniq_pkgs = import_df["FriendlyPackageId"].dropna().astype(str).nunique() if "FriendlyPackageId" in import_df.columns else 0
                        st.caption(f"Shipment batches detected: **{uniq_pkgs}**")

                    st.caption(f"Found **{len(import_df):,}** rows. Preview:")
                    st.dataframe(import_df.head(10), use_container_width=True, hide_index=True)

                    if st.button("⬆️ Import to Supabase", type="primary", key="csn_import_btn"):
                        # Load all existing items for matching
                        all_existing = _csn_get("consignment_items", "?select=id,title,dedup_key")
                        existing_keys2 = {r["dedup_key"] for r in all_existing if r.get("dedup_key")}
                        # Title → id lookup for ledger match-and-update
                        title_to_id = {
                            (r.get("title") or "").strip().lower(): r["id"]
                            for r in all_existing if r.get("title") and r.get("id")
                        }

                        if not is_ledger_csn:
                            # Ensure shipment rows exist for the seller export format
                            pkg_ids2 = import_df["FriendlyPackageId"].dropna().astype(str).unique() if "FriendlyPackageId" in import_df.columns else []
                            existing_ships2 = {s["dc_package_id"] for s in ships_raw if s.get("dc_package_id")}
                            for pkg in pkg_ids2:
                                if str(pkg) not in existing_ships2:
                                    _csn_post("consignment_shipments", {"dc_package_id": str(pkg)}, prefer="return=minimal")

                        fresh_ships2 = _csn_get("consignment_shipments", "?order=id.asc")
                        ship_id_map = {s["dc_package_id"]: s["id"] for s in fresh_ships2 if s.get("dc_package_id")}

                        imported = updated = skipped = failed = 0
                        progress = st.progress(0)
                        total_rows = len(import_df)

                        for _ci, (idx, row) in enumerate(import_df.iterrows()):
                            progress.progress(int((_ci + 1) / max(total_rows, 1) * 100))

                            if is_ledger_csn:
                                title   = str(row.get("Description") or "").strip()
                                gross   = _parse_dc_amount(row.get("GrossAmount", 0))
                                fees    = abs(_parse_dc_amount(row.get("Fees", 0)))
                                net     = _parse_dc_amount(row.get("NetAmount", 0))
                                dedup   = f"dc_ledger|{title}|{round(gross, 2)}"

                                # Try to match an existing card by title and patch it
                                match_id = title_to_id.get(title.lower())
                                if match_id:
                                    ok = _csn_patch("consignment_items", f"id=eq.{match_id}", {
                                        "dc_status":     "Paid",
                                        "dc_sale_price": gross,
                                        "dc_fees":       fees,
                                        "dc_net":        net,
                                    })
                                    if ok:
                                        updated += 1
                                    else:
                                        failed += 1
                                    continue  # never insert a duplicate

                                # No title match — insert as new only if not already in DB
                                if dedup in existing_keys2:
                                    skipped += 1
                                    continue

                                item_row = {
                                    "dc_package_id": None,
                                    "shipment_id":   None,
                                    "title":         title or None,
                                    "dc_status":     "Paid",
                                    "dc_sale_price": gross,
                                    "dc_fees":       fees,
                                    "dc_net":        net,
                                    "dedup_key":     dedup,
                                }
                                res2 = _csn_post("consignment_items", item_row, prefer="return=minimal")
                                if res2 is not None:
                                    imported += 1
                                    existing_keys2.add(dedup)
                                else:
                                    failed += 1

                            else:
                                pkg     = str(row.get("FriendlyPackageId") or "").strip()
                                title   = str(row.get("Title") or "").strip()
                                ending  = str(row.get("EndingDate") or "").strip()[:19]
                                dedup   = f"{pkg}|{title}|{ending}"

                                if dedup in existing_keys2:
                                    skipped += 1
                                    continue

                                item_row = {
                                    "dc_package_id": pkg or None,
                                    "shipment_id":   ship_id_map.get(pkg),
                                    "title":         title or None,
                                    "dc_status":     str(row.get("Status") or "").strip() or None,
                                    "dc_listing_date": str(row.get("ListingDate") or "").strip()[:19] or None,
                                    "dc_ending_date":  ending or None,
                                    "dc_buy_it_now":   _parse_dc_amount(row.get("BuyItNow")),
                                    "dc_sale_price":   _parse_dc_amount(row.get("SalePrice")),
                                    "dc_fees":         _parse_dc_amount(row.get("Fees")),
                                    "dc_net":          _parse_dc_amount(row.get("Net")),
                                    "dc_front_image_url": str(row.get("FrontImageUrl") or "").strip() or None,
                                    "dedup_key":       dedup,
                                }
                                res2 = _csn_post("consignment_items", item_row, prefer="return=minimal")
                                if res2 is not None:
                                    imported += 1
                                    existing_keys2.add(dedup)
                                else:
                                    failed += 1

                        progress.empty()
                        parts = []
                        if updated:  parts.append(f"**{updated}** cards updated with settlement data")
                        if imported: parts.append(f"**{imported}** new cards added")
                        if skipped:  parts.append(f"**{skipped}** already up to date")
                        if failed:   parts.append(f"**{failed}** failed")
                        st.success("✅ " + ", ".join(parts) + "." if parts else "✅ Done.")
                        if imported:
                            st.session_state.pop("csn_data_v2", None)
                            st.rerun()
                except Exception as e:
                    st.error(f"Error reading CSV: {e}")

        # ── DUPLICATES ────────────────────────────────────────────────────────
        with csn_t4:
            st.markdown("### 🔍 Duplicate Finder")
            st.caption(
                "Groups cards with the same card number together. "
                "Keep the row with sale data (Paid / AwaitingPayment); delete the stale $0 copy."
            )

            import re as _re

            # Card-type words that should never count as a player name
            _DUP_STOP = {
                "TOPPS","PANINI","BOWMAN","CHROME","PRIZM","REFRACTOR","ROOKIE","AUTO",
                "AUTOGRAPH","GOLD","SILVER","BLUE","RED","GREEN","PURPLE","ORANGE","BLACK",
                "YELLOW","MOJO","WAVE","CRACKED","GEOMETRIC","REPTILIAN","SHIMMER","AQUA",
                "SAPPHIRE","STERLING","FOIL","RAINBOW","KABOOM","FINEST","OPTIC","DONRUSS",
                "SELECT","MEGA","DRAFT","SERIES","EDITION","LOGOFRACTOR","METEORIC","RISE",
                "SHOWPIECES","COLLECTORS","KIT","EXCLUSIVE","MINI","DIAMOND","PERFORMANCE",
                "PROSPECTS","SANDGLITTER","NOW","SLEEK","FINISHERS","LAZER","RARE","HOLO",
                "FIRST","STAGE","ULTIMATE","ILLUMINATION","PSA","GEM","MINT","CHROME",
                "PRIZM","AUTOS","PICKS","HOUSE","BASKETBALL","FOOTBALL","SOCCER","BASEBALL",
                "SPORTS","CARD","CARDS","CHROME","FRESH","SEASON","TIP","OFF","LAVA","SKY",
            }

            def _card_num(title):
                """Extract card number token like #BCP-63, #MR-3, #75, #17b."""
                m = _re.search(r'#([A-Z0-9]+-[A-Z0-9]+|[A-Z]{1,3}[0-9]+|[0-9]+[A-Z]?)', str(title or ""), _re.IGNORECASE)
                return m.group(1).upper() if m else None

            def _player_hint(title):
                """Return first all-caps word (4+ chars) that isn't a set/type word."""
                for w in _re.findall(r'[A-Z]{4,}', str(title or "")):
                    if w not in _DUP_STOP:
                        return w
                return ""

            # Build groups: key = (card_number, player_hint_if_short_num)
            _groups: dict = {}
            for item in csn_items:
                num = _card_num(item.get("title") or "")
                if not num:
                    continue
                # Short numbers like "75" or "16" need a player name anchor to avoid false matches
                if len(num) <= 3:
                    hint = _player_hint(item.get("title") or "")
                    key = (num, hint)
                else:
                    key = (num, "")
                _groups.setdefault(key, []).append(item)

            dupe_groups = [(k, rows) for k, rows in _groups.items() if len(rows) >= 2]
            dupe_groups.sort(key=lambda x: x[0][0])

            if not csn_items:
                st.info("No cards loaded yet. Import a CSV first.")
            elif not dupe_groups:
                st.success("✅ No duplicates detected across your consignment cards.")
            else:
                st.warning(
                    f"**{len(dupe_groups)} potential duplicate group(s)** found. "
                    "Review each — keep the row with actual sale data, delete the stale $0 copy."
                )

                for (card_num_key, _), rows in dupe_groups:
                    # Header line uses the most informative title
                    best_title = max(rows, key=lambda r: len(r.get("title") or "")).get("title") or "?"
                    label = f"#{card_num_key}  ·  {len(rows)} entries  —  {best_title[:60]}"
                    with st.expander(label, expanded=True):
                        for row in rows:
                            row_id   = row["id"]
                            status   = row.get("dc_status") or "Sent"
                            net      = float(row.get("dc_net") or 0)
                            sale     = float(row.get("dc_sale_price") or 0)
                            pkg      = row.get("dc_package_id") or "—"
                            title    = row.get("title") or "—"
                            has_data = sale > 0

                            badge = "✅ Paid" if status.lower() == "paid" \
                                else "⏳ AwaitingPayment" if "awaiting" in status.lower() \
                                else "📦 Sent"

                            confirm_key = f"csn_del_confirm_{row_id}"

                            col_info, col_btn = st.columns([6, 1])
                            with col_info:
                                if has_data:
                                    st.markdown(
                                        f"**ID {row_id}** &nbsp; {badge} &nbsp; 📦 `{pkg}` &nbsp; "
                                        f"💰 **${net:.2f}** net (${sale:.2f} gross) — ✅ has sale data"
                                    )
                                else:
                                    st.markdown(
                                        f"**ID {row_id}** &nbsp; {badge} &nbsp; 📦 `{pkg}` &nbsp; "
                                        f"$0 — likely the stale copy"
                                    )
                                st.caption(title)

                            with col_btn:
                                if st.session_state.get(confirm_key):
                                    st.markdown("**Sure?**")
                                    if st.button("Yes, delete", key=f"csn_del_yes_{row_id}", type="primary"):
                                        ok = _csn_delete_row("consignment_items", f"id=eq.{row_id}")
                                        st.session_state.pop(confirm_key, None)
                                        if ok:
                                            st.session_state.pop("csn_data_v2", None)
                                            st.success(f"Deleted ID {row_id}")
                                            st.rerun()
                                        else:
                                            st.error("Delete failed — check Supabase connection.")
                                    if st.button("Cancel", key=f"csn_del_no_{row_id}"):
                                        st.session_state.pop(confirm_key, None)
                                        st.rerun()
                                else:
                                    st.button(
                                        "🗑️ Delete",
                                        key=f"csn_del_btn_{row_id}",
                                        on_click=lambda k=confirm_key: st.session_state.update({k: True}),
                                    )
                            st.divider()

            st.divider()
            st.markdown("**SQL — Create Consignment Tables** *(run once in Supabase SQL Editor)*")
            st.code(
                """create table if not exists consignment_shipments (
  id bigint primary key generated always as identity,
  dc_package_id text unique,
  shipped_date date,
  notes text,
  created_at timestamptz default now()
);

create table if not exists consignment_items (
  id bigint primary key generated always as identity,
  shipment_id bigint references consignment_shipments(id),
  dc_package_id text,
  title text,
  dc_status text,
  dc_listing_date text,
  dc_ending_date text,
  dc_buy_it_now numeric,
  dc_sale_price numeric,
  dc_fees numeric,
  dc_net numeric,
  dc_front_image_url text,
  dedup_key text unique,
  created_at timestamptz default now()
);
create index if not exists idx_ci_shipment on consignment_items(shipment_id);""",
                language="sql",
            )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 11 — Purchases (Lot Tracking)
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 9:
    st.markdown("## 📦 Purchases")
    st.caption("Track card lots you buy. SKU prefix (first 2 segments, e.g. MATTSFB-072026) ties every card back to its lot.")

    if not SUPABASE_URL:
        st.warning("Supabase not connected. Configure in sidebar.")
    else:
        def _pur_get(table, params=""):
            url = f"{SUPABASE_URL}/rest/v1/{table}{params}"
            req = urllib.request.Request(url, headers=sb_headers())
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
                    return json.loads(r.read().decode())
            except Exception:
                return []

        _pur_last_error = {"msg": None}

        def _pur_post(table, payload):
            data = json.dumps(payload).encode()
            req = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/{table}",
                data=data,
                headers={**sb_headers(), "Prefer": "return=representation"},
                method="POST",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
                    body = r.read()
                    return json.loads(body.decode()) if body else []
            except urllib.error.HTTPError as e:
                body = e.read().decode("utf-8", errors="replace")
                _pur_last_error["msg"] = f"HTTP {e.code}: {body[:400]}"
                return None
            except Exception as ex:
                _pur_last_error["msg"] = str(ex)
                return None

        def _pur_delete(table, filt):
            req = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/{table}?{filt}",
                headers=sb_headers(),
                method="DELETE",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10):
                    return True
            except Exception:
                return False

        def _pur_patch(table, filt, payload):
            data = json.dumps(payload).encode()
            hdrs = {**sb_headers(), "Content-Type": "application/json", "Prefer": "return=representation"}
            req  = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/{table}?{filt}",
                data=data, headers=hdrs, method="PATCH",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
                    return json.loads(r.read())
            except urllib.error.HTTPError as e:
                _pur_last_error["msg"] = e.read().decode(errors="replace")
                return None
            except Exception as ex:
                _pur_last_error["msg"] = str(ex)
                return None

        def _pur_upsert_lot_cards(rows):
            """Bulk upsert rows into lot_cards (on_conflict: lot_prefix,sku → update title)."""
            if not rows:
                return 0, None
            data = json.dumps(rows).encode()
            hdrs = {
                **sb_headers(),
                "Content-Type": "application/json",
                "Prefer": "resolution=merge-duplicates,return=representation",
            }
            req = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/lot_cards?on_conflict=lot_prefix,sku",
                data=data, headers=hdrs, method="POST",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=30) as r:
                    body = r.read()
                    result = json.loads(body.decode()) if body else []
                    return len(result), None
            except urllib.error.HTTPError as e:
                err = e.read().decode("utf-8", errors="replace")
                return 0, f"HTTP {e.code}: {err[:400]}"
            except Exception as ex:
                return 0, str(ex)

        def _pur_prefix(sku):
            """Extract lot prefix: first 2 dash-separated segments of SKU."""
            parts = str(sku or "").strip().split("-")
            return "-".join(parts[:2]) if len(parts) >= 2 else ""

        if "pur_lots" not in st.session_state:
            st.session_state["pur_lots"] = _pur_get("purchase_lots", "?order=purchase_date.desc")

        lots_data = st.session_state["pur_lots"]

        pur_t1, pur_t2, pur_t3, pur_t4 = st.tabs(["📋 Lots", "⬆️ Import Cards", "📊 P&L by Lot", "🃏 Individual Cards"])

        # ── LOTS ──────────────────────────────────────────────────────────────
        with pur_t1:
            st.markdown("### Purchase Lots")
            st.caption("Each lot has a prefix matching the first 2 segments of your SKU (e.g. `MATTSFB-072026`). All cards with that prefix roll up to this lot.")

            with st.form("pur_new_lot"):
                st.markdown("**Add Lot**")
                p1, p2, p3 = st.columns(3)
                pl_prefix  = p1.text_input("Lot Prefix *", placeholder="MATTSFB-072026")
                pl_source  = p1.text_input("Source", placeholder="Matt's FB Marketplace")
                pl_aliases = p1.text_input("Alias Prefixes", placeholder="RBLOT_-_07, OTHER-ALT", help="Comma-separated alternate SKU prefixes that also belong to this lot")
                pl_date    = p2.date_input("Purchase Date", value=date.today())
                pl_cost    = p2.number_input("Total Cost Paid ($)", min_value=0.0, step=0.01, format="%.2f")
                pl_count   = p2.number_input("Card Count in Lot", min_value=0, step=1, value=0, help="Total cards you received in this lot — used as a check against imported cards")
                pl_notes   = p3.text_area("Notes", height=80, placeholder="What was in the lot, where from, etc.")
                lot_sub    = st.form_submit_button("Add Lot", type="primary")

            if lot_sub:
                prefix_clean = pl_prefix.strip().upper()
                if not prefix_clean or len(prefix_clean.split("-")) < 2:
                    st.error("Prefix must have at least 2 dash-separated segments (e.g. MATTSFB-072026).")
                else:
                    aliases_clean = ",".join([a.strip().upper() for a in pl_aliases.split(",") if a.strip()]) or None
                    res = _pur_post("purchase_lots", {
                        "lot_prefix":      prefix_clean,
                        "source":          pl_source.strip() or None,
                        "purchase_date":   str(pl_date),
                        "total_cost":      float(pl_cost),
                        "card_count":      int(pl_count),
                        "notes":           pl_notes.strip() or None,
                        "alias_prefixes":  aliases_clean,
                    })
                    if res is not None:
                        st.success(f"Lot **{prefix_clean}** added — ${pl_cost:,.2f} paid.")
                        st.session_state.pop("pur_lots", None)
                        st.rerun()
                    else:
                        err = _pur_last_error.get("msg") or "Unknown error"
                        if "does not exist" in err or "42P01" in err:
                            st.error("Table not found — run the SQL block below to create the `purchase_lots` table first.")
                        elif "duplicate" in err.lower() or "unique" in err.lower():
                            st.error(f"Prefix **{prefix_clean}** already exists — delete it first to replace it.")
                        else:
                            st.error(f"Save failed: {err}")

            st.divider()

            if not lots_data:
                st.info("No lots yet. Add one above.")
            else:
                # ── Card CSV loader (persists in session state) ────────────────
                st.markdown("**Load cards from CSV to see them under each lot**")
                lot_csv = st.file_uploader(
                    "Haystack or eBay active listings CSV",
                    type=["csv"], key="pur_lot_csv",
                    help="Upload once — cards stay visible until you reload the page or upload a new file."
                )
                if lot_csv:
                    try:
                        raw_lc = lot_csv.read().decode("utf-8-sig")
                        lines_lc = raw_lc.split("\n")
                        start_lc = 1 if lines_lc and lines_lc[0].strip().startswith("Info,") else 0
                        import io as _io2
                        lc_df = pd.read_csv(_io2.StringIO("\n".join(lines_lc[start_lc:])), encoding="utf-8-sig")
                        lc_df.columns = [c.strip() for c in lc_df.columns]
                        sku_lc = next((c for c in ["CustomLabel","Custom label (SKU)","Custom Label (SKU)","Custom label"] if c in lc_df.columns), None)
                        title_lc = next((c for c in ["*Title","Title"] if c in lc_df.columns), None)
                        price_lc = next((c for c in ["*StartPrice","Current price","Start price"] if c in lc_df.columns), None)
                        if sku_lc:
                            lc_df = lc_df[lc_df[sku_lc].astype(str).str.strip() != ""].copy()
                            lc_df["_prefix"] = lc_df[sku_lc].apply(_pur_prefix)
                            st.session_state["pur_lot_cards"] = {
                                "df": lc_df, "sku_col": sku_lc,
                                "title_col": title_lc, "price_col": price_lc,
                            }
                            st.success(f"Loaded {len(lc_df):,} cards from CSV.")
                        else:
                            st.error("No SKU column found — expected CustomLabel or 'Custom label (SKU)'.")
                    except Exception as lc_e:
                        st.error(f"Error reading CSV: {lc_e}")

                lc_state = st.session_state.get("pur_lot_cards")

                # Load live sales data for all lots (full detail for sold cards table)
                _lot_sales_raw = _pur_get("sales_records", "?select=sku,title,sale_date,gross_revenue,net_proceeds,source&sku=not.is.null&order=sale_date.desc&limit=5000")

                # Load persistent lot inventory from Supabase lot_cards table
                _lot_cards_all = _pur_get("lot_cards", "?select=lot_prefix,sku,title&order=lot_prefix.asc,sku.asc&limit=10000")
                # Group by canonical lot prefix (upper)
                _lot_cards_by_pfx = {}
                for _lc in (_lot_cards_all or []):
                    _lcp = str(_lc.get("lot_prefix") or "").upper()
                    if _lcp not in _lot_cards_by_pfx:
                        _lot_cards_by_pfx[_lcp] = []
                    _lot_cards_by_pfx[_lcp].append(_lc)
                _lot_alias_map = {}
                for _l in lots_data:
                    _lot_alias_map[_l["lot_prefix"].upper()] = _l["lot_prefix"].upper()
                    for _a in (_l.get("alias_prefixes") or "").split(","):
                        _a = _a.strip().upper()
                        if _a:
                            _lot_alias_map[_a] = _l["lot_prefix"].upper()
                _lot_rev = {}
                _lot_sales_by_pfx = {}
                for _s in _lot_sales_raw:
                    _p = _pur_prefix(_s.get("sku", ""))
                    if _p:
                        _canon = _lot_alias_map.get(_p.upper(), _p.upper())
                        if _canon not in _lot_rev:
                            _lot_rev[_canon] = {"net": 0.0, "gross": 0.0, "count": 0}
                            _lot_sales_by_pfx[_canon] = []
                        _lot_rev[_canon]["net"]   += float(_s.get("net_proceeds") or 0)
                        _lot_rev[_canon]["gross"] += float(_s.get("gross_revenue") or 0)
                        _lot_rev[_canon]["count"] += 1
                        _lot_sales_by_pfx[_canon].append(_s)

                st.divider()
                st.markdown("**Your Lots**")
                for lot in lots_data:
                    pfx      = lot["lot_prefix"]
                    expected = int(lot.get("card_count") or 0)
                    cost     = float(lot.get("total_cost") or 0)
                    source   = lot.get("source") or "—"
                    pdate    = lot.get("purchase_date") or "—"
                    _rev     = _lot_rev.get(pfx.upper(), {})
                    sold     = _rev.get("count", 0)
                    net_rev  = _rev.get("net", 0.0)
                    remaining = max(0, expected - sold) if expected > 0 else None
                    pl       = round(net_rev - cost, 2)

                    # Count cards from CSV for this lot
                    lot_cards_df = None
                    card_count_csv = 0
                    if lc_state:
                        mask = lc_state["df"]["_prefix"].str.upper() == pfx.upper()
                        lot_cards_df = lc_state["df"][mask]
                        card_count_csv = len(lot_cards_df)

                    count_label = ""
                    if lc_state:
                        if expected > 0:
                            diff = card_count_csv - expected
                            count_label = f" · {card_count_csv}/{expected} in CSV" + (" ✅" if diff == 0 else f" ({'+'if diff>0 else ''}{diff})")
                        else:
                            count_label = f" · {card_count_csv} in CSV"

                    # Sales tally in header
                    if sold > 0:
                        left_label = f"{remaining} left" if remaining is not None else f"{sold} sold"
                        pl_label = f"{'+'if pl>=0 else ''}{pl:,.0f}"
                        sales_label = f" · {sold} sold / {left_label} · Net ${net_rev:,.0f} · P&L ${pl_label}"
                    else:
                        sales_label = " · no sales yet"

                    # Compute turn/timing metrics
                    sold_records_pre = _lot_sales_by_pfx.get(pfx.upper(), [])
                    cost_per_card   = round(cost / expected, 2) if expected > 0 else None
                    net_per_card    = round(net_rev / sold, 2) if sold > 0 else None
                    roi_pct         = round((net_rev - cost) / cost * 100, 1) if cost > 0 else None
                    proj_net        = round(net_per_card * expected, 2) if (net_per_card and expected > 0) else None
                    proj_roi        = round((proj_net - cost) / cost * 100, 1) if (proj_net and cost > 0) else None

                    # Turn rate and avg days to sell
                    _turn_rate = None   # cards/week
                    _avg_days  = None   # avg days from purchase to sale
                    _days_left = None   # projected weeks to clear remaining
                    if sold_records_pre and pdate and pdate != "—":
                        try:
                            from datetime import datetime as _dt, date as _date
                            _pdate = _dt.strptime(pdate[:10], "%Y-%m-%d").date()
                            _today = _date.today()
                            _weeks_since = max(1, (_today - _pdate).days / 7)
                            _turn_rate = round(sold / _weeks_since, 1)
                            # Avg days from purchase date to each sale
                            _sale_days = []
                            for _s in sold_records_pre:
                                _sd = _s.get("sale_date", "")
                                if _sd:
                                    try:
                                        _sale_days.append((_dt.strptime(_sd[:10], "%Y-%m-%d").date() - _pdate).days)
                                    except Exception:
                                        pass
                            if _sale_days:
                                _avg_days = round(sum(_sale_days) / len(_sale_days), 0)
                            if remaining and _turn_rate and _turn_rate > 0:
                                _days_left = round(remaining / _turn_rate * 7)
                        except Exception:
                            pass

                    header = f"**{pfx}** — {source} · {pdate} · ${cost:,.2f}{sales_label}{count_label}"
                    with st.expander(header, expanded=False):
                        # Row 1: core financials
                        ec1, ec2, ec3, ec4, ec5, ec6 = st.columns(6)
                        ec1.metric("Cost Paid",    f"${cost:,.2f}")
                        ec2.metric("Net Revenue",  f"${net_rev:,.2f}")
                        ec3.metric("P&L",          f"${pl:+,.2f}")
                        ec4.metric("ROI",          f"{roi_pct:+.1f}%" if roi_pct is not None else "—",
                                   help="Based on net revenue received so far vs total cost paid")
                        ec5.metric("Proj ROI",     f"{proj_roi:+.1f}%" if proj_roi is not None else "—",
                                   help="If remaining cards sell at same avg net/card")
                        ec6.metric("$/Card Cost",  f"${cost_per_card:.2f}" if cost_per_card else "—")

                        # Row 2: inventory & pace
                        ep1, ep2, ep3, ep4, ep5, ep6 = st.columns(6)
                        ep1.metric("Cards in Lot", f"{expected}" if expected > 0 else "—")
                        ep2.metric("Sold",         sold)
                        ep3.metric("Left",         remaining if remaining is not None else "—")
                        ep4.metric("$/Card Net",   f"${net_per_card:.2f}" if net_per_card else "—",
                                   help="Average net proceeds per card sold so far")
                        ep5.metric("Turn Rate",    f"{_turn_rate}/wk" if _turn_rate else "—",
                                   help="Cards sold per week since purchase date")
                        ep6.metric("Avg Days to Sell", f"{int(_avg_days)}d" if _avg_days else "—",
                                   help="Average days from lot purchase date to each card's sale date")

                        if _days_left is not None:
                            st.caption(f"📅 At current pace ({_turn_rate}/wk), ~{_days_left} days to clear remaining {remaining} cards.")
                        if lc_state:
                            st.caption(f"Cards in active listings CSV: {card_count_csv}")
                        if lot.get("notes"):
                            st.caption(lot["notes"])

                        # Sold cards table
                        sold_records = _lot_sales_by_pfx.get(pfx.upper(), [])
                        if sold_records:
                            st.divider()
                            st.markdown(f"**✅ Sold Cards ({len(sold_records)})**")
                            sold_df = pd.DataFrame([{
                                "SKU":        s.get("sku", ""),
                                "Title":      s.get("title", "") or "—",
                                "Date":       (s.get("sale_date") or "")[:10] or "—",
                                "Gross ($)":  float(s.get("gross_revenue") or 0),
                                "Net ($)":    float(s.get("net_proceeds") or 0),
                            } for s in sold_records])
                            st.dataframe(sold_df, use_container_width=True, hide_index=True,
                                column_config={
                                    "Gross ($)": st.column_config.NumberColumn(format="$%.2f"),
                                    "Net ($)":   st.column_config.NumberColumn(format="$%.2f"),
                                })

                        # Cross-check: Supabase lot_cards inventory vs sales_records
                        _inv_cards = _lot_cards_by_pfx.get(pfx.upper(), [])
                        if _inv_cards:
                            st.divider()
                            # Build sold SKU lookup: sku.upper() → list of sale records
                            _sold_by_sku = {}
                            for _s in sold_records:
                                _sk = str(_s.get("sku") or "").upper()
                                if _sk:
                                    if _sk not in _sold_by_sku:
                                        _sold_by_sku[_sk] = []
                                    _sold_by_sku[_sk].append(_s)
                            _inv_sold   = sum(1 for c in _inv_cards if str(c.get("sku","")).upper() in _sold_by_sku)
                            _inv_unsold = len(_inv_cards) - _inv_sold
                            st.markdown(f"**📦 Lot Inventory ({len(_inv_cards)} cards · ✅ {_inv_sold} sold · 🟡 {_inv_unsold} outstanding)**")
                            cross_rows = []
                            for _c in _inv_cards:
                                _csku  = str(_c.get("sku") or "")
                                _ctitle = str(_c.get("title") or "—")
                                _sales_for_card = _sold_by_sku.get(_csku.upper(), [])
                                if _sales_for_card:
                                    for _sale in _sales_for_card:
                                        cross_rows.append({
                                            "SKU":      _csku,
                                            "Title":    _ctitle,
                                            "Status":   "✅ Sold",
                                            "Channel":  (_sale.get("source") or "—").replace("ebay","eBay").replace("collx","CollX").replace("dc_sports","DC Sports"),
                                            "Sale Date": (_sale.get("sale_date") or "")[:10] or "—",
                                            "Net ($)":  float(_sale.get("net_proceeds") or 0),
                                        })
                                else:
                                    cross_rows.append({
                                        "SKU":      _csku,
                                        "Title":    _ctitle,
                                        "Status":   "🟡 Outstanding",
                                        "Channel":  "—",
                                        "Sale Date": "—",
                                        "Net ($)":  0.0,
                                    })
                            cross_df = pd.DataFrame(cross_rows)
                            st.dataframe(cross_df, use_container_width=True, hide_index=True,
                                column_config={"Net ($)": st.column_config.NumberColumn(format="$%.2f")})

                            # ── Manual Sale Logger ────────────────────────────
                            _outstanding_cards = [c for c in _inv_cards if str(c.get("sku","")).upper() not in _sold_by_sku]
                            if _outstanding_cards:
                                with st.expander(f"📝 Log a Sale ({len(_outstanding_cards)} outstanding)", expanded=False):
                                    _ls_key = f"ls_{pfx}"
                                    _sku_opts = [f"{c['sku']} — {c.get('title','')}" for c in _outstanding_cards]
                                    _ls_sku_sel = st.selectbox("Card (SKU)", _sku_opts, key=f"{_ls_key}_sku")
                                    _ls_sku = _ls_sku_sel.split(" — ")[0] if _ls_sku_sel else ""
                                    _ls_title = next((c.get("title","") for c in _outstanding_cards if c["sku"] == _ls_sku), "")
                                    ls_c1, ls_c2, ls_c3 = st.columns(3)
                                    _ls_channel = ls_c1.selectbox("Channel", ["collx", "ebay", "dc_sports", "whatnot", "facebook", "other"], key=f"{_ls_key}_ch",
                                                    format_func=lambda x: {"collx":"CollX","ebay":"eBay","dc_sports":"DC Sports","whatnot":"Whatnot","facebook":"Facebook","other":"Other"}.get(x,x))
                                    _ls_date = ls_c2.date_input("Sale Date", value=date.today(), key=f"{_ls_key}_date")
                                    _ls_net  = ls_c3.number_input("Net Proceeds ($)", min_value=0.0, step=0.01, format="%.2f", key=f"{_ls_key}_net")
                                    _ls_gross = st.number_input("Gross Revenue ($) — leave 0 to match net", min_value=0.0, step=0.01, format="%.2f", key=f"{_ls_key}_gross")
                                    if st.button("✅ Save Sale", type="primary", key=f"{_ls_key}_save"):
                                        _gross_final = _ls_gross if _ls_gross > 0 else _ls_net
                                        _fee_final   = round(_gross_final - _ls_net, 2)
                                        _dedup_ls    = f"manual|{_ls_sku}|{_ls_date}"
                                        _rec_ls = {
                                            "source":       _ls_channel,
                                            "sku":          _ls_sku,
                                            "title":        _ls_title or _ls_sku,
                                            "sale_date":    str(_ls_date),
                                            "gross_revenue": _gross_final,
                                            "platform_fee": _fee_final,
                                            "net_proceeds": _ls_net,
                                            "quantity":     1,
                                            "sale_price":   _ls_net,
                                            "status":       "completed",
                                            "dedup_key":    _dedup_ls,
                                        }
                                        _res_ls = _pur_post("sales_records", _rec_ls)
                                        if _res_ls is not None:
                                            st.success(f"✅ Sale logged for {_ls_sku} — ${_ls_net:.2f} net via {_ls_channel}")
                                            st.rerun()
                                        else:
                                            st.error(f"Save failed: {_pur_last_error.get('msg','Unknown error')}")
                        elif lc_state and lot_cards_df is not None and not lot_cards_df.empty:
                            st.divider()
                            st.markdown(f"**📋 Active Listings CSV ({card_count_csv}) — not yet saved to Supabase**")
                            st.caption("Go to Import Cards tab → Save to Supabase to make this persistent.")
                            row_data = {"SKU": lot_cards_df[lc_state["sku_col"]].values}
                            if lc_state["title_col"]: row_data["Title"] = lot_cards_df[lc_state["title_col"]].values
                            if lc_state["price_col"]: row_data["Price ($)"] = lot_cards_df[lc_state["price_col"]].values
                            cfg = {}
                            if lc_state["price_col"]: cfg["Price ($)"] = st.column_config.NumberColumn(format="$%.2f")
                            st.dataframe(pd.DataFrame(row_data), use_container_width=True, hide_index=True, column_config=cfg)
                        elif not _inv_cards:
                            st.info("No inventory saved for this lot yet — use the **Import Cards** tab to upload and save cards.")

                st.markdown("**Edit a lot**")
                edit_opts = ["— select to edit —"] + [l["lot_prefix"] for l in lots_data]
                edit_sel  = st.selectbox("Select lot to edit", edit_opts, key="pur_edit_sel")
                if edit_sel != "— select to edit —":
                    el = next(l for l in lots_data if l["lot_prefix"] == edit_sel)
                    with st.form("pur_edit_lot"):
                        e1, e2, e3 = st.columns(3)
                        e_source   = e1.text_input("Source", value=el.get("source") or "")
                        e_aliases  = e1.text_input("Alias Prefixes", value=el.get("alias_prefixes") or "",
                                        help="Comma-separated alternate SKU prefixes that roll up to this lot (e.g. RBLOT_-_07)")
                        e_date     = e2.date_input("Purchase Date",
                                        value=date.fromisoformat(el["purchase_date"]) if el.get("purchase_date") else date.today())
                        e_cost     = e2.number_input("Total Cost Paid ($)", min_value=0.0, step=0.01,
                                        format="%.2f", value=float(el.get("total_cost") or 0))
                        e_count    = e2.number_input("Card Count in Lot", min_value=0, step=1,
                                        value=int(el.get("card_count") or 0))
                        e_notes    = e3.text_area("Notes", height=80, value=el.get("notes") or "")
                        edit_sub   = st.form_submit_button("💾 Save Changes", type="primary")
                    if edit_sub:
                        aliases_edit = ",".join([a.strip().upper() for a in e_aliases.split(",") if a.strip()]) or None
                        res = _pur_patch("purchase_lots", f"id=eq.{el['id']}", {
                            "source":          e_source.strip() or None,
                            "purchase_date":   str(e_date),
                            "total_cost":      float(e_cost),
                            "card_count":      int(e_count),
                            "notes":           e_notes.strip() or None,
                            "alias_prefixes":  aliases_edit,
                        })
                        if res is not None:
                            st.success(f"**{edit_sel}** updated.")
                            st.session_state.pop("pur_lots", None)
                            st.rerun()
                        else:
                            st.error(f"Update failed: {_pur_last_error.get('msg') or 'Unknown error'}")

                st.divider()
                del_opts = ["— select to delete —"] + [l["lot_prefix"] for l in lots_data]
                del_sel  = st.selectbox("Delete a lot", del_opts, key="pur_del_sel")
                if del_sel != "— select to delete —":
                    lot_id = next(l["id"] for l in lots_data if l["lot_prefix"] == del_sel)
                    if st.button(f"🗑️ Delete {del_sel}", key="pur_del_btn"):
                        _pur_delete("purchase_lots", f"id=eq.{lot_id}")
                        st.success(f"Deleted {del_sel}.")
                        st.session_state.pop("pur_lots", None)
                        st.rerun()

            st.divider()
            st.markdown("**SQL — Run once in Supabase SQL Editor**")
            st.code("""create table if not exists purchase_lots (
  id            bigint primary key generated always as identity,
  lot_prefix    text not null unique,
  source        text,
  purchase_date date,
  total_cost    numeric not null default 0,
  card_count    integer not null default 0,
  notes         text,
  created_at    timestamptz default now()
);

-- If the table already exists, add the card_count column:
alter table purchase_lots add column if not exists card_count integer not null default 0;
alter table purchase_lots add column if not exists alias_prefixes text;

-- Add SKU column to sales_records so sold cards link back to lots
alter table sales_records add column if not exists sku text;
create index if not exists idx_sr_sku on sales_records(sku);

-- lot_cards: persistent inventory per lot (run once)
create table if not exists lot_cards (
  id            bigint primary key generated always as identity,
  lot_prefix    text not null,
  sku           text not null,
  title         text,
  notes         text,
  created_at    timestamptz default now(),
  constraint lot_cards_prefix_sku_unique unique (lot_prefix, sku)
);
create index if not exists idx_lot_cards_prefix on lot_cards(lot_prefix);
create index if not exists idx_lot_cards_sku    on lot_cards(sku);

-- If you ran the old lot_cards SQL without the unique constraint, add it:
alter table lot_cards add constraint if not exists lot_cards_prefix_sku_unique unique (lot_prefix, sku);""", language="sql")

        # ── IMPORT CARDS ──────────────────────────────────────────────────────
        with pur_t2:
            st.markdown("### Import Cards to Lot")
            st.caption("Upload a Haystack or eBay active listings CSV — cards are saved to Supabase by lot prefix so the Lots tab can cross-check sold/outstanding against all channels.")

            # Show current Supabase inventory counts
            _existing_lc = _pur_get("lot_cards", "?select=lot_prefix&limit=10000")
            if _existing_lc:
                from collections import Counter as _Counter
                _lc_counts = _Counter(r.get("lot_prefix","") for r in _existing_lc)
                _lc_info = " · ".join(f"{p}: {n}" for p, n in sorted(_lc_counts.items()))
                st.info(f"📦 **Currently saved in Supabase:** {len(_existing_lc):,} cards across {len(_lc_counts)} lots — {_lc_info}")
            else:
                st.info("📦 No cards saved yet — upload a CSV below to get started.")

            pur_file = st.file_uploader("Haystack or eBay Active Listings CSV", type=["csv"], key="pur_import_file")
            if pur_file:
                try:
                    raw = pur_file.read().decode("utf-8-sig")
                    lines = raw.split("\n")
                    # Skip Haystack Info row if present
                    start = 0
                    if lines and lines[0].strip().startswith("Info,"):
                        start = 1
                    import io as _io
                    card_df_raw = pd.read_csv(_io.StringIO("\n".join(lines[start:])), encoding="utf-8-sig")
                    card_df_raw.columns = [c.strip() for c in card_df_raw.columns]

                    # Find SKU column
                    sku_col = None
                    for candidate in ["CustomLabel", "Custom label (SKU)", "Custom Label (SKU)", "Custom label"]:
                        if candidate in card_df_raw.columns:
                            sku_col = candidate
                            break

                    title_col = None
                    for candidate in ["*Title", "Title"]:
                        if candidate in card_df_raw.columns:
                            title_col = candidate
                            break

                    price_col = None
                    for candidate in ["*StartPrice", "Current price", "Start price"]:
                        if candidate in card_df_raw.columns:
                            price_col = candidate
                            break

                    if not sku_col:
                        st.error("Could not find a SKU column. Expected: `CustomLabel` (Haystack) or `Custom label (SKU)` (eBay).")
                    else:
                        card_df_raw = card_df_raw[card_df_raw[sku_col].astype(str).str.strip() != ""].copy()
                        card_df_raw["_prefix"] = card_df_raw[sku_col].apply(_pur_prefix)
                        card_df_raw = card_df_raw[card_df_raw["_prefix"] != ""]

                        # Build prefix → lot mapping including alias_prefixes
                        known_prefixes = {}
                        for l in lots_data:
                            known_prefixes[l["lot_prefix"].upper()] = l["lot_prefix"]
                            for alias in (l.get("alias_prefixes") or "").split(","):
                                alias = alias.strip().upper()
                                if alias:
                                    known_prefixes[alias] = l["lot_prefix"]
                        lot_card_count = {l["lot_prefix"].upper(): int(l.get("card_count") or 0) for l in lots_data}

                        matched   = card_df_raw["_prefix"].str.upper().isin(known_prefixes).sum()
                        unmatched = len(card_df_raw) - matched
                        c1, c2, c3 = st.columns(3)
                        c1.metric("Cards in CSV", len(card_df_raw))
                        c2.metric("✅ Matched to a lot", matched)
                        c3.metric("⚠️ No lot match", unmatched)

                        # Group by canonical lot prefix (aliases redirect to their parent lot)
                        card_df_raw["_canon_prefix"] = card_df_raw["_prefix"].apply(
                            lambda p: known_prefixes.get(p.upper(), p)
                        )
                        all_groups     = list(card_df_raw.groupby("_canon_prefix", sort=True))
                        known_groups   = [(pfx, grp) for pfx, grp in all_groups if pfx.upper() in {v.upper() for v in known_prefixes.values()}]
                        unknown_groups = [(pfx, grp) for pfx, grp in all_groups if pfx.upper() not in {v.upper() for v in known_prefixes.values()}]

                        if unknown_groups:
                            missing = [pfx for pfx, _ in unknown_groups]
                            st.warning(f"⚠️ These prefixes have no lot — add them in the Lots tab: `{'`, `'.join(missing)}`")

                        st.divider()

                        for pfx, grp in known_groups:
                            expected = lot_card_count.get(pfx.upper(), 0)
                            count    = len(grp)
                            diff     = count - expected if expected > 0 else None
                            badge    = ""
                            if diff is not None:
                                badge = f"  —  {count}/{expected} cards" + (" ✅" if diff == 0 else f" ({'+'if diff>0 else ''}{diff} vs expected)")
                            else:
                                badge = f"  —  {count} cards"
                            with st.expander(f"✅ **{pfx}**{badge}", expanded=(diff not in (None, 0))):
                                row_data = {"SKU": grp[sku_col].values}
                                if title_col: row_data["Title"] = grp[title_col].values
                                if price_col: row_data["Price ($)"] = grp[price_col].values
                                cfg = {}
                                if price_col: cfg["Price ($)"] = st.column_config.NumberColumn(format="$%.2f")
                                st.dataframe(pd.DataFrame(row_data), use_container_width=True, hide_index=True, column_config=cfg)

                        for pfx, grp in unknown_groups:
                            with st.expander(f"⚠️ **{pfx}** — {len(grp)} cards (no matching lot)", expanded=True):
                                row_data = {"SKU": grp[sku_col].values}
                                if title_col: row_data["Title"] = grp[title_col].values
                                if price_col: row_data["Price ($)"] = grp[price_col].values
                                st.dataframe(pd.DataFrame(row_data), use_container_width=True, hide_index=True)

                        # ── Save to Supabase ──────────────────────────────────────
                        st.divider()
                        st.markdown("**💾 Save Inventory to Supabase**")
                        st.caption(f"{matched:,} matched cards will be saved. Re-uploading is safe — existing cards update, new cards add.")
                        if known_groups:
                            if st.button("💾 Save matched cards to Supabase", type="primary", key="pur_save_lc_btn"):
                                rows_to_save = []
                                for pfx2, grp2 in known_groups:
                                    for _, row2 in grp2.iterrows():
                                        rows_to_save.append({
                                            "lot_prefix": pfx2,
                                            "sku":        str(row2[sku_col]).strip(),
                                            "title":      str(row2[title_col]).strip() if title_col else "",
                                        })
                                # Upsert in batches of 500
                                total_saved = 0
                                save_err = None
                                for i in range(0, len(rows_to_save), 500):
                                    batch = rows_to_save[i:i+500]
                                    n, err = _pur_upsert_lot_cards(batch)
                                    if err:
                                        save_err = err
                                        break
                                    total_saved += n
                                if save_err:
                                    st.error(f"Save failed: {save_err}")
                                else:
                                    st.success(f"✅ Saved {total_saved:,} cards to Supabase. Reload the Lots tab to see the cross-check view.")
                                    st.rerun()
                        else:
                            st.warning("No cards matched a lot — fix the lot prefixes first.")

                except Exception as e:
                    st.error(f"Error reading file: {e}")

        # ── P&L BY LOT ────────────────────────────────────────────────────────
        with pur_t3:
            st.markdown("### P&L by Lot")
            st.info("💡 **To see revenue here:** go to the **💰 Sales & P&L** tab → Import → 📦 eBay (upload your eBay sold transactions CSV) or 🃏 CollX. Once imported, sold cards with a matching SKU prefix will automatically roll up to their lot below.")

            if not lots_data:
                st.info("No lots yet — add them in the Lots tab first.")
            else:
                # Load sales records that have a SKU
                sales_with_sku = _pur_get("sales_records", "?select=sku,net_proceeds,gross_revenue,source,sale_date,title&sku=not.is.null&limit=5000")

                # Build alias map: any prefix (main or alias) → canonical lot_prefix upper
                pl_alias_map = {}
                for l in lots_data:
                    pl_alias_map[l["lot_prefix"].upper()] = l["lot_prefix"].upper()
                    for alias in (l.get("alias_prefixes") or "").split(","):
                        alias = alias.strip().upper()
                        if alias:
                            pl_alias_map[alias] = l["lot_prefix"].upper()

                # Group sales by lot prefix (resolving aliases)
                prefix_revenue = {}
                for s in sales_with_sku:
                    prefix = _pur_prefix(s.get("sku", ""))
                    if prefix:
                        canon = pl_alias_map.get(prefix.upper(), prefix.upper())
                        if canon not in prefix_revenue:
                            prefix_revenue[canon] = {"net": 0.0, "gross": 0.0, "count": 0, "sales": []}
                        prefix_revenue[canon]["net"]   += float(s.get("net_proceeds") or 0)
                        prefix_revenue[canon]["gross"] += float(s.get("gross_revenue") or 0)
                        prefix_revenue[canon]["count"] += 1
                        prefix_revenue[canon]["sales"].append(s)

                pl_rows = []
                for lot in lots_data:
                    pfx       = lot["lot_prefix"].upper()
                    cost      = float(lot["total_cost"] or 0)
                    expected  = int(lot.get("card_count") or 0)
                    rev       = prefix_revenue.get(pfx, {})
                    net       = rev.get("net", 0.0)
                    gross     = rev.get("gross", 0.0)
                    sold      = rev.get("count", 0)
                    remaining = max(0, expected - sold) if expected > 0 else None
                    pl        = round(net - cost, 2)
                    turn_pct  = round(sold / expected * 100, 1) if expected > 0 else None
                    cost_per  = round(cost / expected, 2) if expected > 0 else None
                    net_per   = round(net / sold, 2) if sold > 0 else None
                    # Projected total revenue extrapolating from cards already sold
                    proj_rev  = round(net_per * expected, 2) if (net_per and expected > 0) else None
                    proj_pl   = round(proj_rev - cost, 2) if proj_rev is not None else None

                    pl_rows.append({
                        "Lot":             lot["lot_prefix"],
                        "Source":          lot.get("source") or "—",
                        "Date":            (lot.get("purchase_date") or "—")[:10],
                        "Cost ($)":        cost,
                        "Cards":           expected if expected > 0 else "—",
                        "Sold":            sold,
                        "Left":            remaining if remaining is not None else "—",
                        "Turn %":          turn_pct if turn_pct is not None else "—",
                        "Gross Rev ($)":   round(gross, 2),
                        "Net Rev ($)":     round(net, 2),
                        "P&L ($)":         pl,
                        "$/Card Cost":     cost_per if cost_per else "—",
                        "$/Card Net":      net_per if net_per else "—",
                        "Proj Net Rev ($)":proj_rev if proj_rev else "—",
                        "Proj P&L ($)":    proj_pl if proj_pl is not None else "—",
                        "Status":          "✅ Profit" if pl > 0 else ("🔴 Loss" if pl < 0 else "— Even"),
                    })

                pl_df = pd.DataFrame(pl_rows)
                total_cost  = sum(r["Cost ($)"] for r in pl_rows)
                total_gross = sum(r["Gross Rev ($)"] for r in pl_rows)
                total_net   = sum(r["Net Rev ($)"] for r in pl_rows)
                total_pl    = round(total_net - total_cost, 2)
                total_sold  = sum(r["Sold"] for r in pl_rows)
                total_cards = sum(r["Cards"] if isinstance(r["Cards"], int) else 0 for r in pl_rows)
                total_turn  = round(total_sold / total_cards * 100, 1) if total_cards > 0 else 0

                m1, m2, m3, m4, m5 = st.columns(5)
                m1.metric("Total Invested",  f"${total_cost:,.2f}")
                m2.metric("Gross Revenue",   f"${total_gross:,.2f}")
                m3.metric("Net Revenue",     f"${total_net:,.2f}")
                m4.metric("Overall P&L",     f"${total_pl:+,.2f}")
                m5.metric("Inventory Turn",  f"{total_turn:.1f}%",
                          help="% of all cards across all lots that have sold")

                st.divider()
                st.dataframe(
                    pl_df,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Cost ($)":         st.column_config.NumberColumn(format="$%.2f"),
                        "Gross Rev ($)":    st.column_config.NumberColumn(format="$%.2f"),
                        "Net Rev ($)":      st.column_config.NumberColumn(format="$%.2f"),
                        "P&L ($)":          st.column_config.NumberColumn(format="$%.2f"),
                        "$/Card Cost":      st.column_config.NumberColumn(format="$%.2f"),
                        "$/Card Net":       st.column_config.NumberColumn(format="$%.2f"),
                        "Proj Net Rev ($)": st.column_config.NumberColumn(format="$%.2f"),
                        "Proj P&L ($)":     st.column_config.NumberColumn(format="$%.2f"),
                        "Turn %":           st.column_config.NumberColumn(format="%.1f%%"),
                    },
                )

                # Drill-down: click a lot to see its sold cards
                st.divider()
                st.markdown("**Drill into a lot**")
                lot_opts = ["— select —"] + [l["lot_prefix"] for l in lots_data]
                sel_lot  = st.selectbox("Lot", lot_opts, key="pur_drill_sel")
                if sel_lot != "— select —":
                    lot_sales = prefix_revenue.get(sel_lot.upper(), {}).get("sales", [])
                    if not lot_sales:
                        st.info("No sales recorded for this lot yet.")
                    else:
                        drill_rows = [{
                            "Date":   s.get("sale_date", "")[:10],
                            "Title":  s.get("title") or "",
                            "SKU":    s.get("sku") or "",
                            "Channel": s.get("source") or "",
                            "Net ($)": float(s.get("net_proceeds") or 0),
                        } for s in sorted(lot_sales, key=lambda x: x.get("sale_date",""), reverse=True)]
                        st.dataframe(
                            pd.DataFrame(drill_rows),
                            use_container_width=True,
                            hide_index=True,
                            column_config={"Net ($)": st.column_config.NumberColumn(format="$%.2f")},
                        )

        # ── INDIVIDUAL CARDS ──────────────────────────────────────────────────
        with pur_t4:
            st.markdown("### 🃏 Individual Card Purchases")
            st.caption("Track small buys (Whatnot, Facebook, local) that don't need a full lot. Enter a SKU matching your eBay listing — the app auto-links to the sale when it happens.")

            if "cp_data" not in st.session_state:
                st.session_state["cp_data"] = _pur_get("card_purchases", "?order=purchase_date.desc&limit=1000")
            cp_data = st.session_state["cp_data"]

            # ── Add new purchase ──────────────────────────────────────────────
            with st.expander("➕ Add Purchase", expanded=not cp_data):
                with st.form("cp_add_form"):
                    ca1, ca2, ca3 = st.columns(3)
                    cp_sku    = ca1.text_input("SKU", placeholder="WN-073026-00001",
                                    help="Must match the SKU on your eBay listing exactly")
                    cp_title  = ca1.text_input("Description", placeholder="2023 Bowman Chrome Corbin Carroll")
                    cp_source = ca2.selectbox("Source", ["Whatnot", "Facebook", "Local", "eBay", "Other"])
                    cp_date   = ca2.date_input("Purchase Date", value=date.today())
                    cp_cost   = ca2.number_input("Cost Paid ($)", min_value=0.0, step=0.01, format="%.2f")
                    cp_qty    = ca2.number_input("Qty", min_value=1, step=1, value=1,
                                    help="Number of cards in this purchase")
                    cp_notes  = ca3.text_area("Notes", height=80, placeholder="e.g. Auto /25, BGS 9.5")
                    cp_sub    = st.form_submit_button("💾 Add Purchase", type="primary")
                if cp_sub:
                    if not cp_sku.strip():
                        st.error("SKU is required.")
                    elif cp_cost <= 0:
                        st.error("Cost must be greater than $0.")
                    else:
                        res = _pur_post("card_purchases", {
                            "sku":           cp_sku.strip(),
                            "title":         cp_title.strip() or None,
                            "source":        cp_source,
                            "purchase_date": str(cp_date),
                            "cost_paid":     float(cp_cost),
                            "quantity":      int(cp_qty),
                            "notes":         cp_notes.strip() or None,
                        })
                        if res is not None:
                            st.success(f"Added: {cp_sku.strip()}")
                            st.session_state.pop("cp_data", None)
                            st.rerun()
                        else:
                            st.error(f"Save failed: {_pur_last_error.get('msg') or 'Unknown error'}")

            if not cp_data:
                st.info("No individual purchases yet — add one above.")
            else:
                # Load all sales to cross-reference by SKU
                cp_sales_raw = _pur_get("sales_records", "?select=sku,title,sale_date,gross_revenue,net_proceeds,source&sku=not.is.null&limit=5000")
                cp_sales_by_sku = {}
                for _cs in cp_sales_raw:
                    _csk = (_cs.get("sku") or "").strip()
                    if _csk:
                        if _csk not in cp_sales_by_sku:
                            cp_sales_by_sku[_csk] = []
                        cp_sales_by_sku[_csk].append(_cs)

                # Summary metrics
                total_cp_cost = sum(float(c.get("cost_paid") or 0) for c in cp_data)
                total_cp_net  = 0.0
                total_cp_sold = 0
                for c in cp_data:
                    for _s in cp_sales_by_sku.get((c.get("sku") or "").strip(), []):
                        total_cp_net  += float(_s.get("net_proceeds") or 0)
                        total_cp_sold += 1
                total_cp_pl   = round(total_cp_net - total_cp_cost, 2)
                win_count = 0
                for c in cp_data:
                    sks = cp_sales_by_sku.get((c.get("sku") or "").strip(), [])
                    if sks:
                        net = sum(float(_s.get("net_proceeds") or 0) for _s in sks)
                        if net > float(c.get("cost_paid") or 0):
                            win_count += 1
                win_rate = round(win_count / max(1, sum(1 for c in cp_data if cp_sales_by_sku.get((c.get("sku") or "").strip()))) * 100) if total_cp_sold > 0 else None

                sm1, sm2, sm3, sm4, sm5 = st.columns(5)
                sm1.metric("Total Invested", f"${total_cp_cost:,.2f}")
                sm2.metric("Net Recovered",  f"${total_cp_net:,.2f}")
                sm3.metric("Overall P&L",    f"${total_cp_pl:+,.2f}")
                sm4.metric("Cards Sold",     f"{total_cp_sold} / {len(cp_data)}")
                sm5.metric("Win Rate",       f"{win_rate}%" if win_rate is not None else "—",
                           help="% of sold cards where net > cost")

                st.divider()

                # Filter controls
                cf1, cf2, cf3 = st.columns(3)
                _cp_src_opts = ["All Sources"] + sorted(set(c.get("source","") for c in cp_data if c.get("source")))
                _cp_src = cf1.selectbox("Source", _cp_src_opts, key="cp_src_filter")
                _cp_status = cf2.selectbox("Status", ["All", "✅ Sold", "🟡 Unsold"], key="cp_status_filter")
                _cp_result = cf3.selectbox("Result", ["All", "🟢 Profit", "🔴 Loss", "⏳ Pending"], key="cp_result_filter")

                # Build display rows
                cp_rows = []
                for c in cp_data:
                    sku       = (c.get("sku") or "").strip()
                    cost      = float(c.get("cost_paid") or 0)
                    qty       = int(c.get("quantity") or 1)
                    sales     = cp_sales_by_sku.get(sku, [])
                    net_total = sum(float(_s.get("net_proceeds") or 0) for _s in sales)
                    gross_total = sum(float(_s.get("gross_revenue") or 0) for _s in sales)
                    pl        = round(net_total - cost, 2)
                    is_sold   = len(sales) > 0
                    roi       = round(pl / cost * 100, 1) if cost > 0 and is_sold else None
                    status    = "✅ Sold" if is_sold else "🟡 Unsold"
                    result    = ("🟢 Profit" if pl > 0 else "🔴 Loss") if is_sold else "⏳ Pending"
                    sale_date = max((s.get("sale_date","")[:10] for s in sales), default="—") if sales else "—"
                    sale_chan  = ", ".join(sorted(set(s.get("source","") for s in sales))) if sales else "—"
                    days_held = None
                    if c.get("purchase_date") and sale_date != "—":
                        try:
                            from datetime import datetime as _dtt
                            days_held = (_dtt.strptime(sale_date, "%Y-%m-%d") - _dtt.strptime(c["purchase_date"][:10], "%Y-%m-%d")).days
                        except Exception:
                            pass

                    # Apply filters
                    if _cp_src != "All Sources" and c.get("source") != _cp_src:
                        continue
                    if _cp_status == "✅ Sold" and not is_sold:
                        continue
                    if _cp_status == "🟡 Unsold" and is_sold:
                        continue
                    if _cp_result == "🟢 Profit" and result != "🟢 Profit":
                        continue
                    if _cp_result == "🔴 Loss" and result != "🔴 Loss":
                        continue
                    if _cp_result == "⏳ Pending" and result != "⏳ Pending":
                        continue

                    cp_rows.append({
                        "_id":       c["id"],
                        "SKU":       sku,
                        "Description": c.get("title") or "—",
                        "Source":    c.get("source") or "—",
                        "Bought":    (c.get("purchase_date") or "—")[:10],
                        "Cost ($)":  cost,
                        "Qty":       qty,
                        "Gross ($)": round(gross_total, 2) if is_sold else None,
                        "Net ($)":   round(net_total, 2) if is_sold else None,
                        "P&L ($)":   pl if is_sold else None,
                        "ROI %":     roi,
                        "Days Held": days_held,
                        "Sold Via":  sale_chan,
                        "Status":    status,
                    })

                if not cp_rows:
                    st.info("No purchases match the current filters.")
                else:
                    cp_df = pd.DataFrame(cp_rows).drop(columns=["_id"])
                    st.dataframe(cp_df, use_container_width=True, hide_index=True,
                        column_config={
                            "Cost ($)":  st.column_config.NumberColumn(format="$%.2f"),
                            "Gross ($)": st.column_config.NumberColumn(format="$%.2f"),
                            "Net ($)":   st.column_config.NumberColumn(format="$%.2f"),
                            "P&L ($)":   st.column_config.NumberColumn(format="$%.2f"),
                            "ROI %":     st.column_config.NumberColumn(format="%.1f%%"),
                            "Days Held": st.column_config.NumberColumn(format="%d days"),
                        })

                    # Delete
                    st.divider()
                    del_sku = st.selectbox("Delete a purchase", ["— select —"] + [r["SKU"] for r in cp_rows], key="cp_del_sel")
                    if del_sku != "— select —":
                        match = next((c for c in cp_data if (c.get("sku") or "").strip() == del_sku), None)
                        if match and st.button(f"🗑️ Delete {del_sku}", key="cp_del_btn"):
                            _pur_delete("card_purchases", f"id=eq.{match['id']}")
                            st.session_state.pop("cp_data", None)
                            st.rerun()

            st.divider()
            st.markdown("**SQL — Run once in Supabase SQL Editor**")
            st.code("""create table if not exists card_purchases (
  id            bigint primary key generated always as identity,
  sku           text not null,
  title         text,
  source        text,
  purchase_date date,
  cost_paid     numeric not null default 0,
  quantity      integer not null default 1,
  notes         text,
  created_at    timestamptz default now()
);
create index if not exists idx_cp_sku on card_purchases(sku);""", language="sql")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 10 — Sales & P&L
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 10:
    st.markdown("## 💰 Sales & P&L")
    st.caption("Import eBay and CollX sales exports — track gross revenue, platform fees, and net proceeds across all channels.")

    if not SUPABASE_URL:
        st.warning("Supabase not connected. Configure in sidebar to enable Sales & P&L.")
    else:
        # ── Helpers ───────────────────────────────────────────────────────────
        def _sal_get(params=""):
            url = f"{SUPABASE_URL}/rest/v1/sales_records{params}"
            req = urllib.request.Request(url, headers=sb_headers())
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
                    return json.loads(r.read().decode())
            except Exception:
                return []

        _sal_last_error = {"msg": None}

        def _sal_post(payload, prefer="return=minimal"):
            data = json.dumps(payload).encode()
            req = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/sales_records",
                data=data,
                headers={**sb_headers(), "Prefer": prefer},
                method="POST",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10) as r:
                    body = r.read()
                    return json.loads(body.decode()) if body else []
            except urllib.error.HTTPError as e:
                body = e.read().decode("utf-8", errors="replace")
                _sal_last_error["msg"] = f"HTTP {e.code}: {body[:300]}"
                return None
            except Exception as ex:
                _sal_last_error["msg"] = str(ex)
                return None

        def _sal_delete(filt):
            req = urllib.request.Request(
                f"{SUPABASE_URL}/rest/v1/sales_records?{filt}",
                headers=sb_headers(),
                method="DELETE",
            )
            try:
                with urllib.request.urlopen(req, context=ssl_ctx(), timeout=10):
                    return True
            except Exception:
                return False

        def _parse_money(val):
            if val is None or str(val).strip() in ("", "nan"):
                return 0.0
            try:
                return float(str(val).replace("$", "").replace(",", "").strip())
            except ValueError:
                return 0.0

        def _parse_ebay_date(s):
            s = str(s).strip()
            for fmt in ("%b-%d-%y", "%B %d, %Y", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    pass
            return s[:10] if len(s) >= 10 else s

        def _parse_collx_date(s):
            s = str(s).strip()
            for fmt in ("%b %d, %Y", "%B %d, %Y", "%Y-%m-%d", "%m/%d/%Y"):
                try:
                    return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
                except ValueError:
                    pass
            return s[:10] if len(s) >= 10 else s

        # Load all sales records
        if "sal_data" not in st.session_state:
            with st.spinner("Loading sales data…"):
                st.session_state["sal_data"] = _sal_get("?order=sale_date.desc&limit=5000")

        all_sales = st.session_state.get("sal_data", [])

        # ── Snapshot ──────────────────────────────────────────────────────────
        ebay_sales = [s for s in all_sales if s.get("source") == "ebay"]
        collx_sales = [s for s in all_sales if s.get("source") == "collx"]
        dc_sales = [s for s in all_sales if s.get("source") == "dc_sports"]

        total_gross = sum(_parse_money(s.get("gross_revenue")) for s in all_sales)
        total_fees = sum(_parse_money(s.get("platform_fee")) for s in all_sales)
        total_net = sum(_parse_money(s.get("net_proceeds")) for s in all_sales)
        total_txns = len(all_sales)

        sn1, sn2, sn3, sn4, sn5 = st.columns(5)
        sn1.metric("Total Transactions", f"{total_txns:,}")
        sn2.metric("Gross Revenue", f"${total_gross:,.2f}")
        sn3.metric("Platform Fees", f"${total_fees:,.2f}")
        sn4.metric("Net Proceeds", f"${total_net:,.2f}")
        margin = round((total_net / total_gross * 100), 1) if total_gross else 0
        sn5.metric("Keep Rate", f"{margin}%")

        # Per-platform breakdown
        if all_sales:
            st.divider()
            pc1, pc2, pc3 = st.columns(3)
            with pc1:
                eb_gross = sum(_parse_money(s.get("gross_revenue")) for s in ebay_sales)
                eb_net = sum(_parse_money(s.get("net_proceeds")) for s in ebay_sales)
                st.markdown(f"**eBay** — {len(ebay_sales)} sales")
                st.markdown(f"Gross: **${eb_gross:,.2f}** · Net: **${eb_net:,.2f}**")
            with pc2:
                cx_gross = sum(_parse_money(s.get("gross_revenue")) for s in collx_sales)
                cx_net = sum(_parse_money(s.get("net_proceeds")) for s in collx_sales)
                st.markdown(f"**CollX** — {len(collx_sales)} sales")
                st.markdown(f"Gross: **${cx_gross:,.2f}** · Net: **${cx_net:,.2f}**")
            with pc3:
                dc_gross = sum(_parse_money(s.get("gross_revenue")) for s in dc_sales)
                dc_net = sum(_parse_money(s.get("net_proceeds")) for s in dc_sales)
                st.markdown(f"**DC Sports** — {len(dc_sales)} sales")
                st.markdown(f"Gross: **${dc_gross:,.2f}** · Net: **${dc_net:,.2f}**")

        st.divider()

        sal_t1, sal_t2, sal_t3 = st.tabs(["📊 P&L by Month", "📋 Sales Log", "⬆️ Import"])

        # ── P&L BY MONTH ──────────────────────────────────────────────────────
        with sal_t1:
            if not all_sales:
                st.info("No sales yet. Import eBay or CollX data in the Import tab.")
            else:
                year_opts = sorted(set(
                    str(s.get("sale_date") or "")[:4]
                    for s in all_sales
                    if str(s.get("sale_date") or "")[:4].isdigit()
                ), reverse=True)
                sel_year = st.selectbox("Year", year_opts, key="sal_year")

                from collections import defaultdict
                monthly = defaultdict(lambda: {"gross": 0.0, "fees": 0.0, "net": 0.0, "txns": 0})
                for s in all_sales:
                    d = str(s.get("sale_date") or "")
                    if not d[:4] == sel_year:
                        continue
                    month_key = d[:7]
                    monthly[month_key]["gross"] += _parse_money(s.get("gross_revenue"))
                    monthly[month_key]["fees"] += _parse_money(s.get("platform_fee"))
                    monthly[month_key]["net"] += _parse_money(s.get("net_proceeds"))
                    monthly[month_key]["txns"] += 1

                if not monthly:
                    st.info(f"No sales for {sel_year}.")
                else:
                    month_names = {
                        "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
                        "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
                        "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
                    }
                    rows_m = []
                    ytd_gross = ytd_fees = ytd_net = ytd_txns = 0
                    for mk in sorted(monthly.keys()):
                        m = monthly[mk]
                        mn = month_names.get(mk[5:7], mk[5:7])
                        rows_m.append({
                            "Month": f"{mn} {mk[:4]}",
                            "Sales": m["txns"],
                            "Gross ($)": round(m["gross"], 2),
                            "Fees ($)": round(m["fees"], 2),
                            "Net ($)": round(m["net"], 2),
                            "Keep %": round(m["net"] / m["gross"] * 100, 1) if m["gross"] else 0,
                        })
                        ytd_gross += m["gross"]; ytd_fees += m["fees"]
                        ytd_net += m["net"]; ytd_txns += m["txns"]

                    # YTD totals row
                    rows_m.append({
                        "Month": f"— YTD {sel_year} —",
                        "Sales": ytd_txns,
                        "Gross ($)": round(ytd_gross, 2),
                        "Fees ($)": round(ytd_fees, 2),
                        "Net ($)": round(ytd_net, 2),
                        "Keep %": round(ytd_net / ytd_gross * 100, 1) if ytd_gross else 0,
                    })

                    st.dataframe(
                        pd.DataFrame(rows_m),
                        use_container_width=True,
                        hide_index=True,
                        column_config={
                            "Gross ($)": st.column_config.NumberColumn(format="$%.2f"),
                            "Fees ($)": st.column_config.NumberColumn(format="$%.2f"),
                            "Net ($)": st.column_config.NumberColumn(format="$%.2f"),
                            "Keep %": st.column_config.NumberColumn(format="%.1f%%"),
                        },
                    )

                    # Channel mix for the year
                    st.markdown("**Channel Mix**")
                    ch_rows = []
                    for src, label in [("ebay", "eBay"), ("collx", "CollX"), ("dc_sports", "DC Sports")]:
                        src_sales = [s for s in all_sales if s.get("source") == src and str(s.get("sale_date") or "")[:4] == sel_year]
                        if src_sales:
                            sg = sum(_parse_money(s.get("gross_revenue")) for s in src_sales)
                            sn_ = sum(_parse_money(s.get("net_proceeds")) for s in src_sales)
                            sf = sum(_parse_money(s.get("platform_fee")) for s in src_sales)
                            ch_rows.append({"Channel": label, "Sales": len(src_sales), "Gross ($)": round(sg, 2), "Fees ($)": round(sf, 2), "Net ($)": round(sn_, 2)})
                    if ch_rows:
                        st.dataframe(
                            pd.DataFrame(ch_rows),
                            use_container_width=True,
                            hide_index=True,
                            column_config={
                                "Gross ($)": st.column_config.NumberColumn(format="$%.2f"),
                                "Fees ($)": st.column_config.NumberColumn(format="$%.2f"),
                                "Net ($)": st.column_config.NumberColumn(format="$%.2f"),
                            },
                        )

        # ── SALES LOG ─────────────────────────────────────────────────────────
        with sal_t2:
            if not all_sales:
                st.info("No sales yet.")
            else:
                lf1, lf2, lf3 = st.columns(3)
                src_opts = ["All"] + sorted(set(s.get("source", "") for s in all_sales))
                src_filt = lf1.selectbox("Platform", src_opts, key="sal_log_src",
                                          format_func=lambda x: {"ebay": "eBay", "collx": "CollX", "dc_sports": "DC Sports"}.get(x, x))
                status_opts2 = ["All"] + sorted(set(s.get("status", "") or "" for s in all_sales if s.get("status")))
                status_filt2 = lf2.selectbox("Status", status_opts2, key="sal_log_status")
                year_opts2 = ["All"] + sorted(set(str(s.get("sale_date") or "")[:4] for s in all_sales if str(s.get("sale_date") or "")[:4].isdigit()), reverse=True)
                year_filt2 = lf3.selectbox("Year", year_opts2, key="sal_log_year")

                log_filtered = all_sales
                if src_filt != "All":
                    log_filtered = [s for s in log_filtered if s.get("source") == src_filt]
                if status_filt2 != "All":
                    log_filtered = [s for s in log_filtered if s.get("status") == status_filt2]
                if year_filt2 != "All":
                    log_filtered = [s for s in log_filtered if str(s.get("sale_date") or "")[:4] == year_filt2]

                st.caption(f"Showing {len(log_filtered):,} of {len(all_sales):,} sales")

                src_label = {"ebay": "eBay", "collx": "CollX", "dc_sports": "DC Sports"}
                log_rows = []
                for s in log_filtered[:500]:
                    log_rows.append({
                        "Date": str(s.get("sale_date") or "")[:10],
                        "Platform": src_label.get(s.get("source", ""), s.get("source", "")),
                        "Title": s.get("title") or "",
                        "Qty": s.get("quantity") or 1,
                        "Gross ($)": _parse_money(s.get("gross_revenue")),
                        "Fee ($)": _parse_money(s.get("platform_fee")),
                        "Net ($)": _parse_money(s.get("net_proceeds")),
                        "Status": s.get("status") or "",
                    })
                if len(log_filtered) > 500:
                    st.caption("Showing first 500 rows — filter by year or platform to see more.")

                st.dataframe(
                    pd.DataFrame(log_rows),
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Gross ($)": st.column_config.NumberColumn(format="$%.2f"),
                        "Fee ($)": st.column_config.NumberColumn(format="$%.2f"),
                        "Net ($)": st.column_config.NumberColumn(format="$%.2f"),
                    },
                )

                # Export
                if log_rows:
                    csv_out = pd.DataFrame(log_rows).to_csv(index=False).encode()
                    st.download_button("⬇️ Export filtered to CSV", data=csv_out,
                                       file_name=f"sales_export_{date.today()}.csv",
                                       mime="text/csv", key="sal_export")

                st.divider()
                with st.expander("🗑️ Delete sales by platform (re-import to refresh)"):
                    del_src = st.selectbox("Delete all records from", ["— select —", "eBay", "CollX", "DC Sports"], key="sal_del_src")
                    src_map = {"eBay": "ebay", "CollX": "collx", "DC Sports": "dc_sports"}
                    if del_src != "— select —":
                        st.warning(f"This permanently deletes all {del_src} sales records from Supabase. Re-import the CSV to restore.")
                        if st.button(f"Delete all {del_src} records", type="primary", key="sal_del_btn"):
                            _sal_delete(f"source=eq.{src_map[del_src]}")
                            st.success(f"Deleted all {del_src} records.")
                            st.session_state.pop("sal_data", None)
                            st.rerun()

        # ── IMPORT ────────────────────────────────────────────────────────────
        with sal_t3:
            imp_ebay, imp_collx, imp_whatnot, imp_dc, imp_manual = st.tabs(["📦 eBay", "🃏 CollX", "🎥 Whatnot", "🏷️ DC Sports", "✏️ Manual"])

            # ── eBay import ───────────────────────────────────────────────────
            with imp_ebay:
                st.markdown("### Import eBay Sales")
                st.info(
                    "✅ **Recommended: Orders Report** — Seller Hub → Orders → Sold → Download Report → **Orders report**\n\n"
                    "Uses actual Final Value Fees and includes SKU/lot data for automatic lot matching. **This is the best file to upload.**"
                )
                st.caption(
                    "Also accepts two other formats (auto-detected on upload):\n\n"
                    "• **Seller Hub Transactions Report** — Seller Hub → Reports → Downloads → Transactions report. "
                    "Fees are estimated at 12.35% + $0.40 (sales over $10) or $0.30 (under $10). No lot matching.\n\n"
                    "• **eBay Financial Ledger** — Payments → Transactions → Download (TransactionType / GrossAmount / Fees / NetAmount columns). "
                    "Uses actual eBay fees but has no SKU data — no lot matching."
                )
                ebay_file = st.file_uploader("eBay Orders / Transactions / Ledger CSV", type=["csv"], key="sal_ebay_file")
                if ebay_file:
                    try:
                        raw_content = ebay_file.read().decode("utf-8-sig")

                        # Auto-detect format — skip blank preamble rows first
                        lines = raw_content.split("\n")
                        header_idx = 0
                        for i, line in enumerate(lines[:5]):
                            if line.strip().strip(","):
                                header_idx = i
                                break
                        header_line = lines[header_idx]
                        is_ledger = "TransactionType" in header_line and "GrossAmount" in header_line
                        is_orders = ("Order Number" in header_line or "Order number" in header_line) and ("Item Number" in header_line or "Sales Record" in header_line)

                        if is_ledger:
                            ebay_df = pd.read_csv(io.StringIO(raw_content))
                            ebay_df.columns = [c.strip() for c in ebay_df.columns]
                            # Keep only Sale rows; drop totals/blank rows
                            ebay_df = ebay_df[ebay_df["TransactionType"].astype(str).str.strip() == "Sale"].copy()
                            st.info(f"📊 **eBay Financial Ledger detected** — using actual fees from your ledger.")
                            st.caption(f"Found **{len(ebay_df):,}** Sale rows. Preview:")
                        elif is_orders:
                            csv_content = "\n".join(lines[header_idx:])
                            ebay_df = pd.read_csv(io.StringIO(csv_content))
                            ebay_df.columns = [c.strip() for c in ebay_df.columns]
                            ebay_df = ebay_df.dropna(how="all")
                            ebay_df = ebay_df[ebay_df.get("Order Number", ebay_df.iloc[:, 0]).astype(str).str.strip().str.len() > 3]
                            st.info(f"📦 **eBay Orders Report detected** — {len(ebay_df):,} orders with SKU/lot tagging.")
                            st.caption(f"Found **{len(ebay_df):,}** orders. Preview:")
                        else:
                            header_idx2 = 0
                            for i, line in enumerate(lines[:5]):
                                if "Sale Date" in line or "Item Title" in line or "Sales Record" in line:
                                    header_idx2 = i
                                    break
                            csv_content = "\n".join(lines[header_idx2:])
                            ebay_df = pd.read_csv(io.StringIO(csv_content), encoding="utf-8-sig")
                            ebay_df.columns = [c.strip() for c in ebay_df.columns]
                            ebay_df = ebay_df.dropna(how="all")
                            ebay_df = ebay_df[ebay_df.get("Sale Date", ebay_df.iloc[:, 0]).astype(str).str.strip() != ""]
                            st.caption(f"Found **{len(ebay_df):,}** rows. Preview:")

                        st.dataframe(ebay_df.head(5), use_container_width=True, hide_index=True)

                        if st.button("⬆️ Import eBay Sales", type="primary", key="sal_ebay_btn"):
                            existing_raw3 = _sal_get("?select=dedup_key&source=eq.ebay")
                            existing_keys3 = {r["dedup_key"] for r in existing_raw3 if r.get("dedup_key")}

                            imported3 = skipped3 = failed3 = 0
                            prog3 = st.progress(0)
                            total3 = len(ebay_df)

                            for _i3, (idx3, row3) in enumerate(ebay_df.iterrows()):
                                prog3.progress(int((_i3 + 1) / max(total3, 1) * 100))

                                if is_ledger:
                                    title3  = str(row3.get("Description", "") or "").strip()
                                    gross3  = _parse_money(row3.get("GrossAmount", 0))
                                    fee3    = abs(_parse_money(row3.get("Fees", 0)))
                                    net3    = _parse_money(row3.get("NetAmount", 0))
                                    dedup3  = f"ebay_ledger|{title3}|{round(gross3, 2)}"
                                    rec3 = {
                                        "source": "ebay",
                                        "sale_date": None,
                                        "title": title3 or None,
                                        "gross_revenue": gross3,
                                        "platform_fee": fee3,
                                        "net_proceeds": net3,
                                        "quantity": 1,
                                        "sale_price": gross3,
                                        "shipping_collected": 0,
                                        "status": "sold",
                                        "dedup_key": dedup3,
                                    }
                                elif is_orders:
                                    # eBay Orders Report — columns use Title Case
                                    order_num3 = str(row3.get("Order Number", "") or "").strip()
                                    item_num3  = str(row3.get("Item Number", "") or "").strip()
                                    title3     = str(row3.get("Item Title", "") or "").strip()
                                    sku3       = str(row3.get("Custom Label", "") or row3.get("Custom label", "") or "").strip()
                                    sale_date3 = _parse_ebay_date(
                                        row3.get("Sale Date", "") or row3.get("Paid On Date", "") or ""
                                    )
                                    sold_for3  = _parse_money(row3.get("Sold For", 0))
                                    shipping3  = _parse_money(row3.get("Shipping And Handling", 0))
                                    _qty3_raw  = row3.get("Quantity", 1)
                                    try:
                                        qty3 = int(float(_qty3_raw)) if _qty3_raw == _qty3_raw and _qty3_raw not in (None, "") else 1
                                    except (ValueError, TypeError):
                                        qty3 = 1
                                    gross3 = round(sold_for3 * qty3 + shipping3, 2)
                                    _per_item_fee3 = 0.40 if gross3 > 10 else 0.30
                                    fee3 = round(gross3 * 0.1235 + _per_item_fee3, 2)
                                    net3  = round(gross3 - fee3, 2)
                                    dedup3 = f"ebay_orders|{order_num3}|{item_num3}"
                                    rec3 = {
                                        "source": "ebay",
                                        "order_id": order_num3 or None,
                                        "sale_date": sale_date3 or None,
                                        "title": title3 or None,
                                        "item_number": item_num3 or None,
                                        "sku": sku3 or None,
                                        "quantity": qty3,
                                        "sale_price": sold_for3,
                                        "shipping_collected": shipping3,
                                        "gross_revenue": gross3,
                                        "platform_fee": fee3,
                                        "net_proceeds": net3,
                                        "status": "sold",
                                        "dedup_key": dedup3,
                                    }
                                else:
                                    sale_date3 = _parse_ebay_date(row3.get("Sale Date", ""))
                                    item_num3  = str(row3.get("Item Number", "") or "").strip()
                                    sold_for3  = _parse_money(row3.get("Sold For", 0))
                                    shipping3  = _parse_money(row3.get("Shipping And Handling", 0))
                                    title3     = str(row3.get("Item Title", "") or "").strip()
                                    _qty3_raw  = row3.get("Quantity", 1)
                                    try:
                                        qty3 = int(float(_qty3_raw)) if _qty3_raw == _qty3_raw and _qty3_raw not in (None, "") else 1
                                    except (ValueError, TypeError):
                                        qty3 = 1
                                    gross3 = round(sold_for3 + shipping3, 2)
                                    _per_item_fee3 = 0.40 if gross3 > 10 else 0.30
                                    fee3   = round(gross3 * 0.1235 + _per_item_fee3, 2)
                                    net3   = round(gross3 - fee3, 2)
                                    dedup3 = f"ebay|{item_num3}|{sale_date3}|{round(sold_for3, 2)}"
                                    sku3   = str(row3.get("Custom label", "") or row3.get("Custom label (SKU)", "") or "").strip()
                                    rec3 = {
                                        "source": "ebay",
                                        "order_id": str(row3.get("Order Number", "") or "").strip() or None,
                                        "sale_date": sale_date3 or None,
                                        "title": title3 or None,
                                        "item_number": item_num3 or None,
                                        "sku": sku3 or None,
                                        "quantity": qty3,
                                        "sale_price": sold_for3,
                                        "shipping_collected": shipping3,
                                        "gross_revenue": gross3,
                                        "platform_fee": fee3,
                                        "net_proceeds": net3,
                                        "status": str(row3.get("Feedback Received", "") or "sold").strip() or "sold",
                                        "dedup_key": dedup3,
                                    }

                                if dedup3 in existing_keys3:
                                    skipped3 += 1
                                    continue

                                res3 = _sal_post(rec3)
                                if res3 is not None:
                                    imported3 += 1
                                    existing_keys3.add(dedup3)
                                else:
                                    failed3 += 1

                            prog3.empty()
                            msg3 = f"✅ Imported **{imported3}** eBay sales"
                            if skipped3:
                                msg3 += f", skipped **{skipped3}** duplicates"
                            if failed3:
                                msg3 += f", **{failed3}** failed"
                            st.success(msg3 + ".")
                            if not is_ledger:
                                st.caption("Note: eBay fees estimated at 12.35% + $0.40 (>$10 sales) or $0.30 (≤$10). For exact fees, use the eBay Financial Ledger format.")
                            if imported3:
                                st.session_state.pop("sal_data", None)
                                st.rerun()
                    except Exception as e3:
                        st.error(f"Error reading eBay CSV: {e3}")

            # ── CollX import ──────────────────────────────────────────────────
            with imp_collx:
                st.markdown("### Import CollX Orders Export")
                st.caption(
                    "CollX app → Orders → Export. "
                    "Columns: order_number, order_date, buyer_name, total_items, merchandise_value, gross_subtotal, net_proceeds, status. "
                    "Net proceeds are exact from CollX (fees already deducted). Re-importing is safe — duplicates skipped."
                )
                collx_file = st.file_uploader("CollX orders export CSV", type=["csv"], key="sal_collx_file")
                if collx_file:
                    try:
                        collx_df = pd.read_csv(collx_file, encoding="utf-8-sig")
                        collx_df.columns = [c.strip() for c in collx_df.columns]
                        st.caption(f"Found **{len(collx_df):,}** rows. Preview:")
                        st.dataframe(collx_df.head(5), use_container_width=True, hide_index=True)

                        date_range = ""
                        if "order_date" in collx_df.columns:
                            dates4 = collx_df["order_date"].dropna().tolist()
                            if dates4:
                                date_range = f" ({dates4[-1]} → {dates4[0]})"
                        st.caption(f"Date range detected{date_range}")

                        if st.button("⬆️ Import CollX Sales", type="primary", key="sal_collx_btn"):
                            existing_raw4 = _sal_get("?select=dedup_key&source=eq.collx")
                            existing_keys4 = {r["dedup_key"] for r in existing_raw4 if r.get("dedup_key")}

                            imported4 = skipped4 = failed4 = 0
                            prog4 = st.progress(0)
                            total4 = len(collx_df)

                            for idx4, row4 in collx_df.iterrows():
                                prog4.progress(int((idx4 + 1) / max(total4, 1) * 100))
                                order_num4 = str(row4.get("order_number", "") or "").strip()
                                order_date4 = _parse_collx_date(row4.get("order_date", ""))
                                gross4 = _parse_money(row4.get("gross_subtotal", 0))
                                net4 = _parse_money(row4.get("net_proceeds", 0))
                                merch4 = _parse_money(row4.get("merchandise_value", 0))
                                fee4 = round(gross4 - net4, 2)
                                qty4 = int(float(str(row4.get("total_items", 1) or 1)))
                                status4 = str(row4.get("status", "") or "").strip()
                                dedup4 = f"collx|{order_num4}|{order_date4}"

                                if status4.lower() not in ("completed", "complete", ""):
                                    skipped4 += 1
                                    continue

                                if dedup4 in existing_keys4:
                                    skipped4 += 1
                                    continue

                                rec4 = {
                                    "source": "collx",
                                    "order_id": order_num4 or None,
                                    "sale_date": order_date4 or None,
                                    "title": f"CollX order {order_num4} ({qty4} item{'s' if qty4 != 1 else ''})",
                                    "item_number": None,
                                    "quantity": qty4,
                                    "sale_price": merch4,
                                    "shipping_collected": round(gross4 - merch4, 2),
                                    "gross_revenue": gross4,
                                    "platform_fee": fee4,
                                    "net_proceeds": net4,
                                    "status": status4 or "completed",
                                    "dedup_key": dedup4,
                                }
                                res4 = _sal_post(rec4)
                                if res4 is not None:
                                    imported4 += 1
                                    existing_keys4.add(dedup4)
                                else:
                                    failed4 += 1

                            prog4.empty()
                            msg4 = f"✅ Imported **{imported4}** CollX orders"
                            if skipped4:
                                msg4 += f", skipped **{skipped4}** duplicates"
                            if failed4:
                                msg4 += f", **{failed4}** failed"
                            st.success(msg4 + ".")
                            if failed4 and _sal_last_error["msg"]:
                                st.error(f"Last error: {_sal_last_error['msg']}")
                            if imported4:
                                st.session_state.pop("sal_data", None)
                                st.rerun()
                    except Exception as e4:
                        st.error(f"Error reading CollX CSV: {e4}")

            # ── Whatnot import ────────────────────────────────────────────────
            with imp_whatnot:
                st.markdown("### Import Whatnot Seller Earnings")
                st.caption(
                    "Accepts the **WhatNot_Seller_Earnings_XXXX.xlsx** format "
                    "(columns: Month, Created Date, Completed Date, Card / Item, Order ID, Listing ID, Net Earnings, Status). "
                    "Net Earnings is already after Whatnot's fee. Re-importing is safe — duplicates skipped."
                )
                wn_file = st.file_uploader("Whatnot Seller Earnings XLSX", type=["xlsx"], key="sal_wn_file")
                if wn_file:
                    try:
                        import openpyxl as _opxl
                        wb5 = _opxl.load_workbook(wn_file, data_only=True)
                        # Find the right sheet — first sheet with 'Sales' in name or first sheet
                        ws5 = None
                        for sn in wb5.sheetnames:
                            if 'sale' in sn.lower() or 'earning' in sn.lower():
                                ws5 = wb5[sn]
                                break
                        if not ws5:
                            ws5 = wb5.active

                        all_rows5 = list(ws5.iter_rows(values_only=True))
                        # Find header row (has 'Order ID' or 'Net Earnings')
                        header_idx5 = None
                        for i, row in enumerate(all_rows5):
                            if any(str(c or '').strip().lower() in ('order id', 'net earnings') for c in row):
                                header_idx5 = i
                                break

                        if header_idx5 is None:
                            st.error("Could not find header row. Expected columns: Month, Created Date, Card / Item, Order ID, Net Earnings, Status.")
                        else:
                            headers5 = [str(c or '').strip().lower() for c in all_rows5[header_idx5]]
                            data_rows5 = all_rows5[header_idx5 + 1:]

                            def _col5(row, name):
                                try: return row[headers5.index(name)]
                                except ValueError: return None

                            # Filter to real sale rows: order id must be non-None and numeric-ish
                            sale_rows5 = [r for r in data_rows5 if _col5(r, 'order id') and str(_col5(r, 'order id') or '').strip().isdigit()]
                            st.caption(f"Found **{len(sale_rows5)}** sales. Preview:")
                            preview5 = [{"Month": _col5(r,'month'), "Date": _col5(r,'created date'), "Card": str(_col5(r,'card / item') or '')[:60], "Order ID": _col5(r,'order id'), "Net $": _col5(r,'net earnings'), "Status": _col5(r,'status')} for r in sale_rows5[:5]]
                            st.dataframe(pd.DataFrame(preview5), use_container_width=True, hide_index=True)

                            if st.button("⬆️ Import Whatnot Sales", type="primary", key="sal_wn_btn"):
                                existing_raw5 = _sal_get("?select=dedup_key&source=eq.whatnot")
                                existing_keys5 = {r["dedup_key"] for r in existing_raw5 if r.get("dedup_key")}
                                imported5 = skipped5 = failed5 = 0
                                prog5 = st.progress(0)

                                for idx5, row5 in enumerate(sale_rows5):
                                    prog5.progress(int((idx5 + 1) / max(len(sale_rows5), 1) * 100))
                                    order_id5 = str(_col5(row5, 'order id') or '').strip()
                                    dedup5 = f"whatnot|{order_id5}"
                                    if dedup5 in existing_keys5:
                                        skipped5 += 1
                                        continue

                                    created5 = _col5(row5, 'created date')
                                    try:
                                        sale_date5 = pd.to_datetime(str(created5)).strftime('%Y-%m-%d') if created5 else None
                                    except Exception:
                                        sale_date5 = str(created5)[:10] if created5 else None

                                    net5 = float(_col5(row5, 'net earnings') or 0)
                                    title5 = str(_col5(row5, 'card / item') or '').strip()[:200]
                                    status5 = str(_col5(row5, 'status') or 'completed').strip()

                                    rec5 = {
                                        "source": "whatnot",
                                        "order_id": order_id5 or None,
                                        "sale_date": sale_date5,
                                        "title": title5 or None,
                                        "item_number": str(_col5(row5, 'listing id') or '').strip() or None,
                                        "quantity": 1,
                                        "sale_price": net5,
                                        "shipping_collected": 0,
                                        "gross_revenue": net5,
                                        "platform_fee": 0,
                                        "net_proceeds": net5,
                                        "status": status5,
                                        "dedup_key": dedup5,
                                    }
                                    res5 = _sal_post(rec5)
                                    if res5 is not None:
                                        imported5 += 1
                                        existing_keys5.add(dedup5)
                                    else:
                                        failed5 += 1

                                prog5.empty()
                                msg5 = f"✅ Imported **{imported5}** Whatnot sales"
                                if skipped5: msg5 += f", skipped **{skipped5}** duplicates"
                                if failed5: msg5 += f", **{failed5}** failed"
                                st.success(msg5 + ".")
                                st.caption("Note: Whatnot Net Earnings is already after their fee — gross and fee are not broken out separately in their export.")
                                if imported5:
                                    st.session_state.pop("sal_data", None)
                                    st.rerun()
                    except Exception as e5:
                        st.error(f"Error reading Whatnot file: {e5}")

            # ── DC Sports import ───────────────────────────────────────────────
            with imp_dc:
                st.markdown("### Import DC Sports Sales")
                st.caption(
                    "Accepts **two DC Sports export formats** — auto-detected:\n\n"
                    "• **Seller/Consignment CSV** — Seller Dashboard export (Title, Status, SalePrice, Fees, Net, FriendlyPackageId). Only **Paid** rows imported.\n\n"
                    "• **Financial Ledger** — Payments → Download (TransactionType, Description, GrossAmount, Fees, NetAmount). Only **Sale** rows imported. Uses your actual fees."
                )
                dc_sal_file = st.file_uploader("DC Sports CSV export", type=["csv"], key="sal_dc_file")
                if dc_sal_file:
                    try:
                        raw6 = dc_sal_file.read().decode("utf-8-sig")
                        first6 = raw6.split("\n")[0]
                        is_ledger6 = "TransactionType" in first6 and "GrossAmount" in first6

                        dc_df6 = pd.read_csv(io.StringIO(raw6))
                        dc_df6.columns = [c.strip() for c in dc_df6.columns]

                        if is_ledger6:
                            paid_df6 = dc_df6[dc_df6["TransactionType"].astype(str).str.strip() == "Sale"].copy()
                            st.info(f"📊 **DC Sports Financial Ledger detected** — using your actual fees.")
                            st.caption(f"Found **{len(dc_df6):,}** total rows, **{len(paid_df6):,}** Sale rows (will import).")
                        else:
                            paid_df6 = dc_df6[dc_df6.get("Status", dc_df6.iloc[:,1]).astype(str).str.lower() == "paid"].copy() if "Status" in dc_df6.columns else dc_df6
                            st.caption(f"Found **{len(dc_df6):,}** total rows, **{len(paid_df6):,}** Paid (will import as sales).")

                        st.dataframe(paid_df6.head(5), use_container_width=True, hide_index=True)

                        if st.button("⬆️ Import DC Sports Sales", type="primary", key="sal_dc_btn"):
                            existing_raw6 = _sal_get("?select=dedup_key&source=eq.dc_sports")
                            existing_keys6 = {r["dedup_key"] for r in existing_raw6 if r.get("dedup_key")}
                            imported6 = skipped6 = failed6 = 0
                            prog6 = st.progress(0)
                            total6 = len(paid_df6)

                            for _i6, (idx6, row6) in enumerate(paid_df6.iterrows()):
                                prog6.progress(int((_i6 + 1) / max(total6, 1) * 100))

                                if is_ledger6:
                                    title6  = str(row6.get("Description", "") or "").strip()
                                    gross6  = _parse_money(row6.get("GrossAmount", 0))
                                    fees6   = abs(_parse_money(row6.get("Fees", 0)))
                                    net6    = _parse_money(row6.get("NetAmount", 0))
                                    dedup6  = f"dc_sports_ledger|{title6}|{round(gross6, 2)}"
                                    rec6 = {
                                        "source": "dc_sports",
                                        "sale_date": None,
                                        "title": title6 or None,
                                        "quantity": 1,
                                        "sale_price": gross6,
                                        "shipping_collected": 0,
                                        "gross_revenue": gross6,
                                        "platform_fee": fees6,
                                        "net_proceeds": net6,
                                        "status": "paid",
                                        "dedup_key": dedup6,
                                    }
                                else:
                                    pkg6    = str(row6.get("FriendlyPackageId", "") or "").strip()
                                    title6  = str(row6.get("Title", "") or "").strip()
                                    ending6 = str(row6.get("EndingDate", "") or "").strip()[:19]
                                    dedup6  = f"dc_sports|{pkg6}|{title6}|{ending6}"
                                    sale_price6 = _parse_money(row6.get("SalePrice"))
                                    fees6       = _parse_money(row6.get("Fees"))
                                    net6        = _parse_money(row6.get("Net"))
                                    try:
                                        sale_date6 = pd.to_datetime(ending6).strftime('%Y-%m-%d') if ending6 else None
                                    except Exception:
                                        sale_date6 = ending6[:10] if ending6 else None
                                    rec6 = {
                                        "source": "dc_sports",
                                        "order_id": pkg6 or None,
                                        "sale_date": sale_date6,
                                        "title": title6 or None,
                                        "item_number": None,
                                        "quantity": 1,
                                        "sale_price": sale_price6,
                                        "shipping_collected": 0,
                                        "gross_revenue": sale_price6,
                                        "platform_fee": fees6,
                                        "net_proceeds": net6,
                                        "status": "paid",
                                        "dedup_key": dedup6,
                                    }

                                if dedup6 in existing_keys6:
                                    skipped6 += 1
                                    continue

                                res6 = _sal_post(rec6)
                                if res6 is not None:
                                    imported6 += 1
                                    existing_keys6.add(dedup6)
                                else:
                                    failed6 += 1

                            prog6.empty()
                            msg6 = f"✅ Imported **{imported6}** DC Sports sales"
                            if skipped6: msg6 += f", skipped **{skipped6}** duplicates"
                            if failed6: msg6 += f", **{failed6}** failed"
                            st.success(msg6 + ".")
                            if failed6 and _sal_last_error["msg"]:
                                st.error(f"Last error: {_sal_last_error['msg']}")
                            if imported6:
                                st.session_state.pop("sal_data", None)
                                st.rerun()
                    except Exception as e6:
                        st.error(f"Error reading DC Sports CSV: {e6}")

            # ── Manual entry ───────────────────────────────────────────────────
            with imp_manual:
                st.markdown("### Manual Sale Entry")
                st.caption("Use for social media sales, cash sales, show sales, or any platform without an export.")

                with st.form("sal_manual_form"):
                    m1, m2, m3 = st.columns(3)
                    m_platform = m1.selectbox("Platform / Channel", ["Social Media", "Facebook", "Instagram", "Card Show", "Whatnot (manual)", "Other"], key="sal_m_platform")
                    m_date = m2.date_input("Sale Date", value=date.today(), key="sal_m_date")
                    m_status = m3.selectbox("Status", ["completed", "pending"], key="sal_m_status")
                    m_title = st.text_input("Card / Item Description *", placeholder="2025 Topps Chrome Aaron Judge #250 PSA 10", key="sal_m_title")
                    m4, m5, m6 = st.columns(3)
                    m_gross = m4.number_input("Gross / Sale Price ($)", min_value=0.0, step=0.01, format="%.2f", key="sal_m_gross")
                    m_fee = m5.number_input("Platform Fee ($)", min_value=0.0, step=0.01, format="%.2f", key="sal_m_fee")
                    m_net = m6.number_input("Net Proceeds ($)", min_value=0.0, step=0.01, format="%.2f", key="sal_m_net",
                                             help="Leave 0 to auto-calc as Gross − Fee")
                    m_submitted = st.form_submit_button("Add Sale", type="primary")

                if m_submitted:
                    if not m_title.strip():
                        st.error("Item description is required.")
                    else:
                        net_m = float(m_net) if m_net else round(float(m_gross) - float(m_fee), 2)
                        src_m = m_platform.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("/", "_")
                        dedup_m = f"manual|{src_m}|{m_date}|{m_title.strip()[:40]}|{round(float(m_gross),2)}"
                        rec_m = {
                            "source": "manual",
                            "order_id": None,
                            "sale_date": str(m_date),
                            "title": m_title.strip(),
                            "item_number": None,
                            "quantity": 1,
                            "sale_price": float(m_gross),
                            "shipping_collected": 0,
                            "gross_revenue": float(m_gross),
                            "platform_fee": float(m_fee),
                            "net_proceeds": net_m,
                            "status": m_status,
                            "dedup_key": dedup_m,
                        }
                        res_m = _sal_post(rec_m)
                        if res_m is not None:
                            st.success(f"✅ Added: {m_title.strip()} — net ${net_m:.2f}")
                            st.session_state.pop("sal_data", None)
                            st.rerun()
                        else:
                            st.error("Could not save. Check Supabase connection.")

        st.divider()
        st.markdown("**SQL — Create Sales Table** *(run once in Supabase SQL Editor)*")
        st.code(
            """create table if not exists sales_records (
  id bigint primary key generated always as identity,
  source text not null,
  order_id text,
  sale_date date,
  title text,
  item_number text,
  quantity integer default 1,
  sale_price numeric default 0,
  shipping_collected numeric default 0,
  gross_revenue numeric default 0,
  platform_fee numeric default 0,
  net_proceeds numeric default 0,
  status text,
  dedup_key text unique,
  created_at timestamptz default now()
);
create index if not exists idx_sr_date on sales_records(sale_date);
create index if not exists idx_sr_source on sales_records(source);""",
            language="sql",
        )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 12 — Image Prep (eBay corner crops)
# ══════════════════════════════════════════════════════════════════════════════
if _active_tab == 11:
    st.markdown("## 📸 Image Prep")
    st.markdown(
        "Upload a front and back photo — get 8 eBay-ready images automatically: "
        "front, back, 4 front corners, 2 back corners. Download as a ZIP and drag straight into eBay."
    )

    import io as _ip_io
    import zipfile as _ip_zip
    from PIL import Image as _ip_PIL

    _IP_CORNER_PCT  = 0.28   # fraction of card dimension to crop per corner
    _IP_JPEG_Q      = 92     # JPEG quality for output images

    def _ip_corner_crops(img, prefix, corners=4):
        """Return list of (filename, PIL.Image) for corner crops of img."""
        w, h = img.size
        cx = int(w * _IP_CORNER_PCT)
        cy = int(h * _IP_CORNER_PCT)
        regions = [
            ("top_left",     (0,     0,      cx,   cy)),
            ("top_right",    (w-cx,  0,      w,    cy)),
            ("bot_left",     (0,     h-cy,   cx,   h)),
            ("bot_right",    (w-cx,  h-cy,   w,    h)),
        ]
        return [(f"{prefix}_{name}.jpg", img.crop(box)) for name, box in regions[:corners]]

    def _ip_to_bytes(img):
        buf = _ip_io.BytesIO()
        img.convert("RGB").save(buf, format="JPEG", quality=_IP_JPEG_Q)
        return buf.getvalue()

    # ── Upload ────────────────────────────────────────────────────────────────
    _ipc1, _ipc2 = st.columns(2)
    with _ipc1:
        st.markdown("**Front of card**")
        _ip_front_file = st.file_uploader("Upload front", type=["jpg","jpeg","png","webp"], key="ip_front", label_visibility="collapsed")
    with _ipc2:
        st.markdown("**Back of card**")
        _ip_back_file  = st.file_uploader("Upload back",  type=["jpg","jpeg","png","webp"], key="ip_back",  label_visibility="collapsed")

    # ── Optional card label (used in filenames) ────────────────────────────────
    _ip_label = st.text_input("Card label (optional — used in filenames)", placeholder="e.g. 2023 Topps Chrome Acuna PSA10", key="ip_label")
    _ip_slug  = re.sub(r"[^\w]+", "_", _ip_label.strip()).strip("_") if _ip_label.strip() else "card"

    # ── Corner crop size slider ────────────────────────────────────────────────
    _ip_pct = st.slider("Corner crop size (% of card)", min_value=15, max_value=40, value=28, step=1, key="ip_pct",
                        help="28% matches typical grading-quality corner shots. Increase for more detail, decrease for wider view.")
    _IP_CORNER_PCT = _ip_pct / 100

    if _ip_front_file and _ip_back_file:
        _ip_front_img = _ip_PIL.open(_ip_io.BytesIO(_ip_front_file.getvalue()))
        _ip_back_img  = _ip_PIL.open(_ip_io.BytesIO(_ip_back_file.getvalue()))

        # Generate all images
        _ip_all = [
            (f"{_ip_slug}_1_front.jpg",    _ip_front_img),
            (f"{_ip_slug}_2_back.jpg",     _ip_back_img),
        ]
        _ip_all += _ip_corner_crops(_ip_front_img, f"{_ip_slug}_front", corners=4)
        _ip_all += _ip_corner_crops(_ip_back_img,  f"{_ip_slug}_back",  corners=2)

        # ── Preview grid ──────────────────────────────────────────────────────
        st.markdown(f"### Preview — {len(_ip_all)} images")
        _ip_labels = [
            "① Front", "② Back",
            "③ Front — top left", "④ Front — top right",
            "⑤ Front — bot left", "⑥ Front — bot right",
            "⑦ Back — top left",  "⑧ Back — top right",
        ]
        _ip_cols = st.columns(4)
        for _ip_i, (_ip_fname, _ip_img) in enumerate(_ip_all):
            with _ip_cols[_ip_i % 4]:
                st.image(_ip_img, caption=_ip_labels[_ip_i], use_container_width=True)

        # ── Build ZIP ─────────────────────────────────────────────────────────
        _ip_zip_buf = _ip_io.BytesIO()
        with _ip_zip.ZipFile(_ip_zip_buf, "w", _ip_zip.ZIP_DEFLATED) as _ip_zf:
            for _ip_fname, _ip_img in _ip_all:
                _ip_zf.writestr(_ip_fname, _ip_to_bytes(_ip_img))
        _ip_zip_buf.seek(0)

        st.download_button(
            label=f"📥 Download all {len(_ip_all)} images (.zip)",
            data=_ip_zip_buf.getvalue(),
            file_name=f"{_ip_slug}_ebay_images.zip",
            mime="application/zip",
            type="primary",
            key="ip_dl",
        )
        st.caption("Unzip → select all 8 files → drag into eBay listing image uploader. eBay requires a minimum of 2; 8 is the sweet spot.")

    elif _ip_front_file or _ip_back_file:
        st.info("Upload both front and back to generate corner crops.", icon="📷")
    else:
        st.markdown(
            """
            <div style="background:#1a1a2e;border-radius:12px;padding:2rem;text-align:center;color:#aaa;margin-top:1rem;">
            <div style="font-size:3rem;">📷</div>
            <div style="font-size:1.1rem;font-weight:600;color:#e2e8f0;margin:.5rem 0;">Upload front + back to get started</div>
            <div style="font-size:.85rem;">You'll get 8 eBay-ready images: front, back, 4 front corners, 2 back corners</div>
            </div>
            """,
            unsafe_allow_html=True,
        )

# ─── Footer ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    f"""
    <div style="text-align:center; color:#666; font-size:0.78rem; line-height:1.9; padding-bottom:1rem;">
        <strong style="color:#aaa;">Disclaimer</strong><br>
        {APP_NAME} is a research and decision-support tool only. All pricing data (GemRate, eBay)
        is pulled from third-party sources and may be incomplete, delayed, or inaccurate.
        Gem rates and market values fluctuate — always verify data independently before submitting cards for grading.
        All grading decisions and associated costs are solely your responsibility.
        {APP_NAME} assumes no liability for financial outcomes resulting from use of this tool.<br><br>
        ©️ 2026 {APP_NAME}. All rights reserved.
    </div>
    """,
    unsafe_allow_html=True,
)
